"""Build the Disney+ on TMT-MED-01 Sprint Plan workbook.

Mirrors the structure of Nike-R2R-NOAH-Sprint-Plan.xlsx (7 sheets) and
adds two new sheets the engagement scoping requires:
  · Persona Journey Maps  — current vs. target journey per persona, with KPI tie-ins
  · Key Solutions          — 12 productizable solutions distributed across sprints

Output: Disney 02_projects/FY27_Pipeline/subscriber-lifecycle-agentic/deliverables/TMT-MED-01-Sprint-Plan.xlsx
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

OUT_DIR = Path(r"C:\Stage\Clients\Industries\TMT\Global_Media_Entertainment\Walt_Disney\02_projects\FY27_Pipeline\subscriber-lifecycle-agentic\deliverables")
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / "TMT-MED-01-Sprint-Plan.xlsx"

# ---- styling ------------------------------------------------------------

HDR_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="0B0B0F")
HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
SUB_FILL  = PatternFill("solid", fgColor="6648B0")
THIN      = Side(style="thin", color="D4D4D7")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND      = PatternFill("solid", fgColor="F3F3F5")
GOLD_FILL = PatternFill("solid", fgColor="FBF1DC")
TEAL_FILL = PatternFill("solid", fgColor="E4F1EE")
CELL      = Font(name="Calibri", size=10)
CODE      = Font(name="Consolas", size=10, color="6648B0", bold=True)
ITAL      = Font(name="Calibri", size=10, italic=True, color="4A4A55")


def write_hdrs(ws, headers, widths=None, *, row=1, fill=HDR_FILL):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT; c.fill = fill; c.alignment = HDR_ALIGN; c.border = BORDER
        if widths: ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[row].height = 30


def write_row(ws, row_idx, values, *, code_cols=(), bold_cols=(), zebra=True, wrap=True):
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=j, value=v if v is not None else "")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
        c.border = BORDER
        if j in code_cols: c.font = CODE
        elif j in bold_cols: c.font = Font(name="Calibri", size=10, bold=True)
        else: c.font = CELL
        if zebra and row_idx % 2 == 0: c.fill = BAND


# ============================================================================
# SHEETS
# ============================================================================

def sheet_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 96

    sections = [
        ("APEX · TMT-MED-01 · Sprint Plan · Disney+ reference",  "APEX-canonical Media Subscriber Lifecycle service. Reference client: Disney+. This workbook scopes Approach + Wave 1 + Wave 2 in detail and stubs out Wave 3."),
        ("",                                         ""),
        ("How to read this workbook",                ""),
        ("README (this tab)",                        "Document overview · style legend · companion-artifact map."),
        ("Wave Overview",                            "Four engagement tiers — Approach · W1 Foundation · W2 Pilot · W3 Scale & Fuse — with envelopes and exit gates."),
        ("Sprint Plan",                              "Task / Subtask grid with UAC + Use Case binding + RACI + KPI linkage. The operational core."),
        ("User Acceptance",                          "UAC extracted into a focused list — what 'done' means per task; what acceptance evidence looks like."),
        ("Use Case Map",                             "UC-01 through UC-22 as authored in APEX-TMT-MED-01-Walkthrough.docx · sprint anchor · KPI driven."),
        ("Persona RACI",                             "Per-task RACI by named persona role (VP Retention, Audience Lead, etc.). Honors APEX persona discipline."),
        ("KPI Linkage",                              "Each headline KPI tied to driving tasks, exit-criteria targets, and the wave that lands the lift."),
        ("Persona Journey Maps",                     "Current-state vs. target-state journey per persona, with stage-level KPI tie-ins. The narrative the program improves."),
        ("Key Solutions",                            "Twelve productizable assets identified from the use-case set, distributed across the sprint plan, with persona owner and KPI driver."),
        ("",                                         ""),
        ("Style legend",                             ""),
        ("Black header cells",                       "Sheet-level column headers (sticky)."),
        ("Violet code-cell",                         "Identifiers — Sprint IDs, Task IDs, Use Case IDs, Solution IDs, KPI IDs."),
        ("Gold band cells",                          "Wave 2 / Wave 3 forecast tasks, future-wave dependencies."),
        ("Teal band cells",                          "Cross-wave load-bearing connections (W1 → W2, W2 → W3)."),
        ("Italics in tables",                        "Author commentary, scope-boundary notes, dependency callouts."),
        ("",                                         ""),
        ("Companion artifacts",                      ""),
        ("APEX-TMT-MED-01-Walkthrough.docx",  "Executive walkthrough — value-delivery chain · personas · KPIs · waves."),
        ("TMT-MED-01-Resource-Plan.xlsx",                 "Resource loading by phase, role, and sprint (planned next deliverable)."),
        ("TMT-MED-01-ROI-Case.html",                      "Retention-NPV uplift case · sensitivity analysis (planned)."),
        ("APEX-TMT-Reference-Architecture.html",     "TMT Practice reference — six sub-Practices, planes, archetypes."),
        ("APEX-CAF-Integration-Architecture.docx",   "How TMT-MED-01 consumes CAF via A2A — substrate-vs-scenario layering."),
        ("",                                         ""),
        ("Independence posture",                     "All TMT-MED-01 commercial materials use 'Deloitte's Microsoft practice,' 'DMTSP,' 'Microsoft platform capabilities,' 'Microsoft-native deployment.' The terms 'partner,' 'alliance,' and 'strategic alliance' do not appear. Before contracting, confirm Disney's external-audit firm relationship and route the SOW through Deloitte's Independence office."),
    ]
    for i, (left, right) in enumerate(sections, start=1):
        if i == 1:
            ws.cell(row=i, column=1, value=left).font = Font(name="Calibri", bold=True, size=14, color="0B0B0F")
            ws.cell(row=i, column=2, value=right).font = ITAL
        elif left and right:
            ws.cell(row=i, column=1, value=left).font = Font(name="Calibri", bold=True, size=10)
            ws.cell(row=i, column=2, value=right).font = CELL
        else:
            ws.cell(row=i, column=1, value=left).font = Font(name="Calibri", bold=True, size=11, color="6648B0")
            ws.cell(row=i, column=2, value=right).font = CELL
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30
    print("  README")


def sheet_wave_overview(wb):
    ws = wb.create_sheet("Wave Overview")
    ws.cell(row=1, column=1, value="Disney+ on TMT-MED-01 · Four engagement tiers").font = Font(name="Calibri", bold=True, size=14)
    write_hdrs(ws, ["Tier", "Duration", "Sprints", "Scope", "Primary outcome", "Gate"],
               widths=[14, 14, 18, 36, 36, 30], row=3)
    rows = [
        ["Approach",            "3–5 weeks",     "A-S01 – A-S02",
         "Discovery · scenario scoping · stakeholder interviews · SOW build-out",
         "Signed SOW with Independence office approval; risk register; commercial envelope",
         "VP Retention + Streaming COO sign-off"],
        ["Wave 1 — Foundation", "5–8 months",    "W1-S01 – W1-S05",
         "Disney+ NA single market · 3 agents (Engagement-Sense, Risk-Score, Outcome-Attribute) · synthetic data · shadow run",
         "Runnable W1 prototype · audit-row coverage end-to-end · 3 demo paths",
         "VP Retention sign-off on shadow-run accuracy"],
        ["Wave 2 — Pilot",      "8–12 months",   "W2-S01 – W2-S04",
         "Disney+ NA live · 6 agents · 2 cohort families + at-risk · real engagement-event stream · privacy packet",
         "90-day live operating run · first quarterly Privacy/Audit evidence packet",
         "Steering Committee sign-off to scale"],
        ["Wave 3 — Scale & Fuse","12–18 months", "W3-S01 – W3-S0n (stub)",
         "Global · multi-brand (Disney+/Hulu/ESPN+) · cross-streaming bundle fusion · TMT-MED-04/05/06 candidates",
         "Enterprise run-rate · $200M–$400M annualized retention NPV · cross-brand bundle attach",
         "Steady-state operating review"],
    ]
    for i, row in enumerate(rows, start=4):
        write_row(ws, i, row, code_cols=(3,), bold_cols=(1,))
        # Tint each tier row
        if i == 4: fill = PatternFill("solid", fgColor="EDEDF0")
        elif i == 5: fill = PatternFill("solid", fgColor="E4F1EE")
        elif i == 6: fill = PatternFill("solid", fgColor="F6F1E4")
        else: fill = PatternFill("solid", fgColor="F3E0DF")
        ws.cell(row=i, column=1).fill = fill
    print("  Wave Overview")


# ---- sprint plan data ---------------------------------------------------

# Use Case authoritative list (from walkthrough §5)
USE_CASES = [
    ("UC-01", "Streaming-platform event ingest",          "5.1 Engagement Signal Ingest", "W1", "Coverage"),
    ("UC-02", "TMTML.Engagement schema validation",       "5.1 Engagement Signal Ingest", "W1", "Coverage"),
    ("UC-03", "Subscriber lifecycle state computation",   "5.1 Engagement Signal Ingest", "W1", "Coverage"),
    ("UC-04", "At-risk cohort identification (baseline)", "5.2 At-Risk Cohort ID",        "W1", "Precision/Recall"),
    ("UC-05", "Risk-Score model training pipeline",       "5.2 At-Risk Cohort ID",        "W1", "Precision/Recall"),
    ("UC-06", "Confidence + time-to-action calibration",  "5.2 At-Risk Cohort ID",        "W1", "Coverage"),
    ("UC-07", "Cohort segmentation engine",               "5.2 At-Risk Cohort ID",        "W2", "Coverage"),
    ("UC-08", "Content-Match agent (rights-aware)",       "5.3 Content Surface",          "W2", "Loop time"),
    ("UC-09", "Rights-window query service",              "5.3 Content Surface",          "W2", "Loop time"),
    ("UC-10", "Personalized recommendation slate",        "5.3 Content Surface",          "W2", "NPV"),
    ("UC-11", "Programming-conflict detection",           "5.3 Content Surface",          "W2", "NPV"),
    ("UC-12", "Offer-Compose agent",                      "5.4 Retention Offer",          "W2", "NPV"),
    ("UC-13", "Per-cohort offer policy manifest",         "5.4 Retention Offer",          "W2", "Concession Eff."),
    ("UC-14", "Concession-budget service",                "5.4 Retention Offer",          "W2", "Concession Eff."),
    ("UC-15", "LTV computation pipeline",                 "5.4 Retention Offer",          "W2", "NPV"),
    ("UC-16", "Concession-tier mapping service",          "5.4 Retention Offer",          "W2", "Concession Eff."),
    ("UC-17", "Multi-channel dispatcher",                 "5.5 Multi-Channel Dispatch",   "W2", "Loop time"),
    ("UC-18", "Consent-state integration",                "5.5 Multi-Channel Dispatch",   "W2", "Audit-row coverage"),
    ("UC-19", "Adaptive Card HITL flow",                  "5.5 Multi-Channel Dispatch",   "W2", "Loop time"),
    ("UC-20", "Outcome-Attribute agent",                  "5.6 Outcome Attribution",      "W2", "Audit-row coverage"),
    ("UC-21", "NPV measurement engine",                   "5.6 Outcome Attribution",      "W2", "NPV"),
    ("UC-22", "Personalization evidence packet",          "5.6 Outcome Attribution",      "W2", "Audit-row coverage"),
]


# Sprint plan rows: (Wave, Sprint ID, Sprint Name, Task ID, Task, Subtasks, UAC, UCs, R, A, C, I, KPIs)
SPRINT_ROWS = [
    # ---- Approach ----
    ("Approach", "A-S01", "Discovery & Scoping", "A-S01-T1",
     "Disney+ subscriber landscape confirmation",
     "• Catalog Disney+ NA cohort families (engagement-driven, lifecycle-stage, demographic)\n"
     "• Inventory peripheral data: Hulu attach, ESPN+ attach, Disney Bundle membership\n"
     "• Confirm market scope (NA-first, EMEA later)\n"
     "• Map subscriber-data residency boundaries",
     "Cohort inventory signed off by VP Retention and Audience Lead; W1 cohort scope documented with rationale; NA market scope confirmed.",
     "—", "Delivery", "Sponsor", "VP Retention, Audience Lead", "CTO, Privacy Office", "—"),

    ("Approach", "A-S01", "Discovery & Scoping", "A-S01-T2",
     "Streaming platform & source-system state of play",
     "• Confirm streaming platform event-source coverage and event taxonomy\n"
     "• Inventory peripheral systems (CRM, marketing automation, consent platform, BI warehouse)\n"
     "• Map current personalization/recommendation stack\n"
     "• Identify W1 read-only access path for prototype shadow-run",
     "One-page source-system map approved by CTO and Audience Lead; current stack mapped; W1 read-only access plan documented.",
     "UC-01,UC-02", "Impl_Ld", "CTO", "Audience Lead, IT-Streaming", "Sponsor", "Coverage"),

    ("Approach", "A-S01", "Discovery & Scoping", "A-S01-T3",
     "Stakeholder interview series",
     "• VP Retention + retention-org leadership interviews\n"
     "• Audience Lead + segmentation team interviews\n"
     "• Content Programming Director interview\n"
     "• Pricing Strategy + Marketing Operations interviews\n"
     "• Privacy Office + Internal Audit interviews",
     "10+ persona interview log complete; pain points quantified by persona; 5 highest-impact use cases surfaced from the conversations and ranked.",
     "—", "GPL", "Sponsor", "all Disney personas", "DMTSP_lead", "Loop time, Coverage"),

    ("Approach", "A-S02", "SOW Build-Out & Independence", "A-S02-T1",
     "Risk register & mitigation plan",
     "• Populate risk log (Privacy regulator acceptance, retention-org HITL volume, "
     "synthetic-data fidelity, cohort-drift, model bias, channel-throughput)\n"
     "• Identify owner / likelihood / impact / mitigation per risk\n"
     "• Privacy-specific risk track (GDPR/CCPA/state regimes)",
     "Risk register with 10+ identified risks, each with owner, likelihood, impact, mitigation, and acceptance criteria; Privacy track signed off by Privacy Office.",
     "—", "Advisor", "Sponsor", "Privacy Office, IT-Sec", "all", "—"),

    ("Approach", "A-S02", "SOW Build-Out & Independence", "A-S02-T2",
     "Commercial envelope sizing",
     "• Size W1 Foundation envelope based on agent + cohort scope\n"
     "• Size W2 Pilot envelope for 90-day operating run + privacy packet\n"
     "• Stub W3 envelope for global × multi-brand × cross-streaming\n"
     "• Sensitivity analysis on retention-NPV uplift assumptions",
     "3-tier commercial envelope with low/mid/high bands; payback < 24 months at base case; finance-modelable assumptions documented; sensitivity ±20% on key drivers.",
     "—", "Advisor", "Sponsor", "DMTSP_lead, CFO_office", "VP Retention, Streaming COO", "NPV"),

    ("Approach", "A-S02", "SOW Build-Out & Independence", "A-S02-T3",
     "Independence review & SOW drafting",
     "• Confirm Disney external auditor relationship\n"
     "• Route engagement through Deloitte Independence office\n"
     "• Draft SOW with audit-safe structure (Disney-owned manifests/policies/LEDGER)\n"
     "• Privacy-Office review of personalization-evidence shape",
     "Written Independence-office approval received; SOW executed with Disney procurement; Audit Committee notification documented; Privacy Office signoff on evidence shape.",
     "—", "Advisor", "Sponsor", "Independence office, ExtAud, Privacy Office", "DMTSP_lead", "—"),

    # ---- Wave 1 — Foundation ----
    ("Wave 1", "W1-S01", "Foundation — Provisioning & Scaffolding", "W1-S01-T1",
     "Provision Azure AI workspace",
     "• Create Azure AI Foundry workspace\n"
     "• Configure VNet/private endpoints\n"
     "• Deploy model endpoints via AEF LLM Gateway\n"
     "• Configure Entra identity binding for VP Retention",
     "Foundry workspace live; model endpoints reachable via AEF LLM Gateway; Entra binding tested with VP Retention's identity.",
     "—", "Impl_Ld", "CTO", "IT-Sec, AEF team", "Sponsor", "—"),

    ("Wave 1", "W1-S01", "Foundation — Provisioning & Scaffolding", "W1-S01-T2",
     "Stand up local prototype scaffold",
     "• Create React + Node + Postgres scaffold\n"
     "• Wire LangFuse-equivalent local span store\n"
     "• Configure Docker Compose single-command boot\n"
     "• Stand up local Ollama for offline demos",
     "pnpm dev launches on demo laptop; /health returns 200; agent manifests validate against schema; offline Ollama demo path works.",
     "UC-01", "Impl_Ld", "Engineering_Ld", "Backend, Frontend", "Sponsor", "—"),

    ("Wave 1", "W1-S01", "Foundation — Provisioning & Scaffolding", "W1-S01-T3",
     "Manifest scaffolding & Git workflow",
     "• Author 3 W1 agent manifests (Engagement-Sense, Risk-Score, Outcome-Attribute)\n"
     "• Author parent orchestration manifest\n"
     "• Author concession-policy manifest skeleton\n"
     "• Stand up PR-merge discipline for manifest changes",
     "All 4 manifests in Git with PR-template; merge requires 2 reviewers; manifest version-bump rules documented (PATCH/MINOR/MAJOR).",
     "—", "Architect", "Engineering_Ld", "DMTSP_lead, Disney IT", "all", "—"),

    ("Wave 1", "W1-S02", "Foundation — Synthetic Data & Schema", "W1-S02-T1",
     "Synthetic Disney+ subscriber population",
     "• Generate 100K synthetic subscribers across 8 cohort families\n"
     "• Generate 4 quarters of realistic engagement events per subscriber\n"
     "• Inject realistic churn distributions, rights-window scenarios, payment events\n"
     "• Document synthetic-data generation rules for shadow-replay reproducibility",
     "100K-subscriber population validated against statistical realism checks; 4-quarter event stream loaded; reproducibility documented.",
     "UC-02, UC-03", "Data_Eng", "Architect", "Audience Lead", "Sponsor", "Coverage"),

    ("Wave 1", "W1-S02", "Foundation — Synthetic Data & Schema", "W1-S02-T2",
     "TMTML schema implementation",
     "• Implement TMTML.Subscriber, TMTML.Engagement schemas (Zod or equivalent)\n"
     "• Implement TMTML.Content, TMTML.Rights skeleton (W2 fully fills)\n"
     "• Wire schema validation into agent ingest paths\n"
     "• Document schema extension pattern for sub-Practice variants",
     "All W1 schemas validate; ingest agents reject malformed inputs cleanly; extension pattern documented for W2 schema growth.",
     "UC-02", "Architect", "Engineering_Ld", "Data_Eng, Disney IT", "Sponsor", "Coverage"),

    ("Wave 1", "W1-S02", "Foundation — Synthetic Data & Schema", "W1-S02-T3",
     "LEDGER plane scaffolding",
     "• Implement 14-field audit row schema\n"
     "• Implement LEDGER write path from agents\n"
     "• Implement LEDGER read/query API for BI drill-through\n"
     "• Implement evidence-packet export skeleton (real packet ships W2)",
     "Audit rows write end-to-end on shadow runs; query API returns matching rows by trace_id under 100ms; packet skeleton emits placeholder ZIP.",
     "UC-20, UC-22", "Architect", "Engineering_Ld", "DMTSP_lead, Privacy Office", "Sponsor", "Audit-row coverage"),

    ("Wave 1", "W1-S03", "Foundation — Engagement-Sense + Risk-Score", "W1-S03-T1",
     "Engagement-Sense agent",
     "• Implement event ingest from streaming platform feed (synthetic in W1)\n"
     "• Normalize to TMTML.Engagement\n"
     "• Compute subscriber-state transitions per ingestion window\n"
     "• Emit per-decision audit row to LEDGER",
     "Engagement-Sense runs against 100K subscriber population; state transitions computed for all 4 historical quarters; audit rows emit on every state change.",
     "UC-01, UC-03", "Backend", "Architect", "Audience Lead", "Sponsor", "Coverage"),

    ("Wave 1", "W1-S03", "Foundation — Engagement-Sense + Risk-Score", "W1-S03-T2",
     "Risk-Score agent (baseline)",
     "• Implement churn-probability scoring (CAF cap:churn-prediction consumed via A2A)\n"
     "• Implement confidence + time-to-action calibration\n"
     "• Implement HITL-escalation rule (confidence < 0.85)\n"
     "• Wire audit-row emission per scoring decision",
     "Risk-Score runs over the population; precision ≥ 0.78 / recall ≥ 0.72 on shadow data; confidence calibration plot reviewed by Audience Lead; escalations route correctly.",
     "UC-04, UC-05, UC-06", "ML_Eng", "Architect", "Audience Lead, ML platform team", "Sponsor", "Precision/Recall"),

    ("Wave 1", "W1-S03", "Foundation — Engagement-Sense + Risk-Score", "W1-S03-T3",
     "Cohort segmentation skeleton (W2-deepens)",
     "• Implement basic cohort segmentation taxonomy\n"
     "• Wire cohort assignment to Risk-Score outputs\n"
     "• Stub UI for cohort review (Audience Lead)\n"
     "• Document W2 deepening path (multi-cohort, drift detection)",
     "Cohort assignment runs on the 100K population; cohort review UI shows distribution; Audience Lead validates cohort logic against expectations; W2 deepening path documented.",
     "UC-07", "ML_Eng", "Architect", "Audience Lead", "Sponsor", "Coverage"),

    ("Wave 1", "W1-S04", "Foundation — Outcome-Attribute + Audit Trail", "W1-S04-T1",
     "Outcome-Attribute agent",
     "• Implement outcome-window observation logic\n"
     "• Implement attribution-to-trace_id at decision granularity\n"
     "• Compose closeout 14-field audit row\n"
     "• Wire to LEDGER closeout path",
     "Outcome attribution runs end-to-end on synthetic interventions; trace_id binding verified; closeout audit row schema-validates; LEDGER closeout writes succeed under load.",
     "UC-20, UC-21", "Backend", "Architect", "Audience Lead, Privacy Office", "Sponsor", "Audit-row coverage"),

    ("Wave 1", "W1-S04", "Foundation — Outcome-Attribute + Audit Trail", "W1-S04-T2",
     "Adaptive Card mock flow",
     "• Stand up VP Retention's Teams environment (mock for prototype)\n"
     "• Author Adaptive Card template for above-threshold concession\n"
     "• Wire Card → approval → audit-row write\n"
     "• Test on VP Retention's actual device",
     "Adaptive Card fires to VP Retention's device on synthetic concession; one-click approve writes audit row with her named identity; reject + modify flows tested.",
     "UC-19", "Frontend", "Architect", "VP Retention, Disney IT-Teams", "Sponsor", "Loop time"),

    ("Wave 1", "W1-S04", "Foundation — Outcome-Attribute + Audit Trail", "W1-S04-T3",
     "Audit-row drill-through console",
     "• Implement BI dashboard tile → trace_id drill-through\n"
     "• Implement audit-row JSON viewer with version-stamp display\n"
     "• Implement consent-state attestation viewer\n"
     "• Document operator workflow for audit-trail review",
     "Drill-through works from any BI tile to specific audit row; reasoning trace, manifest version, policy version, and consent state all visible; Privacy Office signs off on shape.",
     "UC-22", "Frontend", "Architect", "Privacy Office, Internal Audit", "Sponsor", "Audit-row coverage"),

    ("Wave 1", "W1-S05", "Foundation — Demo & Exit", "W1-S05-T1",
     "Three demo paths authored",
     "• Demo Path A · The intervention loop (synthetic at-risk subscriber → full chain → Adaptive Card → audit row)\n"
     "• Demo Path B · Retention-NPV trending (BI dashboard → cell drill → audit rows → reasoning trace)\n"
     "• Demo Path C · Audit trail (subscriber → full chain of custody → consent attestation)\n"
     "• Each demo narrated, timed under 5 minutes",
     "All 3 demo paths run in under 5 minutes each; narrated for Disney stakeholder audience; offline demo (Ollama) tested end-to-end.",
     "UC-19, UC-20, UC-22", "Delivery", "GPL", "VP Retention, all Disney personas", "DMTSP_lead", "—"),

    ("Wave 1", "W1-S05", "Foundation — Demo & Exit", "W1-S05-T2",
     "Shadow-run replay against historical data",
     "• Execute end-to-end orchestration against 4 quarters of synthetic events\n"
     "• Validate audit-row coverage ≥ 99% on synthetic population\n"
     "• Generate W1 KPI dashboard (precision/recall, coverage, loop-time on shadow data)\n"
     "• Document gaps and W2 close-out plan",
     "Shadow-run completes; audit-row coverage ≥ 99% measured; precision/recall meet W2-pilot targets on shadow data; gap list reviewed and accepted as W2 scope.",
     "UC-04, UC-05, UC-20", "ML_Eng", "Architect", "Audience Lead, VP Retention", "Sponsor", "Precision/Recall, Audit-row coverage"),

    ("Wave 1", "W1-S05", "Foundation — Demo & Exit", "W1-S05-T3",
     "W1 exit review",
     "• Walk shadow-run results with VP Retention\n"
     "• Walk audit-row coverage with Privacy Office and Internal Audit\n"
     "• Confirm W2 commitment from Steering Committee\n"
     "• Sign-off ceremony documented",
     "VP Retention signs off on shadow-run accuracy; Privacy Office signs off on audit-row shape; Steering Committee approves W2 commercial envelope; sign-off documented in engagement log.",
     "—", "GPL", "Sponsor", "VP Retention, Privacy Office, Internal Audit, Steering Committee", "all", "—"),

    # ---- Wave 2 — Pilot ----
    ("Wave 2", "W2-S01", "Pilot — Live Engagement Stream", "W2-S01-T1",
     "Real engagement-event connector",
     "• Replace synthetic event source with Disney+ NA real event stream\n"
     "• Implement back-pressure handling at event-volume scale\n"
     "• Verify TMTML.Engagement schema fidelity against live data\n"
     "• Tune Engagement-Sense throughput",
     "Real event stream ingested for 7 consecutive days; schema fidelity ≥ 99.5%; back-pressure handled without drops; throughput meets daily peak.",
     "UC-01, UC-02", "Data_Eng", "Architect", "Audience Lead, Disney Streaming Eng", "Sponsor", "Coverage"),

    ("Wave 2", "W2-S01", "Pilot — Live Engagement Stream", "W2-S01-T2",
     "Cohort segmentation engine deepening",
     "• Implement multi-cohort family support (8 families for Disney+ NA)\n"
     "• Implement cohort-drift detection (W2 entry into Cohort Drift Watchtower solution)\n"
     "• Wire automated retraining triggers on drift\n"
     "• Audience Lead workbench for cohort definition",
     "All 8 cohort families operating live; drift detection triggers reviewed weekly by Audience Lead; auto-retrain pipeline tested; workbench accepted by Audience Lead.",
     "UC-07", "ML_Eng", "Architect", "Audience Lead, ML platform team", "Sponsor", "Precision/Recall"),

    ("Wave 2", "W2-S01", "Pilot — Live Engagement Stream", "W2-S01-T3",
     "Privacy/consent integration shim",
     "• Wire Disney consent platform → TMT-MED-01 consent-state agent\n"
     "• Implement consent-aware decision routing (no consent = no personalization)\n"
     "• Wire consent-state attestation into audit row\n"
     "• Privacy Office walkthrough of consent decision shape",
     "Consent-state read end-to-end; personalization decisions correctly gated on consent; attestation visible in audit row; Privacy Office signs off on integration.",
     "UC-18", "Backend", "Architect", "Privacy Office, Disney IT", "Sponsor", "Audit-row coverage"),

    ("Wave 2", "W2-S02", "Pilot — Content-Match + Offer-Compose", "W2-S02-T1",
     "Content-Match agent (rights-aware)",
     "• Implement Content-Match consuming CAF cap:content-recommendation via A2A\n"
     "• Wire rights-window query service (TMT-MED-03 service capability)\n"
     "• Implement programming-conflict detection\n"
     "• Wire personalized recommendation slate composition",
     "Content-Match runs against live cohort; rights-window logic validates 100% (no out-of-rights recommendation surfaces); programming-conflict detection catches Content Programming Director's test cases; slate composition tunable per cohort.",
     "UC-08, UC-09, UC-10, UC-11", "ML_Eng", "Architect", "Content Programming, Rights Mgmt Lead", "Sponsor", "Loop time, NPV"),

    ("Wave 2", "W2-S02", "Pilot — Content-Match + Offer-Compose", "W2-S02-T2",
     "Offer-Compose agent",
     "• Implement Offer-Compose with concession-tier mapping\n"
     "• Wire LTV computation pipeline\n"
     "• Implement concession-budget service with daily/weekly budget tracking\n"
     "• Wire HITL-routing rules per commercial envelope tier",
     "Offer-Compose proposes offers within concession budgets; tier mapping correct against policy manifest; HITL routes correctly for above-threshold; LTV computation validated by Pricing Strategy.",
     "UC-12, UC-14, UC-15, UC-16", "ML_Eng", "Architect", "Pricing Strategy, VP Retention", "Sponsor", "NPV, Concession Eff."),

    ("Wave 2", "W2-S02", "Pilot — Content-Match + Offer-Compose", "W2-S02-T3",
     "Concession Policy Studio",
     "• Build declarative policy authoring UI\n"
     "• Implement policy-simulation against historical cohort outcomes\n"
     "• Wire PR-merge workflow for policy changes\n"
     "• Pricing Strategy training on the studio",
     "Pricing Strategy authors a policy change end-to-end; simulation runs against historical data; PR-merge workflow tested; training-of-trainers complete.",
     "UC-13", "Frontend", "Architect", "Pricing Strategy, DMTSP_lead", "all", "Concession Eff."),

    ("Wave 2", "W2-S03", "Pilot — Dispatch + HITL Production Flow", "W2-S03-T1",
     "Multi-channel dispatcher (production)",
     "• Wire push channel API (production)\n"
     "• Wire email channel via marketing-automation API\n"
     "• Wire in-app channel via Disney+ in-app messaging\n"
     "• Implement receipt-aware delivery monitoring + auto-retry",
     "All 3 channels dispatch live; receipt logged for ≥ 98% of dispatches; auto-retry handles transient failures; Marketing Operations Director signs off on dispatch volumes.",
     "UC-17", "Backend", "Architect", "Marketing Ops Director, Disney IT", "Sponsor", "Loop time"),

    ("Wave 2", "W2-S03", "Pilot — Dispatch + HITL Production Flow", "W2-S03-T2",
     "VP Retention Decision Console (production)",
     "• Production Adaptive Card with full concession context\n"
     "• Queue management (queue depth, time-in-queue alerts)\n"
     "• Approval-history dashboard for VP Retention\n"
     "• Designate-approver workflow for VP Retention's deputies",
     "VP Retention uses console live for 30 days; queue depth stays under target; designate-approver flow tested; approval history queryable.",
     "UC-19", "Frontend", "Architect", "VP Retention, retention-org leadership", "Sponsor", "Loop time"),

    ("Wave 2", "W2-S03", "Pilot — Dispatch + HITL Production Flow", "W2-S03-T3",
     "Cohort Drift Watchtower",
     "• Implement drift detection on Risk-Score outputs\n"
     "• Wire alerting to Audience Lead\n"
     "• Implement drift root-cause UI (which features drifted)\n"
     "• Auto-retraining trigger with HITL approval gate",
     "Watchtower running live; first drift event detected and root-caused; Audience Lead signs off on alert quality; auto-retrain trigger tested in shadow mode (no live retrains W2).",
     "UC-05, UC-07", "ML_Eng", "Architect", "Audience Lead, ML platform team", "Sponsor", "Precision/Recall"),

    ("Wave 2", "W2-S04", "Pilot — Privacy Packet & Operating Run", "W2-S04-T1",
     "Privacy/Consent evidence packet",
     "• Implement quarterly evidence-packet generator (real, not skeleton)\n"
     "• Compose manifests + policy + audit-row extract + tie-out\n"
     "• Privacy Office walkthrough of packet shape\n"
     "• External-audit walkthrough of packet shape (audit-firm-dependent)",
     "First quarterly packet delivered to Privacy Office; Privacy Office accepts shape; tie-out reconciles posted decisions to audit-row count; external-audit walkthrough scheduled.",
     "UC-22", "Architect", "DMTSP_lead", "Privacy Office, External Audit", "Sponsor", "Audit-row coverage"),

    ("Wave 2", "W2-S04", "Pilot — Privacy Packet & Operating Run", "W2-S04-T2",
     "90-day operating run",
     "• Run full 6-agent orchestration live for 90 consecutive days\n"
     "• Daily monitoring meeting first 14 days, then weekly\n"
     "• Weekly KPI review with VP Retention\n"
     "• Continuous improvement backlog populated from operations",
     "90 consecutive operating days completed; KPI targets met or gap analysis documented; continuous-improvement backlog with 15+ items prioritized for W3.",
     "all UCs", "Delivery", "VP Retention", "all Disney personas", "DMTSP_lead", "all KPIs"),

    ("Wave 2", "W2-S04", "Pilot — Privacy Packet & Operating Run", "W2-S04-T3",
     "W2 exit review & W3 commit gate",
     "• Steering Committee operating review\n"
     "• Privacy/Audit signoff on Q1 packet\n"
     "• W3 commercial envelope confirmed\n"
     "• Cross-streaming-brand expansion plan reviewed (Hulu attach pilot)",
     "Steering Committee approves W3; Privacy/Audit signoff received; W3 envelope locked; cross-brand pilot scope agreed for W3-S01.",
     "—", "GPL", "Sponsor", "Steering Committee, Privacy Office, External Audit", "all", "all KPIs"),
]


def sheet_sprint_plan(wb):
    ws = wb.create_sheet("Sprint Plan")
    headers = ["Wave", "Sprint ID", "Sprint Name", "Task ID", "Task", "Subtasks",
               "User Acceptance Criteria", "Use Cases", "Responsible", "Accountable",
               "Consulted", "Informed", "KPIs"]
    widths = [9, 10, 28, 12, 30, 60, 60, 14, 11, 12, 22, 18, 18]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "E2"

    for i, row in enumerate(SPRINT_ROWS, start=2):
        write_row(ws, i, row, code_cols=(2, 4, 8))
        # Tint rows by wave
        wave = row[0]
        if wave == "Approach":
            tint = PatternFill("solid", fgColor="EDEDF0")
        elif wave == "Wave 1":
            tint = PatternFill("solid", fgColor="EFF7F5")
        elif wave == "Wave 2":
            tint = PatternFill("solid", fgColor="FBF1DC")
        else:
            tint = None
        if tint:
            ws.cell(row=i, column=1).fill = tint
        ws.row_dimensions[i].height = max(60, 15 * (row[5].count("\n") + 1))
    print(f"  Sprint Plan · {len(SPRINT_ROWS)} tasks")


def sheet_user_acceptance(wb):
    ws = wb.create_sheet("User Acceptance")
    headers = ["Task ID", "Sprint", "Task", "User Acceptance Criteria",
               "Acceptance evidence", "Reviewer"]
    widths = [12, 28, 30, 60, 50, 26]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "C2"

    for i, row in enumerate(SPRINT_ROWS, start=2):
        wave, sid, sname, tid, task, subs, uac, ucs, r, a, c, info, kpis = row
        # Compose acceptance evidence and reviewer based on the task content
        evidence = "Demo screenshot · run log · sign-off email · audit-row sample"
        if "audit" in task.lower() or "audit" in uac.lower():
            evidence = "Audit-row sample · LEDGER query result · Privacy Office signoff"
        if "card" in task.lower() or "HITL" in task:
            evidence = "Adaptive Card screenshot · approval audit row · device test log"
        if "exit" in task.lower():
            evidence = "Sign-off doc · Steering Committee minutes · KPI dashboard"
        reviewer = c.split(",")[0].strip() if c and c != "—" else a
        write_row(ws, i, [tid, sname, task, uac, evidence, reviewer], code_cols=(1,))
        ws.row_dimensions[i].height = 50
    print(f"  User Acceptance · {len(SPRINT_ROWS)} rows")


def sheet_use_case_map(wb):
    ws = wb.create_sheet("Use Case Map")
    headers = ["Use Case ID", "Title", "Agent-level use case (walkthrough §5)",
               "Sprint anchor", "Wave", "KPI driven"]
    widths = [13, 44, 36, 18, 9, 22]
    write_hdrs(ws, headers, widths)

    # Map each UC to its anchor sprint(s) by walking SPRINT_ROWS for UC in column 7
    uc_to_sprints = {}
    for row in SPRINT_ROWS:
        ucs = row[7]
        for uc_token in [t.strip() for t in ucs.split(",")]:
            if uc_token.startswith("UC-"):
                uc_to_sprints.setdefault(uc_token, []).append(row[1])

    for i, (uc_id, title, agent_uc, wave, kpi) in enumerate(USE_CASES, start=2):
        anchor = " · ".join(sorted(set(uc_to_sprints.get(uc_id, []))))
        write_row(ws, i, [uc_id, title, agent_uc, anchor or "(planned)", wave, kpi],
                  code_cols=(1, 4))
        if wave == "W1":
            ws.cell(row=i, column=5).fill = PatternFill("solid", fgColor="EFF7F5")
        elif wave == "W2":
            ws.cell(row=i, column=5).fill = PatternFill("solid", fgColor="FBF1DC")
    print(f"  Use Case Map · {len(USE_CASES)} use cases")


def sheet_persona_raci(wb):
    ws = wb.create_sheet("Persona RACI")
    personas = ["VP Retention", "Audience Lead", "Content Programming",
                "Pricing Strategy", "Marketing Ops", "Privacy Office",
                "External Audit", "Streaming COO (Sponsor)", "DMTSP_lead",
                "Engagement Architect", "Implementation Lead", "ML Engineering",
                "Backend Eng", "Frontend Eng", "Data Engineering", "Disney IT/CTO"]
    headers = ["Task ID", "Task"] + personas
    widths = [12, 32] + [12] * len(personas)
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "C2"

    # RACI legend per task — derived from R/A/C/I columns of SPRINT_ROWS
    role_map = {
        "Sponsor": "Streaming COO (Sponsor)",
        "VP Retention": "VP Retention",
        "VP Retention, retention-org leadership": "VP Retention",
        "GPL": "DMTSP_lead",
        "Delivery": "Engagement Architect",
        "Architect": "Engagement Architect",
        "Impl_Ld": "Implementation Lead",
        "ML_Eng": "ML Engineering",
        "Backend": "Backend Eng",
        "Frontend": "Frontend Eng",
        "Data_Eng": "Data Engineering",
        "Engineering_Ld": "Implementation Lead",
        "DMTSP_lead": "DMTSP_lead",
        "Advisor": "DMTSP_lead",
        "Audience Lead": "Audience Lead",
        "Pricing Strategy": "Pricing Strategy",
        "Pricing Strategy, VP Retention": "Pricing Strategy",
        "Privacy Office": "Privacy Office",
        "ExtAud": "External Audit",
        "External Audit": "External Audit",
        "CTO": "Disney IT/CTO",
        "IT": "Disney IT/CTO",
        "Marketing Ops Director, Disney IT": "Marketing Ops",
        "Content Programming, Rights Mgmt Lead": "Content Programming",
    }

    def map_role(token):
        token = token.strip()
        return role_map.get(token, token)

    for i, row in enumerate(SPRINT_ROWS, start=2):
        wave, sid, sname, tid, task, subs, uac, ucs, R, A, C, I, kpis = row
        line = [tid, task] + [""] * len(personas)
        # parse comma-separated roles in C, I, plus single R / A
        def assign(letter, raw):
            for token in [t.strip() for t in (raw or "").split(",")]:
                if not token or token == "—":
                    continue
                target = map_role(token)
                # Look up persona index
                if target in personas:
                    j = 2 + personas.index(target) + 1
                    existing = line[j-1]
                    line[j-1] = (existing + letter) if existing and letter not in existing else (existing or letter)
                elif target == "all":
                    for p in personas:
                        j = 2 + personas.index(p) + 1
                        line[j-1] = (line[j-1] or letter) if letter == "I" else line[j-1]
                elif target == "all Disney personas":
                    for p in ["VP Retention","Audience Lead","Content Programming","Pricing Strategy","Marketing Ops","Privacy Office"]:
                        j = 2 + personas.index(p) + 1
                        line[j-1] = (line[j-1] or letter) if letter == "I" else line[j-1]
        assign("R", R); assign("A", A); assign("C", C); assign("I", I)

        write_row(ws, i, line, code_cols=(1,))
        # color RACI cells
        for j in range(3, len(line) + 1):
            v = line[j-1]
            cell = ws.cell(row=i, column=j)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=10, bold=True)
            if v == "R":   cell.fill = PatternFill("solid", fgColor="C9E5DC")
            elif v == "A": cell.fill = PatternFill("solid", fgColor="F6D7A1")
            elif v == "C": cell.fill = PatternFill("solid", fgColor="DCE3F4")
            elif v == "I": cell.fill = PatternFill("solid", fgColor="EDEDED")
    print(f"  Persona RACI · {len(SPRINT_ROWS)} task rows × {len(personas)} personas")


def sheet_kpi_linkage(wb):
    ws = wb.create_sheet("KPI Linkage")
    headers = ["KPI ID", "KPI", "Measurement",
               "W2 Pilot target", "W3 Scale target", "Driving sprints / tasks"]
    widths = [10, 32, 32, 18, 18, 36]
    write_hdrs(ws, headers, widths)

    kpis = [
        ("KPI-01", "Intervention loop time",
         "Hours from signal to dispatch", "≤ 24 hrs", "≤ 4 hrs",
         "W1-S03 Risk-Score · W2-S02 Content-Match · W2-S03 Dispatcher · W2-S03 Decision Console"),
        ("KPI-02", "Personalized intervention rate (Coverage)",
         "% of available daily decisions personalized", "35%", "75%",
         "W1-S02 Schemas · W1-S03 Engagement-Sense · W2-S01 Live Stream · W2-S02 Content-Match"),
        ("KPI-03", "Risk-Score precision",
         "P at risk-threshold", "≥ 0.78", "≥ 0.85",
         "W1-S03 Risk-Score baseline · W2-S01 Cohort deepening · W2-S03 Drift Watchtower"),
        ("KPI-04", "Risk-Score recall",
         "R at risk-threshold", "≥ 0.72", "≥ 0.82",
         "W1-S03 Risk-Score baseline · W2-S01 Cohort deepening · W2-S03 Drift Watchtower"),
        ("KPI-05", "Retention NPV per intervention",
         "Audit-row attributed NPV", "+$84", "+$143",
         "W2-S02 Offer-Compose · W2-S02 Content-Match · W2-S02 Concession Studio · W2-S04 90-day run"),
        ("KPI-06", "Concession utilization efficiency",
         "% of concession budget yielding positive NPV", "≥ 70%", "≥ 85%",
         "W2-S02 Offer-Compose · W2-S02 Concession Studio · W2-S04 90-day run"),
        ("KPI-07", "Audit-row coverage",
         "% of personalization decisions attributed", "≥ 98%", "≥ 99.5%",
         "W1-S02 LEDGER · W1-S04 Outcome-Attribute · W1-S04 Audit Console · W2-S04 Privacy Packet"),
        ("KPI-08", "Cross-brand bundle attach lift",
         "Per-cohort attach % above baseline", "+1.8 pp", "+4.2 pp",
         "W3 candidate (TMT-MED-05 Cross-Streaming Brand Bundle)"),
    ]
    for i, row in enumerate(kpis, start=2):
        write_row(ws, i, row, code_cols=(1,))
        ws.row_dimensions[i].height = 48
    print(f"  KPI Linkage · {len(kpis)} KPIs")


def sheet_persona_journey_maps(wb):
    """NEW SHEET: Current vs. target journey for each persona, with KPI tie-ins."""
    ws = wb.create_sheet("Persona Journey Maps")
    headers = ["Persona", "Stage", "Today (current state)",
               "With TMT-MED-01 (target state)", "KPI improved", "Sprint that lands the lift"]
    widths = [22, 26, 50, 50, 22, 22]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "B2"

    journeys = [
        # ---- VP Subscriber Retention ----
        ("VP Subscriber Retention",
         "Wake-up signal review",
         "8:14 AM: pulls 3 BI dashboards, reads overnight Slack threads, scans Excel exports from Audience Lead. ~30 minutes to assemble morning picture.",
         "8:14 AM: glance at Adaptive Card preview on phone — top 3 retention movements ranked by NPV impact, drill-through to specific cohort. ~3 minutes.",
         "Loop time", "W1-S04 Adaptive Card · W2-S03 Decision Console"),
        ("VP Subscriber Retention",
         "Cohort triage",
         "Reviews Audience Lead's prioritization spreadsheet; cross-references against marketing calendar; emails Pricing Strategy on concession headroom.",
         "Reviews queue depth in Decision Console; auto-prioritized by retention-NPV-at-stake; inline view of concession budget remaining.",
         "Loop time, NPV", "W2-S03 Decision Console"),
        ("VP Subscriber Retention",
         "Concession review",
         "Email approval chain for concessions above threshold; 1-3 day turnaround; details reconstructed from cohort exports.",
         "Adaptive Card with full context (subscriber tier, LTV, confidence, prior interventions); one-click approve/modify/reject.",
         "Loop time, Concession Eff.", "W1-S04 Adaptive Card · W2-S03 Decision Console"),
        ("VP Subscriber Retention",
         "Outcome review",
         "Quarterly retrospective in PowerPoint; per-intervention NPV reconstructed manually; attribution debate with Audience Lead.",
         "Daily NPV scorecard with drill-through to specific audit rows; attribution settled at trace_id granularity.",
         "NPV, Audit-row coverage", "W1-S04 Audit Console · W2-S04 Privacy Packet"),

        # ---- Audience Lead ----
        ("Audience Lead",
         "Cohort definition",
         "SQL-author cohort segmentation in BI tool; review with team weekly; commit definitions in shared workbook.",
         "Cohort console with declarative cohort authoring; simulation against historical population; PR-merge for cohort changes.",
         "Coverage", "W2-S01 Cohort Deepening"),
        ("Audience Lead",
         "Risk-Score calibration",
         "Receives quarterly model-performance report; debates retraining priority with ML platform team; no in-loop visibility.",
         "Drift Watchtower flags model-performance issues weekly; in-loop retraining triggers with HITL gate.",
         "Precision/Recall", "W1-S03 Risk-Score · W2-S03 Drift Watchtower"),
        ("Audience Lead",
         "Drift detection",
         "No systematic drift detection today; surprises emerge in retention numbers months later.",
         "Watchtower alerts on drift root-causes (which features drifted, in which cohort); auto-retrain trigger with HITL.",
         "Precision/Recall, NPV", "W2-S03 Drift Watchtower"),

        # ---- Content Programming Director ----
        ("Content Programming Director",
         "Rights window monitoring",
         "Manual rights-window check before campaign builds; fragmented data across rights-management system and content catalog.",
         "Rights-aware Content-Match auto-validates every recommendation against current rights window; programming-conflict detection runs continuously.",
         "Loop time, NPV", "W2-S02 Content-Match · W2-S02 Rights-window service"),
        ("Content Programming Director",
         "Recommendation review",
         "Spot-check campaign recommendations weekly; no per-subscriber visibility; programming feedback loop is quarterly retrospective.",
         "Drill-through from any recommendation slate to reasoning trace; programming-strategy feedback in-loop with W3 fusion.",
         "Coverage, NPV", "W2-S02 Content-Match · W3 fusion (programming feedback)"),

        # ---- Pricing Strategy Lead ----
        ("Pricing Strategy Lead",
         "Concession policy authoring",
         "Annual policy doc + Excel models; quarterly retrospective on concession cost; tier thresholds frozen for 12 months.",
         "Concession Policy Studio: declarative policy authoring + simulation against historical cohorts + PR-merge workflow.",
         "Concession Eff., NPV", "W2-S02 Concession Studio"),
        ("Pricing Strategy Lead",
         "Cost monitoring",
         "Monthly concession-cost reports from Finance; lagging indicator; no per-cohort cost visibility.",
         "Live concession-cost dashboard per cohort; daily budget burn-rate; alert on anomalies.",
         "Concession Eff.", "W2-S02 Offer-Compose · W2-S02 Concession Studio"),

        # ---- Marketing Operations Director ----
        ("Marketing Operations Director",
         "Channel readiness",
         "Manual campaign builds in marketing automation tool; bespoke targeting per campaign; 1-3 day setup time.",
         "Auto-dispatched interventions through pre-configured channels; targeting from cohort engine; setup in seconds.",
         "Loop time", "W2-S03 Multi-channel dispatcher"),
        ("Marketing Operations Director",
         "Dispatch monitoring",
         "Reactive monitoring; failed deliveries surface in customer complaints; manual retry.",
         "Receipt-aware monitoring with auto-retry on transient failures; failed deliveries trigger exception workflow.",
         "Loop time, Coverage", "W2-S03 Multi-channel dispatcher"),

        # ---- Privacy Office ----
        ("Privacy Office",
         "Quarterly review",
         "Reconstructed evidence from system logs and BI exports; 3-week response time on auditor questions; manual sampling.",
         "Generated quarterly evidence packet ZIP; real-time query into LEDGER; sample-from-canonical workflow.",
         "Audit-row coverage", "W1-S04 Audit Console · W2-S04 Privacy Packet"),
        ("Privacy Office",
         "Decision-evidence inquiry",
         "Engineering ticket, log-grep, multi-day turnaround per inquiry.",
         "Direct LEDGER query by trace_id; full version-stamped evidence in seconds.",
         "Audit-row coverage", "W1-S02 LEDGER · W1-S04 Audit Console"),

        # ---- Streaming COO (Sponsor) ----
        ("Streaming COO (Sponsor)",
         "Strategic review",
         "Quarterly retention-NPV retrospective; weak leading indicators; programming-strategy feedback loop is months long.",
         "Daily retention-NPV scorecard + cohort-level breakdown; programming-strategy feedback loop is days.",
         "NPV", "W2-S04 90-day run · W3 fusion"),
    ]
    for i, row in enumerate(journeys, start=2):
        write_row(ws, i, row, bold_cols=(1, 5))
        ws.row_dimensions[i].height = 50
        # Group rows by persona — slight tint variation
        persona = row[0]
        if "VP" in persona:           tint = "FBF1DC"
        elif "Audience" in persona:    tint = "EFF7F5"
        elif "Content" in persona:     tint = "F4ECFB"
        elif "Pricing" in persona:     tint = "FBE8E8"
        elif "Marketing" in persona:   tint = "FBF1DC"
        elif "Privacy" in persona:     tint = "EDEDF0"
        elif "Streaming COO" in persona: tint = "F6F1E4"
        else: tint = None
        if tint:
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=tint)
    print(f"  Persona Journey Maps · {len(journeys)} stage rows")


def sheet_key_solutions(wb):
    """NEW SHEET: 12 productizable solutions distributed across sprints."""
    ws = wb.create_sheet("Key Solutions")
    headers = ["Solution ID", "Solution name", "What it is", "Persona owner",
               "Anchor sprint", "Wave", "KPI driver", "Reusable beyond Disney?"]
    widths = [12, 32, 50, 22, 18, 9, 22, 22]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "C2"

    solutions = [
        ("SOL-01", "Synthetic Subscriber Lab",
         "100K-subscriber synthetic population + 4 quarters of realistic engagement events. Reproducible test environment for shadow-runs and demo paths.",
         "Engagement Architect / Data Eng",
         "W1-S02-T1", "W1", "Coverage",
         "Yes — every TMT-MED-01 engagement reuses the lab pattern with sub-Practice tweaks"),
        ("SOL-02", "Streaming Event Connector Pack",
         "Reusable connector library for streaming-event ingestion: Disney+ event taxonomy + abstract interface for Hulu/ESPN+ extension.",
         "Data Engineering / Disney IT",
         "W1-S03-T1, W2-S01-T1", "W1+W2", "Coverage",
         "Yes — refactors into a TMT Practice-level asset for any streaming engagement"),
        ("SOL-03", "Subscriber Risk Lens",
         "Risk-Score agent + cohort-level dashboard. Audience Lead's primary tool for risk-cohort triage and intervention prioritization.",
         "Audience Lead",
         "W1-S03-T2, W1-S03-T3", "W1", "Precision/Recall, Coverage",
         "Yes — generalizes to any subscriber-driven business (telco, SaaS, gaming)"),
        ("SOL-04", "Outcome Attribution Engine",
         "Outcome-Attribute agent + 14-field audit row schema + LEDGER write/read paths. The substrate for retention-NPV measurement and personalization evidence.",
         "Engagement Architect / Privacy Office",
         "W1-S04-T1, W1-S04-T3", "W1", "Audit-row coverage, NPV",
         "Yes — APEX-canonical; portable to any TMT-MED-* service"),
        ("SOL-05", "Personalization Audit Trail",
         "BI dashboard tile → trace_id drill-through with full reasoning trace, version stamps, and consent-state attestation. Privacy Office's primary surface.",
         "Privacy Office / Internal Audit",
         "W1-S04-T3", "W1", "Audit-row coverage",
         "Yes — extends to every privacy-regulated personalization scenario"),
        ("SOL-06", "Content Surface Personalizer",
         "Content-Match agent + rights-window query service + programming-conflict detection. Rights-aware recommendation slate composition.",
         "Content Programming Director",
         "W2-S02-T1", "W2", "Loop time, NPV",
         "Yes — extends to broadcast/cable rights-aware programming (TMT-MED-03 fusion)"),
        ("SOL-07", "Concession Policy Studio",
         "Declarative policy-authoring UI with simulation against historical cohorts and PR-merge workflow. Pricing Strategy's primary tool.",
         "Pricing Strategy Lead",
         "W2-S02-T3", "W2", "Concession Eff., NPV",
         "Yes — refactors to any commercial-envelope-driven retention scenario"),
        ("SOL-08", "VP Retention Decision Console",
         "Production Adaptive Card flow with queue management, designate-approver workflow, and approval-history dashboard.",
         "VP Subscriber Retention",
         "W2-S03-T2", "W2", "Loop time",
         "Yes — every named-persona HITL gate reuses the pattern"),
        ("SOL-09", "Multi-Channel Dispatcher",
         "Production push + email + in-app dispatcher with receipt-aware monitoring and auto-retry. Consent-aware routing.",
         "Marketing Operations Director",
         "W2-S03-T1", "W2", "Loop time, Coverage",
         "Yes — extends to any subscriber-outreach scenario"),
        ("SOL-10", "Cohort Drift Watchtower",
         "Drift detection on Risk-Score outputs + alerting + drift root-cause UI + auto-retraining trigger with HITL gate.",
         "Audience Lead / ML Engineering",
         "W2-S03-T3", "W2", "Precision/Recall",
         "Yes — generalizes to any agentic-ML scenario with drift exposure"),
        ("SOL-11", "Retention NPV Dashboard",
         "Power BI semantic model + drill-through views per cohort, intervention type, and concession tier. VP Retention and Streaming COO's executive surface.",
         "VP Retention / Streaming COO",
         "W2-S02-T2, W2-S04-T2", "W2", "NPV, Concession Eff.",
         "Yes — APEX BI semantic-model pattern reused across services"),
        ("SOL-12", "Privacy/Consent Evidence Packet Generator",
         "Quarterly evidence-packet ZIP generator: manifests + policy + audit-row extract + tie-out reconciliation. Privacy Office's quarterly artifact.",
         "Privacy Office / DMTSP_lead",
         "W2-S04-T1", "W2", "Audit-row coverage",
         "Yes — analogous to RC-E2E-11 SOX evidence pattern; APEX-canonical"),
    ]
    for i, row in enumerate(solutions, start=2):
        write_row(ws, i, row, code_cols=(1, 5), bold_cols=(2,))
        # Color by wave
        wave_col = row[5]
        if "W1" in wave_col and "W2" not in wave_col:
            tint = "EFF7F5"
        elif wave_col == "W2":
            tint = "FBF1DC"
        else:
            tint = "F6F1E4"  # W1+W2 mixed or W3 candidate
        ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor=tint)
        ws.row_dimensions[i].height = 60
    print(f"  Key Solutions · {len(solutions)} solutions")


# ============================================================================
# BUILD
# ============================================================================

def build():
    print("Building TMT-MED-01 Sprint Plan...")
    wb = Workbook()
    sheet_readme(wb)
    sheet_wave_overview(wb)
    sheet_sprint_plan(wb)
    sheet_user_acceptance(wb)
    sheet_use_case_map(wb)
    sheet_persona_raci(wb)
    sheet_kpi_linkage(wb)
    sheet_persona_journey_maps(wb)
    sheet_key_solutions(wb)

    target = OUT
    try:
        wb.save(str(target))
    except PermissionError:
        n = 2
        while True:
            alt = target.with_name(target.stem + f"-v{n}" + target.suffix)
            if not alt.exists():
                target = alt; break
            n += 1
        wb.save(str(target))
        print(f"\n[notice] {OUT.name} was locked — saved as {target.name} instead.")
    print(f"\nWrote {target}")


if __name__ == "__main__":
    build()

"""Build APEX-Scenario-Chains.xlsx — modeled on AXLE-Scenario-Chains.xlsx.

Pulls data from:
  - APEX-Stacked-Architecture-Narrated.html (APEX_FLOW_DATA · 35 featured × 24 steps)
  - docs/scenarios/_catalog/featured-chains.json (35 chains)
  - docs/scenarios/_catalog/scenario-library.json (723 compact rows)

Five sheets matching AXLE format:
  1. Summary
  2. Scenario Library                 (~758 rows × 9 cols)
  3. Featured Chains                  (35 rows × 14 cols)
  4. Scenario→KPI Chain               (~758 rows × 8 cols)
  5. 24-Step Chain                    (~758 × 24 = ~18,192 rows × 12 cols)

Output: docs/reference/APEX-Scenario-Chains.xlsx
"""
from __future__ import annotations
import io, sys, json, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(r"C:\Stage\Clients\Industries\APEX")
HTML = ROOT / "docs/reference/APEX-Stacked-Architecture-Narrated.html"
FEAT_JSON = ROOT / "docs/scenarios/_catalog/featured-chains.json"
LIB_JSON = ROOT / "docs/scenarios/_catalog/scenario-library.json"
OUT = ROOT / "docs/reference/APEX-Scenario-Chains.xlsx"

# Practices and domain code mapping
PRACTICE_NAMES = {
    "RC":   "Retail & Consumer",
    "HLS":  "Healthcare & Life Sciences",
    "ER":   "Energy & Resources",
    "AXLE": "Industrial Manufacturing",
    "TMT":  "Technology, Media & Telecom",
    "TH":   "Travel & Hospitality",
    "ICE":  "Industrial & Commercial Equipment",
}

DOMAIN_NAMES = {
    "CLIN":   "Clinical Care",
    "NET":    "Network & Infrastructure",
    "ENG":    "Engineering & R&D",
    "SUPCHN": "Supply Chain",
    "ASSET":  "Asset Maintenance",
    "QUAL":   "Quality & Compliance",
    "RISK":   "Risk · Fraud · Security",
    "PRICE":  "Pricing & Revenue",
    "CX":     "Customer Experience",
    "MKTG":   "Marketing & Growth",
    "OPS":    "Operations & Workforce",
    "CHAN":   "Channel · Partner · Dealer",
    "OTHER":  "Other / Cross-cutting",
}

# Styling
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="0B0B0F")
HDR_AL   = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN     = Side(style="thin", color="D4D4D7")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND     = PatternFill("solid", fgColor="F3F3F5")
CELL     = Font(name="Calibri", size=10)
CODE     = Font(name="Consolas", size=10, color="6648B0", bold=True)
CELL_AL  = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR      = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_hdrs(ws, headers, widths, *, row=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_AL; c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[row].height = 32


# -------------------------------------------------------------------------
# Extract APEX_FLOW_DATA from HTML
# -------------------------------------------------------------------------

def extract_flow_data():
    s = HTML.read_text(encoding="utf-8")
    start_idx = s.find("APEX_FLOW_DATA = {")
    eq = s.find("=", start_idx)
    js = s.find("{", eq)
    depth = 0; i = js; in_str = False; esc = False
    while i < len(s):
        ch = s[i]
        if esc:
            esc = False
        elif ch == "\\":
            esc = True
        elif ch == '"' and not esc:
            in_str = not in_str
        elif not in_str:
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    break
        i += 1
    return json.loads(s[js:i+1])


# -------------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------------

def derive_practice(scenario_id: str) -> str:
    """Get practice from scenario_id prefix."""
    for p in PRACTICE_NAMES:
        if scenario_id.startswith(p.lower() + "-") or scenario_id.startswith(p + "-"):
            return p
    return "OTHER"


def derive_domain_from_service(service_code: str) -> str:
    """Map RC-E2E-03 etc. to a domain."""
    # Map service codes → likely domain
    mapping = {
        # RC services
        "RC-E2E-03": "Pricing & Revenue",
        "RC-E2E-04": "Customer Experience",
        "RC-E2E-05": "Operations & Workforce",
        "RC-E2E-06": "Operations & Workforce",
        "RC-E2E-07": "Risk · Fraud · Security",
        "RC-E2E-08": "Marketing & Growth",
        "RC-E2E-09": "Quality & Compliance",
        "RC-E2E-11": "Quality & Compliance",
        # HLS
        "HLS-E2E-02": "Clinical Care",
        "HLS-E2E-04": "Clinical Care",
        "HLS-E2E-06": "Pricing & Revenue",
        "HLS-LS-03":  "Marketing & Growth",
        # ER
        "ER-OG-02":   "Asset Maintenance",
        "ER-OG-03":   "Operations & Workforce",
        "ER-OG-05":   "Operations & Workforce",
        "ER-OG-07":   "Asset Maintenance",
        "ER-MN-02":   "Operations & Workforce",
        "ER-PU-04":   "Network & Infrastructure",
        # AXLE
        "AXLE-Connected-Factory-01": "Asset Maintenance",
        "AXLE-Ops-03":    "Operations & Workforce",
        "AXLE-QMS-02":    "Quality & Compliance",
        "AXLE-Supply-04": "Supply Chain",
        "AXLE-Supply-07": "Supply Chain",
        # TMT
        "TMT-TEL-NET-02": "Network & Infrastructure",
        "TMT-TEL-CC-03":  "Customer Experience",
        "TMT-MED-01":     "Customer Experience",
        "TMT-MED-03":     "Channel · Partner · Dealer",
        "TMT-TEC-04":     "Engineering & R&D",
        # TH
        "TH-AIR-02":  "Operations & Workforce",
        "TH-AIR-04":  "Operations & Workforce",
        "TH-AIR-06":  "Pricing & Revenue",
        "TH-HOT-03":  "Operations & Workforce",
        "TH-HOT-05":  "Pricing & Revenue",
        # ICE
        "ICE-Aftermarket-01": "Operations & Workforce",
        "ICE-Aftermarket-03": "Quality & Compliance",
        "ICE-Aftermarket-05": "Customer Experience",
        "ICE-Connected-06":   "Asset Maintenance",
        "ICE-EaaS-04":        "Operations & Workforce",
    }
    return mapping.get(service_code, "Other / Cross-cutting")


def slugify(text: str) -> str:
    """Convert title to slug."""
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def parse_wave_from_prose(prose: str, wave_marker: str) -> str:
    """Extract wave-specific phrasing from prose if present."""
    if not prose:
        return ""
    parts = re.split(rf"(?i)\b(?:wave\s*)?{wave_marker}\b", prose)
    if len(parts) > 1:
        # Take a snippet after the marker
        return parts[1].split(".")[0].strip(" :;,—") or ""
    return ""


# 24-step canonical sequence (matches APEX_FLOW_DATA structure for featured)
CANONICAL_24_STEPS = [
    ("w1-sor",          "W1", "SOR · System of Record",                       "Integration Plane · SOR",        "source",    "Original system data lives here · agents do not replace the SOR"),
    ("w1-rth",          "W1", "Real-Time Hub · streaming ingest",              "Integration Plane",              "ingest",    "Eventstream / MQTT / OPC-UA bridge from operational systems"),
    ("w1-bronze",       "W1", "Bronze landing · raw + ingest metadata",        "Data Plane · Bronze",            "data",      "Source-aligned · append-only · 5-field ingest metadata"),
    ("w1-tokenizer",    "W1", "Tokenizer / PII handling",                      "Data Plane · Bronze",            "transform", "Hash/redact PII before downstream availability"),
    ("w1-silver",       "W1", "Silver canonical · MERML / SCML / FINML",       "Data Plane · Silver",            "canonical", "Conformed canonical layer agents read against"),
    ("w1-gold",         "W1", "Gold semantic · Power BI Direct Lake",          "Data Plane · Gold",              "semantic",  "Curated KPI marts + semantic models for BI"),
    ("w1-mcp",          "W1", "MCP server · capability registry",              "Runtime Plane",                  "runtime",   "Tool capabilities exposed to agents via JSON-RPC"),
    ("w1-identity",     "W1", "Entra · agent + persona identities",            "Identity Plane",                 "identity",  "Service principals · group-based RBAC · PIM"),
    ("w1-ledger",       "W1", "LEDGER · 14-field audit row store",             "Ledger Plane",                   "evidence",  "Audit-row evidence store · primary substrate for KPIs"),
    ("w1-hitl-surface", "W1", "HITL Adaptive Card surface",                    "Experience Plane",               "hitl",      "Microsoft Teams Adaptive Card · named-persona binding"),
    ("w2-event",        "W2", "Event trigger · scenario start",                "Runtime Plane",                  "trigger",   "Watermark fires · trace_id allocated · orchestrator instantiated"),
    ("w2-orchestrator", "W2", "Parent orchestrator · agent dispatch",          "Runtime Plane · CAF Dynamic",   "orchestrate","Hierarchical or parallel · sequential with HITL gate"),
    ("w2-agent-assess", "W2", "Assess / Sense agent",                          "Decision Plane",                 "agent",     "Characterize the event from telemetry + context"),
    ("w2-agent-classify","W2","Classify / Score agent",                        "Decision Plane",                 "agent",     "Apply policy · classify state · score with confidence"),
    ("w2-agent-quantify","W2","Quantify / Compose agent",                      "Decision Plane",                 "agent",     "Compute $ impact or compose decision with reasoning"),
    ("w2-hitl",         "W2", "HITL gate · Adaptive Card",                     "Experience Plane · HITL",        "hitl",      "Material decisions route to named persona via Teams"),
    ("w2-agent-act",    "W2", "Act + Evidence-Write",                          "Decision Plane · Mutation",      "agent",     "Mutate downstream systems · emit closeout audit row"),
    ("w2-kpi",          "W2", "KPI rollup · Power BI semantic model",          "Experience Plane · BI",          "kpi",       "Drill from KPI → trace_id → reasoning trace"),
    ("w3-scale",        "W3", "Scale · multi-tenant · multi-banner",           "Runtime · Scale",                "scale",     "Cross-banner orchestration · per-banner policy manifests"),
    ("w3-fuse-markdown","W3", "Fuse · adjacent scenarios",                     "Runtime · Fusion",               "fuse",      "Cross-scenario fusion via shared canonical + capability tags"),
    ("w3-fuse-loyalty", "W3", "Fuse · cross-Practice + loyalty",                "Runtime · Fusion",               "fuse",      "Loyalty / cross-practice fusion via A2A swarm"),
    ("w3-purview",      "W3", "Purview · lineage + classification at scale",   "Governance Plane",               "governance","Bronze→Silver→agent→outcome lineage · sensitivity labels"),
    ("w3-ledger-feedback","W3","LEDGER feedback loop · model retraining",      "Ledger · Feedback",              "feedback",  "Audit rows feed retraining · cohort drift watchtower"),
    ("w3-kpi-enterprise","W3","Enterprise KPI · run-rate durable value",       "Experience · Executive",         "kpi",       "Enterprise-scale rollup · run-rate value · Steering Committee"),
]


# -------------------------------------------------------------------------
# Build the workbook
# -------------------------------------------------------------------------

def build():
    print("Loading source data...")
    featured = json.loads(FEAT_JSON.read_text(encoding="utf-8"))
    library = json.loads(LIB_JSON.read_text(encoding="utf-8"))
    flow_data = extract_flow_data()
    print(f"  featured: {len(featured)} chains")
    print(f"  library: {sum(len(v) for v in library.values())} compact rows in {len(library)} practices")
    print(f"  flow_data: {len(flow_data)} scenarios with 24-step flow")

    # Build a unified scenario list combining featured + library
    # featured keys come from flow_data: "rc-cold-chain-excursion", etc.
    # library keys come from JSON: practice → list of {t, s, b, k, kk}
    # featured-chains.json has explicit practice + index_in_practice

    # Build a slug index over the library so we can fuzzy-match featured titles
    # (featured-chains.json titles often carry an extra suffix the library row omits)
    lib_slugs_by_prac: dict[str, dict[str, dict]] = {}
    for prac, rows in library.items():
        lib_slugs_by_prac[prac] = {slugify(r.get("t", "")): r for r in rows}

    def best_lib_slug_match(practice: str, fslug: str) -> str | None:
        """Resolve a featured chain's slug to the closest library slug in the same practice."""
        bucket = lib_slugs_by_prac.get(practice, {})
        if fslug in bucket:
            return fslug
        # Token-overlap scoring; prefer highest overlap, tie-break by shortest extra slug
        f_tokens = set(fslug.split("-"))
        best = None
        best_score = 0.0
        for s in bucket:
            s_tokens = set(s.split("-"))
            if not s_tokens or not f_tokens:
                continue
            overlap = len(f_tokens & s_tokens)
            denom = min(len(f_tokens), len(s_tokens))
            score = overlap / denom if denom else 0.0
            # Score must clear 0.5 to be considered a match
            if score >= 0.5 and (score > best_score or
                                  (score == best_score and best and len(s) < len(best))):
                best_score = score
                best = s
        return best

    # Build featured map: keyed by (practice, *library-aligned* slug) → full chain data
    featured_by_slug = {}
    for f in featured:
        title = f.get("title", "")
        practice = f.get("practice", "OTHER")
        fslug = slugify(title)
        # Map to the library's canonical slug for this scenario (if a fuzzy match exists)
        canonical_slug = best_lib_slug_match(practice, fslug) or fslug
        # Construct flow_data key — should match the existing keys
        # Keys are like "rc-cold-chain-excursion" — build heuristic
        flow_key = None
        for k in flow_data:
            if k.startswith(practice.lower() + "-") and (
                fslug.startswith(k[len(practice) + 1:]) or
                k[len(practice) + 1:].startswith(fslug[:20])
            ):
                flow_key = k
                break
        # Fall back: pick by exact title-slug match
        if flow_key is None:
            for k in flow_data:
                if k.endswith("-" + fslug) or k == practice.lower() + "-" + fslug:
                    flow_key = k
                    break
        # Last-resort fall back: try the canonical (library-aligned) slug against flow_data
        if flow_key is None and canonical_slug != fslug:
            for k in flow_data:
                if k.endswith("-" + canonical_slug) or k == practice.lower() + "-" + canonical_slug:
                    flow_key = k
                    break
        f["_flow_key"] = flow_key
        f["_slug"] = fslug
        f["_canonical_slug"] = canonical_slug
        featured_by_slug[(practice, canonical_slug)] = f

    # Build the master scenario list (758 entries)
    all_scenarios = []  # each: dict with id, title, practice, domain, service_code, brief, kpi, schemas, featured, chain_data
    seq = 0
    for prac, rows in library.items():
        for row in rows:
            seq += 1
            title = row.get("t", "")
            service_code = row.get("s", "")
            brief = row.get("b", "")
            kpi = row.get("k", "")
            sslug = slugify(title)
            # Match against featured
            chain_data = featured_by_slug.get((prac, sslug))
            featured_flag = chain_data is not None
            # Assemble scenario_id like axle-pdm-cnc-spindle (lowercase practice + slug)
            scenario_id = f"{prac.lower()}-{sslug}"
            domain = derive_domain_from_service(service_code)
            schemas = ""  # Not in compact catalog by default
            all_scenarios.append({
                "seq": seq,
                "id": scenario_id,
                "title": title,
                "practice": prac,
                "service_code": service_code,
                "domain": domain,
                "brief": brief,
                "kpi": kpi,
                "schemas": schemas,
                "featured": featured_flag,
                "chain": chain_data,
            })

    print(f"\nBuilding workbook with {len(all_scenarios)} total scenarios...")
    wb = Workbook()

    # ===== Sheet 1: Summary =====
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    rows = [
        ("APEX Scenario Chains — Summary", None, None, None),
        (None, None, None, None),
        ("Source",                               "APEX-Stacked-Architecture-Narrated.html",                None, None),
        ("Source data feeds",                    "featured-chains.json · scenario-library.json · APEX_FLOW_DATA (HTML)", None, None),
        ("Total scenarios",                      str(len(all_scenarios)),  None, None),
        ("Featured chains (full Scenario→KPI · 24-step flow)", str(sum(1 for s in all_scenarios if s['featured'])), None, None),
        ("Catalog scenarios (compact)",          str(sum(1 for s in all_scenarios if not s['featured'])), None, None),
        ("Practices",                            "7 (RC · HLS · ER · AXLE · TMT · TH · ICE)", None, None),
        ("Domains",                              "13 functional domains",  None, None),
        ("Services in catalog",                  "37 unique service IDs",  None, None),
        ("24-step chain rows (758 × 24)",        f"{len(all_scenarios) * 24}",     None, None),
        (None, None, None, None),
        ("Sheet", "What it contains", "Rows (approx)", "Cols"),
        ("Summary", "This metadata sheet", "—", "4"),
        ("Scenario Library", "All scenarios (featured + catalog) with brief + KPI", str(len(all_scenarios)), "9"),
        ("Featured Chains", "Full Scenario → Solution → Use Case → Service → Persona → KPI for the 35 featured chains, plus W1/W2/W3 wave detail", "35", "14"),
        ("Scenario→KPI Chain", "Compact 8-col chain for every scenario (featured + catalog)", str(len(all_scenarios)), "8"),
        ("24-Step Chain", "Per-scenario W1/W2/W3 24-step flow · 24 rows per scenario", f"{len(all_scenarios) * 24}", "12"),
        (None, None, None, None),
        ("Document version", "v1.0 · 2026-05-01", None, None),
        ("Modeled on", "AXLE-Scenario-Chains.xlsx (parallel structure)", None, None),
        ("Independence posture",
         "All APEX commercial materials use 'Deloitte's Microsoft practice,' 'DMTSP,' 'Microsoft platform capabilities,' 'Microsoft-native deployment.' Terms 'partner,' 'alliance,' 'strategic alliance' do not appear.",
         None, None),
    ]
    for i, row in enumerate(rows, start=1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            if i == 1 and j == 1:
                c.font = Font(name="Calibri", size=14, bold=True, color="0B0B0F")
            elif i == 13:
                c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_AL; c.border = BORDER
            elif row[0] and j == 1:
                c.font = Font(name="Calibri", size=10, bold=True)
            else:
                c.font = CELL
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 28
    print(f"  Summary sheet · {len(rows)} rows")

    # ===== Sheet 2: Scenario Library =====
    ws = wb.create_sheet("Scenario Library")
    headers = ["#", "Scenario ID", "Title", "Service Code", "Domain", "Schemas", "Brief (the moment)", "KPI / Outcome", "Featured?"]
    widths = [5, 28, 38, 22, 22, 16, 60, 22, 12]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "C2"

    for i, scen in enumerate(all_scenarios, start=2):
        ws.cell(row=i, column=1, value=scen["seq"]).alignment = CTR
        ws.cell(row=i, column=2, value=scen["id"]).font = CODE
        ws.cell(row=i, column=3, value=scen["title"])
        ws.cell(row=i, column=4, value=scen["service_code"]).font = CODE
        ws.cell(row=i, column=5, value=scen["domain"])
        ws.cell(row=i, column=6, value=scen["schemas"])
        ws.cell(row=i, column=7, value=scen["brief"])
        ws.cell(row=i, column=8, value=scen["kpi"])
        ws.cell(row=i, column=9, value="⭐ Featured" if scen["featured"] else "Catalog")
        for j in range(1, 10):
            c = ws.cell(row=i, column=j); c.border = BORDER
            if c.font is None or c.font.size is None: c.font = CELL
            c.alignment = CELL_AL if j != 1 else CTR
            if i % 2 == 0: c.fill = BAND
    print(f"  Scenario Library · {len(all_scenarios)} rows")

    # ===== Sheet 3: Featured Chains =====
    ws = wb.create_sheet("Featured Chains")
    headers = ["#", "Scenario ID", "Title", "Service", "Domain", "W1 Foundation", "W2 Pilot (you are here)", "W3 Scale & Fuse", "Scenario (the moment)", "Solution (architectural approach)", "Use Case (Wave 2 delivery)", "Service (productized)", "Persona (operator · HITL approver)", "KPI / Outcome"]
    widths = [5, 28, 38, 22, 22, 36, 36, 36, 50, 50, 50, 36, 30, 36]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "C2"

    feat_only = [s for s in all_scenarios if s["featured"]]
    # Sort by practice then index
    feat_only.sort(key=lambda x: (list(PRACTICE_NAMES.keys()).index(x["practice"]) if x["practice"] in PRACTICE_NAMES else 99, x["title"]))
    for i, scen in enumerate(feat_only, start=2):
        chain = scen["chain"] or {}
        flow_key = chain.get("_flow_key")
        flow = flow_data.get(flow_key, {}) if flow_key else {}
        # Synthesize W1/W2/W3 detail from flow steps if available
        w1 = w2 = w3 = ""
        if flow:
            w1_steps = [v.get("title", "") for k, v in flow.items() if v.get("wave") == "W1"]
            w2_steps = [v.get("title", "") for k, v in flow.items() if v.get("wave") == "W2"]
            w3_steps = [v.get("title", "") for k, v in flow.items() if v.get("wave") == "W3"]
            w1 = " · ".join(w1_steps[:5])
            w2 = " · ".join(w2_steps[:5])
            w3 = " · ".join(w3_steps[:5])

        ws.cell(row=i, column=1, value=i-1).alignment = CTR
        ws.cell(row=i, column=2, value=scen["id"]).font = CODE
        ws.cell(row=i, column=3, value=scen["title"])
        ws.cell(row=i, column=4, value=scen["service_code"]).font = CODE
        ws.cell(row=i, column=5, value=scen["domain"])
        ws.cell(row=i, column=6, value=w1)
        ws.cell(row=i, column=7, value=w2)
        ws.cell(row=i, column=8, value=w3)
        ws.cell(row=i, column=9, value=chain.get("Scenario", scen["brief"]))
        ws.cell(row=i, column=10, value=chain.get("Solution", ""))
        ws.cell(row=i, column=11, value=chain.get("Use Case", ""))
        ws.cell(row=i, column=12, value=chain.get("Service", ""))
        ws.cell(row=i, column=13, value=chain.get("Persona", ""))
        ws.cell(row=i, column=14, value=chain.get("KPI", scen["kpi"]))
        for j in range(1, 15):
            c = ws.cell(row=i, column=j); c.border = BORDER
            if c.font is None or c.font.size is None: c.font = CELL
            c.alignment = CELL_AL if j != 1 else CTR
            if i % 2 == 0: c.fill = BAND
    print(f"  Featured Chains · {len(feat_only)} rows")

    # ===== Sheet 4: Scenario → KPI Chain =====
    ws = wb.create_sheet("Scenario→KPI Chain")
    headers = ["Scenario ID", "Scenario (the moment)", "Solution / Agent (ORCH archetype)", "Use Cases (decomposed)", "Service (Deloitte unit)", "Personas (operator · HITL approver)", "KPI (measurable outcome)", "Featured?"]
    widths = [28, 50, 40, 40, 36, 30, 36, 12]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "B2"

    for i, scen in enumerate(all_scenarios, start=2):
        chain = scen["chain"] or {}
        scenario_text = chain.get("Scenario", scen["brief"])
        solution_text = chain.get("Solution", "Standard archetype: hierarchical-root + sequential-with-hitl-gate · 6-agent fleet · APEX canonical")
        use_case = chain.get("Use Case", f"Wave 2 pilot delivery for {scen['title']}")
        service = chain.get("Service", scen["service_code"])
        persona = chain.get("Persona", "Operator with HITL approval authority via Teams Adaptive Card")
        kpi = chain.get("KPI", scen["kpi"])

        ws.cell(row=i, column=1, value=scen["id"]).font = CODE
        ws.cell(row=i, column=2, value=scenario_text)
        ws.cell(row=i, column=3, value=solution_text)
        ws.cell(row=i, column=4, value=use_case)
        ws.cell(row=i, column=5, value=service)
        ws.cell(row=i, column=6, value=persona)
        ws.cell(row=i, column=7, value=kpi)
        ws.cell(row=i, column=8, value="⭐" if scen["featured"] else "")
        for j in range(1, 9):
            c = ws.cell(row=i, column=j); c.border = BORDER
            if c.font is None or c.font.size is None: c.font = CELL
            c.alignment = CELL_AL if j != 8 else CTR
            if i % 2 == 0: c.fill = BAND
    print(f"  Scenario→KPI Chain · {len(all_scenarios)} rows")

    # ===== Sheet 5: 24-Step Chain =====
    ws = wb.create_sheet("24-Step Chain")
    headers = ["Scenario ID", "Title", "Domain", "Wave", "Step #", "Step Key", "Step Title", "Layer", "Kind", "Purpose", "What APEX Does", "Scenario Note"]
    widths = [28, 36, 22, 6, 7, 22, 32, 28, 12, 40, 40, 40]
    write_hdrs(ws, headers, widths)
    ws.freeze_panes = "B2"

    row_idx = 2
    rows_written = 0
    for scen in all_scenarios:
        flow_key = (scen["chain"] or {}).get("_flow_key") if scen["featured"] else None
        flow = flow_data.get(flow_key, {}) if flow_key else {}

        for step_num, step_def in enumerate(CANONICAL_24_STEPS, start=1):
            step_key, wave, step_title, layer, kind, purpose = step_def
            # If we have real flow_data for this scenario, prefer it
            if flow and step_key in flow:
                rs = flow[step_key]
                title = rs.get("title", step_title)
                wave_v = rs.get("wave", wave)
                layer_v = rs.get("layer", layer)
                kind_v = rs.get("kind", kind)
                details = rs.get("details", {}) or {}
                purpose_v = details.get("purpose", purpose)
                what_apex = details.get("what_apex_does", "")
                scen_note = details.get("scenario_note", "")
            else:
                title = step_title
                wave_v = wave
                layer_v = layer
                kind_v = kind
                purpose_v = purpose
                what_apex = "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"
                scen_note = scen["brief"][:120] if scen["brief"] else ""

            ws.cell(row=row_idx, column=1, value=scen["id"]).font = CODE
            ws.cell(row=row_idx, column=2, value=scen["title"])
            ws.cell(row=row_idx, column=3, value=scen["domain"])
            ws.cell(row=row_idx, column=4, value=wave_v)
            ws.cell(row=row_idx, column=5, value=step_num)
            ws.cell(row=row_idx, column=6, value=step_key).font = CODE
            ws.cell(row=row_idx, column=7, value=title)
            ws.cell(row=row_idx, column=8, value=layer_v)
            ws.cell(row=row_idx, column=9, value=kind_v)
            ws.cell(row=row_idx, column=10, value=purpose_v)
            ws.cell(row=row_idx, column=11, value=what_apex)
            ws.cell(row=row_idx, column=12, value=scen_note)
            for j in range(1, 13):
                c = ws.cell(row=row_idx, column=j); c.border = BORDER
                if c.font is None or c.font.size is None: c.font = CELL
                c.alignment = CELL_AL if j not in (4, 5) else CTR
                if row_idx % 2 == 0: c.fill = BAND
            row_idx += 1
            rows_written += 1
    print(f"  24-Step Chain · {rows_written} rows")

    # Save
    OUT.parent.mkdir(parents=True, exist_ok=True)
    target = OUT
    try:
        wb.save(str(target))
    except PermissionError:
        n = 2
        while True:
            alt = target.with_name(target.stem + f"-v{n}" + target.suffix)
            if not alt.exists():
                target = alt; break
            n += 1
        wb.save(str(target))
        print(f"\n[notice] {OUT.name} was locked — saved as {target.name} instead.")
    print(f"\nWrote {target}")
    print(f"  Total: {len(all_scenarios)} scenarios · {rows_written} 24-step rows · 5 sheets")


if __name__ == "__main__":
    build()

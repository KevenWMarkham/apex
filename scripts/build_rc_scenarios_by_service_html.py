"""Build the APEX RC Scenario Chains HTML re-organized by service code.

Reads `APEX-RC-Scenario-Chains.xlsx` and emits a new HTML file where every scenario
is grouped under its 15 service codes (RC-E2E-01 → RC-E2E-15), with an authored
description for every service. Within each service, scenarios are further grouped
by sub-area (preserving domain context as a small chip on each card).

Output mirrors the visual language of `APEX-RC-Scenario-Chains.html` so the two
views feel like a switchable lens on the same catalog.
"""
from __future__ import annotations
import io, sys, html, re, json, os
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook

# Make sure we can import the sibling schema-fields module regardless of cwd
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from apex_rc_schema_fields import SCHEMA_FIELDS, AUDIT_TAIL

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = r"C:\Stage\Clients\Industries\APEX\docs\reference\Retail and Consumer\APEX-RC-Scenario-Chains.xlsx"
OUT  = r"C:\Stage\Clients\Industries\APEX\docs\reference\Retail and Consumer\APEX-RC-Scenario-Chains-by-Service.html"

# ----------------------------------------------------------------------------
# Authored service descriptions (the "(Add description of service)" deliverable)
# ----------------------------------------------------------------------------
# Each description is ~50–80 words and explains: what the service covers,
# the business outcome it drives, and how it composes with adjacent services.
# Authored for an SI/sales audience — concrete, not generic.

SERVICE_DESCRIPTIONS: dict[str, dict] = {
    "RC-E2E-01": {
        "name": "Demand Sensing & Planning",
        "tagline": "Forward-looking forecast and plan generation across the full demand curve.",
        "desc": (
            "Translates leading consumer-intent signals (search · social · weather · macro · "
            "competitive launches) into sellable units by SKU × store × week across baseline, "
            "new-product, seasonal-buy, promotional, and lifecycle-end horizons. Every downstream "
            "service — inventory, pricing, marketing, supply chain — operates on the same governed "
            "forecast, so a single demand correction propagates to every plan automatically."
        ),
        "color": "#0EA5E9",
    },
    "RC-E2E-02": {
        "name": "Inventory & Replenishment",
        "tagline": "Day-by-day position management for in-store, DC, and dark-store inventory.",
        "desc": (
            "Auto-replenishment, allocation, and rebalance under guardrails for OTB, lead time, "
            "and on-shelf availability. Closes the loop between Demand Sensing and Supply Chain so "
            "shelves stay full while working capital stays disciplined. The service that turns "
            "forecast into stock-keeping reality at the SKU × DC × store grain — and the first "
            "service that gets touched by every supply disruption."
        ),
        "color": "#0369A1",
    },
    "RC-E2E-03": {
        "name": "Assortment & Pricing",
        "tagline": "Range, role, and price decisions across the merchandising lifecycle.",
        "desc": (
            "Owns base price, promotion, markdown, clearance, EOL, competitive response, range "
            "architecture, role classification, cluster fit, and white-space discovery. Ties the "
            "merchandising plan to the financial plan with shared guardrails on margin, OTB, and "
            "aggregate weekly spend. The service the Chief Merchant looks at first on Monday "
            "morning — and the anchor of the Q1 Agentic Merch reference scenario."
        ),
        "color": "#6648B0",
    },
    "RC-E2E-04": {
        "name": "Customer Lifecycle & Loyalty",
        "tagline": "Acquisition, engagement, personalization, attribution, retention, advocacy.",
        "desc": (
            "The largest service in the catalog because it carries every customer-facing touchpoint "
            "where a personalized agentic decision moves revenue — site search, recommendations, "
            "cart and checkout, campaign management, content and creative, MTA / MMM, segment "
            "modeling, loyalty tiers, retention, winback, reviews, restaurant loyalty, and "
            "marketplace subscriptions. Spans web, app, store, marketplace, and care channels."
        ),
        "color": "#DB2777",
    },
    "RC-E2E-05": {
        "name": "Store Operations",
        "tagline": "The physical-retail service — shelf, planogram, OSA, task execution, in-store CX.",
        "desc": (
            "Turns merchandising and assortment intent into executed shelf realities through "
            "computer-vision audits, task orchestration, on-shelf-availability monitoring, store "
            "tech telemetry, and in-store experience instrumentation. The service where the rest "
            "of the catalog meets the customer eye-to-shelf — and where many of the highest-ROI "
            "agentic interventions live (compliance, OOS reduction, planogram fidelity)."
        ),
        "color": "#10B981",
    },
    "RC-E2E-06": {
        "name": "Supply Chain & Logistics",
        "tagline": "Vendor collaboration, transportation, warehouse, last-mile, yield, reverse transport.",
        "desc": (
            "Connects upstream supplier ecosystems to downstream fulfillment surfaces. Owns route "
            "optimization, carrier scoring, DC throughput, last-mile delivery promise, yield "
            "management, vendor scorecards, and reverse logistics flow. Every disruption — port, "
            "weather, supplier, carrier — surfaces here first, and every recovery move (rebalance, "
            "secondary supplier, expedited PO) executes through the surfaces this service manages."
        ),
        "color": "#0891B2",
    },
    "RC-E2E-07": {
        "name": "Returns & Refund Integrity",
        "tagline": "The trust perimeter — protects margin and brand from revenue leakage.",
        "desc": (
            "Returns triage, BORIS / BOPIS, refurbishment, salvage, returns fraud, payment fraud, "
            "account takeover, organized retail crime, loss prevention, self-checkout shrink, and "
            "physical security. Defends both intentional revenue leakage (fraud, ORC, ATO) and "
            "unintentional leakage (shrink, returns abuse, salvage write-down). The service that "
            "earns its ROI by what *doesn't* leave the building."
        ),
        "color": "#DC2626",
    },
    "RC-E2E-08": {
        "name": "Workforce & Talent",
        "tagline": "Scheduling, recruiting, onboarding, engagement, attrition, workforce analytics.",
        "desc": (
            "People operations across HQ, field, store, plant, and restaurant — labor planning to "
            "talent retention. Owns demand-driven scheduling, requisition acceleration, onboarding "
            "and training pathway optimization, attrition-risk prediction, engagement signal "
            "synthesis, and restaurant-specific labor mechanics. Tightly composes with Store Ops "
            "and Foodservice on labor demand, and with Finance on wage-and-hour compliance."
        ),
        "color": "#F59E0B",
    },
    "RC-E2E-09": {
        "name": "Product Tracking & Provenance",
        "tagline": "Regulator-grade supply chain visibility — lot trace, recall, provenance.",
        "desc": (
            "Powers cold-chain lot tracing, recall-readiness rehearsal, product authentication, "
            "and provenance certification for sustainability claims. The lowest-volume service in "
            "the catalog by scenario count but among the highest-stakes — every lot trace either "
            "narrowly contains a recall or expensively over-recalls, and every provenance claim is "
            "either auditable or actionable by regulators."
        ),
        "color": "#7C3AED",
    },
    "RC-E2E-10": {
        "name": "Trade Promotion & RGM",
        "tagline": "Manufacturer-to-retailer — translates RGM strategy into executed trade plans.",
        "desc": (
            "Trade-spend effectiveness, deduction management, price-pack architecture, mix "
            "optimization, joint business planning, and promotional funding. The service most "
            "relevant to CPG manufacturers: turns headquarters RGM strategy into store-level "
            "executed plans, recovers leakage through deduction management, and optimizes mix "
            "across pack-price ladders. Composes tightly with Assortment & Pricing on the retailer "
            "side of the trade equation."
        ),
        "color": "#BE185D",
    },
    "RC-E2E-11": {
        "name": "Manufacturing & Plant Ops",
        "tagline": "The CPG plant-floor service — uptime, yield, quality, safety, sustainability.",
        "desc": (
            "OEE monitoring, predictive maintenance, quality, recipe and spec management, plant "
            "safety, plant sustainability, energy and refrigeration, and procurement source. Turns "
            "manufacturing telemetry into uptime, yield, and safety outcomes. Composes with "
            "Sustainability & ESG on Scope-1/2 emissions and with Supply Chain on procurement and "
            "vendor performance signals."
        ),
        "color": "#059669",
    },
    "RC-E2E-12": {
        "name": "Foodservice & Restaurant Ops",
        "tagline": "QSR and casual dining mechanics — menu, kitchen, drive-thru, off-premise.",
        "desc": (
            "Menu engineering, kitchen operations throughput, drive-thru and off-premise channel "
            "optimization, and restaurant-specific workflows. Covers the unique mechanics of "
            "foodservice that don't fit pure retail or pure CPG patterns — order-stream queuing, "
            "minute-grain labor demand, recipe-grade quality, and channel-mix economics across "
            "dine-in, drive-thru, delivery, and curbside."
        ),
        "color": "#EA580C",
    },
    "RC-E2E-13": {
        "name": "Sustainability & ESG",
        "tagline": "Compliance + claim defensibility — every ESG number needs an auditable trail.",
        "desc": (
            "Scope-3 emissions mapping, packaging and waste, water and energy, supplier ESG "
            "scoring, plant sustainability, food-waste shrink reduction, and ESG reporting and "
            "disclosure. Every ESG number that goes to investors or regulators needs an auditable "
            "agentic decision trail; this service produces them. Composes with Manufacturing on "
            "plant-level signals and with Supply Chain on supplier-tier emissions data."
        ),
        "color": "#15803D",
    },
    "RC-E2E-14": {
        "name": "Real Estate & Network",
        "tagline": "Slow-cadence, high-stakes — site, format, lease, capex, closures, restructuring.",
        "desc": (
            "Site selection and trade-area modeling, format and network design, lease and property "
            "management, lease CAM reconciliation, closures and restructuring, and retrofit / "
            "capex planning. Every store touched is multi-year capital and brand exposure, so the "
            "scenarios in this service tend to be lower-volume but higher-impact-per-decision than "
            "almost any other service in the catalog."
        ),
        "color": "#7C2D12",
    },
    "RC-E2E-15": {
        "name": "Finance, Risk & Controls",
        "tagline": "The back-office trust spine — every dollar in/out, every regulator filing, every control.",
        "desc": (
            "Revenue recognition, settlement and reconciliation, FP&amp;A and close acceleration, "
            "tax and compliance, audit and controls, treasury and working capital, food safety and "
            "recipe compliance, cyber and identity, insider threat and privacy, supplier and "
            "vendor management, wage and hour compliance, and counterfeit and brand protection. "
            "The second-largest service by scenario count — every other service ultimately settles "
            "into a row this one signs off on."
        ),
        "color": "#1F2937",
    },
}

# ----------------------------------------------------------------------------
# Read scenarios from xlsx
# ----------------------------------------------------------------------------
print(f"Reading {XLSX}")
wb = load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["Scenario Library"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print(f"Headers: {header}")

# Index columns by header name for safety
H = {h: i for i, h in enumerate(header)}
COL_ID       = H["Scenario ID"]
COL_TITLE    = H["Title"]
COL_DOMAIN   = H["Domain"]
COL_SUBAREA  = H["Sub-Area"]
COL_SVC_CODE = H["Service Code"]
COL_SVC_NAME = H["Service"]
COL_SCHEMAS  = H["Schemas"]
COL_BRIEF    = H["Brief (the moment)"]
COL_KPI      = H["KPI / Outcome"]
COL_PERSONA  = H["Persona"]
COL_FEAT     = H["Featured?"]

scenarios: list[dict] = []
for r in rows[1:]:
    if not r[COL_ID]:
        continue
    scenarios.append({
        "id":       r[COL_ID] or "",
        "title":    r[COL_TITLE] or "",
        "domain":   r[COL_DOMAIN] or "",
        "subarea":  r[COL_SUBAREA] or "",
        "svc_code": r[COL_SVC_CODE] or "",
        "svc_name": r[COL_SVC_NAME] or "",
        "schemas":  r[COL_SCHEMAS] or "",
        "brief":    r[COL_BRIEF] or "",
        "kpi":      r[COL_KPI] or "",
        "persona":  r[COL_PERSONA] or "",
        "featured": "featured" in str(r[COL_FEAT] or "").strip().lower() or "★" in str(r[COL_FEAT] or ""),
    })

print(f"Loaded {len(scenarios)} scenarios")

# Featured chains for the top section
ws_f = wb["Featured Chains"]
f_rows = list(ws_f.iter_rows(values_only=True))
fh = f_rows[0]
FH = {h: i for i, h in enumerate(fh)}
featured_chains = []
for r in f_rows[1:]:
    if not r[FH["Scenario ID"]]:
        continue
    featured_chains.append({
        "id":      r[FH["Scenario ID"]],
        "title":   r[FH["Title"]],
        "svc":     r[FH["Service"]],
        "domain":  r[FH["Domain"]],
        "subarea": r[FH["Sub-Area"]],
        "w1":      r[FH["W1 Foundation"]],
        "w2":      r[FH["W2 Pilot (you are here)"]],
        "w3":      r[FH["W3 Scale & Fuse"]],
        "moment":  r[FH["Scenario (the moment)"]],
        "persona": r[FH["Persona (operator · HITL approver)"]],
        "kpi":     r[FH["KPI / Outcome"]],
    })
print(f"Loaded {len(featured_chains)} featured chains")

# ----------------------------------------------------------------------------
# Group scenarios by service > sub-area
# ----------------------------------------------------------------------------
by_svc: dict[str, dict[str, list[dict]]] = OrderedDict()
for code in sorted(SERVICE_DESCRIPTIONS.keys()):
    by_svc[code] = OrderedDict()

for s in scenarios:
    code = s["svc_code"]
    if code not in by_svc:
        by_svc[code] = OrderedDict()
    sub = s["subarea"]
    by_svc[code].setdefault(sub, []).append(s)

# Sort sub-areas alphabetically and scenarios within each sub-area by title
for code, subs in by_svc.items():
    by_svc[code] = OrderedDict(sorted(subs.items()))
    for sub in by_svc[code]:
        by_svc[code][sub].sort(key=lambda s: s["title"].lower())

svc_counts = {c: sum(len(v) for v in subs.values()) for c, subs in by_svc.items()}
svc_featured_counts = {c: sum(1 for v in subs.values() for s in v if s["featured"]) for c, subs in by_svc.items()}

# ----------------------------------------------------------------------------
# Render HTML
# ----------------------------------------------------------------------------
def esc(x) -> str:
    if x is None:
        return ""
    return html.escape(str(x))

def slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")

def short_chain(text: str) -> str:
    """Compact a long chain string by collapsing repeating boilerplate."""
    if not text:
        return ""
    return text.replace("Real-Time Hub", "RTH").replace("Raw Landing", "Bronze Raw")

# Stat counts
total_scenarios = len(scenarios)
total_services = len(by_svc)
total_subareas = sum(len(v) for v in by_svc.values())
total_featured = sum(1 for s in scenarios if s["featured"])
total_domains = len({s["domain"] for s in scenarios})

# ----------------------------------------------------------------------------
# Begin HTML
# ----------------------------------------------------------------------------
out: list[str] = []
ap = out.append

ap("""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>APEX · Retail &amp; Consumer Scenario Chains — by Service</title>
<style>
  :root {
    --ink: #0B0B0F;
    --ink-2: #1F2024;
    --ink-3: #4A4D55;
    --bg: #FAFAFC;
    --bg-card: #FFFFFF;
    --line: #E4E4E7;
    --line-2: #D4D4D7;
    --accent: #6648B0;
    --accent-2: #8E70D8;
    --kpi: #047857;
    --w1: #0EA5E9;
    --w2: #6648B0;
    --w3: #10B981;
    --hitl: #F59E0B;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; }
  body {
    font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, 'Helvetica Neue', sans-serif;
    background: var(--bg);
    color: var(--ink);
    font-size: 14px;
    line-height: 1.5;
  }
  header {
    background: linear-gradient(135deg, #0B0B0F 0%, #1F2024 60%, #6648B0 100%);
    color: #FFF;
    padding: 36px 32px 28px;
  }
  header h1 { margin: 0 0 8px 0; font-size: 30px; font-weight: 700; letter-spacing: -0.02em; }
  header .sub { font-size: 14px; opacity: 0.85; max-width: 980px; }
  header .stats { margin-top: 18px; display: flex; gap: 14px; flex-wrap: wrap; }
  header .stat { background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.18); border-radius: 10px; padding: 10px 14px; }
  header .stat .v { font-size: 22px; font-weight: 700; }
  header .stat .l { font-size: 11px; opacity: 0.78; text-transform: uppercase; letter-spacing: 0.04em; }

  .controls {
    background: var(--bg-card);
    border-bottom: 1px solid var(--line);
    padding: 14px 32px;
    position: sticky;
    top: 0;
    z-index: 50;
    display: flex;
    gap: 10px;
    align-items: center;
    flex-wrap: wrap;
  }
  .controls .search-wrap { flex: 1 1 280px; min-width: 240px; position: relative; }
  .controls input[type="search"] { width: 100%; padding: 10px 12px; border: 1px solid var(--line-2); border-radius: 8px; font-size: 13px; font-family: inherit; box-sizing: border-box; }
  .controls input[type="search"]:focus { outline: 2px solid var(--accent); outline-offset: -1px; border-color: var(--accent); }
  .suggestions {
    position: absolute;
    top: calc(100% + 4px);
    left: 0;
    right: 0;
    background: #FFF;
    border: 1px solid var(--line-2);
    border-radius: 8px;
    box-shadow: 0 8px 24px rgba(15, 23, 42, .12);
    max-height: 360px;
    overflow-y: auto;
    z-index: 60;
  }
  .suggestions[hidden] { display: none; }
  .sugg-item {
    padding: 10px 12px;
    border-bottom: 1px solid var(--line);
    cursor: pointer;
    font-size: 12px;
  }
  .sugg-item:last-child { border-bottom: none; }
  .sugg-item:hover, .sugg-item.active { background: #F5F3FF; }
  .sugg-title { font-weight: 600; color: var(--ink); font-size: 13px; line-height: 1.35; margin-bottom: 3px; }
  .sugg-title mark { background: #FEF08A; color: var(--ink); padding: 0 1px; border-radius: 2px; }
  .sugg-meta { color: var(--ink-3); font-size: 11px; }
  .sugg-meta .sugg-id { font-family: ui-monospace, "SF Mono", Menlo, monospace; color: var(--ink-2); }
  .sugg-meta .sugg-svc { color: var(--accent); font-weight: 600; }
  .sugg-empty { padding: 12px; color: var(--ink-3); font-size: 12px; font-style: italic; }
  .controls select { padding: 9px 10px; border: 1px solid var(--line-2); border-radius: 8px; font-size: 13px; background: #FFF; font-family: inherit; }
  .controls .toggles { display: flex; gap: 6px; }
  .controls .toggle { border: 1px solid var(--line-2); background: #FFF; border-radius: 8px; padding: 8px 12px; font-size: 12px; cursor: pointer; color: var(--ink-2); }
  .controls .toggle.active { background: var(--ink); color: #FFF; border-color: var(--ink); }
  .controls .btn {
    border: 1px solid var(--line-2);
    background: #FFF;
    border-radius: 8px;
    padding: 9px 14px;
    font-size: 12px;
    cursor: pointer;
    color: var(--ink-2);
    font-weight: 600;
    font-family: inherit;
  }
  .controls .btn:hover { background: var(--bg); }
  .controls .btn.primary { background: var(--accent); color: #FFF; border-color: var(--accent); }
  .controls .btn.primary:hover { background: var(--accent-2); border-color: var(--accent-2); }
  .controls .btn-clear { color: var(--ink-3); }
  .controls .filter-state {
    font-size: 11px;
    color: var(--accent);
    background: #F5F3FF;
    border: 1px solid #DDD6FE;
    border-radius: 999px;
    padding: 3px 10px;
    font-weight: 600;
  }
  .controls .filter-state.hidden { display: none; }

  main { padding: 28px 32px 60px; }

  /* Flat results view — applied when any filter / search / sub-area / featured-only is active.
     We collapse the service / sub-area boxes via display:contents so every visible card
     flows into a single contiguous grid at the <main> level. The TOC and Featured-Chains
     section also hide so the user only sees what matched. */
  body.filtering main {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(420px, 1fr));
    gap: 12px;
    padding: 24px 32px 60px;
  }
  body.filtering .toc,
  body.filtering .featured-section,
  body.filtering .service-head,
  body.filtering .subarea-head { display: none !important; }
  body.filtering .service,
  body.filtering .subarea,
  body.filtering .cards { display: contents; }
  body.filtering .service.hidden,
  body.filtering .subarea.hidden { display: contents; } /* container hidden state irrelevant in flat view */
  body.filtering .results-banner {
    grid-column: 1 / -1;
    background: var(--bg-card);
    border: 1px solid var(--line);
    border-left: 3px solid var(--accent);
    border-radius: 10px;
    padding: 12px 16px;
    margin-bottom: 4px;
    font-size: 13px;
    color: var(--ink-2);
  }
  body.filtering .results-banner b { color: var(--ink); }
  body:not(.filtering) .results-banner { display: none; }

  /* Service blocks (the new primary grouping) */
  .service {
    margin-bottom: 36px;
    background: var(--bg-card);
    border: 1px solid var(--line);
    border-radius: 14px;
    overflow: hidden;
  }
  .service-head {
    padding: 22px 26px 18px;
    color: #FFF;
    background: var(--svc-color, #1F2024);
  }
  .service-head .top {
    display: flex;
    align-items: baseline;
    gap: 14px;
    flex-wrap: wrap;
    margin-bottom: 6px;
  }
  .service-head .code {
    font-family: 'Cascadia Code', Consolas, monospace;
    background: rgba(255,255,255,0.18);
    border: 1px solid rgba(255,255,255,0.28);
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 0.02em;
  }
  .service-head h2 {
    margin: 0;
    font-size: 22px;
    font-weight: 700;
    letter-spacing: -0.01em;
    flex: 1;
  }
  .service-head .badge-row {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
  }
  .service-head .badge {
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.28);
    border-radius: 999px;
    padding: 3px 11px;
    font-size: 11px;
    font-weight: 600;
  }
  .service-head .tagline {
    font-size: 13px;
    font-style: italic;
    opacity: 0.92;
    margin: 0 0 10px 0;
  }
  .service-head .desc {
    font-size: 13px;
    line-height: 1.55;
    opacity: 0.94;
    max-width: 1100px;
  }

  .subarea {
    padding: 14px 22px 6px;
    border-top: 1px dashed var(--line);
  }
  .subarea:first-of-type { border-top: none; }
  .subarea-head {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 10px;
  }
  .subarea-head h3 {
    margin: 0;
    font-size: 14px;
    font-weight: 700;
    color: var(--ink-2);
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }
  .subarea-head .pill {
    background: var(--bg);
    border: 1px solid var(--line-2);
    border-radius: 999px;
    padding: 2px 10px;
    font-size: 11px;
    color: var(--ink-3);
    font-weight: 600;
  }
  .subarea-head .domain-pill {
    background: #F5F3FF;
    border: 1px solid #DDD6FE;
    border-radius: 999px;
    padding: 2px 10px;
    font-size: 11px;
    color: #5B21B6;
    font-weight: 600;
  }

  .cards {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(420px, 1fr));
    gap: 12px;
    padding-bottom: 12px;
  }
  .card {
    border: 1px solid var(--line);
    border-radius: 10px;
    padding: 14px 16px;
    background: var(--bg-card);
    transition: box-shadow 0.18s, transform 0.18s;
    position: relative;
  }
  .card:hover { box-shadow: 0 4px 18px rgba(11,11,15,0.07); transform: translateY(-1px); }
  .card.featured {
    border-left: 3px solid var(--accent);
    background: linear-gradient(180deg, #FBFAFF 0%, #FFFFFF 30%);
  }
  .card .row1 { display: flex; justify-content: space-between; align-items: flex-start; gap: 10px; margin-bottom: 6px; }
  .card h4 { margin: 0; font-size: 14px; font-weight: 700; color: var(--ink); letter-spacing: -0.005em; flex: 1; }
  .card .star { color: var(--accent); font-size: 14px; flex-shrink: 0; }
  .card .id { font-family: 'Cascadia Code', Consolas, monospace; color: var(--accent); font-size: 11px; font-weight: 600; margin-bottom: 8px; }
  .card .brief { font-size: 12px; color: var(--ink-2); margin-bottom: 10px; line-height: 1.5; }
  .card .kpi { color: var(--kpi); font-weight: 700; font-size: 12px; background: #ECFDF5; border: 1px solid #A7F3D0; border-radius: 6px; padding: 4px 8px; display: inline-block; margin-bottom: 8px; }
  .card .meta { display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 8px; }
  .card .meta .tag { font-size: 10px; background: var(--bg); border: 1px solid var(--line); border-radius: 4px; padding: 2px 6px; color: var(--ink-3); font-weight: 500; }
  .card .meta .tag.dom { background: #F5F3FF; border-color: #DDD6FE; color: #5B21B6; }
  .card .schemas { font-size: 10px; color: var(--ink-3); margin-bottom: 6px; font-family: 'Cascadia Code', Consolas, monospace; display: flex; flex-wrap: wrap; gap: 4px; }
  .card .schema-tag {
    display: inline-block; padding: 1px 6px; border-radius: 4px;
    background: #F1F5F9; border: 1px solid #E2E8F0; color: var(--ink-2);
    font-family: 'Cascadia Code', Consolas, monospace; font-size: 10px;
    line-height: 1.6;
  }
  .card .schema-tag.clickable { cursor: pointer; background: #EFF6FF; border-color: #BFDBFE; color: #1E40AF; }
  .card .schema-tag.clickable:hover { background: #DBEAFE; border-color: #93C5FD; color: #1E3A8A; }
  .card .schema-tag .ns { color: var(--ink-3); font-weight: 500; }
  .card .schema-tag.clickable .ns { color: #2563EB; }
  .card .persona { font-size: 11px; color: var(--ink-3); margin-top: 4px; }

  .featured-section {
    background: var(--bg-card);
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 22px;
    margin-bottom: 36px;
  }
  .featured-section h2 { font-size: 19px; margin: 0 0 14px 0; color: var(--ink); }
  .featured-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(540px, 1fr)); gap: 14px; }
  .featured-chain {
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 16px;
    background: linear-gradient(180deg, #FBFAFF, #FFFFFF);
  }
  .featured-chain .head { display: flex; justify-content: space-between; margin-bottom: 8px; gap: 10px; }
  .featured-chain .head h4 { margin: 0; font-size: 15px; font-weight: 700; color: var(--ink); }
  .featured-chain .head .svc { background: var(--ink); color: #FFF; border-radius: 6px; padding: 2px 8px; font-size: 11px; font-family: Consolas, monospace; font-weight: 600; flex-shrink: 0; height: fit-content; }
  .featured-chain .moment { font-size: 12px; color: var(--ink-2); margin: 6px 0 12px; font-style: italic; }
  .waves { display: grid; grid-template-columns: 60px 1fr; gap: 4px 10px; align-items: start; font-size: 11px; }
  .waves .lbl { font-weight: 700; color: #FFF; border-radius: 4px; padding: 4px 6px; text-align: center; }
  .waves .lbl.w1 { background: var(--w1); }
  .waves .lbl.w2 { background: var(--w2); }
  .waves .lbl.w3 { background: var(--w3); }
  .waves .body { color: var(--ink-2); padding: 4px 0; }
  .featured-chain .kpi-row { margin-top: 10px; padding-top: 10px; border-top: 1px dashed var(--line); display: flex; justify-content: space-between; align-items: center; gap: 10px; }
  .featured-chain .kpi-row .kpi { color: var(--kpi); font-weight: 700; font-size: 12px; background: #ECFDF5; border: 1px solid #A7F3D0; border-radius: 6px; padding: 4px 8px; }
  .featured-chain .kpi-row .persona { font-size: 11px; color: var(--ink-3); }

  .toc { background: var(--bg-card); border: 1px solid var(--line); border-radius: 12px; padding: 18px 22px; margin-bottom: 28px; }
  .toc h3 { margin: 0 0 12px 0; font-size: 13px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--ink-3); }
  .toc-list { display: grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap: 8px 14px; }
  .toc-list a { color: var(--ink-2); text-decoration: none; font-size: 13px; padding: 6px 10px; border: 1px solid var(--line); border-left: 3px solid var(--svc-color, var(--accent)); border-radius: 6px; display: flex; justify-content: space-between; align-items: center; gap: 8px; }
  .toc-list a:hover { background: var(--bg); }
  .toc-list a .lbl { font-weight: 600; }
  .toc-list a .code { font-family: Consolas, monospace; font-size: 11px; color: var(--ink-3); margin-right: 6px; }
  .toc-list a .ct { color: var(--ink-3); font-size: 11px; font-weight: 600; }

  footer { text-align: center; color: var(--ink-3); padding: 30px 20px; font-size: 12px; border-top: 1px solid var(--line); background: var(--bg-card); }
  .hidden { display: none !important; }

  /* ----- Card click affordance + modal narrator ------------------------- */
  .card { cursor: pointer; }
  .card:focus { outline: 2px solid var(--accent); outline-offset: 2px; }

  .modal-backdrop {
    position: fixed; inset: 0;
    background: rgba(15, 23, 42, .55);
    z-index: 200;
    display: flex; align-items: center; justify-content: center;
    padding: 24px;
    backdrop-filter: blur(2px);
  }
  .modal-backdrop[hidden] { display: none; }
  .modal {
    background: #FFF;
    border-radius: 14px;
    box-shadow: 0 20px 60px rgba(15, 23, 42, .35);
    max-width: 720px;
    width: 100%;
    max-height: calc(100vh - 48px);
    overflow-y: auto;
    padding: 28px 32px 24px;
    position: relative;
    animation: modal-in .18s ease-out;
  }
  @keyframes modal-in {
    from { transform: translateY(8px); opacity: 0; }
    to   { transform: translateY(0);   opacity: 1; }
  }
  .modal-close {
    position: absolute; top: 12px; right: 14px;
    background: transparent; border: none;
    font-size: 24px; line-height: 1;
    color: var(--ink-3); cursor: pointer;
    width: 32px; height: 32px; border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
  }
  .modal-close:hover { background: var(--bg); color: var(--ink); }
  .modal-tags {
    display: flex; gap: 8px; flex-wrap: wrap; align-items: center;
    margin-bottom: 8px;
  }
  .modal-tags .modal-id {
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 11px; color: var(--ink-2);
    background: var(--bg); border: 1px solid var(--line);
    padding: 2px 8px; border-radius: 999px;
  }
  .modal-tags .modal-svc {
    font-size: 11px; color: var(--accent); font-weight: 700;
    background: #F5F3FF; border: 1px solid #DDD6FE;
    padding: 2px 10px; border-radius: 999px;
  }
  .modal-tags .modal-feat {
    font-size: 11px; color: #92400E; font-weight: 700;
    background: #FEF3C7; border: 1px solid #FDE68A;
    padding: 2px 10px; border-radius: 999px;
  }
  .modal h2 {
    font-size: 22px; line-height: 1.3; margin: 4px 0 14px;
    color: var(--ink); font-weight: 700;
  }
  .modal-meta {
    display: flex; flex-wrap: wrap; gap: 16px;
    font-size: 12px; color: var(--ink-2);
    padding: 10px 14px; background: var(--bg); border-radius: 8px;
    margin-bottom: 18px;
  }
  .modal-meta b { color: var(--ink-3); font-weight: 600; margin-right: 4px; }
  .modal h3 {
    font-size: 12px; text-transform: uppercase; letter-spacing: .06em;
    color: var(--ink-3); font-weight: 700; margin: 14px 0 6px;
  }
  .modal p { font-size: 14px; line-height: 1.55; color: var(--ink); margin: 0 0 10px; }
  .modal-kpi p {
    background: #ECFDF5; border-left: 3px solid #10B981;
    padding: 10px 14px; border-radius: 6px; color: #064E3B;
  }
  .modal-narrative {
    background: linear-gradient(180deg, #FAFAF9 0%, #FFFFFF 100%);
    border-left: 3px solid var(--accent);
    border-radius: 8px;
    padding: 18px 22px 14px;
    margin-bottom: 14px;
  }
  .modal-narrative p {
    font-size: 14.5px; line-height: 1.7;
    color: var(--ink); margin: 0 0 12px;
  }
  .modal-narrative p:last-child { margin-bottom: 0; font-style: italic; color: var(--ink-2); font-size: 13px; }
  .modal-narrative p:first-child::first-letter {
    font-size: 36px; font-weight: 800; color: var(--accent);
    float: left; line-height: .95;
    padding: 2px 8px 0 0;
  }
  .modal-rawfacts {
    margin: 8px 0 4px; font-size: 12px;
  }
  .modal-rawfacts summary {
    cursor: pointer; color: var(--ink-3); font-weight: 600;
    padding: 6px 0; user-select: none;
  }
  .modal-rawfacts summary:hover { color: var(--accent); }
  .modal-rawfacts[open] summary { color: var(--accent); }
  .rawfacts-grid {
    background: var(--bg); border-radius: 6px;
    padding: 10px 14px; margin-top: 4px;
    display: flex; flex-direction: column; gap: 8px;
    color: var(--ink-2); line-height: 1.55;
  }
  .rawfacts-grid b { color: var(--ink-3); font-weight: 700; margin-right: 4px; }
  .rawfacts-grid .rf-kpi b { color: #047857; }

  .player {
    margin-top: 20px;
    background: linear-gradient(135deg, #1F2937 0%, #0F172A 100%);
    color: #F9FAFB;
    border-radius: 12px;
    padding: 14px 18px;
  }
  .player-head {
    display: flex; justify-content: space-between; align-items: center;
    font-size: 11px; margin-bottom: 10px;
  }
  .player-voice { color: #C7D2FE; }
  .player-voice b { color: #FFF; font-weight: 600; }
  .player-state {
    text-transform: uppercase; letter-spacing: .08em; font-weight: 700;
    color: #94A3B8; font-size: 10px;
    padding: 2px 10px; border-radius: 999px; background: rgba(148, 163, 184, .15);
  }
  .player-state.speaking { color: #6EE7B7; background: rgba(110, 231, 183, .15); }
  .player-state.paused   { color: #FCD34D; background: rgba(252, 211, 77, .15); }
  .player-controls {
    display: flex; gap: 10px; align-items: center;
  }
  .play-btn {
    background: rgba(255, 255, 255, .08);
    border: 1px solid rgba(255, 255, 255, .15);
    color: #FFF;
    width: 44px; height: 44px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    cursor: pointer;
    transition: background .12s, transform .12s;
  }
  .play-btn:hover:not(:disabled) { background: rgba(255, 255, 255, .18); transform: scale(1.05); }
  .play-btn:active:not(:disabled) { transform: scale(.95); }
  .play-btn:disabled { opacity: .4; cursor: not-allowed; }
  .play-btn.primary { background: var(--accent); border-color: var(--accent); width: 52px; height: 52px; }
  .play-btn.primary:hover:not(:disabled) { background: var(--accent-2); border-color: var(--accent-2); }
  .play-btn svg { width: 18px; height: 18px; fill: currentColor; }
  .play-btn.primary svg { width: 22px; height: 22px; }
  .player-rate {
    margin-left: auto; font-size: 11px; color: #94A3B8;
    display: flex; align-items: center; gap: 8px;
  }
  .player-rate input[type="range"] { width: 90px; accent-color: var(--accent); }

  /* ----- Schema entity tree-view modal --------------------------------- */
  .schema-modal {
    background: #FFF;
    border-radius: 14px;
    box-shadow: 0 20px 60px rgba(15, 23, 42, .35);
    max-width: 820px;
    width: 100%;
    max-height: calc(100vh - 48px);
    overflow-y: auto;
    padding: 26px 30px 24px;
    position: relative;
    animation: modal-in .18s ease-out;
  }
  .schema-modal .modal-close {
    position: absolute; top: 12px; right: 14px;
    background: transparent; border: none;
    font-size: 24px; line-height: 1;
    color: var(--ink-3); cursor: pointer;
    width: 32px; height: 32px; border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
  }
  .schema-modal .modal-close:hover { background: var(--bg); color: var(--ink); }
  .schema-modal .sm-tags {
    display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin-bottom: 8px;
  }
  .schema-modal .sm-ns {
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 11px; font-weight: 700;
    background: #F0F9FF; color: #0369A1; border: 1px solid #BAE6FD;
    padding: 2px 10px; border-radius: 999px;
  }
  .schema-modal .sm-domain {
    font-size: 11px; color: var(--ink-3);
    background: var(--bg); border: 1px solid var(--line);
    padding: 2px 10px; border-radius: 999px;
  }
  .schema-modal .sm-layer {
    font-size: 11px; font-weight: 700;
    padding: 2px 10px; border-radius: 999px;
  }
  .schema-modal .sm-layer.bronze { background: #FEF3C7; color: #92400E; border: 1px solid #FDE68A; }
  .schema-modal .sm-layer.silver { background: #F1F5F9; color: #334155; border: 1px solid #CBD5E1; }
  .schema-modal .sm-layer.gold   { background: #FEF9C3; color: #854D0E; border: 1px solid #FDE047; }
  .schema-modal h2 {
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 22px; line-height: 1.3; margin: 4px 0 10px;
    color: var(--ink); font-weight: 700; letter-spacing: -0.01em;
  }
  .schema-modal .sm-desc {
    font-size: 13.5px; line-height: 1.6; color: var(--ink-2);
    margin-bottom: 14px;
  }
  .schema-modal .sm-pk {
    background: linear-gradient(180deg, #FEF9C3, #FFFFFF);
    border: 1px solid #FDE68A; border-radius: 8px;
    padding: 10px 14px; margin-bottom: 14px;
    font-size: 12px; color: #713F12;
  }
  .schema-modal .sm-pk b { color: #854D0E; text-transform: uppercase; letter-spacing: 0.06em; font-size: 10px; margin-right: 8px; }
  .schema-modal .sm-pk code {
    font-family: 'Cascadia Code', Consolas, monospace; font-size: 12px;
    background: #FEF9C3; padding: 2px 6px; border-radius: 4px;
    border: 1px solid #FDE68A; color: #713F12; margin-right: 6px;
  }
  .schema-modal h3.sm-section {
    font-size: 11px; text-transform: uppercase; letter-spacing: .08em;
    color: var(--ink-3); font-weight: 700; margin: 16px 0 8px;
    padding-bottom: 4px; border-bottom: 1px solid var(--line);
  }

  .sm-tree { display: flex; flex-direction: column; gap: 2px; }
  .sm-node {
    border: 1px solid transparent;
    border-radius: 6px;
    padding: 0;
  }
  .sm-node-row {
    display: grid;
    grid-template-columns: 18px minmax(160px, 1.4fr) 110px auto;
    gap: 10px;
    align-items: baseline;
    padding: 7px 10px;
    border-radius: 6px;
    cursor: pointer;
    transition: background .12s;
  }
  .sm-node-row:hover { background: #F8FAFC; }
  .sm-node.expanded > .sm-node-row { background: #F1F5F9; border-bottom-left-radius: 0; border-bottom-right-radius: 0; }
  .sm-tw {
    color: var(--ink-3); font-size: 11px; user-select: none;
    transition: transform .12s ease;
  }
  .sm-node.expanded > .sm-node-row .sm-tw { transform: rotate(90deg); color: var(--accent); }
  .sm-name {
    font-family: 'Cascadia Code', Consolas, monospace;
    font-size: 12.5px; color: var(--ink); font-weight: 600;
    word-break: break-word;
  }
  .sm-name.is-key { color: #854D0E; }
  .sm-type {
    font-family: 'Cascadia Code', Consolas, monospace;
    font-size: 11px; color: #475569;
    background: #F1F5F9; border: 1px solid #E2E8F0; border-radius: 4px;
    padding: 1px 7px; justify-self: start;
  }
  .sm-badges { display: flex; flex-wrap: wrap; gap: 4px; justify-self: end; }
  .sm-badge {
    font-size: 10px; font-weight: 700;
    padding: 1px 7px; border-radius: 4px;
    text-transform: uppercase; letter-spacing: 0.05em;
  }
  .sm-badge.pk    { background: #FEF9C3; color: #713F12; border: 1px solid #FDE68A; }
  .sm-badge.req   { background: #FEE2E2; color: #991B1B; border: 1px solid #FECACA; }
  .sm-badge.null  { background: #F1F5F9; color: #475569; border: 1px solid #CBD5E1; }
  .sm-badge.pii   { background: #FCE7F3; color: #9D174D; border: 1px solid #FBCFE8; }
  .sm-badge.pci   { background: #FEE2E2; color: #991B1B; border: 1px solid #FECACA; }
  .sm-badge.fin   { background: #DBEAFE; color: #1E3A8A; border: 1px solid #93C5FD; }
  .sm-badge.regul { background: #E0E7FF; color: #3730A3; border: 1px solid #C7D2FE; }
  .sm-badge.audit { background: #ECFDF5; color: #065F46; border: 1px solid #A7F3D0; }
  .sm-detail {
    border: 1px solid var(--line); border-top: none;
    background: #FFFFFF;
    border-bottom-left-radius: 6px; border-bottom-right-radius: 6px;
    padding: 10px 14px 12px 38px;
    margin-bottom: 4px;
    font-size: 12.5px; line-height: 1.55; color: var(--ink-2);
  }
  .sm-detail-desc { margin-bottom: 6px; }
  .sm-detail-enum {
    display: flex; flex-wrap: wrap; gap: 4px; margin-top: 4px;
    align-items: baseline;
  }
  .sm-detail-enum .lbl {
    font-size: 10px; text-transform: uppercase; letter-spacing: 0.06em;
    color: var(--ink-3); font-weight: 700; margin-right: 4px;
  }
  .sm-detail-enum code {
    font-family: 'Cascadia Code', Consolas, monospace; font-size: 11px;
    background: #F0F9FF; color: #0369A1;
    border: 1px solid #BAE6FD; border-radius: 4px;
    padding: 1px 6px;
  }
  .sm-group-label {
    font-family: 'Cascadia Code', Consolas, monospace;
    font-size: 11px; color: var(--ink-3); font-weight: 600;
    margin: 14px 0 4px; padding-left: 4px;
  }
  .sm-empty {
    background: var(--bg); border: 1px dashed var(--line);
    border-radius: 8px; padding: 14px; font-size: 12.5px;
    color: var(--ink-3); text-align: center;
  }
  .sm-actions {
    display: flex; justify-content: space-between; align-items: center;
    margin-top: 16px; padding-top: 12px; border-top: 1px solid var(--line);
    font-size: 11px; color: var(--ink-3);
  }
  .sm-expand-toggle {
    background: transparent; border: 1px solid var(--line);
    color: var(--ink-2); border-radius: 6px;
    padding: 4px 10px; font-size: 11px; cursor: pointer;
    font-weight: 600;
  }
  .sm-expand-toggle:hover { background: var(--bg); color: var(--ink); }

  body.modal-open { overflow: hidden; }
</style>
</head>
<body>
""")

# Header
ap(f"""<header>
  <h1>APEX · Retail &amp; Consumer Scenario Chains — sorted by Service</h1>
  <div class="sub">Industry-specific catalog of agentic scenarios, organized by the 15 Deloitte RC service codes (RC-E2E-01 → RC-E2E-15). Every service carries a description of what it covers, the business outcome it drives, and how it composes with adjacent services. Domain and sub-area context is preserved on every card. Modeled on <code>APEX-RC-Scenario-Chains.xlsx</code>.</div>
  <div class="stats">
    <div class="stat"><div class="v">{total_scenarios}</div><div class="l">Total scenarios</div></div>
    <div class="stat"><div class="v">{total_services}</div><div class="l">Services (RC-E2E-XX)</div></div>
    <div class="stat"><div class="v">{total_subareas}</div><div class="l">Service × sub-area pairs</div></div>
    <div class="stat"><div class="v">{total_domains}</div><div class="l">Domains spanned</div></div>
    <div class="stat"><div class="v">{total_featured}</div><div class="l">Featured chains</div></div>
  </div>
</header>""")

# Controls — search input + Search/Clear buttons + service / sub-area / domain filters + toggle
# Build the sub-area → list of service codes map so the sub-area dropdown can be filtered
# dynamically when the user picks a service.
sub_to_svcs: dict[str, set[str]] = {}
sub_to_count: dict[str, int] = {}
for s in scenarios:
    sub = s["subarea"]
    if not sub:
        continue
    sub_to_svcs.setdefault(sub, set()).add(s["svc_code"])
    sub_to_count[sub] = sub_to_count.get(sub, 0) + 1

ap('<div class="controls">')
ap('  <div class="search-wrap">')
ap('    <input type="search" id="q" autocomplete="off" placeholder="Search title · ID · KPI · brief · domain · sub-area · persona — press Enter or click Search">')
ap('    <div class="suggestions" id="suggestions" hidden></div>')
ap('  </div>')
ap('  <button class="btn primary" id="btn-search" type="button">Search</button>')
ap('  <button class="btn btn-clear" id="btn-clear" type="button">Clear</button>')
ap('  <select id="svc-filter"><option value="">All services</option>')
for code in by_svc:
    sd = SERVICE_DESCRIPTIONS.get(code, {"name": code})
    ap(f'    <option value="{code}">{code} — {esc(sd["name"])}</option>')
ap('  </select>')
# Sub-area filter — every option carries data-services so we can hide non-matching sub-areas
# when a service is selected. Sub-areas are listed alphabetically.
ap('  <select id="subarea-filter"><option value="">All sub-areas</option>')
for sub in sorted(sub_to_svcs.keys()):
    svcs = ",".join(sorted(sub_to_svcs[sub]))
    cnt = sub_to_count[sub]
    ap(f'    <option value="{esc(sub)}" data-services="{svcs}">{esc(sub)} ({cnt})</option>')
ap('  </select>')
# Domain filter as tertiary
domains_in = sorted({s["domain"] for s in scenarios})
ap('  <select id="domain-filter"><option value="">All domains</option>')
for d in domains_in:
    ap(f'    <option value="{esc(d)}">{esc(d)}</option>')
ap('  </select>')
ap('  <div class="toggles"><button class="toggle active" data-mode="all">All</button><button class="toggle" data-mode="featured">Featured only</button></div>')
ap('  <span class="filter-state hidden" id="filter-state">Flat view · filtered</span>')
ap('  <span id="result-count" style="margin-left:auto;font-size:12px;color:var(--ink-3);font-weight:600;"></span>')
ap('</div>')

# Main
ap('<main>')

# Results banner — shows above the flat result grid only when filtering is active.
# Tells the user exactly what their filter set is and how many cards matched, so they
# don't have to read past empty section headers to understand the result.
ap('<div class="results-banner" id="results-banner"></div>')

# TOC
ap('<div class="toc"><h3>Services</h3><div class="toc-list">')
for code in by_svc:
    sd = SERVICE_DESCRIPTIONS.get(code, {"name": code, "color": "#6648B0"})
    cnt = svc_counts[code]
    fcnt = svc_featured_counts[code]
    fmark = f' · ★ {fcnt}' if fcnt else ''
    ap(f'  <a href="#svc-{code.lower()}" style="--svc-color:{sd["color"]}"><span><span class="code">{code}</span><span class="lbl">{esc(sd["name"])}</span></span><span class="ct">{cnt}{fmark}</span></a>')
ap('</div></div>')

# Featured chains section (preserved, sorted by service)
featured_chains_sorted = sorted(featured_chains, key=lambda c: (c["svc"] or "", c["title"] or ""))
ap('<div class="featured-section">')
ap('  <h2>★ Featured Chains — Wave 1 / Wave 2 / Wave 3 detail</h2>')
ap('  <div class="featured-grid">')
for fc in featured_chains_sorted:
    ap('    <div class="featured-chain">')
    ap(f'      <div class="head"><h4>{esc(fc["title"])}</h4><span class="svc">{esc(fc["svc"])}</span></div>')
    ap(f'      <div class="moment">{esc(fc["moment"])}</div>')
    ap('      <div class="waves">')
    ap(f'        <span class="lbl w1">W1</span><div class="body">{esc(short_chain(fc["w1"]))}</div>')
    ap(f'        <span class="lbl w2">W2</span><div class="body">{esc(short_chain(fc["w2"]))}</div>')
    ap(f'        <span class="lbl w3">W3</span><div class="body">{esc(short_chain(fc["w3"]))}</div>')
    ap('      </div>')
    ap('      <div class="kpi-row">')
    ap(f'        <span class="kpi">{esc(fc["kpi"])}</span>')
    ap(f'        <span class="persona">{esc(fc["persona"])}</span>')
    ap('      </div>')
    ap('    </div>')
ap('  </div>')
ap('</div>')

# Service sections
for code, subs in by_svc.items():
    sd = SERVICE_DESCRIPTIONS.get(code, {"name": code, "tagline": "", "desc": "", "color": "#1F2024"})
    cnt = svc_counts[code]
    fcnt = svc_featured_counts[code]
    n_sub = len(subs)
    n_dom = len({s["domain"] for ss in subs.values() for s in ss})

    ap(f'<div class="service" id="svc-{code.lower()}" data-svc="{code}" style="--svc-color:{sd["color"]}">')
    ap('  <div class="service-head">')
    ap('    <div class="top">')
    ap(f'      <span class="code">{code}</span>')
    ap(f'      <h2>{esc(sd["name"])}</h2>')
    ap('      <div class="badge-row">')
    ap(f'        <span class="badge">{cnt} scenarios</span>')
    ap(f'        <span class="badge">{n_sub} sub-areas</span>')
    ap(f'        <span class="badge">{n_dom} domains</span>')
    if fcnt:
        ap(f'        <span class="badge">★ {fcnt} featured</span>')
    ap('      </div>')
    ap('    </div>')
    if sd.get("tagline"):
        ap(f'    <p class="tagline">{esc(sd["tagline"])}</p>')
    if sd.get("desc"):
        ap(f'    <p class="desc">{sd["desc"]}</p>')  # desc may contain &amp; entity, leave intact
    ap('  </div>')

    for sub, sc_list in subs.items():
        # Within a sub-area, the cards may span multiple domains (rare) — show domain on each card
        ap('    <div class="subarea">')
        ap('      <div class="subarea-head">')
        ap(f'        <h3>{esc(sub)}</h3>')
        ap(f'        <span class="pill">{len(sc_list)} scenarios</span>')
        # If all scenarios in this sub-area share one domain, show it once at the head for readability
        sub_domains = {s["domain"] for s in sc_list}
        if len(sub_domains) == 1:
            ap(f'        <span class="domain-pill">{esc(next(iter(sub_domains)))}</span>')
        ap('      </div>')
        ap('      <div class="cards">')
        for s in sc_list:
            search_text = " ".join([
                s["title"], s["id"], s["brief"], s["kpi"], s["domain"], s["subarea"],
                s["svc_code"], s["svc_name"], s["persona"]
            ]).lower()
            cls = "card featured" if s["featured"] else "card"
            ap(f'        <div class="{cls}" data-text="{esc(search_text)}" data-title="{esc(s["title"])}" data-id="{esc(s["id"])}" data-svc="{esc(s["svc_code"])}" data-svcname="{esc(s["svc_name"])}" data-domain="{esc(s["domain"])}" data-subarea="{esc(s["subarea"])}" data-persona="{esc(s["persona"])}" data-brief="{esc(s["brief"])}" data-kpi="{esc(s["kpi"])}" data-featured="{1 if s["featured"] else 0}">')
            ap('          <div class="row1">')
            ap(f'            <h4>{esc(s["title"])}</h4>')
            if s["featured"]:
                ap('            <span class="star" title="Featured chain">★</span>')
            ap('          </div>')
            ap(f'          <div class="id">{esc(s["id"])}</div>')
            ap(f'          <div class="brief">{esc(s["brief"])}</div>')
            if s["kpi"]:
                ap(f'          <div class="kpi">{esc(s["kpi"])}</div>')
            ap('          <div class="meta">')
            if s["svc_code"]:
                ap(f'            <span class="tag">{esc(s["svc_code"])}</span>')
            if s["svc_name"]:
                ap(f'            <span class="tag">{esc(s["svc_name"])}</span>')
            if s["domain"]:
                ap(f'            <span class="tag dom">{esc(s["domain"])}</span>')
            ap('          </div>')
            if s["schemas"]:
                # Render each entity as its own chip — clickable when we have field definitions for it.
                ap('          <div class="schemas">')
                for tok in re.split(r"[·;,]", s["schemas"]):
                    tok = tok.strip()
                    if not tok:
                        continue
                    has_def = tok in SCHEMA_FIELDS
                    cls = "schema-tag clickable" if has_def else "schema-tag"
                    if "." in tok:
                        ns, _, ent = tok.partition(".")
                        inner = f'<span class="ns">{esc(ns)}.</span>{esc(ent)}'
                    else:
                        inner = esc(tok)
                    if has_def:
                        ap(f'            <span class="{cls}" data-schema="{esc(tok)}" tabindex="0" role="button" aria-label="View schema {esc(tok)}">{inner}</span>')
                    else:
                        ap(f'            <span class="{cls}" title="No field definition available">{inner}</span>')
                ap('          </div>')
            if s["persona"]:
                ap(f'          <div class="persona">👤 {esc(s["persona"])}</div>')
            ap('        </div>')
        ap('      </div>')  # /.cards
        ap('    </div>')    # /.subarea

    ap('</div>')  # /.service

ap('</main>')

ap("""<footer>
  APEX · Retail &amp; Consumer Scenario Chains — by Service · 502 scenarios · 15 services · 109 sub-areas · 35 featured chains
</footer>

<!-- Scenario narrator modal -->
<div class="modal-backdrop" id="modal-backdrop" hidden>
  <div class="modal" role="dialog" aria-modal="true" aria-labelledby="m-title">
    <button class="modal-close" id="modal-close" type="button" aria-label="Close">&times;</button>
    <div class="modal-tags">
      <span class="modal-id" id="m-id"></span>
      <span class="modal-svc" id="m-svc"></span>
      <span class="modal-feat" id="m-feat" hidden>&#9733; Featured chain</span>
    </div>
    <h2 id="m-title"></h2>
    <div class="modal-meta">
      <span><b>Sub-area</b> <span id="m-sub"></span></span>
      <span><b>Domain</b> <span id="m-dom"></span></span>
      <span><b>Persona</b> <span id="m-persona"></span></span>
    </div>
    <h3>Narrative</h3>
    <div class="modal-narrative" id="m-narrative"></div>
    <details class="modal-rawfacts">
      <summary>Source facts</summary>
      <div class="rawfacts-grid">
        <div><b>The moment.</b> <span id="m-brief"></span></div>
        <div class="rf-kpi" id="m-kpi-wrap" hidden><b>KPI / outcome.</b> <span id="m-kpi"></span></div>
      </div>
    </details>
    <div class="player">
      <div class="player-head">
        <div class="player-voice"><span id="m-voice-name">Loading voice&hellip;</span></div>
        <div class="player-state" id="m-state">Idle</div>
      </div>
      <div class="player-controls">
        <button class="play-btn primary" id="btn-play" type="button" aria-label="Play">
          <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M8 5v14l11-7z"/></svg>
        </button>
        <button class="play-btn" id="btn-pause" type="button" aria-label="Pause" disabled>
          <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M6 5h4v14H6zM14 5h4v14h-4z"/></svg>
        </button>
        <button class="play-btn" id="btn-stop" type="button" aria-label="Stop" disabled>
          <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M6 6h12v12H6z"/></svg>
        </button>
        <div class="player-rate">
          <span>Speed</span>
          <input type="range" id="m-rate" min="0.7" max="1.4" step="0.05" value="0.95" aria-label="Playback speed">
          <span id="m-rate-val">0.95x</span>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- Schema entity tree-view modal -->
<div class="modal-backdrop" id="schema-backdrop" hidden>
  <div class="schema-modal" role="dialog" aria-modal="true" aria-labelledby="sm-title">
    <button class="modal-close" id="schema-close" type="button" aria-label="Close">&times;</button>
    <div class="sm-tags">
      <span class="sm-ns" id="sm-ns"></span>
      <span class="sm-domain" id="sm-domain"></span>
      <span class="sm-layer" id="sm-layer"></span>
    </div>
    <h2 id="sm-title"></h2>
    <div class="sm-desc" id="sm-desc"></div>
    <div class="sm-pk" id="sm-pk"></div>
    <h3 class="sm-section">Fields</h3>
    <div class="sm-tree" id="sm-tree"></div>
    <div class="sm-actions">
      <span id="sm-count"></span>
      <button class="sm-expand-toggle" id="sm-expand-all" type="button">Expand all</button>
    </div>
  </div>
</div>""")

# JavaScript: search button + Enter, service / sub-area / domain filters, featured toggle,
# and flat-view rendering whenever any filter is active.
# First, embed schema field data as window-level globals so the main IIFE can read them.
ap("<script>")
ap("window.SCHEMA_FIELDS = " + json.dumps(SCHEMA_FIELDS, ensure_ascii=False) + ";")
ap("window.AUDIT_TAIL = "   + json.dumps(AUDIT_TAIL,   ensure_ascii=False) + ";")
ap("</script>")
ap("""<script>
(function () {
  const SCHEMA_FIELDS = window.SCHEMA_FIELDS || {};
  const AUDIT_TAIL    = window.AUDIT_TAIL    || [];
  const q          = document.getElementById('q');
  const btnSearch  = document.getElementById('btn-search');
  const btnClear   = document.getElementById('btn-clear');
  const svcF       = document.getElementById('svc-filter');
  const subF       = document.getElementById('subarea-filter');
  const domF       = document.getElementById('domain-filter');
  const toggles    = document.querySelectorAll('.toggle');
  const counter    = document.getElementById('result-count');
  const banner     = document.getElementById('results-banner');
  const stateChip  = document.getElementById('filter-state');
  let mode = 'all';
  let activeNeedle = '';   // only updated when user clicks Search / presses Enter

  function norm(s) { return (s || '').toLowerCase(); }

  function fuzzy(needle, hay) {
    if (!needle) return true;
    needle = norm(needle); hay = norm(hay);
    if (hay.indexOf(needle) >= 0) return true;
    // Loose char-order match for typo tolerance
    let i = 0;
    for (let c of hay) { if (c === needle[i]) i++; if (i === needle.length) return true; }
    return false;
  }

  // Stairstep cascade across all filters: every dropdown is mutually constrained
  // by every OTHER active filter, so impossible combinations are removed from the
  // pickers themselves. We compute, for each filter, the set of values that would
  // still yield at least one card if that filter were the next one chosen — given
  // everything else currently selected. Options outside that set get hidden, and
  // any selection that becomes invalid is reset. Iterate to fixed point because
  // resetting one selection can re-validate options on another dropdown.
  function cardMatchesExcept(card, excludeKey) {
    const svcSel = svcF.value, subSel = subF.value, domSel = domF.value;
    if (excludeKey !== 'svc'    && svcSel    && card.dataset.svc !== svcSel)            return false;
    if (excludeKey !== 'sub'    && subSel    && card.dataset.subarea !== subSel)        return false;
    if (excludeKey !== 'dom'    && domSel    && card.dataset.domain !== domSel)         return false;
    if (excludeKey !== 'feat'   && mode === 'featured' && card.dataset.featured !== '1') return false;
    if (excludeKey !== 'search' && activeNeedle && !fuzzy(activeNeedle, card.dataset.text)) return false;
    return true;
  }

  const ALL_CARDS = document.querySelectorAll('.card');

  function refreshDropdowns() {
    let changed = true, iter = 0;
    while (changed && iter < 5) {
      changed = false; iter++;
      const validSvc = new Set(), validSub = new Set(), validDom = new Set();
      ALL_CARDS.forEach(card => {
        if (cardMatchesExcept(card, 'svc')) validSvc.add(card.dataset.svc);
        if (cardMatchesExcept(card, 'sub')) validSub.add(card.dataset.subarea);
        if (cardMatchesExcept(card, 'dom')) validDom.add(card.dataset.domain);
      });
      if (svcF.value && !validSvc.has(svcF.value)) { svcF.value = ''; changed = true; }
      if (subF.value && !validSub.has(subF.value)) { subF.value = ''; changed = true; }
      if (domF.value && !validDom.has(domF.value)) { domF.value = ''; changed = true; }
      if (!changed) {
        // Hide options that wouldn't match anything under current sibling filters
        svcF.querySelectorAll('option').forEach(o => { if (o.value) o.hidden = !validSvc.has(o.value); });
        subF.querySelectorAll('option').forEach(o => { if (o.value) o.hidden = !validSub.has(o.value); });
        domF.querySelectorAll('option').forEach(o => { if (o.value) o.hidden = !validDom.has(o.value); });
      }
    }
  }

  function applyAndRefresh() {
    refreshDropdowns();
    apply();
  }

  function isFiltering() {
    return !!(activeNeedle || svcF.value || subF.value || domF.value || mode === 'featured');
  }

  function apply() {
    const needle = activeNeedle;
    const svcSel = svcF.value;
    const subSel = subF.value;
    const domSel = domF.value;
    let visible = 0;

    document.querySelectorAll('.service').forEach(sec => {
      let secVisible = false;
      const secSvc = sec.dataset.svc;
      // Service-level fast-skip only when not flattening (in flat view we still need
      // each card to be evaluated, but display:contents collapses the boxes)
      sec.querySelectorAll('.subarea').forEach(sub => {
        let subVisible = false;
        sub.querySelectorAll('.card').forEach(card => {
          const matchSvc  = !svcSel || card.dataset.svc === svcSel;
          const matchSub  = !subSel || card.dataset.subarea === subSel;
          const matchDom  = !domSel || card.dataset.domain === domSel;
          const matchTxt  = fuzzy(needle, card.dataset.text);
          const matchFeat = mode === 'all' || card.dataset.featured === '1';
          if (matchSvc && matchSub && matchDom && matchTxt && matchFeat) {
            card.classList.remove('hidden');
            subVisible = true;
            secVisible = true;
            visible++;
          } else {
            card.classList.add('hidden');
          }
        });
        sub.classList.toggle('hidden', !subVisible);
      });
      sec.classList.toggle('hidden', !secVisible);
    });

    counter.textContent = visible + ' scenarios';

    // Toggle flat view + render the results banner whenever a filter is active.
    const filtering = isFiltering();
    document.body.classList.toggle('filtering', filtering);
    stateChip.classList.toggle('hidden', !filtering);
    if (filtering) {
      const parts = [];
      if (svcSel) parts.push('Service <b>' + svcSel + '</b>');
      if (subSel) parts.push('Sub-area <b>' + escapeHtml(subSel) + '</b>');
      if (domSel) parts.push('Domain <b>' + escapeHtml(domSel) + '</b>');
      if (mode === 'featured') parts.push('<b>Featured only</b>');
      if (needle) parts.push('Search <b>"' + escapeHtml(needle) + '"</b>');
      banner.innerHTML =
        '<b>' + visible + '</b> scenarios match · ' +
        (parts.length ? parts.join(' · ') : 'no filters') +
        ' &nbsp;·&nbsp; <a href="#" id="banner-clear" style="color:var(--accent);text-decoration:none;font-weight:600;">Clear all</a>';
      const bclear = document.getElementById('banner-clear');
      if (bclear) bclear.addEventListener('click', (e) => { e.preventDefault(); clearAll(); });
    } else {
      banner.innerHTML = '';
    }
  }

  function escapeHtml(s) {
    return String(s).replace(/[&<>"']/g, c => ({
      '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'
    }[c]));
  }

  function runSearch() {
    activeNeedle = q.value.trim();
    applyAndRefresh();
  }

  function clearAll() {
    q.value = '';
    activeNeedle = '';
    svcF.value = '';
    subF.value = '';
    domF.value = '';
    toggles.forEach(x => x.classList.remove('active'));
    document.querySelector('.toggle[data-mode="all"]').classList.add('active');
    mode = 'all';
    applyAndRefresh();
    window.scrollTo({ top: 0, behavior: 'smooth' });
  }

  // ---- Autocomplete suggestions popdown ----------------------------------
  const sugg = document.getElementById('suggestions');
  const MAX_SUGG = 8;
  let suggIdx = -1;

  function selectSugg(el) {
    if (!el) return;
    q.value = el.dataset.title;
    activeNeedle = el.dataset.title.trim();
    closeSugg();
    applyAndRefresh();
  }

  function closeSugg() {
    sugg.hidden = true;
    sugg.innerHTML = '';
    suggIdx = -1;
  }

  function highlightSugg(idx) {
    const items = sugg.querySelectorAll('.sugg-item');
    items.forEach((el, i) => el.classList.toggle('active', i === idx));
    suggIdx = idx;
    if (idx >= 0 && items[idx]) items[idx].scrollIntoView({ block: 'nearest' });
  }

  function highlightMatch(s, needle) {
    const i = s.toLowerCase().indexOf(needle.toLowerCase());
    if (i < 0) return escapeHtml(s);
    return escapeHtml(s.slice(0, i)) +
           '<mark>' + escapeHtml(s.slice(i, i + needle.length)) + '</mark>' +
           escapeHtml(s.slice(i + needle.length));
  }

  function renderSuggestions(needle) {
    if (!needle || needle.length < 2) { closeSugg(); return; }
    const n = needle.toLowerCase();
    const prefix = [], contains = [], body = [], fuzzyHits = [];
    ALL_CARDS.forEach(card => {
      // Only suggest cards that match the OTHER active filters (svc / sub / dom / featured)
      if (!cardMatchesExcept(card, 'search')) return;
      const t = (card.dataset.title || '').toLowerCase();
      const text = card.dataset.text || '';
      if (t.startsWith(n))           prefix.push(card);
      else if (t.indexOf(n) >= 0)    contains.push(card);
      else if (text.indexOf(n) >= 0) body.push(card);
      else if (fuzzy(n, text))       fuzzyHits.push(card);
    });
    const items = prefix.concat(contains, body, fuzzyHits).slice(0, MAX_SUGG);
    if (!items.length) {
      sugg.innerHTML = '<div class="sugg-empty">No scenarios match &ldquo;' + escapeHtml(needle) + '&rdquo;</div>';
      sugg.hidden = false;
      suggIdx = -1;
      return;
    }
    sugg.innerHTML = items.map((c) =>
      '<div class="sugg-item" data-title="' + escapeHtml(c.dataset.title) + '">' +
        '<div class="sugg-title">' + highlightMatch(c.dataset.title, needle) + '</div>' +
        '<div class="sugg-meta">' +
          '<span class="sugg-id">' + escapeHtml(c.dataset.id || '') + '</span>' +
          ' &middot; <span class="sugg-svc">' + escapeHtml(c.dataset.svc) + '</span>' +
          ' &middot; <span class="sugg-sub">' + escapeHtml(c.dataset.subarea) + '</span>' +
        '</div>' +
      '</div>'
    ).join('');
    sugg.hidden = false;
    suggIdx = -1;
    sugg.querySelectorAll('.sugg-item').forEach((el) => {
      // mousedown (not click) so the input's blur doesn't fire first and close us
      el.addEventListener('mousedown', (e) => { e.preventDefault(); selectSugg(el); });
      el.addEventListener('mouseenter', () => {
        const all = sugg.querySelectorAll('.sugg-item');
        const idx = Array.prototype.indexOf.call(all, el);
        highlightSugg(idx);
      });
    });
  }

  // Wire up events — every change runs the cascade so dropdowns mutually constrain
  btnSearch.addEventListener('click', runSearch);
  btnClear.addEventListener('click', clearAll);
  q.addEventListener('input', () => renderSuggestions(q.value.trim()));
  q.addEventListener('focus', () => { if (q.value.trim().length >= 2) renderSuggestions(q.value.trim()); });
  q.addEventListener('blur', () => { setTimeout(closeSugg, 150); });
  q.addEventListener('keydown', (e) => {
    const open = !sugg.hidden;
    const items = open ? sugg.querySelectorAll('.sugg-item') : [];
    if (open && e.key === 'ArrowDown') { e.preventDefault(); highlightSugg(Math.min(suggIdx + 1, items.length - 1)); return; }
    if (open && e.key === 'ArrowUp')   { e.preventDefault(); highlightSugg(Math.max(suggIdx - 1, 0)); return; }
    if (e.key === 'Enter') {
      e.preventDefault();
      if (open && suggIdx >= 0 && items[suggIdx]) selectSugg(items[suggIdx]);
      else { closeSugg(); runSearch(); }
      return;
    }
    if (e.key === 'Escape') {
      if (open) { closeSugg(); }
      else { clearAll(); }
    }
  });
  svcF.addEventListener('change', applyAndRefresh);
  subF.addEventListener('change', applyAndRefresh);
  domF.addEventListener('change', applyAndRefresh);
  toggles.forEach(t => t.addEventListener('click', () => {
    toggles.forEach(x => x.classList.remove('active'));
    t.classList.add('active');
    mode = t.dataset.mode;
    applyAndRefresh();
  }));

  applyAndRefresh();

  // ---- Modal narrator (Brian-voice TTS) ----------------------------------
  const backdrop = document.getElementById('modal-backdrop');
  const mClose   = document.getElementById('modal-close');
  const mTitle   = document.getElementById('m-title');
  const mId      = document.getElementById('m-id');
  const mSvc     = document.getElementById('m-svc');
  const mFeat    = document.getElementById('m-feat');
  const mSub     = document.getElementById('m-sub');
  const mDom     = document.getElementById('m-dom');
  const mPersona = document.getElementById('m-persona');
  const mBrief   = document.getElementById('m-brief');
  const mKpiWrap = document.getElementById('m-kpi-wrap');
  const mKpi     = document.getElementById('m-kpi');
  const mNarr    = document.getElementById('m-narrative');
  const mVoice   = document.getElementById('m-voice-name');
  const mState   = document.getElementById('m-state');
  const btnPlay  = document.getElementById('btn-play');
  const btnPause = document.getElementById('btn-pause');
  const btnStop  = document.getElementById('btn-stop');
  const rateInp  = document.getElementById('m-rate');

  let chosenVoice = null;
  let utterance   = null;
  let currentText = '';
  let currentCard = null;
  const synth = window.speechSynthesis;

  function pickVoice() {
    if (!synth) return;
    const voices = synth.getVoices();
    if (!voices.length) return;
    // Prefer Amazon-Polly-style "Brian" if it's been registered; otherwise UK English males.
    const prefs = [
      v => /brian/i.test(v.name),
      v => /daniel/i.test(v.name) && /en[-_]GB/i.test(v.lang),
      v => /george/i.test(v.name) && /en[-_]GB/i.test(v.lang),
      v => /ryan/i.test(v.name)   && /en[-_]GB/i.test(v.lang),
      v => /arthur/i.test(v.name) && /en[-_]GB/i.test(v.lang),
      v => /\bmale\b/i.test(v.name) && /en[-_]GB/i.test(v.lang),
      v => /en[-_]GB/i.test(v.lang),
      v => /^en/i.test(v.lang),
    ];
    for (const test of prefs) {
      const found = voices.find(test);
      if (found) { chosenVoice = found; break; }
    }
    if (!chosenVoice) chosenVoice = voices[0];
    if (mVoice) mVoice.innerHTML = 'Voice: <b>' + chosenVoice.name + '</b> &middot; ' + chosenVoice.lang;
  }
  if (synth) {
    pickVoice();
    synth.onvoiceschanged = pickVoice;
  } else if (mVoice) {
    mVoice.textContent = 'Speech synthesis not supported in this browser';
  }

  // ---- Verbose Brian-voice narrative engine ------------------------------
  // Each service gets a story-suitable framing: what the service does, the verb
  // that describes how it engages a trigger, and the kind of pain it spares
  // the operator from. The narrative engine weaves these with the scenario's
  // brief, KPI, and persona to produce a 4-5 paragraph walkthrough that reads
  // like a Brian-voice walkthrough — not a recital of facts.
  const SERVICE_NARRATIVE = {
    "RC-E2E-01": { short: "Demand Sensing & Planning", purpose: "translates leading consumer-intent signals into sellable units by SKU, store, and week — so every downstream plan operates on the same governed forecast", verb: "reads the demand signal", pain: "stale forecasts and surprise stock-outs that everyone saw coming" },
    "RC-E2E-02": { short: "Inventory & Replenishment", purpose: "manages day-by-day position across stores, DCs, and dark stores, under guardrails for open-to-buy, lead time, and on-shelf availability", verb: "reads the position and the shock", pain: "empty shelves on one hand and bloated working capital on the other" },
    "RC-E2E-03": { short: "Assortment & Pricing", purpose: "owns base price, promotion, markdown, clearance, and competitive response — tying the merchandising plan to the financial plan with shared guardrails on margin and open-to-buy", verb: "reads the price-and-mix signal", pain: "leaving margin on the table while the competitor moves first" },
    "RC-E2E-04": { short: "Customer Lifecycle & Loyalty", purpose: "carries every customer-facing touchpoint where a personalized agentic decision moves revenue — site search, recommendations, campaigns, loyalty, retention, winback", verb: "reads the customer signal", pain: "leaking lifetime value one disappointing journey at a time" },
    "RC-E2E-05": { short: "Store Operations", purpose: "turns merchandising and assortment intent into executed shelf realities through computer-vision audits, task orchestration, and on-shelf-availability monitoring", verb: "reads the shelf signal", pain: "the gap between planogram intent and what the customer actually sees" },
    "RC-E2E-06": { short: "Supply Chain & Logistics", purpose: "connects upstream supplier ecosystems to downstream fulfillment surfaces — owning routing, carrier scoring, DC throughput, last-mile promise, and reverse logistics", verb: "reads the disruption", pain: "a port event, a weather event, a supplier slip — and the recovery is always manual" },
    "RC-E2E-07": { short: "Returns & Refund Integrity", purpose: "defends the trust perimeter — returns triage, refund integrity, fraud, organized retail crime, account takeover, shrink, and salvage", verb: "reads the leakage signal", pain: "revenue walking out the door, intentionally or otherwise" },
    "RC-E2E-08": { short: "Workforce & Talent", purpose: "runs people operations across HQ, field, store, plant, and restaurant — labor planning, recruiting, onboarding, attrition risk, engagement", verb: "reads the labor signal", pain: "a schedule that doesn't match the demand and talent that walks before week six" },
    "RC-E2E-09": { short: "Product Tracking & Provenance", purpose: "powers regulator-grade visibility — cold-chain lot tracing, recall readiness, product authentication, and provenance certification", verb: "reads the provenance signal", pain: "an over-broad recall on one side, a regulator finding on the other" },
    "RC-E2E-10": { short: "Trade Promotion & RGM", purpose: "translates manufacturer revenue-growth strategy into executed trade plans — trade-spend effectiveness, deduction recovery, price-pack architecture, joint business planning", verb: "reads the trade signal", pain: "trade dollars that don't lift, deductions that don't get recovered" },
    "RC-E2E-11": { short: "Manufacturing & Plant Ops", purpose: "turns plant-floor telemetry into uptime, yield, and safety outcomes — overall equipment effectiveness, predictive maintenance, quality, recipe and spec management, and plant sustainability", verb: "reads the line signal", pain: "an unplanned stop that costs a shift and a quality miss that costs a brand" },
    "RC-E2E-12": { short: "Foodservice & Restaurant Ops", purpose: "runs the unique mechanics of QSR and casual dining — menu engineering, kitchen throughput, drive-thru, off-premise, minute-grain labor demand", verb: "reads the restaurant signal", pain: "a drive-thru line that grows past three minutes and an inventory that doesn't match the lunch rush" },
    "RC-E2E-13": { short: "Sustainability & ESG", purpose: "produces auditable agentic decision trails for every ESG number — Scope-3 emissions, packaging, water, energy, supplier ESG scoring, food-waste reduction", verb: "reads the ESG signal", pain: "claims that don't survive an audit and disclosures that don't reconcile" },
    "RC-E2E-14": { short: "Real Estate & Network", purpose: "runs site, format, lease, capex, and closure decisions — every store touched is multi-year capital and brand exposure", verb: "reads the network signal", pain: "a lease renegotiated late and a closure that blindsides a community" },
    "RC-E2E-15": { short: "Finance, Risk & Controls", purpose: "runs the back-office trust spine — every dollar in and out, every regulator filing, every control, every cyber and identity event", verb: "reads the financial-control signal", pain: "a control that fired late and a reconciliation that didn't tie" }
  };

  const OPENERS = [
    "Picture this.",
    "Set the scene.",
    "Here's the situation.",
    "Walk into this Tuesday morning.",
    "Think about how this used to play out."
  ];

  const STAKES = [
    "Without the right system underneath it, the response is improvised — and improvisation at scale is how shelves go empty, how margin walks out the door, how the audit trail goes dark.",
    "In the old world, this turns into a war room — phone calls, spreadsheets, and a Monday-morning post-mortem about why nobody saw it coming.",
    "The legacy answer is a fire-drill and a hope. The auditable answer is a chain.",
    "Without an agentic spine, the cost of this moment is paid in either revenue or trust — usually both."
  ];

  const CODAS = [
    "That's the chain.",
    "That's the play.",
    "That's how the chain runs end to end.",
    "That's APEX, doing what APEX is built to do."
  ];

  function hashIdx(s, mod) {
    let h = 0;
    for (let i = 0; i < s.length; i++) { h = (h * 31 + s.charCodeAt(i)) | 0; }
    return Math.abs(h) % mod;
  }

  function articleFor(noun) {
    const n = (noun || '').trim();
    return /^[aeiouAEIOU]/.test(n) ? 'an' : 'a';
  }

  function endPunctuated(s) {
    s = (s || '').trim();
    if (!s) return '';
    return /[.!?]$/.test(s) ? s : s + '.';
  }

  function buildNarrativeParagraphs(card) {
    const title    = (card.dataset.title    || '').trim();
    const code     = (card.dataset.svc      || '').trim();
    const svcName  = (card.dataset.svcname  || '').trim();
    const sub      = (card.dataset.subarea  || '').trim();
    const dom      = (card.dataset.domain   || '').trim();
    const persona  = (card.dataset.persona  || '').trim();
    const brief    = (card.dataset.brief    || '').trim();
    const kpi      = (card.dataset.kpi      || '').trim();
    const featured = card.dataset.featured === '1';
    const id       = (card.dataset.id || '').trim();

    const svcInfo = SERVICE_NARRATIVE[code] || {
      short: svcName || 'this service area',
      purpose: 'orchestrates the right action with full evidence and approval discipline',
      verb: 'reads the signal',
      pain: 'a manual fire-drill nobody wants to run again'
    };

    const seed = id || title || svcName;
    const opener = OPENERS[hashIdx(seed, OPENERS.length)];
    const stakes = STAKES[hashIdx(seed + 'stakes', STAKES.length)];
    const coda   = CODAS[hashIdx(seed + 'coda', CODAS.length)];
    const personaPhrase = persona ? persona : 'the operator on point';
    const personaLower  = personaPhrase.toLowerCase();

    const paras = [];

    // 1. Setup — persona, world, the legacy pain
    let p1 = opener + ' ';
    if (persona) {
      p1 += "You're " + articleFor(persona) + ' ' + personaLower + '. ';
    } else {
      p1 += "You're the operator on point. ";
    }
    if (dom && sub && sub.toLowerCase() !== dom.toLowerCase()) {
      p1 += 'Your world is ' + dom.toLowerCase() + ', and the corner of it that matters today is ' + sub.toLowerCase() + '. ';
    } else if (dom) {
      p1 += 'Your world is ' + dom.toLowerCase() + '. ';
    } else if (sub) {
      p1 += 'The corner of the business that matters today is ' + sub.toLowerCase() + '. ';
    }
    p1 += "And in the old world, the moment we're about to walk through means " + svcInfo.pain + '.';
    if (featured) {
      p1 += " This one is a featured chain — the kind APEX is built to handle end to end, with the human still in the loop where it matters.";
    }
    paras.push(p1);

    // 2. The moment — quote the brief, frame it as the trigger
    let p2 = "Here's the moment. " + endPunctuated(brief || 'The trigger fires from the source signals this service watches around the clock.') + ' ';
    p2 += "That's the trigger. That's the thing that lights up the screen at the wrong time of day, that pages the on-call, that turns a Tuesday into a war room. ";
    p2 += stakes;
    paras.push(p2);

    // 3. The APEX play — what the service does
    let p3 = 'This is where ' + svcInfo.short + ' earns its keep. The service ' + svcInfo.purpose + '. ';
    p3 += 'When the trigger fires, APEX ' + svcInfo.verb + ', classifies it against the policies that govern this play, ';
    p3 += 'quantifies the financial impact in the units the business actually runs on, and proposes a specific course of action. ';
    p3 += 'It comes back to ' + personaLower + ' as a Done — Your Call — Watch brief, with the evidence already attached: ';
    p3 += 'the source signal, the policy match, the dollars at stake, the alternatives considered, and the recommended move. ';
    p3 += 'Approve, edit, or reject. The system writes the WORM audit row, fires the action, and moves on to the next signal.';
    paras.push(p3);

    // 4. Outcome — KPI as business outcome, not vanity metric
    let p4 = 'What good looks like? ';
    if (kpi) { p4 += endPunctuated(kpi) + ' '; }
    else     { p4 += 'The right move, made fast, with the evidence on file. '; }
    p4 += "That's not a vanity metric. That's " + personaLower + ' spending the day on the next ten decisions instead of relitigating this one. ';
    p4 += "That's working capital staying disciplined while the customer experience stays whole. ";
    p4 += "That's APEX doing the heavy lift so the human signs off and moves on — every time, with the same evidence trail, every chain, every play.";
    paras.push(p4);

    // 5. Coda
    paras.push('Scenario ' + (id || '—') + '. ' + endPunctuated(title) + ' ' + coda);

    return paras;
  }


  function setState(s) {
    mState.textContent = s.charAt(0).toUpperCase() + s.slice(1);
    mState.classList.remove('speaking', 'paused');
    if (s === 'speaking') mState.classList.add('speaking');
    if (s === 'paused')   mState.classList.add('paused');
    btnPlay.disabled  = (s === 'speaking');
    btnPause.disabled = (s !== 'speaking');
    btnStop.disabled  = (s === 'idle');
  }

  function play() {
    if (!synth) return;
    if (synth.paused && utterance) {
      synth.resume();
      setState('speaking');
      return;
    }
    if (synth.speaking) return;
    utterance = new SpeechSynthesisUtterance(currentText);
    if (chosenVoice) utterance.voice = chosenVoice;
    utterance.rate  = parseFloat(rateInp.value || '0.95');
    utterance.pitch = 1.0;
    utterance.onstart = () => setState('speaking');
    utterance.onend   = () => { setState('idle'); utterance = null; };
    utterance.onerror = () => { setState('idle'); utterance = null; };
    synth.speak(utterance);
    setState('speaking');
  }
  function pause() {
    if (!synth || !synth.speaking) return;
    synth.pause();
    setState('paused');
  }
  function stop() {
    if (!synth) return;
    synth.cancel();
    utterance = null;
    setState('idle');
  }

  btnPlay.addEventListener('click', play);
  btnPause.addEventListener('click', pause);
  btnStop.addEventListener('click', stop);
  rateInp.addEventListener('input', () => {
    document.getElementById('m-rate-val').textContent = parseFloat(rateInp.value).toFixed(2) + 'x';
    if (synth && synth.speaking) {
      // Restart at new rate from the beginning of the narrative
      stop();
      play();
    }
  });

  function openModal(card) {
    currentCard = card;
    const paras = buildNarrativeParagraphs(card);
    currentText = paras.join('  ');   // TTS reads the verbose narrative
    // Render multi-paragraph narrative
    mNarr.innerHTML = paras.map(p => '<p>' + escapeHtml(p) + '</p>').join('');
    mTitle.textContent   = card.dataset.title || '';
    mId.textContent      = card.dataset.id    || '';
    mSvc.textContent     = (card.dataset.svc || '') + (card.dataset.svcname ? '  -  ' + card.dataset.svcname : '');
    mSub.textContent     = card.dataset.subarea || '-';
    mDom.textContent     = card.dataset.domain  || '-';
    mPersona.textContent = card.dataset.persona || '-';
    mBrief.textContent   = card.dataset.brief   || '';
    const kpi = card.dataset.kpi || '';
    mKpiWrap.hidden = !kpi;
    mKpi.textContent = kpi;
    mFeat.hidden = card.dataset.featured !== '1';
    backdrop.hidden = false;
    document.body.classList.add('modal-open');
    setState('idle');
    utterance = null;
    // Auto-stop any speech in progress when opening a different scenario
    if (synth) synth.cancel();
    // Focus close button for keyboard users
    setTimeout(() => mClose.focus(), 50);
  }

  function closeModal() {
    stop();
    backdrop.hidden = true;
    document.body.classList.remove('modal-open');
    currentCard = null;
    currentText = '';
  }

  // Bind clicks once across all cards
  document.querySelectorAll('.card').forEach(card => {
    card.tabIndex = 0;
    card.addEventListener('click', (e) => {
      // Schema-tag clicks open the schema modal — let them through without opening the scenario modal
      if (e.target.closest('.schema-tag')) return;
      // Don't open if user is selecting text
      if (window.getSelection && window.getSelection().toString().length > 0) return;
      openModal(card);
    });
    card.addEventListener('keydown', (e) => {
      if (e.target.closest('.schema-tag')) return;
      if (e.key === 'Enter' || e.key === ' ') { e.preventDefault(); openModal(card); }
    });
  });

  mClose.addEventListener('click', closeModal);
  backdrop.addEventListener('click', (e) => { if (e.target === backdrop) closeModal(); });
  document.addEventListener('keydown', (e) => {
    if (!backdrop.hidden && e.key === 'Escape') closeModal();
  });

  // ---- Schema entity tree-view modal -----------------------------------
  // Click a schema chip on any card → open a modal that shows every field
  // for that entity, grouped (key fields, business fields, classified fields,
  // audit/lineage tail), with type chips, nullability, and enum values.
  const sBackdrop  = document.getElementById('schema-backdrop');
  const sClose     = document.getElementById('schema-close');
  const sNs        = document.getElementById('sm-ns');
  const sDomain    = document.getElementById('sm-domain');
  const sLayer     = document.getElementById('sm-layer');
  const sTitle     = document.getElementById('sm-title');
  const sDesc      = document.getElementById('sm-desc');
  const sPk        = document.getElementById('sm-pk');
  const sTree      = document.getElementById('sm-tree');
  const sCount     = document.getElementById('sm-count');
  const sExpandAll = document.getElementById('sm-expand-all');
  let sAllExpanded = false;

  function classifyBadge(c) {
    if (!c) return null;
    const k = String(c).toLowerCase();
    if (k.indexOf('pii')   >= 0) return { cls: 'pii',   label: 'PII' };
    if (k.indexOf('pci')   >= 0) return { cls: 'pci',   label: 'PCI' };
    if (k.indexOf('finan') >= 0 || k.indexOf('sox') >= 0) return { cls: 'fin', label: 'FIN' };
    if (k.indexOf('reg')   >= 0 || k.indexOf('hipaa') >= 0) return { cls: 'regul', label: 'REGUL' };
    return { cls: 'regul', label: String(c).toUpperCase().slice(0, 8) };
  }

  function fieldRow(f, opts) {
    opts = opts || {};
    const pk = !!f.key || (opts.pkSet && opts.pkSet.has(f.name));
    const nameCls = 'sm-name' + (pk ? ' is-key' : '');
    const badges = [];
    if (pk) badges.push('<span class="sm-badge pk" title="Primary key">PK</span>');
    if (!f.nullable && !pk) badges.push('<span class="sm-badge req" title="Required (NOT NULL)">REQ</span>');
    if (f.nullable) badges.push('<span class="sm-badge null" title="Nullable">NULL</span>');
    const cb = classifyBadge(f.classification);
    if (cb) badges.push('<span class="sm-badge ' + cb.cls + '" title="Classification: ' + escapeHtml(String(f.classification)) + '">' + cb.label + '</span>');
    if (opts.audit) badges.push('<span class="sm-badge audit" title="Shared audit/lineage tail">AUDIT</span>');

    const hasDetail = (f.desc && f.desc.length) || (f.enum && f.enum.length);
    const tw = hasDetail ? '▶' : '';

    let detail = '';
    if (hasDetail) {
      let parts = '';
      if (f.desc) parts += '<div class="sm-detail-desc">' + escapeHtml(f.desc) + '</div>';
      if (f.enum && f.enum.length) {
        parts += '<div class="sm-detail-enum"><span class="lbl">Enum</span>' +
                 f.enum.map(v => '<code>' + escapeHtml(String(v)) + '</code>').join('') +
                 '</div>';
      }
      detail = '<div class="sm-detail" hidden>' + parts + '</div>';
    }

    return '<div class="sm-node' + (hasDetail ? ' has-detail' : '') + '">' +
             '<div class="sm-node-row">' +
               '<span class="sm-tw">' + tw + '</span>' +
               '<span class="' + nameCls + '">' + escapeHtml(f.name) + '</span>' +
               '<span class="sm-type">' + escapeHtml(f.type || '?') + '</span>' +
               '<span class="sm-badges">' + badges.join('') + '</span>' +
             '</div>' +
             detail +
           '</div>';
  }

  function renderTree(entity) {
    const fields = entity.fields || [];
    const pkSet = new Set(entity.primary_key || []);
    // Group: key fields first (PK or any field with key=true), then business, then classified, then audit tail
    const keyFields = [];
    const classified = [];
    const business = [];
    fields.forEach(f => {
      if (pkSet.has(f.name) || f.key) keyFields.push(f);
      else if (f.classification)      classified.push(f);
      else                            business.push(f);
    });

    let html = '';
    if (keyFields.length) {
      html += '<div class="sm-group-label">Identifiers &amp; keys (' + keyFields.length + ')</div>';
      keyFields.forEach(f => { html += fieldRow(f, { pkSet }); });
    }
    if (business.length) {
      html += '<div class="sm-group-label">Business attributes (' + business.length + ')</div>';
      business.forEach(f => { html += fieldRow(f, { pkSet }); });
    }
    if (classified.length) {
      html += '<div class="sm-group-label">Classified / sensitive (' + classified.length + ')</div>';
      classified.forEach(f => { html += fieldRow(f, { pkSet }); });
    }
    html += '<div class="sm-group-label">Audit &amp; lineage tail (' + AUDIT_TAIL.length + ')</div>';
    AUDIT_TAIL.forEach(f => { html += fieldRow(f, { audit: true }); });

    sTree.innerHTML = html;
    sCount.textContent = fields.length + ' fields + ' + AUDIT_TAIL.length + ' audit fields';

    // Expand-on-click for any node that has detail content
    sTree.querySelectorAll('.sm-node.has-detail').forEach(node => {
      const row = node.querySelector('.sm-node-row');
      const det = node.querySelector('.sm-detail');
      row.addEventListener('click', () => {
        const open = !node.classList.contains('expanded');
        node.classList.toggle('expanded', open);
        det.hidden = !open;
      });
    });
  }

  function openSchemaModal(key) {
    const entity = SCHEMA_FIELDS[key];
    if (!entity) return;
    sNs.textContent     = entity.namespace || key.split('.')[0] || '';
    sDomain.textContent = entity.domain || '';
    sLayer.textContent  = entity.layer  || '';
    sLayer.className    = 'sm-layer ' + (String(entity.layer || '').toLowerCase());
    sTitle.textContent  = key;
    sDesc.textContent   = entity.description || '';
    const pk = entity.primary_key || [];
    if (pk.length) {
      sPk.innerHTML = '<b>Primary key</b>' + pk.map(f => '<code>' + escapeHtml(f) + '</code>').join('');
      sPk.hidden = false;
    } else {
      sPk.hidden = true;
    }
    renderTree(entity);
    sAllExpanded = false;
    sExpandAll.textContent = 'Expand all';
    sBackdrop.hidden = false;
    document.body.classList.add('modal-open');
    setTimeout(() => sClose.focus(), 50);
  }

  function closeSchemaModal() {
    sBackdrop.hidden = true;
    // Only release modal-open if the scenario modal isn't also open
    if (backdrop.hidden) document.body.classList.remove('modal-open');
  }

  sClose.addEventListener('click', closeSchemaModal);
  sBackdrop.addEventListener('click', (e) => { if (e.target === sBackdrop) closeSchemaModal(); });
  document.addEventListener('keydown', (e) => {
    if (!sBackdrop.hidden && e.key === 'Escape') { closeSchemaModal(); e.stopPropagation(); }
  });
  sExpandAll.addEventListener('click', () => {
    sAllExpanded = !sAllExpanded;
    sExpandAll.textContent = sAllExpanded ? 'Collapse all' : 'Expand all';
    sTree.querySelectorAll('.sm-node.has-detail').forEach(node => {
      node.classList.toggle('expanded', sAllExpanded);
      const det = node.querySelector('.sm-detail');
      if (det) det.hidden = !sAllExpanded;
    });
  });

  // Bind schema-tag clicks (delegated, so it works in flat view too).
  document.addEventListener('click', (e) => {
    const tag = e.target.closest('.schema-tag.clickable');
    if (!tag) return;
    e.preventDefault();
    e.stopPropagation();
    const key = tag.dataset.schema;
    if (key) openSchemaModal(key);
  });
  document.addEventListener('keydown', (e) => {
    if (e.key !== 'Enter' && e.key !== ' ') return;
    const tag = e.target.closest && e.target.closest('.schema-tag.clickable');
    if (!tag) return;
    e.preventDefault();
    e.stopPropagation();
    const key = tag.dataset.schema;
    if (key) openSchemaModal(key);
  });

})();
</script>
</body>
</html>
""")

# Write file
output = "\n".join(out)
with open(OUT, "w", encoding="utf-8") as fh:
    fh.write(output)

print(f"Wrote: {OUT}")
print(f"Bytes: {len(output):,}")
print(f"Services: {len(by_svc)}")
print(f"Scenarios: {sum(svc_counts.values())}")
print(f"Featured: {sum(svc_featured_counts.values())}")

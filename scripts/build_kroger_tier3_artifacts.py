"""Tier-3 artifacts:
  8. Risk Register (XLSX)
  9. Stakeholder Map / RACI (XLSX)
  10. Pitch Deck (HTML)
"""
from __future__ import annotations
import io, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
OUT = Path(r"C:\Stage\Clients\Industries\Consumer\Retail\Kroger\02_projects\FY27_Pipeline\assortment-pricing-agentic\deliverables")
OUT.mkdir(parents=True, exist_ok=True)

HDR_FONT=Font(name="Calibri",size=11,bold=True,color="FFFFFF")
HDR_FILL=PatternFill("solid",fgColor="0B0B0F")
THIN=Side(style="thin",color="D4D4D7")
BOR=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
BAND_FILL=PatternFill("solid",fgColor="F3F3F5")
CELL=Font(name="Calibri",size=10)
CODE=Font(name="Consolas",size=10,color="6648B0",bold=True)
HDR_AL = Alignment(horizontal="left",vertical="center",wrap_text=True)
CELL_AL = Alignment(horizontal="left",vertical="top",wrap_text=True)
CTR = Alignment(horizontal="center",vertical="center",wrap_text=True)

def write_hdrs(ws, headers, widths, row=1):
    for j,h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h); c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=HDR_AL; c.border=BOR
        ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[row].height = 30


# =========================================================================
# 8. RISK REGISTER
# =========================================================================
def build_risk_register():
    print("Building Risk Register xlsx...")
    wb = Workbook()
    ws = wb.active; ws.title = "README"
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 92
    rows = [
        ("APEX · RC-E2E-03 · Risk Register · Kroger", "Identified risks across Approach + Wave 1 + Wave 2 + Wave 3 with owner, likelihood, impact, mitigation, status."),
        ("",""),
        ("How to read this workbook",""),
        ("Risk Register","All identified risks with owner, likelihood, impact, mitigation, status. Sorted by inherent risk score."),
        ("Risk Categories","Risks grouped by category (technical, regulatory, organizational, commercial, security, third-party)."),
        ("Mitigation Plan","Per-risk mitigation steps with owners and target dates."),
        ("Risk Heatmap","2x2 likelihood-vs-impact grid with risk IDs plotted."),
        ("",""),
        ("Likelihood scale","1=Low (<10%) · 2=Possible (10-30%) · 3=Likely (30-60%) · 4=High (60-90%) · 5=Near-certain (>90%)"),
        ("Impact scale","1=Negligible · 2=Minor · 3=Moderate · 4=Major · 5=Critical (engagement-threatening)"),
        ("Inherent risk score","Likelihood × Impact (1-25); Residual is post-mitigation"),
        ("",""),
        ("Status legend","OPEN=identified, no action yet · IN_PROGRESS=mitigation underway · MITIGATED=residual within tolerance · CLOSED=no longer applicable"),
        ("Independence posture","All RC-E2E-03 commercial materials use Deloitte's Microsoft practice/DMTSP/Microsoft platform capabilities/Microsoft-native deployment language."),
    ]
    for i,(l,r) in enumerate(rows, start=1):
        c1 = ws.cell(row=i, column=1, value=l); c2 = ws.cell(row=i, column=2, value=r)
        if i == 1: c1.font = Font(bold=True, size=14); c2.font = Font(italic=True, size=10, color="4A4A55")
        elif l and r: c1.font = Font(bold=True, size=10); c2.font = CELL
        else: c1.font = Font(bold=True, size=11, color="6648B0"); c2.font = CELL
        c2.alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[i].height = 28

    # ===== Risk Register sheet =====
    ws = wb.create_sheet("Risk Register")
    write_hdrs(ws, ["Risk ID","Category","Title","Description","Owner","L","I","Score",
                    "Mitigation","Residual","Target date","Status"],
               [9,12,28,52,14,4,4,7,52,9,12,12])
    ws.freeze_panes = "C2"

    RISKS = [
      ("R-01","Regulatory","FDA enforcement-discretion timing on FSMA 204",
       "FDA may extend or vary enforcement-discretion timing for FSMA 204 final rule, affecting urgency narrative. Compliance required regardless but delivery cadence may shift.",
       "Food Safety Lead",2,3,6,
       "Track FDA announcements quarterly · build to current effective date regardless · communicate to Steering Committee on any timing change",
       3,"Quarterly review","IN_PROGRESS"),
      ("R-02","Technical","Refrigeration telemetry data quality below SLO",
       "EMS gateway intermittent connectivity could push BRZ.RC.REFRIG_TELEMETRY completeness below 99.5% SLO, blocking Assess agent.",
       "Facilities Engineering",3,4,12,
       "MQTT bridge with offline buffering · per-fixture data-quality monitor · alert on store completeness < 99% · incident workflow with Facilities Eng",
       6,"W1-S03","IN_PROGRESS"),
      ("R-03","Regulatory","Auditor acceptance of FSMA evidence packet format",
       "Internal Audit and external auditors may request format changes that require regenerating evidence packet generator.",
       "Engagement Architect",3,3,9,
       "Joint design session with Internal Audit + auditor in W1 · iterate packet format with auditor feedback before W2 · anchor on auditor-already-accepted SOC-2-adjacent patterns",
       3,"W1-S04","IN_PROGRESS"),
      ("R-04","Organizational","Store Operations Lead HITL volume saturation",
       "If materiality thresholds tuned too aggressive, Store Ops Leads may be overwhelmed by Adaptive Cards, eroding adoption.",
       "Pricing Manager + VP Retention",4,4,16,
       "Tier-tuned to ~5-10% of events at steady state · digest mode for low-priority · designate-approver workflow · weekly volume review during W2 pilot",
       6,"W2-S03","OPEN"),
      ("R-05","Technical","Synthetic-data fidelity gap (W1 prototype vs. live data)",
       "W1 shadow-run uses synthetic excursion patterns; live W2 data may reveal classification-accuracy degradation.",
       "ML Engineering",3,3,9,
       "Synthetic-data generation calibrated against historical 4 quarters of Cincinnati Kroger data · shadow-run 30 days against live Cincinnati cluster before W2 cutover · model recalibration window in W2-S01",
       4,"W1-S05","OPEN"),
      ("R-06","Organizational","Cohort drift in Risk-Score over time",
       "Risk-Score model drift on cohort populations could degrade precision/recall below thresholds.",
       "ML Engineering + Audience Lead",4,3,12,
       "Cohort Drift Watchtower (UC-22) running W2 · drift alerting · auto-retrain trigger with HITL gate · quarterly model performance review",
       6,"W2-S03","OPEN"),
      ("R-07","Technical","Model bias across demographics or banner mix",
       "Risk-Score or elasticity models may exhibit performance gaps across demographic segments or banner-specific patterns.",
       "ML Engineering + Privacy",2,4,8,
       "Bias-evaluation suite at model deployment · per-banner performance breakdown reviewed weekly · Privacy Office sign-off before live W2 use",
       4,"W2-S02","OPEN"),
      ("R-08","Commercial","Concession budget overrun vs. policy",
       "If Quantify agent recommends markdown above per-cohort concession budget, total concession spend exceeds quarterly target.",
       "Pricing Manager",3,3,9,
       "Concession Policy Studio enforces per-cohort daily/weekly budgets · live concession-cost dashboard · alert at 80% budget consumed · auto-escalate to Pricing Director",
       3,"W2-S02","OPEN"),
      ("R-09","Third-party","Vendor-portal API availability (Emerson · Lennox · Hussmann · Hill PHOENIX)",
       "Refrigeration vendor portals may impose rate limits or experience downtime affecting REFRIG_SERVICE ticket creation and feed.",
       "Engineering Architect + Facilities Engineering",3,2,6,
       "15-min polling cycle within rate limits · graceful degradation on vendor outage (queue tickets, retry) · vendor-account-manager escalation path",
       2,"W1-S03","IN_PROGRESS"),
      ("R-10","Security","SAP S/4HANA / WMS connector access scope creep",
       "Initial read-only access scope may need to expand for W2 production action; security review may delay Wave 2 timing.",
       "IT-Security + Engineering Architect",3,3,9,
       "Architecture review with IT-Sec at W1 entry · privileged-access management for W2 write operations · approved access scope documented in SOW · Independence-office aware",
       4,"W1-S01","IN_PROGRESS"),
      ("R-11","Organizational","Independence-office requirements impact engagement structure",
       "Deloitte's audit relationship with Kroger may require specific engagement structure (manifests/policies as Kroger artifacts; Deloitte advisory only).",
       "DMTSP_lead + Independence office",2,5,10,
       "Independence-office routing in Approach stage · SOW structured to ensure manifests/policies/LEDGER are Kroger artifacts · Audit Committee notification documented",
       2,"A-S02","IN_PROGRESS"),
      ("R-12","Technical","Bronze schema drift breaks Silver",
       "Source system schema changes (POS, pricing, vendor portals) could break Silver canonical layer.",
       "Data Engineering",3,3,9,
       "Additive-only schema evolution discipline · nightly schema-drift detection in DQ pack · auto-PR on net-new fields · weekly review by Data Eng",
       3,"W1-S02","IN_PROGRESS"),
      ("R-13","Commercial","Wave 3 envelope underestimated due to multi-banner complexity",
       "10-banner orchestration may require more effort than estimated, pushing W3 envelope higher.",
       "Engagement Architect",3,3,9,
       "W2 exit-review includes W3 sizing recalibration · per-banner cost benchmarks established · 1.35x high estimate carries cushion",
       6,"W2-S04","OPEN"),
      ("R-14","Organizational","Cross-functional change management (Merch + Store Ops + Food Safety)",
       "RC-E2E-03 spans 3 functional organizations; lack of unified governance can derail execution.",
       "Engagement Architect + Sponsor",4,4,16,
       "Cross-functional Steering Committee with rep from all 3 orgs · monthly review cadence · KPI ownership matrix in Persona RACI · joint training program",
       8,"Approach + W1","OPEN"),
      ("R-15","Regulatory","Food Traceability List expansion mid-engagement",
       "FDA may add categories to FTL during engagement, expanding RC-E2E-09 scope.",
       "Food Safety Lead + Engineering Architect",2,3,6,
       "Quarterly FTL review · BRZ.RC.LOT_TRACE schema accommodates new categories without break · ftl_category field is extensible",
       2,"Quarterly","OPEN"),
      ("R-16","Third-party","84.51° loyalty platform API changes",
       "Read-only loyalty feed may change format or availability if 84.51° platform evolves.",
       "Data Engineering + 84.51° relationship owner",2,2,4,
       "Stable contract negotiated with 84.51° · schema change notice ≥ 30 days · loyalty data is calibration-only (W1+) — graceful degradation",
       2,"Ongoing","OPEN"),
      ("R-17","Technical","LEDGER PostgreSQL index lag",
       "PG index lookup for trace_id may lag behind Lakehouse writes during peak load.",
       "Engineering Architect",3,2,6,
       "PG indexed on trace_id · async hot-path with eventual consistency · failover to Lakehouse-only mode for low-latency consumers",
       2,"W2-S01","IN_PROGRESS"),
      ("R-18","Security","Privacy/consent state desync vs. 84.51° platform",
       "Loyalty consent state could lag, allowing personalization on consent-withdrawn households.",
       "Privacy Office + Data Engineering",2,4,8,
       "Daily consent-state sync · personalization decisions check consent at decision-time · Purview audit of any decision on withdrawn-consent household · DPA reviewed",
       2,"W1-S04","IN_PROGRESS"),
      ("R-19","Organizational","Disposition-policy disagreement (Pricing Manager vs. Food Safety Lead)",
       "Pricing-side and food-safety-side dispositions may conflict on borderline cases.",
       "VP Retention",3,3,9,
       "Materiality policy explicitly prioritizes Food Safety Lead on destroy-all and FSMA-evidence-gap events · escalation matrix documented",
       3,"W2-S02","OPEN"),
      ("R-20","Regulatory","Cross-state inspection variability (state PUC + state-health-departments)",
       "Different state inspection workflows may require state-specific evidence formats.",
       "Food Safety Lead",2,3,6,
       "Default packet aligns to FDA FSMA 204 spec · state-specific packet variants documented · per-state walkthrough during W3",
       3,"W3","OPEN"),
    ]
    for i, r in enumerate(RISKS, start=2):
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=v); c.border=BOR; c.alignment=CELL_AL
            c.font = CELL
            if j == 1: c.font = CODE
            if j in (6,7,8,10): c.alignment = CTR
            if i % 2 == 0: c.fill = BAND_FILL
        # Color score by severity
        score = r[7]
        if score >= 16: ws.cell(row=i, column=8).fill = PatternFill("solid", fgColor="F8D5DA")
        elif score >= 9: ws.cell(row=i, column=8).fill = PatternFill("solid", fgColor="FBF1DC")
        elif score >= 5: ws.cell(row=i, column=8).fill = PatternFill("solid", fgColor="EFF7F5")
        else: ws.cell(row=i, column=8).fill = PatternFill("solid", fgColor="EDEDED")
        ws.row_dimensions[i].height = max(60, 14*(r[3].count("\n")+r[8].count("\n")+3))

    # ===== Risk Heatmap =====
    ws = wb.create_sheet("Risk Heatmap")
    ws.cell(row=1, column=1, value="Risk Heatmap · Likelihood × Impact").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Cells contain risk IDs · color by score").font = Font(italic=True, size=10, color="4A4A55")
    write_hdrs(ws, ["", "I=1 (Negligible)", "I=2 (Minor)", "I=3 (Moderate)", "I=4 (Major)", "I=5 (Critical)"],
               [12,18,18,18,18,18], row=4)
    likelihoods = ["L=5 (Near-certain)","L=4 (High)","L=3 (Likely)","L=2 (Possible)","L=1 (Low)"]
    # Bucket risks by L,I
    cells = {(l,i):[] for l in range(1,6) for i in range(1,6)}
    for r in RISKS:
        cells[(r[5], r[6])].append(r[0])
    for row_idx, lname in enumerate(likelihoods, start=5):
        c = ws.cell(row=row_idx, column=1, value=lname); c.font = Font(bold=True, size=10); c.alignment=CTR; c.border=BOR
        c.fill = HDR_FILL
        for col_idx, i in enumerate(range(1, 6), start=2):
            l = 6 - (row_idx - 4)  # invert to map row 5 = L=5
            ids = cells.get((l, i), [])
            cell_value = "\n".join(ids) if ids else ""
            score = l * i
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = CTR; cell.border = BOR; cell.font = CODE
            if score >= 16: cell.fill = PatternFill("solid", fgColor="F8D5DA")
            elif score >= 9: cell.fill = PatternFill("solid", fgColor="FBF1DC")
            elif score >= 5: cell.fill = PatternFill("solid", fgColor="EFF7F5")
            else: cell.fill = PatternFill("solid", fgColor="EDEDED")
        ws.row_dimensions[row_idx].height = 50

    # ===== Risk Categories =====
    ws = wb.create_sheet("Risk Categories")
    write_hdrs(ws, ["Category","Count","High-score risks (>=12)","Owner"],
               [16,8,32,30])
    cat_summary = {}
    for r in RISKS:
        cat_summary.setdefault(r[1], []).append(r)
    rr = 2
    for cat, items in sorted(cat_summary.items()):
        high = [x[0] for x in items if x[7] >= 12]
        owners = ", ".join(sorted(set(x[4] for x in items)))
        ws.cell(row=rr, column=1, value=cat).font = Font(bold=True, size=10)
        ws.cell(row=rr, column=2, value=len(items)).alignment = CTR
        ws.cell(row=rr, column=3, value=", ".join(high) if high else "—")
        ws.cell(row=rr, column=4, value=owners)
        for j in range(1, 5):
            cc = ws.cell(row=rr, column=j); cc.border = BOR
            if rr % 2 == 0: cc.fill = BAND_FILL
            if cc.font.size is None: cc.font = CELL
        rr += 1

    target = OUT / "APEX-RC-E2E-03-Kroger-Risk-Register.xlsx"
    try: wb.save(str(target))
    except PermissionError:
        n = 2
        while True:
            alt = target.with_name(target.stem + f"-v{n}" + target.suffix)
            if not alt.exists(): target = alt; break
            n += 1
        wb.save(str(target))
    print(f"  wrote {target.name}")


# =========================================================================
# 9. STAKEHOLDER MAP / RACI (XLSX)
# =========================================================================
def build_stakeholder_map():
    print("\nBuilding Stakeholder Map xlsx...")
    wb = Workbook()
    ws = wb.active; ws.title = "README"
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 92
    rows = [
        ("APEX · RC-E2E-03 · Stakeholder Map · Kroger","Stakeholders, decision-makers, sponsors, blockers, allies. RACI matrix by major decision. Communication cadence."),
        ("",""),
        ("How to read this workbook",""),
        ("Stakeholder Map","Named (or named-role) stakeholders with role, influence, posture, our position."),
        ("Decision RACI","Major decisions × stakeholders mapped to RACI roles."),
        ("Communication Plan","Cadence and surface for each stakeholder."),
        ("Power-Interest Grid","2x2 visual grid of stakeholder positioning."),
        ("",""),
        ("Posture legend","CHAMPION=actively advocating · ALLY=supportive · NEUTRAL=undecided · SKEPTIC=requires conversion · BLOCKER=actively opposing"),
        ("Influence legend","HIGH=can move/block decision · MEDIUM=influences others who can · LOW=affected but limited influence"),
        ("Engagement priority","T1=heavy direct engagement · T2=monthly touchpoint · T3=regular updates · T4=informational only"),
        ("",""),
        ("Note","All names are placeholder/role-based until Account Team validates with Kroger relationship contacts. Engagement Manager owns the live version of this document."),
    ]
    for i,(l,r) in enumerate(rows, start=1):
        c1 = ws.cell(row=i, column=1, value=l); c2 = ws.cell(row=i, column=2, value=r)
        if i == 1: c1.font = Font(bold=True, size=14); c2.font = Font(italic=True, size=10, color="4A4A55")
        elif l and r: c1.font = Font(bold=True, size=10); c2.font = CELL
        else: c1.font = Font(bold=True, size=11, color="6648B0"); c2.font = CELL
        c2.alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[i].height = 28

    # ===== Stakeholder Map =====
    ws = wb.create_sheet("Stakeholder Map")
    write_hdrs(ws, ["#","Stakeholder","Role / title","Org","Influence","Posture",
                    "Our position","Engagement priority","Notes"],
               [4, 22, 30, 18, 11, 12, 36, 14, 36])
    ws.freeze_panes = "B2"

    STKS = [
      (1,"VP Merchandising / Chief Merchant","Chief Merchant (TBD-name)","Merchandising","HIGH","NEUTRAL",
       "Strategic sponsor candidate","T1","Decision-maker on RC-E2E-03 anchor service. Account team to validate name."),
      (2,"Senior Category Manager — Dairy","Category Manager","Merchandising · Dairy","HIGH","NEUTRAL",
       "Primary persona · daily user","T1","Key reference for Wave-1 design; champion if onboarded early"),
      (3,"VP Subscriber Retention / Loyalty","VP Loyalty","Customer / Loyalty","MEDIUM","NEUTRAL",
       "Tier-2 attach (RC-E2E-04) sponsor","T2","RC-E2E-04 fusion in W3 depends on this relationship"),
      (4,"Pricing Director","Pricing Director","Merchandising · Pricing","HIGH","ALLY",
       "Concession-policy authority · likely champion","T1","Pricing studio adoption depends here"),
      (5,"VP Quality &amp; Regulatory","VP Quality","Quality / Food Safety","HIGH","CHAMPION",
       "FSMA 204 readiness driver","T1","Strongest FSMA driver · co-anchor on RC-E2E-09"),
      (6,"Food Safety Compliance Lead","Food Safety Lead","Quality","MEDIUM","CHAMPION",
       "Co-primary persona · daily evidence consumer","T1","FSMA 204 deep collaboration"),
      (7,"COO","COO","Operations","HIGH","NEUTRAL",
       "Final authority on enterprise-scale W3 commit","T1","Steering Committee chair candidate"),
      (8,"CFO","CFO","Finance","HIGH","NEUTRAL",
       "ROI signoff · Audit Committee link","T1","ROI Case is the lever for this stakeholder"),
      (9,"District Manager (Cincinnati)","District Manager","Operations","MEDIUM","ALLY",
       "Wave-1 anchor cluster sponsor","T2","Champion for Cincinnati pilot · drives Store Ops Lead onboarding"),
      (10,"Store Operations Lead — Store #428","Store Ops Lead","Operations","LOW","NEUTRAL",
       "Executional primary persona","T2","Wave-1 reference user · vivid testimonial source"),
      (11,"CIO","CIO","Technology","HIGH","NEUTRAL",
       "Microsoft-native architecture sign-off","T1","CAF integration acceptance pathway"),
      (12,"CDO / Chief Data Officer","CDO","Data / 84.51°","HIGH","ALLY",
       "84.51° platform integration · loyalty data feeds","T1","Privacy + consent boundary owner"),
      (13,"VP Engineering · 84.51°","84.51° VP Engineering","84.51° subsidiary","MEDIUM","NEUTRAL",
       "Loyalty platform technical interface","T2","Read-only loyalty feed contract owner"),
      (14,"VP Finance Systems","VP Finance Systems","Technology · IT-Finance","MEDIUM","NEUTRAL",
       "SAP S/4HANA connector access","T2","Wave-2 production posting access"),
      (15,"VP Supply Chain Technology","VP SCM Tech","Technology · IT-SCM","MEDIUM","NEUTRAL",
       "Manhattan WMS + receipt CDC integration","T2","Wave-1 ingest dependency"),
      (16,"VP Internal Audit","VP Internal Audit","Audit","HIGH","NEUTRAL",
       "Evidence packet acceptance · risk-control validation","T1","Privacy/Audit walkthrough partner"),
      (17,"External Audit PIC (PwC)","Audit Partner","External Audit","MEDIUM","NEUTRAL",
       "Quarterly evidence walkthrough · independence-office partner","T2","Independence-office routing required"),
      (18,"Vendor Manager — Dairy","Vendor Manager","Procurement","MEDIUM","ALLY",
       "Supplier scorecard consumer","T3","Win at QBR demonstration"),
      (19,"VP Procurement / CPO","CPO","Procurement","MEDIUM","NEUTRAL",
       "Supplier-readiness program owner","T2","FSMA-tier supplier qualification"),
      (20,"Marketing Operations Director","Marketing Ops","Marketing / Loyalty","MEDIUM","NEUTRAL",
       "Channel-dispatch infrastructure","T3","RC-E2E-04 fusion stakeholder W3"),
      (21,"VP IT Security / CISO","CISO","Security","HIGH","NEUTRAL",
       "Architecture review · access scope","T1","Wave-1 architecture-review gate"),
      (22,"Privacy Office Lead","Privacy Lead","Legal / Privacy","HIGH","ALLY",
       "Consent governance · personalization-evidence","T1","Strong ally given FSMA + privacy alignment"),
      (23,"FDA Relations Lead","FDA Relations","External Affairs","MEDIUM","ALLY",
       "FDA inspection coordination · evidence packet recipient","T2","Champion for evidence-packet quality"),
      (24,"Audit Committee chair","Board · Audit Committee","Board","HIGH","NEUTRAL",
       "Oversight on FSMA + control posture","T2","Briefed via VP Internal Audit and CFO"),
      (25,"VP Communications / IR","VP IR","Communications","LOW","NEUTRAL",
       "Public messaging on FSMA readiness","T4","Briefed when packet exercises with auditors complete"),
    ]
    for i, s in enumerate(STKS, start=2):
        for j, v in enumerate(s, start=1):
            c = ws.cell(row=i, column=j, value=v); c.border=BOR; c.alignment=CELL_AL
            c.font = CELL
            if j == 1: c.alignment = CTR
            if j == 2: c.font = Font(bold=True, size=10)
            if i % 2 == 0: c.fill = BAND_FILL
        # Color posture
        posture = s[5]
        if posture == "CHAMPION": ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor="C9E5DC")
        elif posture == "ALLY": ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor="EFF7F5")
        elif posture == "NEUTRAL": ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor="EDEDED")
        elif posture == "SKEPTIC": ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor="FBF1DC")
        elif posture == "BLOCKER": ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor="F8D5DA")
        ws.row_dimensions[i].height = 42

    # ===== Decision RACI =====
    ws = wb.create_sheet("Decision RACI")
    decisions = [
        "Approach SOW signoff",
        "Wave 1 Foundation commit",
        "W1 cohort + entity scope",
        "Materiality policy approval (W1)",
        "Adaptive Card persona binding (Entra)",
        "FSMA 204 evidence packet format",
        "Wave 1 → Wave 2 gate review",
        "W2 cohort family selection",
        "Concession Policy Studio adoption",
        "Multi-channel dispatch go-live (W2)",
        "Wave 2 → Wave 3 gate review",
        "Multi-banner enterprise rollout commit",
        "Cross-scenario fusion (W3) commit",
        "External-audit walkthrough acceptance",
        "Privacy/Consent governance signoff",
    ]
    pers = ["Chief Merchant","Cat Mgr Dairy","Pricing Dir","VP Quality","Food Safety Lead",
            "COO","CFO","CIO","CDO","Internal Audit","Privacy Lead","CISO","DMTSP Lead","Acct Team"]
    write_hdrs(ws, ["Decision"] + pers, [38] + [11]*len(pers))
    ws.freeze_panes = "B2"
    # RACI map: simplified per decision
    raci_data = {
        "Approach SOW signoff":          ["A","I","C","A","C", "R","A","C","I","I","I","C","R","R"],
        "Wave 1 Foundation commit":     ["A","C","C","C","C", "A","A","C","C","I","I","I","R","R"],
        "W1 cohort + entity scope":      ["A","R","R","C","C", "C","I","I","C","I","I","I","R","C"],
        "Materiality policy approval (W1)": ["A","C","R","C","C","C","I","I","I","I","I","I","R","I"],
        "Adaptive Card persona binding (Entra)": ["I","C","I","I","I","I","I","R","C","I","I","R","A","C"],
        "FSMA 204 evidence packet format": ["I","I","I","R","R", "I","C","C","C","R","R","C","A","I"],
        "Wave 1 → Wave 2 gate review":  ["A","R","R","R","R", "A","C","C","C","C","C","C","R","C"],
        "W2 cohort family selection":   ["A","R","C","R","R", "C","I","I","C","C","C","I","R","C"],
        "Concession Policy Studio adoption": ["C","C","R","I","I","C","C","I","I","I","I","I","R","C"],
        "Multi-channel dispatch go-live (W2)": ["I","I","C","I","I","C","I","R","C","C","C","R","A","C"],
        "Wave 2 → Wave 3 gate review":  ["A","R","R","R","R", "A","A","C","C","R","R","C","R","R"],
        "Multi-banner enterprise rollout commit": ["A","C","C","C","C","A","A","C","C","C","C","C","R","R"],
        "Cross-scenario fusion (W3) commit": ["A","C","C","R","R","A","C","C","C","C","C","C","R","R"],
        "External-audit walkthrough acceptance": ["I","I","I","R","R","C","R","I","I","R","C","I","A","C"],
        "Privacy/Consent governance signoff": ["I","I","I","C","C","I","C","C","R","R","R","R","A","I"],
    }
    for i, dec in enumerate(decisions, start=2):
        ws.cell(row=i, column=1, value=dec).font = Font(bold=True, size=10)
        ws.cell(row=i, column=1).border = BOR
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        for j, val in enumerate(raci_data.get(dec, ["—"]*len(pers)), start=2):
            c = ws.cell(row=i, column=j, value=val); c.alignment = CTR; c.font = Font(bold=True, size=10); c.border = BOR
            if val == "R":   c.fill = PatternFill("solid", fgColor="C9E5DC")
            elif val == "A": c.fill = PatternFill("solid", fgColor="F6D7A1")
            elif val == "C": c.fill = PatternFill("solid", fgColor="DCE3F4")
            elif val == "I": c.fill = PatternFill("solid", fgColor="EDEDED")
        ws.row_dimensions[i].height = 30

    # ===== Communication Plan =====
    ws = wb.create_sheet("Communication Plan")
    write_hdrs(ws, ["Stakeholder","Cadence","Surface","Owner","Notes"],
               [22, 14, 24, 18, 36])
    plan = [
        ("Steering Committee","Monthly","Live exec briefing + Power BI scorecard","DMTSP Lead","Committee chair = COO; rotating sponsors"),
        ("Chief Merchant","Bi-weekly","Live walkthrough + dashboard","DMTSP Lead","Account Team owns relationship"),
        ("Category Manager — Dairy","Weekly during W1","Working session","Engagement Architect","Vivid testimonial source"),
        ("VP Quality + Food Safety Lead","Bi-weekly","Working session + monthly exec","Engagement Architect","FSMA 204 readiness review"),
        ("CFO","Monthly","ROI Case dashboard + briefing","Engagement Architect + DMTSP Lead","Finance-modelable assumptions"),
        ("CIO","Monthly","Architecture review + tech-debt review","Engagement Architect","CAF integration alignment"),
        ("CDO","Monthly","Data + privacy review","Engagement Architect","84.51° integration"),
        ("Internal Audit + External Audit","Quarterly","Joint evidence-packet walkthrough","Engagement Architect","Independence-office routing"),
        ("Privacy Office","Bi-weekly during W2","Working session","Engagement Architect","Consent governance"),
        ("CISO","Monthly","Security review","Engagement Architect","Access scope review"),
        ("Audit Committee","Quarterly","Briefing via VP Internal Audit + CFO","CFO + DMTSP Lead","FSMA + control posture"),
        ("All other stakeholders","As-needed","Email + dashboard access","DMTSP PMO","Per Engagement priority tier"),
    ]
    for i, p in enumerate(plan, start=2):
        for j, v in enumerate(p, start=1):
            c = ws.cell(row=i, column=j, value=v); c.border=BOR; c.alignment=CELL_AL; c.font=CELL
            if j == 1: c.font = Font(bold=True, size=10)
            if i % 2 == 0: c.fill = BAND_FILL
        ws.row_dimensions[i].height = 32

    target = OUT / "APEX-RC-E2E-03-Kroger-Stakeholder-Map.xlsx"
    try: wb.save(str(target))
    except PermissionError:
        n = 2
        while True:
            alt = target.with_name(target.stem + f"-v{n}" + target.suffix)
            if not alt.exists(): target = alt; break
            n += 1
        wb.save(str(target))
    print(f"  wrote {target.name}")


# =========================================================================
# 10. PITCH DECK (HTML)
# =========================================================================
PITCH_DECK = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>APEX · RC-E2E-03 · Kroger · Pitch Deck</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:ital,opsz,wght@0,9..144,400;0,9..144,500;0,9..144,600&family=IBM+Plex+Sans:wght@0,300;400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root {
  --bg-0: #07090F; --bg-1: #0C1119; --bg-2: #141B27; --bg-3: #1C2433;
  --line: rgba(255,255,255,0.10); --line-strong: rgba(255,255,255,0.20);
  --ink-0: #F4EFE5; --ink-1: #D9D2C3; --ink-2: #9A9383; --ink-3: #6B6757;
  --gold: #E3B657; --teal: #3DD9C4; --crimson: #D94B5A; --violet: #9B7BE0;
  --sky: #6FB6E5; --amber: #E8A33D;
}
body.light {
  --bg-0: #F5F1E8; --bg-1: #EEE7D7; --bg-2: #E5DCC7; --bg-3: #D9CFB6;
  --line: rgba(20,14,4,0.10); --line-strong: rgba(20,14,4,0.22);
  --ink-0: #1A1508; --ink-1: #3D3420; --ink-2: #6B5F45; --ink-3: #8E8367;
  --gold: #8E6A1E; --teal: #147A6C; --crimson: #A22A39;
  --violet: #6548B0; --sky: #3E7FAE; --amber: #B8791F;
}
* { box-sizing: border-box; }
html, body { margin: 0; padding: 0; }
body {
  background: var(--bg-0); color: var(--ink-0);
  font-family: "IBM Plex Sans", system-ui, sans-serif;
  font-size: 16px; line-height: 1.55;
  transition: background .5s, color .5s;
}
.theme-toggle {
  position: fixed; top: 24px; right: 24px; z-index: 100;
  display: flex; align-items: center; gap: 8px;
  background: var(--bg-1); border: 1px solid var(--line-strong);
  border-radius: 999px; padding: 8px 14px; cursor: pointer;
  font-family: "JetBrains Mono", monospace;
  font-size: 11px; letter-spacing: .12em; text-transform: uppercase;
  color: var(--ink-1);
}

.slide {
  min-height: 100vh; padding: 80px 56px;
  position: relative; z-index: 1;
  display: flex; flex-direction: column; justify-content: center;
  border-bottom: 1px solid var(--line);
}
.slide-head {
  display: flex; justify-content: space-between; align-items: center;
  margin-bottom: 48px;
  font-family: "JetBrains Mono", monospace; font-size: 11px;
  letter-spacing: .14em; text-transform: uppercase; color: var(--ink-3);
}
.slide-num {
  font-family: "JetBrains Mono", monospace;
  font-size: 11px; color: var(--ink-3);
}
h1 {
  font-family: "Fraunces", serif; font-weight: 500;
  font-size: clamp(48px, 6vw, 96px); line-height: 1.04;
  letter-spacing: -0.02em; margin: 0 0 24px;
}
h1 em { font-style: italic; font-weight: 400; color: var(--amber); }
h2 {
  font-family: "Fraunces", serif; font-weight: 500;
  font-size: clamp(36px, 4.5vw, 60px); line-height: 1.08;
  letter-spacing: -0.015em; margin: 0 0 24px;
}
h3 {
  font-family: "Fraunces", serif; font-weight: 500;
  font-size: 24px; margin: 0 0 12px;
}
.lead {
  font-size: 22px; color: var(--ink-1); max-width: 880px;
  font-weight: 300;
}
.giant-num {
  font-family: "Fraunces", serif; font-weight: 500;
  font-size: clamp(120px, 16vw, 220px); line-height: 0.95;
  color: var(--gold); margin: 24px 0;
  letter-spacing: -0.04em;
}
.three-col {
  display: grid; grid-template-columns: repeat(3, 1fr); gap: 24px;
  margin-top: 32px;
}
@media (max-width: 1100px) { .three-col { grid-template-columns: 1fr; } }
.col {
  background: var(--bg-1); border: 1px solid var(--line);
  border-top: 4px solid var(--accent); border-radius: 12px;
  padding: 28px;
}
.col h3 {
  font-family: "Fraunces", serif; font-size: 22px;
  margin: 0 0 8px; color: var(--ink-0);
}
.col-eye {
  font-family: "JetBrains Mono", monospace; font-size: 10.5px;
  letter-spacing: .14em; text-transform: uppercase;
  color: var(--accent); margin-bottom: 14px;
}
.col p { color: var(--ink-1); font-size: 14.5px; }
.col ul { margin: 12px 0 0; padding-left: 20px; color: var(--ink-1); font-size: 14px; }
.col ul li { margin-bottom: 6px; }

.metrics-row {
  display: grid; grid-template-columns: repeat(4, 1fr); gap: 18px;
  margin-top: 32px;
}
@media (max-width: 900px) { .metrics-row { grid-template-columns: 1fr 1fr; } }
.metric {
  background: var(--bg-1); border: 1px solid var(--line);
  border-radius: 12px; padding: 24px;
}
.metric .lbl {
  font-family: "JetBrains Mono", monospace; font-size: 10.5px;
  letter-spacing: .14em; text-transform: uppercase;
  color: var(--ink-3); margin-bottom: 8px;
}
.metric .val {
  font-family: "Fraunces", serif; font-size: 36px; font-weight: 500;
  color: var(--accent); line-height: 1;
}
.metric .delta { font-size: 12.5px; color: var(--ink-2); margin-top: 6px; }

.agent-row {
  display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px;
  margin-top: 28px;
}
@media (max-width: 1100px) { .agent-row { grid-template-columns: 1fr 1fr; } }
.agent {
  background: var(--bg-1); border: 1px solid var(--line);
  border-radius: 10px; padding: 20px; text-align: center;
}
.agent .num {
  font-family: "JetBrains Mono", monospace; font-size: 11px;
  letter-spacing: .14em; color: var(--gold); margin-bottom: 10px;
}
.agent h4 {
  font-family: "Fraunces", serif; font-size: 20px; font-weight: 500;
  margin: 0 0 8px;
}
.agent p { color: var(--ink-2); font-size: 13px; margin: 0; }

.wave-cards {
  display: grid; grid-template-columns: repeat(3, 1fr); gap: 18px;
  margin-top: 28px;
}
@media (max-width: 900px) { .wave-cards { grid-template-columns: 1fr; } }
.wave-card {
  background: var(--bg-1); border: 1px solid var(--line);
  border-left: 5px solid var(--accent); border-radius: 12px;
  padding: 28px;
}
.wave-card .wname {
  font-family: "JetBrains Mono", monospace; font-size: 11px;
  letter-spacing: .14em; color: var(--accent); margin-bottom: 8px;
}
.wave-card h3 { font-size: 24px; margin: 0 0 12px; }
.wave-card .scope { color: var(--ink-1); font-size: 14px; margin-bottom: 14px; }
.wave-card .env {
  font-family: "JetBrains Mono", monospace; font-size: 13px;
  color: var(--gold); padding-top: 12px; border-top: 1px solid var(--line);
}

.cta-block {
  background: linear-gradient(135deg, rgba(227,182,87,0.15), rgba(155,123,224,0.10));
  border: 1px solid var(--gold); border-radius: 16px;
  padding: 40px 44px; margin-top: 32px;
}
.cta-block h3 { font-size: 28px; margin: 0 0 14px; }
.cta-block p { color: var(--ink-0); font-size: 17px; max-width: 780px; }
.cta-list {
  display: grid; grid-template-columns: 1fr 1fr; gap: 20px;
  margin-top: 28px;
}
@media (max-width: 900px) { .cta-list { grid-template-columns: 1fr; } }
.cta-item {
  background: rgba(0,0,0,0.2); border-radius: 8px; padding: 18px 22px;
}
body.light .cta-item { background: rgba(255,255,255,0.5); }
.cta-item .step {
  font-family: "JetBrains Mono", monospace; font-size: 11px;
  letter-spacing: .14em; color: var(--gold); margin-bottom: 4px;
}
.cta-item h4 {
  font-family: "Fraunces", serif; font-size: 18px; font-weight: 500;
  margin: 0 0 6px;
}
.cta-item p { font-size: 14px; color: var(--ink-1); margin: 0; }

.signoff {
  font-family: "Fraunces", serif; font-size: 28px;
  margin-top: 48px; color: var(--ink-0);
}
.signoff-mono {
  font-family: "JetBrains Mono", monospace; font-size: 12px;
  color: var(--ink-2); margin-top: 8px;
}
</style>
</head>
<body>

<div class="theme-toggle" onclick="toggleTheme()">
  <span id="themeIcon">☀</span>
  <span id="themeLabel">Light</span>
</div>

<!-- Slide 1 — Cover -->
<div class="slide">
  <div class="slide-head"><div>APEX · RC-E2E-03 · KROGER</div><div class="slide-num">01 / 16</div></div>
  <div>
    <h1>Cold-chain markdown intelligence at <em>2,800-store</em> scale.</h1>
    <p class="lead">
      A pitch for The Kroger Co. — APEX RC-E2E-03 Assortment &amp; Pricing as the Wave-1 anchor for
      Grocery Merchandising. Co-anchored with RC-E2E-09 Product Tracking for FSMA 204 compliance.
    </p>
    <div class="signoff">— Deloitte Microsoft Technology Solutions &amp; Practices</div>
    <div class="signoff-mono">v1.0 · 27 Apr 2026 · 16 slides · ~25 min walk + Q&amp;A</div>
  </div>
</div>

<!-- Slide 2 — The problem -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 2 · THE PROBLEM</div><div class="slide-num">02 / 16</div></div>
  <h2>The 52-minute window that costs Kroger <em>~$1.1B/yr</em>.</h2>
  <p class="lead">Every day across 2,800+ stores, a refrigerated fixture breaches setpoint. The first
  customer touches the dairy case in 52 minutes. Today, getting from event to disposition takes 17
  minutes — leaving 35 minutes for execution and zero for analytical work.</p>
  <div class="metrics-row">
    <div class="metric" style="--accent: var(--crimson);">
      <div class="lbl">Annual fresh-shrink exposure</div>
      <div class="val">~$1.1B</div>
      <div class="delta">~2.5% of $45B fresh-food revenue</div>
    </div>
    <div class="metric" style="--accent: var(--ember);">
      <div class="lbl">Decision-loop today</div>
      <div class="val">17 min</div>
      <div class="delta">vs. 52-min pre-opening window</div>
    </div>
    <div class="metric" style="--accent: var(--violet);">
      <div class="lbl">Manager attention lost</div>
      <div class="val">5.2 hrs</div>
      <div class="delta">per shift on info-gathering, not floor</div>
    </div>
    <div class="metric" style="--accent: var(--gold);">
      <div class="lbl">FSMA-204 audit gap</div>
      <div class="val">7%</div>
      <div class="delta">of events lack complete evidence chain</div>
    </div>
  </div>
</div>

<!-- Slide 3 — The opportunity -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 3 · THE OPPORTUNITY</div><div class="slide-num">03 / 16</div></div>
  <h2>What APEX converts the problem into.</h2>
  <div style="text-align: center;">
    <div class="giant-num">~$200M</div>
    <p class="lead" style="text-align: center; margin: 0 auto;">annualized W3 run-rate shrink avoided · 18% reduction across 12.4M events/year · base case ROI</p>
  </div>
  <p style="margin-top: 32px; color: var(--ink-2); font-size: 14px; text-align: center;">
    See ROI Case · Sensitivity bounds · Conservative $95M / Optimistic $340M
  </p>
</div>

<!-- Slide 4 — What we deliver -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 4 · WHAT WE DELIVER</div><div class="slide-num">04 / 16</div></div>
  <h2>RC-E2E-03 — five purpose-built agents, one decision loop.</h2>
  <p class="lead">A choreographed agent fleet that takes a refrigeration excursion at 6:08 AM and
  produces a Store Operations Lead's approved disposition at 6:09 AM — with FSMA 204 evidence
  attached.</p>
  <div class="agent-row">
    <div class="agent" style="border-color: var(--sky);"><div class="num">01</div><h4>Assess</h4><p>Characterize the excursion event in real time</p></div>
    <div class="agent" style="border-color: var(--violet);"><div class="num">02</div><h4>Classify</h4><p>Per-SKU disposition vs. FSMA 204 policy</p></div>
    <div class="agent" style="border-color: var(--amber);"><div class="num">03</div><h4>Quantify</h4><p>$ impact: markdown vs. destroy vs. sell-through</p></div>
    <div class="agent" style="border-color: var(--gold);"><div class="num">04</div><h4>Approve</h4><p>HITL via Teams Adaptive Card to Store Ops Lead</p></div>
    <div class="agent" style="border-color: var(--teal);"><div class="num">05</div><h4>Act</h4><p>POS markdown · vendor ticket · FSMA log — parallel</p></div>
  </div>
</div>

<!-- Slide 5 — Why Kroger is the reference -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 5 · WHY KROGER IS THE RC REFERENCE</div><div class="slide-num">05 / 16</div></div>
  <h2>Four characteristics no other US grocer matches at scale.</h2>
  <div class="three-col">
    <div class="col" style="--accent: var(--gold);">
      <div class="col-eye">Multi-banner</div>
      <h3>10+ banner footprint</h3>
      <p>Kroger · Ralphs · Fred Meyer · King Soopers · Smith's · Fry's · Harris Teeter · Mariano's · QFC · Pick 'n Save · Roundy's · with distinct local pricing zones.</p>
    </div>
    <div class="col" style="--accent: var(--teal);">
      <div class="col-eye">Pricing DNA</div>
      <h3>84.51° subsidiary</h3>
      <p>Sophisticated customer-data-driven pricing in the muscle memory of the merchant org. APEX complements rather than replaces.</p>
    </div>
    <div class="col" style="--accent: var(--violet);">
      <div class="col-eye">Fresh intensity</div>
      <h3>~$45B fresh exposure</h3>
      <p>30% of revenue is dairy/deli/meat/produce/bakery. Even 1% shrink reduction = ~$450M of free cash. RC-E2E-03 targets 18%.</p>
    </div>
  </div>
  <div style="height: 18px;"></div>
  <div class="three-col" style="grid-template-columns: 1fr 1fr;">
    <div class="col" style="--accent: var(--amber);">
      <div class="col-eye">FSMA 204 mandate</div>
      <h3>Effective 2026</h3>
      <p>FDA Final Rule lot-traceability for FTL categories. Compliance is operationally non-optional. RC-E2E-09 makes it auditable by construction.</p>
    </div>
    <div class="col" style="--accent: var(--crimson);">
      <div class="col-eye">Loyalty footprint</div>
      <h3>60M+ Plus households</h3>
      <p>Provides the demand-elasticity signal density that pricing decisions actually need to converge.</p>
    </div>
  </div>
</div>

<!-- Slide 6 — Architecture in 30 seconds -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 6 · ARCHITECTURE</div><div class="slide-num">06 / 16</div></div>
  <h2>Microsoft-native, CAF-rideable.</h2>
  <p class="lead">Five planes from Experience down to LEDGER. Twelve Bronze tables. MERML + SCML
  Silver canonical. CAF substrate underneath.</p>
  <div class="three-col">
    <div class="col" style="--accent: var(--sky);">
      <div class="col-eye">Bronze landing</div>
      <h3>12 SOR tables</h3>
      <ul>
        <li>POS_TXN · ITEM_MASTER · PRICING · PROMO_MARKDOWN</li>
        <li>INVENTORY · REFRIG_TELEMETRY · REFRIG_SERVICE</li>
        <li>FOOD_SAFETY · LOT_TRACE · VENDOR_MASTER · RECEIPT</li>
        <li>LOYALTY_HH (read-only from 84.51°)</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--teal);">
      <div class="col-eye">Silver canonical</div>
      <h3>MERML + SCML</h3>
      <ul>
        <li>MERML.Price · Markdown · Elasticity · Promotion</li>
        <li>SCML.Lot · Inventory · Fixture · Receipt · Disposition</li>
        <li>DIM.Item · Vendor · Household</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--violet);">
      <div class="col-eye">LEDGER + CAF</div>
      <h3>Audit row + agent runtime</h3>
      <ul>
        <li>14-field audit row per decision</li>
        <li>FSMA 204 quarterly evidence packet</li>
        <li>CAF AEF LLM Gateway + A2A Swarm + Dynamic Orchestrator</li>
        <li>Microsoft Fabric · Foundry · Teams · Entra · Power BI</li>
      </ul>
    </div>
  </div>
</div>

<!-- Slide 7 — KPIs -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 7 · KPIS</div><div class="slide-num">07 / 16</div></div>
  <h2>What we measure and what we commit to.</h2>
  <div class="metrics-row">
    <div class="metric" style="--accent: var(--gold);">
      <div class="lbl">Per-event shrink avoided</div>
      <div class="val">$1,313</div>
      <div class="delta">destroy-all baseline → marked-down</div>
    </div>
    <div class="metric" style="--accent: var(--teal);">
      <div class="lbl">Time-to-decision</div>
      <div class="val">90 sec</div>
      <div class="delta">from 17-min current state</div>
    </div>
    <div class="metric" style="--accent: var(--violet);">
      <div class="lbl">Manager attention reclaimed</div>
      <div class="val">+5.2 hrs</div>
      <div class="delta">per shift per store</div>
    </div>
    <div class="metric" style="--accent: var(--amber);">
      <div class="lbl">FSMA-204 audit-row coverage</div>
      <div class="val">99.5%</div>
      <div class="delta">trace_id-attributed evidence</div>
    </div>
  </div>
  <div style="height: 18px;"></div>
  <div class="metrics-row">
    <div class="metric" style="--accent: var(--sky);">
      <div class="lbl">Auto-execute rate (W3)</div>
      <div class="val">75%</div>
      <div class="delta">25% routed to HITL</div>
    </div>
    <div class="metric" style="--accent: var(--crimson);">
      <div class="lbl">Policy-conformance rate</div>
      <div class="val">≥95%</div>
      <div class="delta">vs. 71% manual baseline</div>
    </div>
    <div class="metric" style="--accent: var(--teal);">
      <div class="lbl">Vendor-ticket rate</div>
      <div class="val">≥92%</div>
      <div class="delta">qualifying events with ticket</div>
    </div>
    <div class="metric" style="--accent: var(--gold);">
      <div class="lbl">Network shrink avoided</div>
      <div class="val">~$200M</div>
      <div class="delta">annualized W3 run-rate</div>
    </div>
  </div>
</div>

<!-- Slide 8 — Three waves -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 8 · THREE WAVES</div><div class="slide-num">08 / 16</div></div>
  <h2>From single-banner pilot to enterprise scale.</h2>
  <div class="wave-cards">
    <div class="wave-card" style="--accent: var(--sky);">
      <div class="wname">Wave 1 · Foundation</div>
      <h3>Single-banner anchor</h3>
      <div class="scope">Cincinnati Kroger ~50 stores · 3 of 5 agents · shadow-run · audit-row e2e · Adaptive Card live · LEDGER scaffolding</div>
      <div class="env">5–8 months · $1.2M – $3.5M</div>
    </div>
    <div class="wave-card" style="--accent: var(--teal);">
      <div class="wname">Wave 2 · Pilot</div>
      <h3>250-store anchor cluster</h3>
      <div class="scope">Multi-banner · 5 agents live · 90-day operating run · first quarterly FSMA evidence packet · Power BI dashboards · Concession Policy Studio</div>
      <div class="env">8–12 months · $5M – $12M</div>
    </div>
    <div class="wave-card" style="--accent: var(--gold);">
      <div class="wname">Wave 3 · Scale &amp; Fuse</div>
      <h3>Enterprise 2,800+ stores</h3>
      <div class="scope">All banners · multi-banner orchestration · cross-scenario fusion (E2E-11/12/13) · enterprise FSMA 204 coverage</div>
      <div class="env">12–18 months · $18M – $45M</div>
    </div>
  </div>
</div>

<!-- Slide 9 — Personas -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 9 · PERSONAS</div><div class="slide-num">09 / 16</div></div>
  <h2>Five named roles. Each gets a measurable lift.</h2>
  <div class="three-col">
    <div class="col" style="--accent: var(--amber);">
      <div class="col-eye">Strategic primary</div>
      <h3>Category Manager · Sarah</h3>
      <ul>
        <li>Category-quality data: 12 days → 24 hrs</li>
        <li>QBR prep: 3 hrs → 20 min</li>
        <li>Policy iteration: annual → weekly</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--ember);">
      <div class="col-eye">Executional primary</div>
      <h3>Store Ops Lead · Marisol</h3>
      <ul>
        <li>Time-to-decision: 17 min → 90 sec</li>
        <li>Manager attention: +5.2 hrs/shift</li>
        <li>FSMA log latency: 24 hrs → real-time</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--violet);">
      <div class="col-eye">Co-primary · FSMA</div>
      <h3>Food Safety Lead · Daniel</h3>
      <ul>
        <li>Audit-row coverage: 62% → 99.5%</li>
        <li>FDA inspection prep: 3 wks → 72 hrs</li>
        <li>Recall response: days → seconds</li>
      </ul>
    </div>
  </div>
  <div style="height: 18px;"></div>
  <div class="three-col" style="grid-template-columns: 1fr 1fr;">
    <div class="col" style="--accent: var(--gold);">
      <div class="col-eye">Pricing policy author</div>
      <h3>Pricing Manager · Yvonne</h3>
      <ul>
        <li>Policy iteration: annual → weekly</li>
        <li>Per-cohort visibility: none → live</li>
        <li>Concession-decision latency: days → seconds</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--teal);">
      <div class="col-eye">Vendor management</div>
      <h3>Vendor Manager · Marcus</h3>
      <ul>
        <li>QBR prep: 6 hrs → 20 min</li>
        <li>Vendor-finding disputability: high → low</li>
        <li>Trade-spend ROI visibility: qualitative → quantitative</li>
      </ul>
    </div>
  </div>
</div>

<!-- Slide 10 — Why now -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 10 · WHY NOW</div><div class="slide-num">10 / 16</div></div>
  <h2>Three converging pressures land on perishables.</h2>
  <div class="three-col">
    <div class="col" style="--accent: var(--crimson);">
      <div class="col-eye">Force 1</div>
      <h3>FSMA 204 effective 2026</h3>
      <p>FDA Food Safety Modernization Act §204(d) lot-traceability becomes operationally non-optional for FTL categories. ~1,900 SKUs at Kroger. Compliance must be by construction.</p>
    </div>
    <div class="col" style="--accent: var(--gold);">
      <div class="col-eye">Force 2</div>
      <h3>Margin compression on shelf-stable</h3>
      <p>Center-store margin under pressure makes perishables-shrink reduction the single largest free-cash-flow lever available. ~$1.1B annual exposure on fresh.</p>
    </div>
    <div class="col" style="--accent: var(--teal);">
      <div class="col-eye">Force 3</div>
      <h3>Agentic AI threshold crossed</h3>
      <p>FY26 inflection — agent-orchestration patterns are enterprise-viable for store-by-store decisioning at sub-second latency. Microsoft-native deployment is the right substrate.</p>
    </div>
  </div>
</div>

<!-- Slide 11 — Independence -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 11 · INDEPENDENCE POSTURE</div><div class="slide-num">11 / 16</div></div>
  <h2>Designed to preserve independence at every audit-firm relationship.</h2>
  <div style="background: var(--bg-1); border-left: 4px solid var(--gold); padding: 28px 32px; border-radius: 8px; max-width: 980px;">
    <h3 style="font-size: 22px; margin: 0 0 14px;">The pattern is independence-safe by design</h3>
    <p style="font-size: 15px; color: var(--ink-1); margin-bottom: 16px;">RC-E2E-03 manifests, policy manifests, and LEDGER are <strong style="color: var(--ink-0);">Kroger artifacts</strong>. Deloitte's role is advisory and implementation enablement. Before any engagement at an audit-client, the SOW is routed through Deloitte's Independence office.</p>
    <p style="font-size: 14px; color: var(--ink-2); font-family: 'JetBrains Mono', monospace; line-height: 1.55;">All RC-E2E-03 commercial materials use:<br>
      "Deloitte's Microsoft practice" · "DMTSP" · "Microsoft platform capabilities" · "Microsoft-native deployment"<br><br>
      The terms "partner," "alliance," and "strategic alliance" do not appear.</p>
  </div>
</div>

<!-- Slide 12 — Service portfolio -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 12 · BEYOND RC-E2E-03</div><div class="slide-num">12 / 16</div></div>
  <h2>Tier-2 attaches and W3 fusion.</h2>
  <p class="lead">RC-E2E-03 anchors Wave 1. RC-E2E-09 co-anchors. Tier-2 services attach at W2/W3 once the foundation is running.</p>
  <div class="three-col">
    <div class="col" style="--accent: var(--gold);">
      <div class="col-eye">Tier 1 · Anchors</div>
      <h3>Today's pitch</h3>
      <ul>
        <li><strong>RC-E2E-03</strong> Assortment &amp; Pricing</li>
        <li><strong>RC-E2E-09</strong> Product Tracking (FSMA 204)</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--teal);">
      <div class="col-eye">Tier 2 · Attach W2/W3</div>
      <h3>High-attach</h3>
      <ul>
        <li>RC-E2E-04 Customer Lifecycle &amp; Loyalty (Plus + 84.51°)</li>
        <li>RC-E2E-08 Customer Acquisition (promo-driven trial)</li>
        <li>RC-E2E-05 Store Operations</li>
      </ul>
    </div>
    <div class="col" style="--accent: var(--violet);">
      <div class="col-eye">W3 candidates</div>
      <h3>Cross-scenario fusion</h3>
      <ul>
        <li>RC-E2E-11 Perishables Integrity Response</li>
        <li>RC-E2E-12 Regulator-Grade Evidence Service</li>
        <li>RC-E2E-13 Manager Time-Return Service</li>
      </ul>
    </div>
  </div>
</div>

<!-- Slide 13 — Risk and how we mitigate -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 13 · RISKS &amp; MITIGATIONS</div><div class="slide-num">13 / 16</div></div>
  <h2>The four risks we anticipate — and how we plan against each.</h2>
  <div class="three-col">
    <div class="col" style="--accent: var(--crimson);">
      <div class="col-eye">Risk · Auditor acceptance</div>
      <h3>FSMA evidence packet</h3>
      <p>Joint design session with Internal Audit and external auditor in W1. Iterate format with auditor feedback before W2. Anchor on auditor-already-accepted SOC-2-adjacent patterns.</p>
    </div>
    <div class="col" style="--accent: var(--ember);">
      <div class="col-eye">Risk · HITL volume</div>
      <h3>Tier-tuned thresholds</h3>
      <p>Materiality policy explicitly sized to ~5–10% of events at steady state. Designate-approver workflow for VP Retention's deputies. Weekly volume review during W2.</p>
    </div>
    <div class="col" style="--accent: var(--violet);">
      <div class="col-eye">Risk · Cross-functional governance</div>
      <h3>Steering Committee</h3>
      <p>Cross-functional Steering with Merch + Store Ops + Food Safety reps. Monthly cadence. KPI ownership matrix in Persona RACI. Joint training program.</p>
    </div>
  </div>
  <p style="margin-top: 24px; color: var(--ink-2); font-family: 'JetBrains Mono', monospace; font-size: 12px;">
    Full Risk Register · 20 identified risks · APEX-RC-E2E-03-Kroger-Risk-Register.xlsx
  </p>
</div>

<!-- Slide 14 — Roadmap -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 14 · 12-MONTH ROADMAP</div><div class="slide-num">14 / 16</div></div>
  <h2>From Approach signoff to first FSMA evidence packet.</h2>
  <div style="background: var(--bg-1); border: 1px solid var(--line); border-radius: 12px; padding: 28px; margin-top: 28px;">
    <table style="width:100%; border-collapse: collapse; font-size: 14px;">
      <tr style="border-bottom: 1px solid var(--line);">
        <td style="padding: 10px 14px; font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--gold); letter-spacing: .14em;">Q1</td>
        <td style="padding: 10px 14px; color: var(--ink-1);">Approach + W1-S01 · workspace + scaffolding · stakeholder interviews · SOW signoff · Independence routing</td>
      </tr>
      <tr style="border-bottom: 1px solid var(--line);">
        <td style="padding: 10px 14px; font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--gold); letter-spacing: .14em;">Q2</td>
        <td style="padding: 10px 14px; color: var(--ink-1);">W1-S02 + W1-S03 · Bronze landing on 5 streaming feeds · Assess + Risk-Score agents · Cincinnati synthetic data lab</td>
      </tr>
      <tr style="border-bottom: 1px solid var(--line);">
        <td style="padding: 10px 14px; font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--gold); letter-spacing: .14em;">Q3</td>
        <td style="padding: 10px 14px; color: var(--ink-1);">W1-S04 + W1-S05 · LEDGER scaffolding · Adaptive Card mock · 3 demo paths · shadow-run validation · W1 exit gate</td>
      </tr>
      <tr style="border-bottom: 1px solid var(--line);">
        <td style="padding: 10px 14px; font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--teal); letter-spacing: .14em;">Q4</td>
        <td style="padding: 10px 14px; color: var(--ink-1);">W2-S01 · live engagement-event stream · cohort segmentation · privacy/consent integration · 250-store cutover prep</td>
      </tr>
      <tr style="border-bottom: 1px solid var(--line);">
        <td style="padding: 10px 14px; font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--teal); letter-spacing: .14em;">Q5</td>
        <td style="padding: 10px 14px; color: var(--ink-1);">W2-S02 · Content-Match + Offer-Compose live · Concession Policy Studio · markdown-NPV trending dashboard</td>
      </tr>
      <tr>
        <td style="padding: 10px 14px; font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--teal); letter-spacing: .14em;">Q6</td>
        <td style="padding: 10px 14px; color: var(--ink-1);">W2-S03 + W2-S04 · all 5 agents live · 90-day run · FIRST quarterly FSMA evidence packet to FDA Relations · W2 exit gate</td>
      </tr>
    </table>
  </div>
</div>

<!-- Slide 15 — What we need from Kroger -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 15 · WHAT WE NEED FROM KROGER</div><div class="slide-num">15 / 16</div></div>
  <h2>The five Approach-stage commitments.</h2>
  <div class="three-col" style="grid-template-columns: 1fr 1fr;">
    <div class="col" style="--accent: var(--gold);">
      <div class="col-eye">Commitment 1</div>
      <h3>Stakeholder access</h3>
      <p>Working sessions with Chief Merchant, Pricing Director, VP Quality + Food Safety Lead, CIO, CDO, and CISO in the 4-week Approach window.</p>
    </div>
    <div class="col" style="--accent: var(--teal);">
      <div class="col-eye">Commitment 2</div>
      <h3>Read-only data access</h3>
      <p>Cincinnati banner cluster: SAP S/4HANA · Manhattan WMS · Revionics · 84.51° loyalty feed · refrigeration EMS · Trace One/FoodLogiQ.</p>
    </div>
    <div class="col" style="--accent: var(--violet);">
      <div class="col-eye">Commitment 3</div>
      <h3>External-audit walkthrough</h3>
      <p>Joint design session with Internal Audit + external auditor in W1 for FSMA evidence packet format. We bring agenda; Kroger brings auditor.</p>
    </div>
    <div class="col" style="--accent: var(--amber);">
      <div class="col-eye">Commitment 4</div>
      <h3>Steering Committee charter</h3>
      <p>Cross-functional Steering with Merch + Store Ops + Food Safety + IT representation. Monthly cadence. We facilitate; Kroger chairs (recommend COO).</p>
    </div>
  </div>
  <div style="height: 18px;"></div>
  <div style="display: grid; grid-template-columns: 1fr; max-width: 60%;">
    <div class="col" style="--accent: var(--crimson);">
      <div class="col-eye">Commitment 5</div>
      <h3>Privacy &amp; Independence routing</h3>
      <p>Privacy Office walkthrough on consent governance + personalization-evidence shape. SOW routed through Deloitte's Independence office in concert with Kroger's external-audit relationship.</p>
    </div>
  </div>
</div>

<!-- Slide 16 — Next steps -->
<div class="slide">
  <div class="slide-head"><div>SLIDE 16 · NEXT STEPS</div><div class="slide-num">16 / 16</div></div>
  <h2>From this conversation to engagement signoff in <em>5 weeks</em>.</h2>
  <div class="cta-block">
    <h3>Recommended path</h3>
    <p>Approach stage scoped at $200K–$500K. 3–5 weeks. Risk-managed entry point that converts to W1 commit when both parties are aligned.</p>
    <div class="cta-list">
      <div class="cta-item">
        <div class="step">Week 1</div>
        <h4>Stakeholder discovery</h4>
        <p>10+ persona interviews. Pain-point quantification. Banner-scope identification.</p>
      </div>
      <div class="cta-item">
        <div class="step">Week 2</div>
        <h4>Source-system state-of-play</h4>
        <p>SAP, WMS, Revionics, 84.51°, EMS, Trace One coverage and access path.</p>
      </div>
      <div class="cta-item">
        <div class="step">Week 3</div>
        <h4>Risk register + commercial sizing</h4>
        <p>20-risk identification, mid-band envelope sizing, ROI sensitivity bounds.</p>
      </div>
      <div class="cta-item">
        <div class="step">Week 4</div>
        <h4>SOW + Independence routing</h4>
        <p>Independence-office signoff, SOW execution, Audit Committee notification, Wave-1 kickoff scheduled.</p>
      </div>
    </div>
  </div>
  <div class="signoff" style="margin-top: 36px;">— Deloitte Microsoft Technology Solutions &amp; Practices</div>
  <div class="signoff-mono">v1.0 · 27 Apr 2026 · APEX RC-E2E-03 + RC-E2E-09 · Kroger reference engagement</div>
</div>

<script>
function toggleTheme() {
  const body = document.body;
  body.classList.toggle('light');
  const icon = document.getElementById('themeIcon'), label = document.getElementById('themeLabel');
  if (body.classList.contains('light')) { icon.textContent = '☾'; label.textContent = 'Dark'; localStorage.setItem('apex-theme', 'light'); }
  else { icon.textContent = '☀'; label.textContent = 'Light'; localStorage.setItem('apex-theme', 'dark'); }
}
(function() {
  const saved = localStorage.getItem('apex-theme');
  if (saved === 'light') {
    document.body.classList.add('light');
    document.getElementById('themeIcon').textContent = '☾';
    document.getElementById('themeLabel').textContent = 'Dark';
  }
})();
</script>
</body></html>
"""

print("\nWriting Pitch Deck html...")
target = OUT / "APEX-RC-E2E-03-Kroger-Pitch-Deck.html"
target.write_text(PITCH_DECK, encoding="utf-8")
print(f"  wrote {target.name}")

build_risk_register()
build_stakeholder_map()
print("\nTier 3 complete.")

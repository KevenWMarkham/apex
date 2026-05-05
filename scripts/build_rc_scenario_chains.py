"""Build APEX-RC-Scenario-Chains.xlsx + .html — Retail & Consumer industry scenario library.

Generates 500+ R&C scenarios broken out by Domain + Sub-Area, mirroring the format of
docs/reference/APEX-Scenario-Chains.xlsx. Produces:
  - docs/reference/Retail and Consumer/APEX-RC-Scenario-Chains.xlsx
  - docs/reference/Retail and Consumer/APEX-RC-Scenario-Chains.html
"""
from __future__ import annotations
import io, sys, html, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT   = Path(r"C:\Stage\Clients\Industries\APEX")
OUTDIR = ROOT / "docs/reference/Retail and Consumer"
OUTDIR.mkdir(parents=True, exist_ok=True)
OUT_XL = OUTDIR / "APEX-RC-Scenario-Chains.xlsx"
OUT_HT = OUTDIR / "APEX-RC-Scenario-Chains.html"

# -------------------------------------------------------------------------
# Service code → Deloitte unit mapping
# -------------------------------------------------------------------------
SERVICES = {
    "RC-E2E-01": "Demand Sensing & Planning",
    "RC-E2E-02": "Inventory & Replenishment",
    "RC-E2E-03": "Assortment & Pricing",
    "RC-E2E-04": "Customer Lifecycle & Loyalty",
    "RC-E2E-05": "Store Operations",
    "RC-E2E-06": "Supply Chain & Logistics",
    "RC-E2E-07": "Returns & Refund Integrity",
    "RC-E2E-08": "Workforce & Talent",
    "RC-E2E-09": "Product Tracking & Provenance",
    "RC-E2E-10": "Trade Promotion & RGM",
    "RC-E2E-11": "Manufacturing & Plant Ops",
    "RC-E2E-12": "Foodservice & Restaurant Ops",
    "RC-E2E-13": "Sustainability & ESG",
    "RC-E2E-14": "Real Estate & Network",
    "RC-E2E-15": "Finance, Risk & Controls",
}

PERSONA_BY_DOMAIN = {
    "Merchandising & Assortment":   "Daniel Chen · Merchandising Director",
    "Pricing & Promotion":          "Priya Shah · Pricing Strategy Lead",
    "Marketing & Customer Experience": "Aisha Williams · CMO Office",
    "Loyalty & CRM":                "Maya Patel · Loyalty & CRM Director",
    "Supply Chain & Logistics":     "Marcus Greene · VP Supply Chain",
    "Store Operations":             "Jamie O'Connor · Store Operations Lead",
    "Workforce & Talent":           "Lena Kowalski · Field HR Director",
    "E-commerce & Digital":         "Elena Vasquez · Director Digital Commerce",
    "Returns & Reverse Logistics":  "Rebecca Hall · Returns Operations Manager",
    "Risk · Fraud · Security":      "Omar Haddad · Loss Prevention Lead",
    "Finance & Revenue":            "Katherine Boyd · Controller",
    "Sustainability & ESG":         "Nikhil Rao · ESG Program Lead",
    "Real Estate & Network":        "Devon Park · Real Estate Director",
    "Manufacturing & CPG Plant Ops": "Sara Lindberg · Plant Operations Lead",
    "Trade Promotion & RGM":        "Carlos Mendez · Revenue Growth Lead",
    "New Product Introduction":     "Ben Tanaka · Innovation Lead",
    "Foodservice & Restaurant Ops": "Riley Barnes · Restaurant Operations Director",
}

# Schemas referenced per domain
SCHEMA_BY_DOMAIN = {
    "Merchandising & Assortment":   "MERML.Item · MERML.Assortment · MERML.Vendor",
    "Pricing & Promotion":          "MERML.Price · MERML.Markdown · MERML.Promotion · MERML.Elasticity",
    "Marketing & Customer Experience": "CXML.Campaign · CXML.Audience · CXML.Touchpoint",
    "Loyalty & CRM":                "CXML.LoyaltyMember · CXML.Offer · CXML.Tier",
    "Supply Chain & Logistics":     "SCML.Inventory · SCML.Shipment · SCML.Lot",
    "Store Operations":             "SCML.Inventory · MERML.Planogram · OPSML.Task",
    "Workforce & Talent":           "OPSML.Schedule · OPSML.Worker · OPSML.LaborForecast",
    "E-commerce & Digital":         "CXML.Session · CXML.Cart · MERML.Item",
    "Returns & Reverse Logistics":  "SCML.Return · MERML.Transaction · CXML.Customer",
    "Risk · Fraud · Security":      "RISKML.Event · CXML.Customer · MERML.Transaction",
    "Finance & Revenue":            "FINML.GL · FINML.Settlement · FINML.Tax",
    "Sustainability & ESG":         "ESGML.Emission · SCML.Lot · MERML.Item",
    "Real Estate & Network":        "OPSML.Site · FINML.Lease · OPSML.Format",
    "Manufacturing & CPG Plant Ops": "MFGML.Asset · MFGML.Run · MFGML.Quality",
    "Trade Promotion & RGM":        "MERML.Promotion · FINML.TradeSpend · MERML.Deduction",
    "New Product Introduction":     "MERML.Item · MERML.Concept · CXML.Audience",
    "Foodservice & Restaurant Ops": "MERML.Recipe · OPSML.KitchenRun · CXML.Order",
}

# -------------------------------------------------------------------------
# Domain → Sub-Area → list of (slug, title, brief, KPI, service_code)
# Each scenario is unique. Targeting 500+.
# -------------------------------------------------------------------------

CATALOG = {
"Merchandising & Assortment": {
    "Assortment Planning": [
        ("range-architecture-by-cluster", "Range architecture by store cluster", "Cluster-aware range architecture with delist/keep/test recommendations per region.", "+3.1pp gross margin", "RC-E2E-03"),
        ("assortment-rationalization-tail", "Tail-SKU rationalization", "SKU-level contribution-margin analysis with delist candidate list for long-tail items.", "+2.4pp gross margin", "RC-E2E-03"),
        ("good-better-best-tiering", "Good-better-best price tiering", "Tier coverage analysis across price ladders with gap-fill recommendations.", "+1.8pp basket value", "RC-E2E-03"),
        ("local-assortment-personalization", "Local assortment personalization", "Demographic and demand-pattern-aware local SKU swaps for neighborhood stores.", "+4.2% same-store sales", "RC-E2E-03"),
        ("seasonal-buy-quantity-optimization", "Seasonal buy-quantity optimization", "Demand-curve-aware seasonal buy quantity with hedged depth across SKUs.", "−18% end-of-season markdown", "RC-E2E-01"),
        ("white-space-opportunity-detection", "White-space opportunity detection", "Cross-banner sales-gap analysis with new-item entry recommendations.", "+$11M incremental revenue", "RC-E2E-03"),
        ("cluster-redefinition", "Store-cluster redefinition", "Behavioral cluster re-segmentation using transaction and demographic shift data.", "+2.7pp assortment fit", "RC-E2E-03"),
        ("dead-net-cost-rebalance", "Dead-net cost rebalance", "Dead-net-cost-aware assortment with vendor mix shift to maximize gross profit.", "+1.6pp gross margin", "RC-E2E-03"),
        ("category-role-realignment", "Category role realignment", "Traffic / destination / convenience role reclassification with assortment depth shift.", "+5% category traffic", "RC-E2E-03"),
        ("micro-fulfillment-assortment", "Micro-fulfillment assortment fit", "Dark-store / micro-fulfillment SKU-fit and pick-density optimization.", "+22% pick rate", "RC-E2E-03"),
    ],
    "Category Management": [
        ("category-line-review-acceleration", "Category line-review acceleration", "Generative line-review prep: vendor scorecards, gap analysis, term modeling.", "−40% line-review prep time", "RC-E2E-03"),
        ("vendor-scorecard-automation", "Vendor scorecard automation", "Service / fill / quality / dead-net composite vendor scorecard with action triggers.", "+1.4pp fill rate", "RC-E2E-03"),
        ("category-margin-leak-detection", "Category margin-leak detection", "Multi-factor margin-erosion attribution: cost, mix, markdown, freight, shrink.", "$6.4M leakage recovered", "RC-E2E-03"),
        ("category-co-management-private", "Category co-management with private label", "Brand / private-label balance optimization in category architecture.", "+2.1pp own-brand penetration", "RC-E2E-03"),
        ("term-negotiation-prep-agent", "Term-negotiation prep agent", "Vendor-term benchmark with simulated counter-offer P&L impact.", "+0.9pp net cost improvement", "RC-E2E-03"),
        ("hierarchy-cleansing-master-data", "Hierarchy cleansing & master data", "Item-master hierarchy harmonization with duplicate detection and re-attribution.", "92% master-data quality score", "RC-E2E-09"),
        ("brand-block-spacing-analysis", "Brand-block spacing analysis", "Brand-block share-of-shelf alignment to share-of-market with re-block recommendation.", "+3.3% category sales", "RC-E2E-03"),
        ("category-captain-collab-workflow", "Category-captain collaboration workflow", "Vendor captain co-planning workflow with audit-row trace and IP ringfence.", "−30% planning cycle", "RC-E2E-03"),
        ("attribute-enrichment-from-supplier", "Attribute enrichment from supplier feeds", "Auto-extract structured attributes from supplier docs/spec sheets to enrich item master.", "+87% attribute coverage", "RC-E2E-09"),
    ],
    "Space & Planogram": [
        ("planogram-compliance-monitoring", "Planogram compliance monitoring", "CV-assisted shelf-image compliance check with remediation tasks for store ops.", "+91% compliance rate", "RC-E2E-05"),
        ("planogram-auto-generation", "Planogram auto-generation", "Velocity / margin / facing-rule-aware planogram generation per store cluster.", "−65% plano cycle time", "RC-E2E-05"),
        ("end-cap-effectiveness", "End-cap effectiveness", "End-cap incrementality measurement with rotation-cadence recommendation.", "+18% end-cap lift", "RC-E2E-03"),
        ("micro-space-localization", "Micro-space localization", "Per-store micro-space tweaks (last 5% facings) using local sell-through.", "+1.2% same-store comp", "RC-E2E-05"),
        ("space-to-sales-rebalance", "Space-to-sales rebalance", "Quarterly category space-to-sales index with reset queue.", "+2.5pp category margin", "RC-E2E-05"),
        ("ada-accessibility-audit-shelves", "ADA accessibility audit on shelves", "Shelf-height and cross-aisle accessibility audit with corrective task list.", "100% ADA compliance", "RC-E2E-05"),
        ("cooler-door-optimization", "Cooler-door optimization", "Cold-vault SKU placement using basket affinity and dwell heatmap.", "+5% cold-beverage attach", "RC-E2E-05"),
        ("front-end-checkout-merch", "Front-end checkout merch", "Front-end fixture cycle and SKU rotation tied to dwell-time and impulse lift.", "+11% impulse-buy lift", "RC-E2E-05"),
    ],
    "New Item Introduction": [
        ("new-item-launch-readiness", "New-item launch readiness", "Cross-functional launch checklist with gating signals and slip-risk score.", "−27% launch defects", "RC-E2E-03"),
        ("new-item-velocity-prediction", "New-item velocity prediction", "Pre-launch velocity prediction using analog-item embeddings and panel data.", "+14% forecast accuracy", "RC-E2E-01"),
        ("first-90-day-tracking", "First-90-day tracking", "New-item performance tracking with delist / scale / pivot recommendation.", "−22% failed-item rate", "RC-E2E-03"),
        ("new-item-shelf-set-acceleration", "New-item shelf-set acceleration", "Auto-generate shelf-set plan to compress slot-day-to-shelf time.", "−40% time-to-shelf", "RC-E2E-05"),
        ("supplier-onboarding-data-completeness", "Supplier onboarding data completeness", "Supplier-side data and spec completeness check before slot-allocation.", "+95% onboarding completeness", "RC-E2E-09"),
    ],
    "Private Label & Own Brand": [
        ("private-label-margin-modeling", "Private-label margin modeling", "PL margin and price-pack architecture with co-pack capacity check.", "+3.8pp PL margin", "RC-E2E-03"),
        ("pl-substitution-modeling", "Private-label substitution modeling", "Modeling of cannibalization and incrementality of new PL launches.", "+1.4pp PL share", "RC-E2E-03"),
        ("pl-brand-equity-tracker", "PL brand-equity tracker", "Cross-channel PL sentiment, NPS, and equity-vs-NB benchmark.", "+9pt PL equity", "RC-E2E-04"),
        ("pl-quality-complaint-routing", "PL quality complaint routing", "Customer complaint routing to PL co-packer with corrective action workflow.", "−35% complaint resolution time", "RC-E2E-09"),
    ],
    "Vendor Collaboration": [
        ("vendor-fill-rate-collab", "Vendor fill-rate collaboration", "Joint fill-rate scorecard with weekly OTIF action queue.", "+4.2pp OTIF", "RC-E2E-06"),
        ("vendor-funded-promo-effectiveness", "Vendor-funded promo effectiveness", "Trade-funded promo ROI with auto-deduction reconciliation.", "+12% trade-spend ROI", "RC-E2E-10"),
        ("dropship-vendor-onboarding", "Dropship-vendor onboarding", "Dropship vendor scorecard, SLA gating, and chargeback automation.", "+2.1pp dropship OTD", "RC-E2E-06"),
        ("collaborative-forecasting-cpfr", "Collaborative forecasting / CPFR", "Joint demand forecast share-out with vendor with reconciliation workflow.", "+2.8pp forecast accuracy", "RC-E2E-01"),
        ("vendor-compliance-chargeback", "Vendor compliance chargeback", "Auto-detect and chargeback compliance violations with audit trail.", "$4.8M recovered", "RC-E2E-15"),
    ],
},
"Pricing & Promotion": {
    "Base Pricing": [
        ("zone-base-pricing-optimization", "Zone-base price optimization", "Per-zone base-price elasticity model with margin-vs-traffic balance.", "+1.7pp margin", "RC-E2E-03"),
        ("regular-price-architecture", "Regular price architecture", "Price-line and ladder cleanup to reduce psychological-point gaps.", "+0.9pp basket value", "RC-E2E-03"),
        ("kvi-key-value-item-management", "Key-value-item management", "KVI list dynamic refresh with price-image impact tracking.", "+2.4pp price image", "RC-E2E-03"),
        ("price-rounding-strategy", "Price-rounding strategy", "Rounding-rule rebalance to maximize perceived value vs realized margin.", "+0.4pp margin", "RC-E2E-03"),
        ("competitor-price-match-rules", "Competitor price-match rules", "Auto-match-rule tuning with revenue-protection guardrails.", "−7% match-program cost", "RC-E2E-03"),
        ("price-image-monitoring", "Price-image monitoring", "Cross-banner price-image index against shopper perception survey.", "+5pt price-image index", "RC-E2E-04"),
        ("dual-pricing-online-store-alignment", "Dual-pricing online vs store alignment", "Channel-price gap monitoring with reconciliation workflow.", "−42% channel mismatch alerts", "RC-E2E-08"),
    ],
    "Markdown & Clearance": [
        ("dynamic-markdown-optimization-rc", "Dynamic markdown optimization", "Per-SKU-per-store markdown with simulated P&L impact and HITL gate.", "+3.2pp gross margin", "RC-E2E-03"),
        ("seasonal-clearance-cadence", "Seasonal-clearance cadence", "Multi-step markdown cadence to clear seasonal stock at target margin.", "+19% sell-through at target margin", "RC-E2E-03"),
        ("aged-inventory-trigger", "Aged-inventory markdown trigger", "Days-on-hand threshold-based markdown trigger with vendor co-fund check.", "−28% aged-stock days", "RC-E2E-03"),
        ("damaged-merch-clearance-routing", "Damaged-merchandise clearance routing", "Auto-route damaged items to clearance, donation, salvage, or destroy.", "+$2.1M recovered", "RC-E2E-07"),
        ("price-flex-by-store-velocity", "Price flex by store velocity", "Per-store markdown depth based on local velocity and competitor proximity.", "+1.9pp clearance margin", "RC-E2E-03"),
        ("markdown-lever-coordination-promo", "Markdown / promo lever coordination", "Coordinate markdown and promo levers to avoid stack-discount margin erosion.", "+1.1pp blended margin", "RC-E2E-03"),
        ("perishable-rolling-markdown", "Perishable rolling markdown", "Code-date-aware rolling markdown for perishables with shrink-cost tradeoff.", "−24% perishable shrink", "RC-E2E-06"),
    ],
    "Promotion Optimization": [
        ("promo-mix-optimization", "Promo mix optimization", "Mix-of-mechanic optimization (BOGO / pct-off / bundle) for max incremental margin.", "+8% promo ROI", "RC-E2E-03"),
        ("circular-page-allocation", "Circular page allocation", "Page-feature SKU allocation modeled on incrementality and vendor co-fund.", "+11% circular incrementality", "RC-E2E-03"),
        ("promo-cannibalization-detection", "Promo cannibalization detection", "Halo / cannibalization decomposition across substitute and complement SKUs.", "+1.6pp net promo lift", "RC-E2E-03"),
        ("promotion-fatigue-monitoring", "Promotion-fatigue monitoring", "Detect SKU and category promo-fatigue and redirect investment.", "+6% promo lift recovered", "RC-E2E-03"),
        ("digital-coupon-targeting", "Digital coupon targeting", "Targeted digital-coupon distribution by elasticity and lifetime-value tier.", "+14% redemption / member", "RC-E2E-04"),
        ("promo-creative-effectiveness", "Promo creative effectiveness", "Creative-element-level (image, copy, offer) impact attribution.", "+12% click-through rate", "RC-E2E-04"),
        ("burnout-rest-cycle-management", "Promo burnout rest-cycle management", "Mandatory rest-week scheduling for over-promoted SKUs.", "+9% promo lift recovery", "RC-E2E-03"),
        ("supplier-funded-promo-uplift", "Supplier-funded promo uplift", "Net-of-trade-funding promo ROI with deduction-tie-out.", "+7% net trade ROI", "RC-E2E-10"),
    ],
    "Competitive Pricing": [
        ("competitive-shelf-price-scrape", "Competitive shelf-price scrape", "Multi-banner shelf-price scrape with price-rule-violation alert.", "+1.3pp margin protected", "RC-E2E-03"),
        ("competitor-promo-detection", "Competitor promo detection", "Detect competitor promos and trigger countermeasure workflow within hours.", "−18% share-loss events", "RC-E2E-03"),
        ("market-basket-price-comparison", "Market-basket price comparison", "Cross-banner basket-level price index for competitive positioning.", "+3pt price-position index", "RC-E2E-03"),
        ("local-competitor-elasticity-tuning", "Local-competitor elasticity tuning", "Local elasticity recalibration when new competitor opens nearby.", "−6% local share-loss", "RC-E2E-03"),
        ("min-advertised-price-enforcement", "Minimum-advertised-price enforcement", "MAP-violation detection on third-party marketplaces with takedown workflow.", "+88% MAP compliance", "RC-E2E-15"),
    ],
    "Personalized Pricing": [
        ("personalized-offer-targeting", "Personalized offer targeting", "Per-member offer targeting using elasticity and engagement signals.", "+22% redemption rate", "RC-E2E-04"),
        ("loyalty-tier-pricing-experiments", "Loyalty-tier pricing experiments", "A/B test of tier-specific pricing for retention and trade-up.", "+1.4pp top-tier retention", "RC-E2E-04"),
        ("ltv-aware-discount-depth", "LTV-aware discount depth", "Discount depth modulated by LTV-tier rather than blanket SKU rule.", "+9% LTV / discount $", "RC-E2E-04"),
        ("opt-in-personal-prices", "Opt-in personal prices", "Member-opt-in personalized base prices with consent-trail audit.", "+11% opt-in member sales", "RC-E2E-04"),
    ],
    "Yield Management": [
        ("dynamic-perishable-yield", "Dynamic perishable yield", "Real-time perishable price adjustment based on shelf-life and demand.", "−21% perishable shrink", "RC-E2E-06"),
        ("event-day-yield-management", "Event-day yield management", "Game-day, weather-event, and holiday yield-management per store.", "+8% revenue per event", "RC-E2E-03"),
        ("convention-center-event-pricing", "Convention-center event pricing", "Event-driven price-flex for QSR locations near convention venues.", "+12% event-day sales", "RC-E2E-12"),
        ("daypart-yield-management", "Daypart yield management", "Daypart-specific yield-management for foodservice and grocery deli.", "+9% daypart revenue", "RC-E2E-12"),
    ],
    "Promo Compliance & Audit": [
        ("promo-execution-compliance", "Promo-execution compliance", "Store-level promo-execution compliance with corrective task list.", "+92% execution compliance", "RC-E2E-05"),
        ("promo-margin-leakage", "Promo margin-leakage detection", "Detect promo-related margin leakage at SKU / store level.", "+1.1pp blended margin", "RC-E2E-03"),
        ("promo-creative-rights-audit", "Promo creative rights audit", "Audit promo creative for talent / IP / co-brand rights compliance.", "+99% rights compliance", "RC-E2E-15"),
    ],
},
"Marketing & Customer Experience": {
    "Campaign Management": [
        ("campaign-orchestration-engine", "Campaign orchestration engine", "Multi-touch campaign orchestration with consent and frequency caps.", "+18% campaign throughput", "RC-E2E-04"),
        ("creative-asset-generation", "Creative-asset generation", "Generative creative production tied to brand library and market localization.", "−55% creative production cost", "RC-E2E-04"),
        ("media-mix-modeling", "Media-mix modeling", "Bayesian MMM with channel saturation and adstock modeling.", "+12% media efficiency", "RC-E2E-04"),
        ("attribution-modeling-multi-touch", "Multi-touch attribution modeling", "Multi-touch attribution across paid, owned, earned channels.", "+22% marketing ROI", "RC-E2E-04"),
        ("creative-fatigue-detection", "Creative fatigue detection", "Detect creative fatigue and trigger refresh workflow with brand-safe variants.", "+9% engagement recovery", "RC-E2E-04"),
        ("campaign-incrementality-test-design", "Campaign incrementality-test design", "Auto-generate ghost-cell holdout designs and read-outs.", "+15% measurable lift", "RC-E2E-04"),
        ("dynamic-content-syndication", "Dynamic content syndication", "Auto-syndicate localized content variants across owned channels.", "+27% content reach", "RC-E2E-04"),
    ],
    "Personalization": [
        ("real-time-recommendation", "Real-time recommendation", "On-site recommender tied to in-session intent and assortment availability.", "+14% AOV uplift", "RC-E2E-04"),
        ("category-affinity-personalization", "Category-affinity personalization", "Cross-category affinity model for cross-sell and category expansion.", "+8% category trip rate", "RC-E2E-04"),
        ("personalized-homepage-merchandising", "Personalized homepage merchandising", "Per-segment homepage SKU rotation with merchandising guardrails.", "+11% homepage CTR", "RC-E2E-04"),
        ("life-stage-segmentation", "Life-stage segmentation", "Identify life-stage shifts (new parent, mover, retiree) and switch journeys.", "+19% segment-relevance score", "RC-E2E-04"),
        ("opt-in-permissioned-personalization", "Opt-in permissioned personalization", "Personalization tied to consent ledger with data-handling audit row.", "+8pp consent rate", "RC-E2E-04"),
        ("next-best-offer-engine", "Next-best-offer engine", "NBO scoring engine integrated with offer-management catalog.", "+17% offer redemption", "RC-E2E-04"),
        ("dynamic-email-content-blocks", "Dynamic email content blocks", "Per-recipient dynamic email blocks with realtime inventory check.", "+13% email-driven revenue", "RC-E2E-04"),
    ],
    "Content & Creative": [
        ("brand-safe-creative-generation", "Brand-safe creative generation", "Generative creative gated by brand voice and asset-rights checker.", "−60% creative cycle time", "RC-E2E-04"),
        ("ugc-content-curation", "UGC content curation", "User-generated-content sourcing with rights and brand-safety scoring.", "+18% UGC-attributed conversion", "RC-E2E-04"),
        ("dam-asset-tagging-automation", "DAM asset-tagging automation", "Auto-tag and classify DAM assets for retrievability and compliance.", "+72% asset reuse", "RC-E2E-04"),
        ("video-thumbnail-optimization", "Video-thumbnail optimization", "Thumbnail-image variant testing for view-through rate optimization.", "+9% video CTR", "RC-E2E-04"),
        ("multilingual-localization", "Multilingual content localization", "Brand-glossary-aware localization across markets with reviewer workflow.", "+24% market-launch speed", "RC-E2E-04"),
    ],
    "Attribution & MMM": [
        ("incrementality-test-orchestration", "Incrementality-test orchestration", "Always-on geo-holdout incrementality testing across channels.", "+15% measured incremental ROAS", "RC-E2E-04"),
        ("bayesian-mmm", "Bayesian media-mix model", "Hierarchical Bayesian MMM with retailer-segment priors.", "+12% allocation efficiency", "RC-E2E-04"),
        ("touchpoint-decay-modeling", "Touchpoint decay modeling", "Exposure-decay modeling per channel for attribution windowing.", "+7% attribution accuracy", "RC-E2E-04"),
        ("offline-online-attribution-stitch", "Offline-online attribution stitch", "Stitch in-store transactions to digital exposures via loyalty match.", "+19% offline ROAS visibility", "RC-E2E-04"),
        ("retail-media-network-roi", "Retail-media network ROI", "RMN exposure-to-purchase attribution with margin-aware ROAS.", "+21% RMN net margin", "RC-E2E-04"),
    ],
    "Brand Health & Voice of Customer": [
        ("brand-equity-social-signal", "Brand-equity social-signal tracking", "Cross-platform brand-sentiment and share-of-voice tracking.", "+6pt brand score", "RC-E2E-04"),
        ("brand-health-sentiment", "Brand-health sentiment tracking", "Cross-channel brand-health and competitive sentiment tracking.", "+7pt NPS", "RC-E2E-04"),
        ("review-summarization-categorization", "Review summarization & categorization", "Summarize and categorize product reviews to feed merch and NPD.", "−45% review-handling cost", "RC-E2E-04"),
        ("complaint-root-cause-clustering", "Complaint root-cause clustering", "Cluster complaints by root-cause for quality and supplier action.", "−18% recurring complaints", "RC-E2E-04"),
        ("nps-driver-decomposition", "NPS driver decomposition", "Decompose NPS into causal drivers (price, service, quality, availability).", "+3pt NPS via targeted action", "RC-E2E-04"),
        ("crisis-detection-social", "Crisis-detection social monitor", "Detect emerging brand-crisis signals and trigger comms workflow.", "−72% crisis-response latency", "RC-E2E-04"),
    ],
    "Customer Service & Care": [
        ("contact-center-deflection", "Contact-center deflection", "Smart contact-center deflection to self-service for routine inquiries.", "−27% contact volume", "RC-E2E-04"),
        ("agent-assist-copilot", "Agent-assist Copilot", "Real-time agent-assist Copilot with knowledge-base retrieval.", "−32% AHT", "RC-E2E-04"),
        ("voice-of-customer-ingestion", "Voice-of-customer ingestion", "Multi-channel VOC ingestion with theme extraction and routing.", "+18% VOC throughput", "RC-E2E-04"),
        ("escalation-pattern-detection", "Escalation-pattern detection", "Detect escalation patterns early and route to senior agent.", "+22pt escalation NPS", "RC-E2E-04"),
        ("self-service-knowledge-curation", "Self-service knowledge curation", "Auto-curate self-service knowledge from agent and ticket signals.", "+14% self-service success", "RC-E2E-04"),
        ("post-issue-recovery-comms", "Post-issue recovery comms", "Post-issue recovery communications with empathy-tuned messaging.", "+24% post-issue retention", "RC-E2E-04"),
    ],
},
"Loyalty & CRM": {
    "Membership Acquisition": [
        ("acquisition-channel-attribution-loyalty", "Loyalty acquisition channel attribution", "Channel-level attribution of loyalty signups with cost-per-active-member.", "−18% CPA per active", "RC-E2E-04"),
        ("acquisition-incentive-elasticity", "Acquisition incentive elasticity", "Optimize signup-incentive depth by channel and segment.", "+14% activation rate", "RC-E2E-04"),
        ("dormant-account-reactivation", "Dormant-account reactivation", "Reactivation campaigns for dormant signups with retention-likelihood scoring.", "+9% reactivation rate", "RC-E2E-04"),
        ("partner-loyalty-coalition", "Partner-loyalty coalition", "Coalition-loyalty cross-program acquisition with revenue-share reconciliation.", "+22% incremental signups", "RC-E2E-04"),
    ],
    "Tier & Status Management": [
        ("tier-upgrade-prediction", "Tier-upgrade prediction", "Predict tier-up candidates and run nudge campaigns.", "+12% tier-up rate", "RC-E2E-04"),
        ("status-extension-modeling", "Status-extension modeling", "Model status-extension impact on retention and revenue.", "+8% top-tier retention", "RC-E2E-04"),
        ("tier-economics-rebalance", "Tier economics rebalance", "Rebalance tier-benefit economics to align cost with member value.", "−6% per-member benefit cost", "RC-E2E-04"),
        ("benefit-utilization-tracking", "Benefit-utilization tracking", "Track benefit utilization to surface dormant or overused benefits.", "+11% benefit-driven retention", "RC-E2E-04"),
    ],
    "Retention & Winback": [
        ("loyalty-churn-prediction-winback", "Loyalty-churn prediction & winback", "Top-tier silent-churn prediction with personalized winback offers.", "−22% top-tier churn", "RC-E2E-04"),
        ("at-risk-member-flag", "At-risk member flag", "Real-time at-risk flag based on engagement and transaction patterns.", "+17% retention save rate", "RC-E2E-04"),
        ("winback-offer-depth-optimization", "Winback offer-depth optimization", "Optimize winback offer depth by churn risk and LTV potential.", "+24% winback ROI", "RC-E2E-04"),
        ("post-purchase-onboarding", "Post-purchase onboarding", "Smart post-purchase onboarding sequences for category penetration.", "+13% category cross-shop", "RC-E2E-04"),
        ("churn-cause-decomposition", "Churn-cause decomposition", "Decompose churn into competitive, life-event, service-failure, value drivers.", "+9% targeted save rate", "RC-E2E-04"),
    ],
    "Lifetime Value": [
        ("ltv-prediction-modeling", "LTV prediction modeling", "Member-level LTV prediction with confidence intervals.", "+18% LTV / acquisition $", "RC-E2E-04"),
        ("ltv-driver-attribution", "LTV driver attribution", "Decompose LTV drivers (frequency, basket, longevity, cross-shop).", "+12% LTV growth", "RC-E2E-04"),
        ("share-of-wallet-modeling", "Share-of-wallet modeling", "Estimate competitive share-of-wallet to size growth headroom.", "+6pp SoW", "RC-E2E-04"),
    ],
    "Rewards & Offers": [
        ("offer-engine-personalization", "Offer-engine personalization", "Per-member offer engine tuned to elasticity and engagement.", "+19% offer redemption", "RC-E2E-04"),
        ("reward-redemption-balance", "Reward-redemption balance", "Balance breakage / liability vs engagement in reward economy.", "−8% rewards liability", "RC-E2E-15"),
        ("partner-reward-redemption", "Partner-reward redemption", "Partner-reward catalog with real-time inventory and reconciliation.", "+24% partner redemption", "RC-E2E-04"),
        ("offer-orchestration-omnichannel", "Offer orchestration omnichannel", "Cross-channel single-offer orchestration with attribution clarity.", "+13% net offer lift", "RC-E2E-04"),
    ],
    "Segment Modeling": [
        ("rfm-segment-evolution", "RFM segment evolution", "Auto-update RFM segments and trigger segment-shift campaigns.", "+9% segment-driven revenue", "RC-E2E-04"),
        ("micro-segmentation-engine", "Micro-segmentation engine", "Sub-segment discovery for hyper-targeted campaigns.", "+15% campaign relevance", "RC-E2E-04"),
        ("look-alike-modeling", "Look-alike modeling", "High-value member look-alike modeling for acquisition.", "+18% qualified leads", "RC-E2E-04"),
        ("propensity-to-buy-modeling", "Propensity-to-buy modeling", "Per-category propensity-to-buy scoring for offer targeting.", "+13% conversion lift", "RC-E2E-04"),
        ("household-graph-stitching", "Household graph stitching", "Cross-device / cross-channel household graph stitching with privacy guardrails.", "+22% household match rate", "RC-E2E-04"),
    ],
    "Loyalty Program Analytics": [
        ("loyalty-program-roi-attribution", "Loyalty-program ROI attribution", "Attribute incremental sales / margin to loyalty-program investment.", "+18% measured loyalty ROI", "RC-E2E-04"),
        ("partner-coalition-share-tracking", "Partner-coalition share tracking", "Partner-coalition spend-share tracking with revenue-share recon.", "+11% coalition net revenue", "RC-E2E-04"),
        ("loyalty-program-redesign-modeling", "Loyalty-program redesign modeling", "Multi-scenario loyalty-program redesign modeling.", "+14% program-economics improvement", "RC-E2E-04"),
    ],
},
"Supply Chain & Logistics": {
    "Demand Forecasting": [
        ("baseline-demand-forecasting", "Baseline demand forecasting", "Hierarchical demand forecast with reconciliation across DC / store / SKU.", "+3.2pp forecast accuracy", "RC-E2E-01"),
        ("new-product-demand-sensing", "New-product demand sensing", "Pre-launch demand-curve generation using analog item embeddings.", "+14% NPI forecast accuracy", "RC-E2E-01"),
        ("event-driven-demand-shaping", "Event-driven demand shaping", "Holiday / weather / event-driven demand shaping with override workflow.", "+5% event-week accuracy", "RC-E2E-01"),
        ("promotional-demand-decomposition", "Promotional demand decomposition", "Decompose baseline / promotional / halo / cannibalization signal in forecast.", "+2.8pp promo accuracy", "RC-E2E-01"),
        ("weather-driven-demand", "Weather-driven demand", "Weather-conditional demand for weather-sensitive categories.", "+6% weather-cat accuracy", "RC-E2E-01"),
        ("intermittent-demand-modeling", "Intermittent-demand modeling", "Croston-class modeling for slow-moving SKUs with safety-stock optimization.", "−12% slow-mover stockout", "RC-E2E-01"),
        ("price-elasticity-feedback-loop", "Price-elasticity feedback loop", "Continuous elasticity refresh from realized POS into forecast.", "+1.4pp price-driven accuracy", "RC-E2E-01"),
        ("hierarchical-reconciliation", "Hierarchical forecast reconciliation", "Top-down / bottom-up / middle-out reconciliation with audit trail.", "+2.1pp blended accuracy", "RC-E2E-01"),
        ("market-event-detection", "Market-event detection", "Detect anomalous demand events (recall, viral product) and recompute.", "−24% event-week stockouts", "RC-E2E-01"),
    ],
    "Inventory Management": [
        ("safety-stock-rightsizing", "Safety-stock rightsizing", "Per-SKU-per-DC safety-stock rightsizing balancing service level vs cost.", "−8% safety-stock cost", "RC-E2E-02"),
        ("inventory-positioning-network", "Network inventory positioning", "Cross-DC inventory positioning to align with demand pattern.", "+2.4pp service level", "RC-E2E-02"),
        ("phantom-inventory-detection", "Phantom-inventory detection", "POS-vs-perpetual reconciliation to detect phantom on-hand.", "−14% perpetual-inventory variance", "RC-E2E-02"),
        ("excess-obsolescence-flagging", "Excess & obsolescence flagging", "Proactive flagging of E&O risk with action recommendation.", "−18% E&O write-off", "RC-E2E-02"),
        ("multi-echelon-inventory", "Multi-echelon inventory", "Multi-echelon optimization across DC / cross-dock / store.", "+1.8pp service / inventory$", "RC-E2E-02"),
        ("seasonal-pre-position-strategy", "Seasonal pre-position strategy", "Pre-position seasonal inventory to capture pre-peak demand.", "+7% seasonal sell-through", "RC-E2E-02"),
        ("inventory-snapshot-discrepancy", "Inventory snapshot discrepancy", "Periodic-physical-vs-perpetual gap analysis with shrink decomposition.", "−21% adjustment volume", "RC-E2E-02"),
    ],
    "Replenishment": [
        ("auto-replenishment-tuning", "Auto-replenishment tuning", "Replenishment-rule auto-tuning with override governance.", "+1.6pp in-stock rate", "RC-E2E-02"),
        ("ladder-replenishment-promo", "Ladder replenishment for promos", "Promo-uplift-aware replenishment with stage-based load to store.", "+9% promo in-stock", "RC-E2E-02"),
        ("fresh-replenishment-perishable", "Fresh-replenishment for perishables", "Perishable replenishment with shelf-life and demand-curve modeling.", "−18% perishable shrink", "RC-E2E-02"),
        ("vendor-managed-inventory", "Vendor-managed inventory", "VMI program governance with vendor scorecard and visibility.", "+1.2pp VMI fill rate", "RC-E2E-02"),
        ("store-level-replen-min-max", "Store-level replen min/max", "Per-store min/max tuning with seasonality and trip-pattern overlays.", "+1.4pp on-shelf availability", "RC-E2E-02"),
        ("dropship-direct-fulfillment", "Dropship direct fulfillment", "Dropship-vendor SLA management with auto-rerouting on miss.", "+2.4pp dropship OTD", "RC-E2E-06"),
    ],
    "Warehouse & DC": [
        ("dc-receive-yard-orchestration", "DC receive yard orchestration", "Yard-and-dock orchestration to compress unload-to-stocked time.", "−22% dock-dwell time", "RC-E2E-06"),
        ("slotting-velocity-optimization", "Slotting velocity optimization", "Velocity-aware slotting with travel-distance minimization.", "+11% pick productivity", "RC-E2E-06"),
        ("warehouse-task-interleaving", "Warehouse task interleaving", "Pick / putaway / cycle-count task interleaving for efficiency.", "+9% labor productivity", "RC-E2E-08"),
        ("voice-and-vision-pick-validation", "Voice and vision pick validation", "Voice-and-vision pick validation for accuracy without speed loss.", "+18% pick accuracy", "RC-E2E-08"),
        ("dc-labor-forecasting", "DC labor forecasting", "Inbound / outbound volume-driven DC labor forecast.", "−11% labor-cost / unit", "RC-E2E-08"),
        ("inbound-load-prep-optimization", "Inbound load-prep optimization", "Load build / break-down planning for cross-dock efficiency.", "+8% trailer utilization", "RC-E2E-06"),
        ("returns-receiving-disposition", "Returns receiving disposition", "Disposition decisioning at DC returns receive (resell / refurb / salvage).", "+$3.2M recovery / yr", "RC-E2E-07"),
    ],
    "Transportation": [
        ("dynamic-route-optimization", "Dynamic route optimization", "Real-time route optimization with traffic, weather, and ETA inputs.", "+12% on-time delivery", "RC-E2E-06"),
        ("backhaul-truck-loading", "Backhaul truck-loading optimization", "Route-aware backhaul load planning to reduce empty miles.", "+12% backhaul utilization", "RC-E2E-06"),
        ("carrier-scorecard-otd", "Carrier scorecard / OTD", "Carrier OTD / damage scorecard with auto-tendering rule update.", "+2.8pp on-time-in-full", "RC-E2E-06"),
        ("freight-spend-leak-detection", "Freight-spend leak detection", "Audit freight invoices for leak (duplicate, accessorial misuse, mis-rated).", "$2.4M recovered", "RC-E2E-15"),
        ("appointment-scheduling-optimization", "Appointment-scheduling optimization", "Yard-appointment optimization with carrier-fairness logic.", "−16% detention cost", "RC-E2E-06"),
        ("intermodal-mode-shift", "Intermodal mode-shift", "Mode-shift modeling for cost-vs-service tradeoff (truck / rail / intermodal).", "−9% transportation cost", "RC-E2E-06"),
        ("carbon-aware-routing", "Carbon-aware routing", "Carbon-emission-aware route choice with cost-vs-Scope3 tradeoff.", "−14% Scope-3 transport emissions", "RC-E2E-13"),
    ],
    "Last-Mile & Fulfillment": [
        ("last-mile-orchestration", "Last-mile orchestration", "Multi-carrier last-mile orchestration with cost-vs-service routing.", "−11% last-mile cost", "RC-E2E-06"),
        ("delivery-window-prediction", "Delivery-window prediction", "Predictive narrow-window estimates for shopper transparency.", "+7pt delivery NPS", "RC-E2E-06"),
        ("ship-from-store-orchestration", "Ship-from-store orchestration", "Ship-from-store node selection with capacity and labor inputs.", "+18% SFS unit volume", "RC-E2E-06"),
        ("locker-pickup-network-management", "Locker / pickup network management", "Locker / BOPIS network capacity and SLA management.", "+9% pickup utilization", "RC-E2E-06"),
        ("white-glove-delivery-scheduling", "White-glove delivery scheduling", "Multi-touch white-glove install scheduling with installer skill match.", "+12% on-time install", "RC-E2E-06"),
    ],
    "Supplier & Vendor Management": [
        ("supplier-risk-scoring", "Supplier risk scoring", "Multi-factor supplier-risk scoring (financial, geo, compliance, quality).", "−24% high-risk concentration", "RC-E2E-15"),
        ("supplier-onboarding-automation", "Supplier-onboarding automation", "Auto-onboarding with KYC / sanctions / capability assessment.", "−42% onboarding cycle", "RC-E2E-09"),
        ("dual-source-coverage-monitor", "Dual-source-coverage monitor", "Dual-source / single-source coverage by category with risk dashboard.", "+18% category coverage", "RC-E2E-15"),
        ("supplier-cost-benchmark", "Supplier-cost benchmark", "Should-cost benchmark with negotiation prep and savings target.", "+1.1pp net cost improvement", "RC-E2E-03"),
    ],
    "Cold Chain": [
        ("cold-chain-excursion-mid-shift", "Cold-chain excursion mid-shift", "Real-time perishable temperature-excursion triage with markdown decisioning.", "−18% shrink cost", "RC-E2E-03"),
        ("cold-chain-lot-trace", "Cold-chain lot trace", "FSMA-204 / KDE-aware lot-level trace with end-to-end visibility.", "+99% trace coverage", "RC-E2E-09"),
        ("reefer-trailer-telemetry", "Reefer-trailer telemetry", "Reefer trailer telemetry with proactive failure-mode detection.", "−12% in-transit excursions", "RC-E2E-06"),
        ("warehouse-cold-room-monitoring", "Warehouse cold-room monitoring", "DC cold-room temperature monitoring with anomaly investigation queue.", "−18% cold-room downtime", "RC-E2E-11"),
        ("cold-chain-audit-rehearsal", "Cold-chain audit rehearsal", "Quarterly cold-chain audit rehearsal with auto-evidence pack.", "−45% audit-prep cost", "RC-E2E-15"),
    ],
},
"Store Operations": {
    "Shelf & On-Shelf Availability": [
        ("on-shelf-availability-oos", "On-shelf availability / OOS reduction", "Store-network OSA monitoring with prioritized shelf-tasking.", "−34% OOS rate KVI", "RC-E2E-05"),
        ("shelf-cv-audit", "Shelf computer-vision audit", "CV-based shelf audit at scale with planogram-deviation flagging.", "+91% planogram compliance", "RC-E2E-05"),
        ("phantom-out-detection", "Phantom-out detection", "Detect phantom out-of-stocks where perpetual says in-stock but shelf is empty.", "+8% recovered sales", "RC-E2E-05"),
        ("shelf-replen-task-prioritization", "Shelf replen-task prioritization", "Top-N shelf-replen task generation per associate per shift.", "+23% replen efficiency", "RC-E2E-05"),
        ("shelf-edge-pricing-mismatch", "Shelf-edge pricing-mismatch", "Detect shelf-edge label vs system-price mismatch with corrective task.", "−92% pricing-error events", "RC-E2E-05"),
    ],
    "Task Management & Compliance": [
        ("store-task-engine", "Store task engine", "Centralized task engine with smart prioritization by impact and shift.", "+18% task completion", "RC-E2E-05"),
        ("compliance-audit-automation", "Compliance audit automation", "Multi-regulation compliance audit with evidence pack and remediation queue.", "+95% audit pass rate", "RC-E2E-15"),
        ("safety-walk-automation", "Safety-walk automation", "Daily safety walk with photo-evidence and incident-trend dashboard.", "−24% incident rate", "RC-E2E-05"),
        ("opening-closing-checklist", "Opening / closing checklist", "Smart opening / closing checklist tied to store traffic and event calendar.", "+8% open-on-time rate", "RC-E2E-05"),
        ("temperature-log-automation", "Temperature-log automation", "Automated temperature-log capture for FSMA / HACCP compliance.", "−72% manual logging", "RC-E2E-09"),
    ],
    "Self-Checkout & Front-End": [
        ("self-checkout-shrink-detection", "Self-checkout shrink detection", "CV-assisted SCO shrink detection with intervention escalation.", "−22% SCO shrink", "RC-E2E-07"),
        ("scan-avoidance-detection", "Scan-avoidance detection", "Detect scan-avoidance behavior at SCO with audit-row trace.", "−34% scan-miss events", "RC-E2E-07"),
        ("front-end-queue-balancing", "Front-end queue balancing", "Dynamic queue balancing across staffed and SCO lanes.", "+9% checkout throughput", "RC-E2E-05"),
        ("age-verification-workflow", "Age-verification workflow", "Age-restricted item flow with associate-approval and audit trail.", "+99% age-verification compliance", "RC-E2E-15"),
        ("checkout-error-resolution-coach", "Checkout error-resolution coach", "Real-time checkout-error resolution with associate coaching prompts.", "−18% intervention time", "RC-E2E-08"),
    ],
    "Loss Prevention": [
        ("orc-organized-retail-crime", "Organized retail-crime ring detection", "Cross-store ORC ring detection using transaction and CV signals.", "$11M recovered Y1", "RC-E2E-07"),
        ("internal-shrink-detection", "Internal-shrink detection", "Associate-shrink detection using transaction-anomaly patterns.", "−14% internal shrink", "RC-E2E-07"),
        ("shrink-cause-decomposition", "Shrink-cause decomposition", "Decompose shrink into theft / damage / process / unknown root causes.", "−9% shrink rate", "RC-E2E-07"),
        ("safe-cash-management", "Safe / cash management", "Cash-handling exception detection with auto-investigation queue.", "−26% cash-variance events", "RC-E2E-07"),
        ("merchandise-protection-analytics", "Merchandise-protection analytics", "EAS / spider-wrap / case-lock effectiveness analytics by category.", "+7% protected-merch sell-through", "RC-E2E-07"),
        ("returns-no-receipt-fraud", "No-receipt-returns fraud", "No-receipt-returns abuse pattern detection with store-level coaching.", "+8% recovered loss", "RC-E2E-07"),
    ],
    "Energy & Refrigeration": [
        ("refrigeration-asset-monitoring", "Refrigeration-asset monitoring", "Predictive refrigeration-asset failure detection with work-order auto-create.", "−27% refrigeration downtime", "RC-E2E-11"),
        ("hvac-runtime-optimization", "HVAC runtime optimization", "HVAC duty-cycle optimization tied to occupancy and weather.", "−12% energy spend / store", "RC-E2E-13"),
        ("lighting-controls-optimization", "Lighting-controls optimization", "Daylight-aware lighting controls per store fixture profile.", "−18% lighting spend", "RC-E2E-13"),
        ("refrigerant-leak-detection", "Refrigerant-leak detection", "Refrigerant-leak detection for compliance and emissions reduction.", "−42% refrigerant loss", "RC-E2E-13"),
        ("smart-defrost-cycle-tuning", "Smart defrost-cycle tuning", "Defrost cycle tuning based on rack-load and ambient conditions.", "−9% refrigeration kWh", "RC-E2E-13"),
    ],
    "In-Store Customer Experience": [
        ("in-store-traffic-analytics", "In-store traffic analytics", "Foot-traffic and dwell-pattern analytics with merchandising actions.", "+11% category dwell time", "RC-E2E-05"),
        ("queue-prediction-staffing", "Queue prediction & staffing", "Forecast checkout queues 30 minutes ahead and re-deploy staff.", "+6pt customer NPS", "RC-E2E-08"),
        ("customer-help-button-routing", "Customer-help button routing", "Smart help-button-press routing to nearest qualified associate.", "−42% help-call response time", "RC-E2E-05"),
        ("fitting-room-attendance", "Fitting-room attendance", "Fitting-room attendance and conversion analytics with associate prompts.", "+14% try-on conversion", "RC-E2E-05"),
        ("ev-charger-uptime-management", "EV-charger uptime management", "Store-network EV-charger uptime with auto-ticket on degradation.", "+99% uptime", "RC-E2E-14"),
        ("in-store-event-orchestration", "In-store event orchestration", "In-store event / sampling / demo orchestration with associate prompts.", "+22% event attendee conversion", "RC-E2E-05"),
        ("clienteling-associate-companion", "Clienteling associate companion", "Clienteling companion app with VIP-customer profile and product recommendations.", "+18% clientele-attributed sales", "RC-E2E-05"),
    ],
    "Store Tech & Endpoints": [
        ("electronic-shelf-label-rollout", "Electronic shelf-label rollout", "Electronic-shelf-label rollout with price-update orchestration.", "−68% price-update labor", "RC-E2E-05"),
        ("digital-signage-orchestration", "Digital-signage orchestration", "Digital-signage orchestration with daypart and inventory awareness.", "+13% signage-driven sales", "RC-E2E-05"),
        ("rfid-inventory-accuracy", "RFID inventory accuracy", "RFID-driven inventory-accuracy program with re-count cadence.", "+22pp inventory accuracy", "RC-E2E-02"),
        ("store-device-fleet-mgmt", "Store device fleet management", "Store-device fleet management with proactive break-fix workflow.", "+99% device uptime", "RC-E2E-05"),
    ],
},
"Workforce & Talent": {
    "Scheduling & Labor": [
        ("labor-forecast-and-scheduling", "Labor forecast & scheduling", "Demand-driven labor forecast with shift assignment and skill match.", "−7% labor cost / sale", "RC-E2E-08"),
        ("schedule-fairness-monitoring", "Schedule-fairness monitoring", "Detect bias / fairness issues in schedule generation with corrective workflow.", "+12% schedule-fairness score", "RC-E2E-08"),
        ("flex-shift-marketplace", "Flex-shift marketplace", "Cross-store shift marketplace with skill and certification matching.", "+24% shift-fill rate", "RC-E2E-08"),
        ("call-out-recovery", "Call-out recovery", "Real-time call-out recovery with skill-aware backfill recommendation.", "−18% under-coverage hours", "RC-E2E-08"),
        ("overtime-prediction-prevention", "Overtime prediction & prevention", "Predict overtime risk and trigger preventive shift-rebalance.", "−15% overtime spend", "RC-E2E-08"),
        ("split-shift-detection", "Split-shift detection", "Detect undesirable split shifts and offer alternatives.", "+9pt schedule-satisfaction", "RC-E2E-08"),
    ],
    "Recruiting & Hiring": [
        ("hourly-recruiting-funnel", "Hourly recruiting funnel", "Hourly-role funnel with apply-to-hire optimization and bias monitoring.", "−21% time-to-hire", "RC-E2E-08"),
        ("internal-mobility-engine", "Internal-mobility engine", "Internal mobility engine matching skills to open store roles.", "+18% internal-fill rate", "RC-E2E-08"),
        ("hiring-event-orchestration", "Hiring-event orchestration", "Multi-store hiring event orchestration with same-day-offer flow.", "+34% same-day offers", "RC-E2E-08"),
        ("seasonal-surge-hiring", "Seasonal surge hiring", "Seasonal surge plan with funnel scaling and onboarding throughput.", "+27% peak-period coverage", "RC-E2E-08"),
        ("background-screening-orchestration", "Background-screening orchestration", "Coordinated background screening with auto-status updates.", "−12% candidate drop-off", "RC-E2E-08"),
    ],
    "Onboarding & Training": [
        ("onboarding-content-orchestration", "Onboarding content orchestration", "Personalized onboarding sequences by role and shift.", "+11% 30-day retention", "RC-E2E-08"),
        ("digital-coach-frontline", "Digital coach for frontline", "On-shift digital coach delivering micro-learning at point of need.", "+14% task accuracy", "RC-E2E-08"),
        ("compliance-training-tracking", "Compliance-training tracking", "Compliance-training completion tracking with reminder cascade.", "+98% on-time completion", "RC-E2E-15"),
        ("certification-renewal-tracking", "Certification-renewal tracking", "Renewal tracking for food-handler, alcohol, forklift certifications.", "+99% certification compliance", "RC-E2E-15"),
        ("manager-readiness-assessment", "Manager-readiness assessment", "Multi-source manager-readiness signal for promotion decisions.", "+18% manager success rate", "RC-E2E-08"),
    ],
    "Engagement & Retention": [
        ("frontline-pulse-survey", "Frontline pulse survey", "Pulse-survey program with always-on listening and action queue.", "+8pt eNPS", "RC-E2E-08"),
        ("attrition-risk-prediction", "Attrition-risk prediction", "Per-associate attrition-risk model with manager-action prompts.", "+11% retention save", "RC-E2E-08"),
        ("recognition-and-rewards", "Recognition and rewards", "Peer / manager recognition with rewards-marketplace integration.", "+13% engagement score", "RC-E2E-08"),
        ("manager-coaching-cadence", "Manager-coaching cadence", "Coaching-cadence tracking with quality-check sampling.", "+9% coaching adherence", "RC-E2E-08"),
        ("first-90-days-retention-program", "First-90-days retention program", "Targeted intervention program for new-hire 90-day retention.", "+21% 90-day retention", "RC-E2E-08"),
    ],
    "Wage & Hour Compliance": [
        ("predictive-wage-hour-violation", "Predictive wage-hour violation", "Detect risky scheduling patterns before they become violations.", "−87% wage-hour complaints", "RC-E2E-15"),
        ("predictive-scheduling-law", "Predictive-scheduling-law compliance", "Predictive-scheduling jurisdiction compliance with auto-adjustment.", "+99% jurisdictional compliance", "RC-E2E-15"),
        ("meal-break-compliance", "Meal-break compliance", "Meal / rest-break compliance monitoring with manager action queue.", "−92% break-violation events", "RC-E2E-15"),
        ("minor-employee-restrictions", "Minor-employee restrictions", "Minor-employee work-hour-restriction enforcement with audit row.", "+99% minor-rule compliance", "RC-E2E-15"),
        ("ftw-fair-workweek-compliance", "Fair Workweek compliance", "Fair-Workweek-jurisdiction compliance with auto-adjustment workflow.", "+99% Fair-Workweek compliance", "RC-E2E-15"),
    ],
    "Workforce Analytics": [
        ("dei-equity-analytics", "DEI / equity analytics", "DEI / pay-equity analytics with corrective-action recommendations.", "+24% equity-gap reduction", "RC-E2E-08"),
        ("succession-planning-engine", "Succession-planning engine", "Succession-planning analytics for critical roles with readiness signal.", "+18% succession-coverage rate", "RC-E2E-08"),
        ("skill-gap-analysis", "Skill-gap analysis", "Workforce skill-gap analysis tied to digital-transformation roadmap.", "+11% role-skill match", "RC-E2E-08"),
        ("absence-pattern-detection", "Absence-pattern detection", "Detect absence patterns and surface targeted interventions.", "−14% unscheduled absence", "RC-E2E-08"),
    ],
},
"E-commerce & Digital": {
    "Site Search & Discovery": [
        ("search-relevance-tuning", "Search relevance tuning", "Continuous search-relevance tuning with click-stream feedback.", "+13% search conversion", "RC-E2E-04"),
        ("zero-result-recovery", "Zero-result recovery", "Detect zero-result queries and route to recovery (synonyms, synonyms, similar).", "−42% null-search rate", "RC-E2E-04"),
        ("query-intent-classification", "Query-intent classification", "Classify search-query intent (browse / specific / problem) for tailored UX.", "+9% search-to-add-to-cart", "RC-E2E-04"),
        ("merch-pinned-positions", "Merch pinned positions", "Merchandiser-pinned search positions with auto-fade and rotation.", "+11% pinned-SKU sales", "RC-E2E-04"),
        ("typo-and-synonym-management", "Typo and synonym management", "Auto-discover typos and synonyms from search log and add to dictionary.", "+22% search recall", "RC-E2E-04"),
        ("vector-semantic-search", "Vector semantic search", "Embedding-based semantic search for descriptive / vibe queries.", "+18% relevance score", "RC-E2E-04"),
    ],
    "Recommendations": [
        ("homepage-recommender", "Homepage recommender", "Personalized homepage recommendations tied to inventory availability.", "+11% homepage CTR", "RC-E2E-04"),
        ("pdp-related-items", "PDP related items", "Related-item carousel with constraint-aware (in-stock, in-zone) scoring.", "+8% PDP attach", "RC-E2E-04"),
        ("cart-recommender", "Cart recommender", "Cart upsell / cross-sell with shipping-threshold incentive logic.", "+9% basket size", "RC-E2E-04"),
        ("post-purchase-recommender", "Post-purchase recommender", "Post-purchase recommendations tied to delivery window.", "+7% repeat purchase", "RC-E2E-04"),
        ("complete-the-look-recommender", "Complete-the-look recommender", "Apparel / home complete-the-look recommendations with image-similarity.", "+12% AOV", "RC-E2E-04"),
    ],
    "Cart & Checkout": [
        ("cart-abandonment-recovery", "Cart-abandonment recovery", "Multi-touch cart-abandonment recovery with friction-aware messaging.", "+17% abandonment recovery", "RC-E2E-04"),
        ("checkout-friction-detection", "Checkout-friction detection", "Detect specific checkout-step friction (shipping, payment, account).", "+9% completion rate", "RC-E2E-04"),
        ("payment-orchestration", "Payment orchestration", "Payment-routing optimization for auth rate and cost.", "+1.4pp auth rate", "RC-E2E-15"),
        ("address-validation-resolution", "Address validation & resolution", "Address validation with incentive to fix typos before submit.", "−24% address-correction cost", "RC-E2E-04"),
        ("dynamic-shipping-incentive", "Dynamic shipping incentive", "Conditional shipping incentive tied to LTV / margin tradeoff.", "+11% conversion / margin$", "RC-E2E-04"),
        ("buy-now-pay-later-attach", "BNPL attach optimization", "BNPL provider attach for AOV uplift with risk-of-loss check.", "+13% AOV via BNPL", "RC-E2E-04"),
    ],
    "Mobile App & Experience": [
        ("mobile-app-store-mode", "Mobile app store mode", "In-store mobile-app mode with aisle wayfinding and self-scan.", "+18% app-engaged customer spend", "RC-E2E-04"),
        ("push-notification-orchestration", "Push-notification orchestration", "Per-member push-notification orchestration with frequency caps.", "+9% engagement", "RC-E2E-04"),
        ("in-app-loyalty-experience", "In-app loyalty experience", "Loyalty-tied in-app experience with personalized challenges.", "+14% app DAU", "RC-E2E-04"),
        ("wishlist-engagement", "Wishlist engagement", "Wishlist re-engagement with price-drop and back-in-stock signals.", "+11% wishlist conversion", "RC-E2E-04"),
        ("scan-and-go-experience", "Scan-and-go experience", "Scan-and-go in-app checkout with shrink-control patterns.", "+8% trip frequency", "RC-E2E-05"),
    ],
    "Reviews & UGC": [
        ("review-velocity-engagement", "Review-velocity engagement", "Programmatic review solicitation tuned to review-velocity targets.", "+27% review submission rate", "RC-E2E-04"),
        ("review-fraud-detection", "Review-fraud detection", "Detect fake / coordinated reviews with takedown workflow.", "+98% review trust score", "RC-E2E-07"),
        ("review-summary-generation", "Review-summary generation", "Auto-generate review summaries on PDPs from review corpus.", "+11% PDP conversion", "RC-E2E-04"),
        ("video-review-program", "Video-review program", "Video-review program with creator incentives and rights management.", "+19% video-attached PDP conversion", "RC-E2E-04"),
    ],
    "SEO / SEM": [
        ("seo-content-gap-analysis", "SEO content-gap analysis", "Detect search-volume content-gaps and prioritize content production.", "+22% organic traffic", "RC-E2E-04"),
        ("paid-search-bidding-optimization", "Paid-search bidding optimization", "Margin-aware paid-search bidding with portfolio-level budget caps.", "+14% paid-search ROAS", "RC-E2E-04"),
        ("brand-vs-non-brand-allocation", "Brand vs non-brand allocation", "Brand-vs-non-brand spend allocation tied to incremental sales.", "+7% incremental SEM ROAS", "RC-E2E-04"),
        ("structured-data-coverage", "Structured-data coverage", "PDP structured-data coverage for rich-result eligibility.", "+11% rich-result CTR", "RC-E2E-04"),
    ],
    "Marketplace & Subscriptions": [
        ("marketplace-seller-onboarding", "Marketplace seller onboarding", "Third-party seller onboarding with KYC and quality scoring.", "+34% seller throughput", "RC-E2E-04"),
        ("marketplace-listing-quality", "Marketplace listing quality", "3P listing quality scoring with seller-action queue.", "+18% 3P listing conversion", "RC-E2E-04"),
        ("subscription-anchor-optimization", "Subscription anchor optimization", "Subscription cadence and bundle optimization for retention.", "+12% subscription retention", "RC-E2E-04"),
        ("subscription-pause-save", "Subscription pause-and-save flow", "Detect subscription cancel intent and offer pause / skip alternatives.", "+19% retention save rate", "RC-E2E-04"),
        ("marketplace-buy-box-optimization", "Marketplace buy-box optimization", "3P marketplace buy-box win-rate optimization with margin guardrails.", "+17% buy-box win rate", "RC-E2E-04"),
        ("recurring-billing-resolution", "Recurring-billing resolution", "Recurring-billing failed-payment resolution with smart-retry logic.", "+8pp payment-recovery rate", "RC-E2E-15"),
    ],
    "Connected Commerce": [
        ("livestream-commerce-orchestration", "Livestream commerce orchestration", "Livestream commerce session orchestration with inventory and creator management.", "+34% livestream conversion", "RC-E2E-04"),
        ("social-commerce-attribution", "Social commerce attribution", "Social-commerce attribution stitching from impression to purchase.", "+21% social ROAS visibility", "RC-E2E-04"),
        ("connected-tv-shoppable", "Connected-TV shoppable ads", "Connected-TV shoppable-ad orchestration with real-time inventory.", "+18% CTV-attributed revenue", "RC-E2E-04"),
        ("voice-commerce-experience", "Voice-commerce experience", "Voice-assistant commerce experience with reorder simplification.", "+9% repeat reorder rate", "RC-E2E-04"),
    ],
},
"Returns & Reverse Logistics": {
    "Returns Triage": [
        ("returns-fraud-detection", "Returns-fraud detection", "Graph-based returns-fraud detection across customer / payment / device / SKU.", "+52% fraud detection rate", "RC-E2E-07"),
        ("returns-policy-personalization", "Returns-policy personalization", "Risk-aware returns policy personalization (free / paid / denied).", "−14% return-cost / order", "RC-E2E-07"),
        ("returns-routing-optimization", "Returns-routing optimization", "Route returns to nearest disposition node (resell / refurb / salvage).", "+22% recovery / unit", "RC-E2E-07"),
        ("returns-reason-code-quality", "Returns reason-code quality", "Capture-side reason-code quality with NPI / merch feedback loop.", "+39% reason-code accuracy", "RC-E2E-07"),
    ],
    "BORIS / BOPIS": [
        ("boris-handling-orchestration", "BORIS handling orchestration", "Buy-online-return-in-store handling with associate-task generation.", "−18% BORIS handling time", "RC-E2E-07"),
        ("bopis-pickup-optimization", "BOPIS pickup optimization", "BOPIS pick / hold / handoff orchestration with shopper-ETA visibility.", "+11% BOPIS NPS", "RC-E2E-06"),
        ("curbside-pickup-orchestration", "Curbside pickup orchestration", "Curbside pickup arrival / handoff orchestration with door-time tracking.", "−24% curbside wait time", "RC-E2E-06"),
    ],
    "Refurbishment & Salvage": [
        ("refurb-grading-automation", "Refurb-grading automation", "Auto-grade returned items using image and sensor data for refurb routing.", "+31% refurb yield", "RC-E2E-07"),
        ("salvage-resale-routing", "Salvage / resale routing", "Route to liquidator / open-box / charity based on recovery max.", "+$4.2M recovery / yr", "RC-E2E-07"),
        ("warranty-claim-workflow", "Warranty-claim workflow", "Warranty-claim disposition workflow with vendor-funded recovery.", "+$1.8M warranty recovery", "RC-E2E-07"),
    ],
    "Reverse Transportation": [
        ("reverse-load-consolidation", "Reverse load consolidation", "Consolidate reverse-flow shipments to reduce per-unit cost.", "−12% reverse-freight cost", "RC-E2E-06"),
        ("returns-receive-yard-staging", "Returns receive yard staging", "Yard / dock staging for returns flow to compress dock-to-disposition time.", "−28% disposition cycle time", "RC-E2E-06"),
        ("third-party-returns-aggregator", "Third-party returns aggregator", "Third-party returns aggregator network with cost-vs-experience routing.", "−14% reverse-cost / unit", "RC-E2E-06"),
        ("packageless-return-network", "Package-less return network", "Package-less return drop-off network management with locker integration.", "+22% package-less return adoption", "RC-E2E-06"),
    ],
    "Customer Returns Experience": [
        ("self-service-return-portal", "Self-service return portal", "Self-service return portal with pre-validation and reason capture.", "+34% portal-initiated returns", "RC-E2E-04"),
        ("returnless-refund-decisioning", "Returnless-refund decisioning", "Returnless-refund decisioning balancing customer experience vs cost.", "+8pt return NPS", "RC-E2E-07"),
        ("partial-refund-flexible-resolution", "Partial-refund flexible resolution", "Partial-refund / flex-resolution offers balancing satisfaction vs cost.", "+12% retention after issue", "RC-E2E-07"),
        ("exchange-vs-refund-optimization", "Exchange vs refund optimization", "Exchange-incentive design to convert refund intent to exchange.", "+18% exchange-vs-refund mix", "RC-E2E-07"),
    ],
},
"Risk · Fraud · Security": {
    "Returns Fraud": [
        ("serial-returner-detection", "Serial-returner detection", "Detect serial-returner pattern across customer / device / payment graph.", "+38% returner-pattern detection", "RC-E2E-07"),
        ("wardrobing-fraud-detection", "Wardrobing-fraud detection", "Wardrobing pattern detection with risk-aware return policy.", "+22% wardrobing prevention", "RC-E2E-07"),
        ("receipt-fabrication-detection", "Receipt-fabrication detection", "Detect fabricated / altered receipts at returns desk.", "+89% receipt-validity check", "RC-E2E-07"),
    ],
    "Payment & Auth Fraud": [
        ("card-not-present-fraud", "Card-not-present fraud", "CNP fraud detection with adaptive risk-rule tuning.", "−27% CNP losses", "RC-E2E-07"),
        ("auth-rate-fraud-tradeoff", "Auth-rate / fraud tradeoff", "Optimize the auth-rate / fraud-loss tradeoff per channel and segment.", "+1.6pp auth, flat losses", "RC-E2E-15"),
        ("ach-bank-fraud-detection", "ACH / bank-fraud detection", "ACH-based payment fraud with same-day-clawback workflow.", "−42% ACH-fraud loss", "RC-E2E-07"),
        ("loyalty-tender-misuse", "Loyalty-tender misuse", "Detect loyalty-tender misuse (point-balance abuse, account takeover).", "+58% misuse detection", "RC-E2E-07"),
    ],
    "Loyalty & Account Fraud": [
        ("account-takeover-detection", "Account-takeover detection", "Detect ATO across login, behavioral, and transactional anomalies.", "−72% ATO losses", "RC-E2E-07"),
        ("synthetic-identity-detection", "Synthetic-identity detection", "Detect synthetic identities at signup and payment-instrument stages.", "+47% synthetic detection", "RC-E2E-07"),
        ("multi-account-coupon-abuse", "Multi-account coupon abuse", "Detect multi-account abuse of acquisition / coupon offers.", "+$2.1M recovered", "RC-E2E-07"),
    ],
    "Organized Retail Crime": [
        ("orc-ring-mapping", "ORC ring mapping", "Map ORC rings across stores using transaction and CV signals.", "$11M+ recovered Y1", "RC-E2E-07"),
        ("boost-bag-sled-detection", "Boost-bag / sled detection", "Detect boost-bag / sled signals via CV and POS-anomaly correlation.", "+34% ORC apprehension", "RC-E2E-07"),
        ("flash-mob-event-detection", "Flash-mob event detection", "Detect coordinated flash-mob theft with cross-store signal correlation.", "−68% flash-mob loss", "RC-E2E-07"),
    ],
    "Cyber & Identity": [
        ("pos-malware-detection", "POS-malware detection", "POS / payment-terminal malware detection with auto-isolation.", "+99% endpoint coverage", "RC-E2E-15"),
        ("third-party-vendor-risk", "Third-party vendor risk", "Third-party vendor cyber-risk monitoring with contract-trigger workflow.", "−18% vendor-risk concentration", "RC-E2E-15"),
        ("data-loss-prevention", "Data-loss prevention", "DLP for PII / payment / loyalty data with audit trail.", "+99% DLP coverage", "RC-E2E-15"),
        ("phishing-and-bec-defense", "Phishing & BEC defense", "Phishing / BEC defense with auto-quarantine and user coaching.", "−72% phishing click-through", "RC-E2E-15"),
    ],
    "Insider Threat & Privacy": [
        ("insider-threat-monitoring", "Insider-threat monitoring", "Detect insider-threat patterns in HR / IT / finance access.", "+38% insider-pattern detection", "RC-E2E-15"),
        ("privacy-rights-fulfillment", "Privacy rights fulfillment", "Auto-fulfill GDPR / CCPA rights requests with audit row.", "+99% on-time fulfillment", "RC-E2E-15"),
        ("data-classification-discovery", "Data classification & discovery", "Discover and classify sensitive data across the data estate.", "+92% sensitive-data coverage", "RC-E2E-15"),
        ("consent-ledger-monitoring", "Consent-ledger monitoring", "Cross-channel consent-ledger monitoring with revocation flow.", "+99% consent-trail integrity", "RC-E2E-15"),
        ("vendor-data-sharing-control", "Vendor data-sharing control", "Vendor data-sharing control with purpose-limitation enforcement.", "+99% vendor-purpose compliance", "RC-E2E-15"),
    ],
    "Counterfeit & Brand Protection": [
        ("counterfeit-detection-marketplace", "Counterfeit detection (marketplace)", "Counterfeit / unauthorized-seller detection on third-party marketplaces.", "+84% takedown rate", "RC-E2E-15"),
        ("gray-market-diversion-detection", "Gray-market diversion detection", "Gray-market diversion detection with channel-control enforcement.", "+18% gray-market reduction", "RC-E2E-15"),
        ("brand-impersonation-domain-monitor", "Brand-impersonation domain monitor", "Brand-impersonation domain detection with takedown orchestration.", "+72% impersonation takedown rate", "RC-E2E-15"),
    ],
    "Physical Security": [
        ("incident-response-store", "Incident response (store)", "In-store incident-response orchestration with corporate escalation.", "−48% response time", "RC-E2E-07"),
        ("workplace-violence-prevention", "Workplace-violence prevention", "Workplace-violence-pattern detection with intervention pathways.", "−24% incident severity", "RC-E2E-07"),
        ("active-shooter-readiness", "Active-shooter readiness", "Active-shooter readiness program with drills and post-drill analytics.", "+99% drill participation", "RC-E2E-07"),
    ],
},
"Finance & Revenue": {
    "Revenue Recognition": [
        ("rev-rec-multi-element", "Multi-element revenue recognition", "Multi-element arrangement rev-rec with auto-allocation engine.", "+99% rev-rec automation", "RC-E2E-15"),
        ("subscription-revenue-recognition", "Subscription revenue recognition", "Subscription-arrangement rev-rec with deferred-balance management.", "+18% close-cycle compression", "RC-E2E-15"),
        ("gift-card-breakage-tracking", "Gift-card breakage tracking", "Gift-card breakage estimation with audit-defensible methodology.", "+$2.4M breakage recognition", "RC-E2E-15"),
        ("loyalty-program-liability", "Loyalty-program liability", "Loyalty-liability estimation with breakage and burn-rate updates.", "+12% liability accuracy", "RC-E2E-15"),
    ],
    "Settlement & Reconciliation": [
        ("payment-settlement-reconciliation", "Payment settlement reconciliation", "Multi-processor settlement reconciliation with exception queue.", "−72% manual recon hours", "RC-E2E-15"),
        ("interchange-leak-detection", "Interchange leak detection", "Interchange-fee leak detection with downgrade-cause attribution.", "$1.8M recovered / yr", "RC-E2E-15"),
        ("vendor-payment-recon", "Vendor payment reconciliation", "AP-vendor reconciliation with auto-match across PO / GR / invoice.", "+89% auto-match rate", "RC-E2E-15"),
        ("chargeback-orchestration", "Chargeback orchestration", "Chargeback-dispute orchestration with evidence-pack assembly.", "+22% chargeback win rate", "RC-E2E-15"),
    ],
    "FP&A": [
        ("close-acceleration-flux-explain", "Close acceleration / flux explain", "Auto-flux variance explanations with drill-down evidence.", "−42% close-cycle time", "RC-E2E-15"),
        ("rolling-forecast-driver-based", "Rolling forecast / driver-based", "Driver-based rolling forecast with what-if scenario modeling.", "+24% forecast cycle speed", "RC-E2E-15"),
        ("category-pnl-attribution", "Category P&L attribution", "Category-level P&L attribution including freight / shrink / markdown.", "+18% P&L transparency", "RC-E2E-15"),
        ("scenario-planning-engine", "Scenario-planning engine", "Multi-scenario planning with macro-driver overlays.", "+36% scenario-cycle speed", "RC-E2E-15"),
    ],
    "Tax & Compliance": [
        ("sales-tax-jurisdiction-monitor", "Sales-tax jurisdiction monitor", "Multi-jurisdiction sales-tax rule monitor with system-rule update.", "+99% jurisdictional accuracy", "RC-E2E-15"),
        ("nexus-monitoring-marketplace", "Nexus monitoring (marketplace)", "Marketplace-driven nexus exposure monitoring across states.", "+99% nexus-coverage compliance", "RC-E2E-15"),
        ("sugar-and-bag-tax-handling", "Sugar / bag-tax handling", "City-level surcharge / bag-tax automation with audit trail.", "+99% surcharge accuracy", "RC-E2E-15"),
        ("transfer-pricing-monitoring", "Transfer-pricing monitoring", "Inter-company transfer-pricing monitoring with documentation pack.", "+18% TP-doc compliance", "RC-E2E-15"),
    ],
    "Audit & Controls": [
        ("audit-evidence-orchestration", "Audit-evidence orchestration", "Auto-assemble audit evidence from controls and source systems.", "−54% audit-prep time", "RC-E2E-15"),
        ("sox-control-monitoring", "SOX control monitoring", "Continuous SOX-control monitoring with deficiency-detection alerts.", "+99% control coverage", "RC-E2E-15"),
        ("itgc-evidence-capture", "ITGC evidence capture", "ITGC evidence capture across access / change / operations domains.", "+92% control-evidence completeness", "RC-E2E-15"),
        ("vendor-master-data-controls", "Vendor master-data controls", "Vendor-master-data integrity controls with anomaly detection.", "−91% master-data exception", "RC-E2E-15"),
        ("segregation-of-duties-monitoring", "Segregation-of-duties monitoring", "Continuous SoD monitoring with conflict-resolution workflow.", "+99% SoD compliance", "RC-E2E-15"),
        ("internal-audit-coverage-planning", "Internal-audit coverage planning", "Risk-based internal-audit coverage planning with annual roadmap.", "+22% audit-coverage efficiency", "RC-E2E-15"),
    ],
    "Treasury & Working Capital": [
        ("cash-forecasting-treasury", "Cash forecasting (treasury)", "Driver-based cash forecasting with bank-feed reconciliation.", "+18% forecast accuracy", "RC-E2E-15"),
        ("ar-aging-collections", "AR aging & collections", "AR aging analytics with prioritized collections action queue.", "−14% DSO", "RC-E2E-15"),
        ("ap-payment-strategy", "AP payment strategy", "AP payment-timing strategy balancing cash and supplier relationships.", "+8 days DPO", "RC-E2E-15"),
        ("working-capital-decomposition", "Working-capital decomposition", "Decompose working-capital movement into root-cause levers.", "+22% working-capital efficiency", "RC-E2E-15"),
    ],
},
"Sustainability & ESG": {
    "Scope 3 Emissions": [
        ("scope3-emissions-mapping", "Scope-3 emissions mapping", "Map and quantify Scope-3 emissions across categories.", "+95% category coverage", "RC-E2E-13"),
        ("supplier-emissions-data-quality", "Supplier-emissions data quality", "Drive supplier-emissions data quality with automated chase.", "+38% data-quality score", "RC-E2E-13"),
        ("scope3-reduction-roadmap", "Scope-3 reduction roadmap", "Reduction-roadmap modeling tied to corporate-target glidepath.", "−18% Scope-3 by 2030", "RC-E2E-13"),
        ("inbound-freight-emissions", "Inbound-freight emissions", "Inbound-freight emissions modeling with mode-shift levers.", "−12% inbound-freight emissions", "RC-E2E-13"),
    ],
    "Packaging & Waste": [
        ("packaging-circularity-tracker", "Packaging-circularity tracker", "Track packaging recyclability / reuse with corrective action.", "+24% recyclable packaging share", "RC-E2E-13"),
        ("food-waste-shrink-reduction", "Food-waste shrink reduction", "Food-waste shrink reduction with donation / mark-down sequencing.", "−21% food waste", "RC-E2E-13"),
        ("plastic-reduction-monitoring", "Plastic-reduction monitoring", "Plastic-content monitoring with regulatory-jurisdiction overlay.", "+27% reduction commitment progress", "RC-E2E-13"),
        ("epr-reporting-automation", "EPR reporting automation", "EPR (extended-producer-responsibility) reporting automation.", "+99% EPR-report accuracy", "RC-E2E-13"),
    ],
    "Water & Energy": [
        ("water-usage-monitoring", "Water-usage monitoring", "Site-level water-usage monitoring with leak detection.", "−12% water spend", "RC-E2E-13"),
        ("renewable-energy-purchase", "Renewable-energy purchase", "Renewable-energy procurement and Scope-2 reporting.", "+27% renewable share", "RC-E2E-13"),
        ("scope2-market-based-reporting", "Scope-2 market-based reporting", "Market-based vs location-based Scope-2 reporting with audit pack.", "+99% reporting accuracy", "RC-E2E-13"),
    ],
    "Supplier ESG": [
        ("supplier-esg-scorecard", "Supplier ESG scorecard", "Multi-pillar supplier ESG scorecard with action queue.", "+34% supplier-ESG-data coverage", "RC-E2E-13"),
        ("forced-labor-due-diligence", "Forced-labor due diligence", "Forced-labor / supply-chain transparency due diligence.", "+99% due-diligence coverage", "RC-E2E-13"),
        ("conflict-minerals-tracking", "Conflict-minerals tracking", "Conflict-minerals reporting with smelter-level tracing.", "+99% smelter-level coverage", "RC-E2E-13"),
    ],
    "Product Provenance": [
        ("product-provenance-trace", "Product-provenance trace", "Product-provenance trace from origin to shelf with consumer transparency.", "+92% trace coverage", "RC-E2E-09"),
        ("organic-claim-substantiation", "Organic-claim substantiation", "Organic-claim substantiation with certification-evidence trail.", "+99% claim-substantiation rate", "RC-E2E-09"),
        ("fair-trade-certification-tracking", "Fair-trade certification tracking", "Fair-trade certification tracking with renewal cadence.", "+99% certification compliance", "RC-E2E-09"),
        ("regenerative-ag-sourcing", "Regenerative-ag sourcing", "Regenerative-agriculture sourcing tracking with farmer-program tier.", "+18% regen-ag share", "RC-E2E-09"),
        ("animal-welfare-claim-tracking", "Animal-welfare claim tracking", "Animal-welfare claim substantiation with auditable supplier evidence.", "+99% claim-substantiation rate", "RC-E2E-09"),
    ],
    "ESG Reporting & Disclosure": [
        ("csrd-eu-reporting", "CSRD / EU reporting", "CSRD reporting orchestration with double-materiality assessment.", "+99% on-time disclosure", "RC-E2E-13"),
        ("sec-climate-disclosure", "SEC climate disclosure", "SEC climate-disclosure orchestration with audit-ready evidence pack.", "+99% disclosure accuracy", "RC-E2E-13"),
        ("tcfd-aligned-disclosure", "TCFD-aligned disclosure", "TCFD-aligned climate-risk disclosure with scenario analysis.", "+18% climate-risk transparency", "RC-E2E-13"),
        ("sasb-industry-disclosure", "SASB industry disclosure", "SASB industry-specific disclosure orchestration.", "+99% SASB completeness", "RC-E2E-13"),
        ("esg-investor-engagement-tracker", "ESG investor-engagement tracker", "ESG investor-engagement and stewardship tracking.", "+22% investor-engagement quality", "RC-E2E-13"),
    ],
},
"Real Estate & Network": {
    "Site Selection": [
        ("site-selection-trade-area", "Site selection / trade area", "Trade-area modeling with cannibalization and demographic overlay.", "+18% site-roi accuracy", "RC-E2E-14"),
        ("competitive-cannibalization-modeling", "Competitive / cannibalization modeling", "Self-cannibalization and competitive-overlap modeling.", "+11% net-new sales", "RC-E2E-14"),
        ("demographic-shift-detection", "Demographic-shift detection", "Detect demographic shifts impacting trade-area economics.", "+8% trade-area accuracy", "RC-E2E-14"),
    ],
    "Format & Network": [
        ("format-portfolio-rebalance", "Format-portfolio rebalance", "Format-mix rebalance (full / small / urban / drive-thru) by trade area.", "+5% comp sales", "RC-E2E-14"),
        ("network-coverage-optimization", "Network-coverage optimization", "Network-coverage optimization for delivery / pickup / fulfillment.", "+12% delivery coverage", "RC-E2E-14"),
        ("dark-store-network-design", "Dark-store network design", "Dark-store network design tied to e-commerce demand patterns.", "+22% e-commerce node throughput", "RC-E2E-14"),
    ],
    "Lease & Property Management": [
        ("lease-renewal-decision", "Lease-renewal decision", "Lease-renewal decision with site-economics and trade-area outlook.", "+9% renewal-decision quality", "RC-E2E-14"),
        ("lease-cam-recon", "Lease CAM reconciliation", "CAM-charge audit and reconciliation with landlord communications.", "$2.1M recovered / yr", "RC-E2E-15"),
        ("kick-out-clause-monitoring", "Kick-out clause monitoring", "Kick-out clause and co-tenancy clause monitoring.", "+99% clause-trigger coverage", "RC-E2E-14"),
        ("lease-abstract-extraction", "Lease abstract extraction", "Auto-extract lease abstracts from PDFs with audit trail.", "−68% abstraction effort", "RC-E2E-14"),
    ],
    "Closures & Restructuring": [
        ("store-closure-planning", "Store-closure planning", "Closure planning with inventory liquidation and customer-handoff.", "+11% closure-recovery $", "RC-E2E-14"),
        ("liquidation-merchandise-flow", "Liquidation merchandise flow", "Liquidation flow optimization with vendor / liquidator routing.", "+18% liquidation recovery", "RC-E2E-14"),
        ("customer-transfer-program", "Customer-transfer program", "Customer-transfer program to nearby stores with retention focus.", "+22% transfer retention", "RC-E2E-14"),
        ("post-closure-asset-disposition", "Post-closure asset disposition", "Post-closure fixture / equipment asset disposition with recovery.", "+14% asset-recovery $", "RC-E2E-14"),
    ],
    "Retrofit & Capex Planning": [
        ("retrofit-capex-prioritization", "Retrofit capex prioritization", "Capex retrofit prioritization tied to ROI and ESG impact.", "+17% capex ROI", "RC-E2E-14"),
        ("smart-store-rollout-planning", "Smart-store rollout planning", "Smart-store-tech rollout planning with sequencing and ROI.", "+22% sequencing efficiency", "RC-E2E-14"),
        ("equipment-fleet-replacement", "Equipment-fleet replacement", "Refrigeration / HVAC fleet replacement program with rebate optimization.", "$3.4M rebate captured", "RC-E2E-14"),
        ("permitting-acceleration", "Permitting acceleration", "Permitting-process acceleration via auto-generated submission packs.", "−42% permitting cycle", "RC-E2E-14"),
    ],
},
"Manufacturing & CPG Plant Ops": {
    "OEE & Plant Performance": [
        ("oee-monitoring-line", "OEE monitoring per line", "Real-time OEE monitoring with loss-cause attribution.", "+4.2pp OEE", "RC-E2E-11"),
        ("changeover-time-reduction", "Changeover-time reduction", "Changeover-time reduction with SMED-pattern coaching.", "−18% changeover time", "RC-E2E-11"),
        ("yield-loss-attribution", "Yield-loss attribution", "Yield-loss attribution to root cause across material / process / quality.", "+1.8pp yield", "RC-E2E-11"),
        ("digital-andon-escalation", "Digital andon escalation", "Digital-andon escalation with auto-routing to qualified responder.", "−22% downtime per event", "RC-E2E-11"),
        ("plant-energy-intensity", "Plant energy intensity", "Plant energy-intensity per unit produced with reduction levers.", "−9% energy / unit", "RC-E2E-13"),
    ],
    "Predictive Maintenance": [
        ("predictive-maintenance-vibration", "Predictive maintenance / vibration", "Vibration-anomaly-based PdM for rotating equipment.", "−27% unplanned downtime", "RC-E2E-11"),
        ("predictive-maintenance-thermal", "Predictive maintenance / thermal", "Thermal-imaging PdM for electrical and motor assets.", "−24% thermal-fault failures", "RC-E2E-11"),
        ("oil-analysis-monitoring", "Oil-analysis monitoring", "Oil-analysis-driven PdM with auto-work-order creation.", "−18% lubricated-asset failures", "RC-E2E-11"),
        ("spare-parts-optimization", "Spare-parts optimization", "Spare-parts inventory optimization tied to PdM forecast.", "−14% spare-parts holding", "RC-E2E-11"),
    ],
    "Quality": [
        ("inline-quality-cv-detection", "Inline quality / CV detection", "Inline CV-based defect detection with auto-quarantine.", "+38% defect detection", "RC-E2E-11"),
        ("spc-process-control", "SPC process control", "Statistical-process-control with auto-out-of-control detection.", "+27% process-stability score", "RC-E2E-11"),
        ("supplier-incoming-quality", "Supplier-incoming quality", "Supplier-incoming quality with auto-disposition routing.", "−18% incoming-defect cost", "RC-E2E-11"),
        ("recall-readiness-rehearsal", "Recall-readiness rehearsal", "Recall rehearsal with trace-down precision under one hour.", "−72% recall-prep time", "RC-E2E-09"),
        ("haccp-ccp-monitoring", "HACCP / CCP monitoring", "Real-time HACCP critical-control-point monitoring with auto-hold.", "+99% HACCP compliance", "RC-E2E-11"),
    ],
    "Recipe & Spec Management": [
        ("recipe-version-management", "Recipe version management", "Multi-site recipe version management with rollout governance.", "+99% recipe-compliance", "RC-E2E-11"),
        ("ingredient-substitution-impact", "Ingredient-substitution impact", "Substitute-ingredient impact modeling on cost / quality / labeling.", "+22% subst-decision speed", "RC-E2E-11"),
        ("allergen-labeling-control", "Allergen labeling control", "Allergen control across recipe / line / packaging with cross-contact monitoring.", "+99% allergen-label accuracy", "RC-E2E-09"),
        ("clean-label-formulation", "Clean-label formulation", "Clean-label formulation modeling with regulatory and consumer overlays.", "+18% clean-label-launch speed", "RC-E2E-11"),
    ],
    "Plant Safety": [
        ("ehs-near-miss-reporting", "EHS near-miss reporting", "Near-miss reporting with frictionless capture and trend analytics.", "+58% near-miss reporting", "RC-E2E-11"),
        ("hazard-walk-cv-assist", "Hazard-walk CV assist", "CV-assisted hazard walks with prioritized remediation queue.", "−24% incident rate", "RC-E2E-11"),
        ("loto-procedure-compliance", "LOTO procedure compliance", "LOTO procedure-compliance monitoring with exception alerts.", "+99% LOTO compliance", "RC-E2E-11"),
        ("workplace-ergonomics-monitoring", "Workplace ergonomics monitoring", "Workplace-ergonomics monitoring with intervention recommendations.", "−18% MSD claims", "RC-E2E-11"),
    ],
    "Plant Sustainability": [
        ("plant-water-stewardship", "Plant water stewardship", "Plant-level water-stewardship program with reduction targets.", "−14% water / unit", "RC-E2E-13"),
        ("zero-waste-to-landfill", "Zero-waste-to-landfill", "Zero-waste-to-landfill program with by-product valorization.", "+92% diversion rate", "RC-E2E-13"),
        ("greenhouse-gas-plant-reduction", "GHG plant-level reduction", "Plant-level GHG reduction program with verified-data reporting.", "−22% Scope-1 emissions", "RC-E2E-13"),
    ],
    "Procurement & Source": [
        ("raw-material-cost-volatility", "Raw-material cost volatility", "Raw-material cost-volatility monitoring with hedge recommendation.", "−12% input-cost variance", "RC-E2E-11"),
        ("ingredient-substitution-tradeoff", "Ingredient-substitution tradeoff", "Ingredient-substitution tradeoff modeling on cost / quality / claim.", "+1.6pp gross margin", "RC-E2E-11"),
        ("co-packer-network-optimization", "Co-packer network optimization", "Co-packer network design with capacity / capability / cost balance.", "+18% co-pack throughput", "RC-E2E-11"),
        ("ingredient-supplier-risk", "Ingredient-supplier risk", "Ingredient-supplier risk monitoring with dual-source coverage check.", "−24% single-source exposure", "RC-E2E-15"),
    ],
},
"Trade Promotion & RGM": {
    "Trade Spend Effectiveness": [
        ("trade-spend-effectiveness", "Trade-spend effectiveness", "Trade-spend ROI with retailer-segment and event-type decomposition.", "+12% trade ROI", "RC-E2E-10"),
        ("trade-spend-allocation-engine", "Trade-spend allocation engine", "Allocate trade spend across retailers and events to maximize ROI.", "+18% allocation efficiency", "RC-E2E-10"),
        ("trade-promotion-pre-post-eval", "Trade promotion pre/post evaluation", "Pre-event commitment vs post-event delivered analysis.", "+9% deliver-vs-commit accuracy", "RC-E2E-10"),
    ],
    "Deduction Management": [
        ("deduction-management-automation", "Deduction-management automation", "Auto-classify and resolve deductions with backup-doc retrieval.", "−54% deduction handling time", "RC-E2E-10"),
        ("invalid-deduction-chargeback", "Invalid-deduction chargeback", "Identify invalid deductions and pursue with retailer evidence pack.", "+$3.4M recovery / yr", "RC-E2E-10"),
        ("deduction-root-cause-prevention", "Deduction root-cause prevention", "Prevent recurring deductions via root-cause analysis.", "−18% recurring deductions", "RC-E2E-10"),
    ],
    "Price-Pack Architecture": [
        ("price-pack-architecture", "Price-pack architecture", "Price-pack architecture optimization across channels.", "+1.6pp net revenue", "RC-E2E-10"),
        ("channel-pack-differentiation", "Channel-pack differentiation", "Channel-specific pack differentiation to manage rate-by-channel.", "+11% channel-rate spread", "RC-E2E-10"),
        ("everyday-vs-promo-pricing-balance", "Everyday vs promo pricing balance", "Balance everyday and promo pricing for net-revenue maximization.", "+1.2pp net revenue", "RC-E2E-10"),
    ],
    "Mix Optimization": [
        ("mix-optimization-product", "Product mix optimization", "Optimize product mix to balance volume / margin / revenue growth.", "+1.4pp blended margin", "RC-E2E-10"),
        ("channel-mix-optimization", "Channel mix optimization", "Optimize channel mix (mass / club / e-comm / convenience) for net revenue.", "+9% channel net revenue", "RC-E2E-10"),
        ("customer-mix-optimization", "Customer mix optimization", "Customer-level mix optimization for portfolio strategy.", "+1.8pp portfolio margin", "RC-E2E-10"),
    ],
    "Joint Business Planning": [
        ("jbp-acceleration-with-retailer", "JBP acceleration with retailer", "Generative JBP-prep with retailer-specific levers and ROI modeling.", "−42% JBP cycle time", "RC-E2E-10"),
        ("jbp-commitment-tracking", "JBP commitment tracking", "Track JBP commitments and surface deviations with action queue.", "+22% commitment-delivery rate", "RC-E2E-10"),
        ("retailer-scorecard-rgm", "Retailer scorecard (RGM)", "Retailer scorecard tying volume / mix / margin to RGM strategy.", "+1.4pp net rate-per-customer", "RC-E2E-10"),
        ("retailer-trade-mix-modeling", "Retailer trade-mix modeling", "Retailer-specific trade-mix modeling for negotiated annual plan.", "+9% retailer-net revenue", "RC-E2E-10"),
    ],
    "Promotional Funding": [
        ("trade-fund-leakage-detection", "Trade-fund leakage detection", "Trade-fund-leakage detection with corrective-action workflow.", "$2.4M recovered / yr", "RC-E2E-10"),
        ("scan-deal-vs-off-invoice-mix", "Scan-deal vs off-invoice mix", "Scan-deal vs off-invoice mix optimization for compliance and cost.", "+8% trade-fund efficiency", "RC-E2E-10"),
        ("vendor-funded-display-roi", "Vendor-funded display ROI", "Vendor-funded display ROI with planogram-compliance audit.", "+14% display ROI", "RC-E2E-10"),
    ],
},
"New Product Introduction": {
    "Concept & Test": [
        ("concept-test-orchestration", "Concept-test orchestration", "Concept-test orchestration with consumer-panel and digital-listening signals.", "−22% concept-test cycle", "RC-E2E-04"),
        ("npd-pipeline-prioritization", "NPD pipeline prioritization", "NPD pipeline prioritization based on launch-success modeling.", "+18% pipeline ROI", "RC-E2E-04"),
        ("first-mover-vs-fast-follower", "First-mover vs fast-follower modeling", "First-mover vs fast-follower modeling for category-entry decisions.", "+11% category-entry success", "RC-E2E-04"),
    ],
    "Launch Planning": [
        ("launch-readiness-scoring", "Launch-readiness scoring", "Launch-readiness scoring across supply / marketing / retailer dimensions.", "+27% on-time launch rate", "RC-E2E-01"),
        ("launch-supply-allocation", "Launch supply allocation", "Constrained-supply allocation across retailers at launch.", "−18% retailer-allocation friction", "RC-E2E-02"),
        ("launch-promo-design", "Launch promo design", "Launch promo design with trade-funded incentive modeling.", "+22% launch-velocity uplift", "RC-E2E-10"),
    ],
    "In-Life Performance": [
        ("in-life-performance-tracking", "In-life performance tracking", "Multi-source in-life performance tracking with course-correction.", "+14% NPI revenue / launch", "RC-E2E-04"),
        ("range-life-cycle-management", "Range life-cycle management", "Range life-cycle management with EOL / refresh / extension decisions.", "+1.6pp portfolio gross margin", "RC-E2E-03"),
        ("npi-trade-up-cannibalization", "NPI trade-up / cannibalization", "Trade-up vs cannibalization decomposition for new variants.", "+9% net incremental revenue", "RC-E2E-03"),
    ],
    "EOL Decisioning": [
        ("eol-decision-support", "EOL decision support", "End-of-life decisioning with sell-through / liquidation modeling.", "+22% EOL recovery / unit", "RC-E2E-03"),
        ("end-of-life-merchandise-flow", "End-of-life merchandise flow", "EOL merchandise flow with retailer / liquidator routing.", "+11% blended recovery", "RC-E2E-07"),
        ("end-of-life-consumer-comms", "End-of-life consumer comms", "EOL consumer communications with successor-product transition path.", "+18% successor adoption", "RC-E2E-04"),
        ("legacy-sku-spare-parts-window", "Legacy-SKU spare-parts window", "Legacy-SKU spare-parts window planning for warranty obligations.", "+99% warranty-coverage compliance", "RC-E2E-09"),
    ],
    "Innovation Pipeline": [
        ("innovation-pipeline-portfolio", "Innovation pipeline portfolio view", "Cross-stage innovation pipeline portfolio view with risk-adjusted value.", "+22% pipeline NPV", "RC-E2E-04"),
        ("ideation-to-shelf-acceleration", "Ideation-to-shelf acceleration", "Ideation-to-shelf cycle compression via parallel-stage orchestration.", "−34% time-to-shelf", "RC-E2E-04"),
        ("trend-signal-detection", "Trend signal detection", "Detect emerging consumer trends via social and search-trend signals.", "+18% on-trend launch rate", "RC-E2E-04"),
        ("co-creation-with-consumers", "Co-creation with consumers", "Consumer co-creation programs with idea-curation and selection.", "+24% on-trend NPI", "RC-E2E-04"),
    ],
},
"Foodservice & Restaurant Ops": {
    "Menu Engineering": [
        ("menu-engineering-margin-mix", "Menu engineering margin/mix", "Menu engineering across margin / mix / popularity quadrants.", "+2.8pp menu margin", "RC-E2E-12"),
        ("menu-localization", "Menu localization", "Per-region menu localization with substitution impact modeling.", "+5% same-store sales", "RC-E2E-12"),
        ("menu-substitution-impact", "Menu-substitution impact", "Item-substitution impact modeling on cost / quality / experience.", "+1.6pp food cost", "RC-E2E-12"),
        ("limited-time-offer-design", "Limited-time-offer design", "LTO design with margin and operational complexity overlay.", "+11% LTO take rate", "RC-E2E-12"),
    ],
    "Kitchen Operations": [
        ("kitchen-station-balancing", "Kitchen station balancing", "Real-time kitchen station load balancing with order routing.", "−18% ticket time", "RC-E2E-12"),
        ("food-prep-yield-tracking", "Food-prep yield tracking", "Food-prep yield tracking with waste-cause attribution.", "−14% prep waste", "RC-E2E-12"),
        ("equipment-uptime-monitoring", "Equipment-uptime monitoring", "Kitchen-equipment uptime monitoring with PdM.", "+9pp equipment uptime", "RC-E2E-11"),
        ("freshness-forecast-prep-pacing", "Freshness forecast / prep pacing", "Demand-driven prep pacing for freshness windows.", "+13% freshness score", "RC-E2E-12"),
    ],
    "Drive-Thru & Off-Premise": [
        ("drive-thru-time-optimization", "Drive-thru time optimization", "Drive-thru cycle-time optimization with order-priority routing.", "−22% drive-thru cycle time", "RC-E2E-12"),
        ("drive-thru-personalized-menu", "Drive-thru personalized menu", "Personalized menu boards by daypart / weather / loyalty signal.", "+8% drive-thru AOV", "RC-E2E-12"),
        ("delivery-aggregator-management", "Delivery aggregator management", "Multi-aggregator orchestration with margin-aware acceptance rules.", "+11% aggregator margin", "RC-E2E-12"),
        ("ghost-kitchen-orchestration", "Ghost-kitchen orchestration", "Ghost / virtual kitchen orchestration across brands per kitchen.", "+18% kitchen utilization", "RC-E2E-12"),
    ],
    "Restaurant Loyalty & Engagement": [
        ("rest-loyalty-frequency-driver", "Restaurant loyalty frequency driver", "Restaurant-loyalty visit-frequency driver with personalized challenges.", "+12% visit frequency", "RC-E2E-04"),
        ("guest-recovery-service-failure", "Guest-recovery service failure", "Guest-recovery for service failures with empowered offer-engine.", "+22pt service-recovery NPS", "RC-E2E-04"),
    ],
    "Food Safety & Recipe Compliance": [
        ("rest-temp-monitoring-compliance", "Restaurant temp monitoring compliance", "Restaurant temperature-monitoring compliance with auto-action.", "+99% temp-log compliance", "RC-E2E-15"),
        ("allergen-handling-procedure", "Allergen-handling procedure compliance", "Allergen-handling procedure-compliance monitoring with audit row.", "+99% allergen-handling compliance", "RC-E2E-15"),
        ("daily-line-check-compliance", "Daily line-check compliance", "Daily line-check compliance with photo-evidence and trend analytics.", "+24% line-check completion", "RC-E2E-15"),
        ("food-safety-incident-response", "Food-safety incident response", "Food-safety-incident response orchestration with regulatory reporting.", "−54% incident-response time", "RC-E2E-15"),
        ("supplier-food-safety-audit", "Supplier food-safety audit", "Supplier food-safety audit cadence with corrective-action queue.", "+22% supplier-audit pass rate", "RC-E2E-09"),
    ],
    "Restaurant Workforce": [
        ("hourly-labor-rest-scheduling", "Hourly-labor restaurant scheduling", "Hourly-labor restaurant scheduling with drive-thru / kitchen station match.", "+11% labor productivity", "RC-E2E-08"),
        ("crew-cross-training-management", "Crew cross-training management", "Crew cross-training matrix and shift-station-fit optimization.", "+18% station-fit rate", "RC-E2E-08"),
        ("manager-shift-handoff", "Manager shift-handoff orchestration", "Manager shift-handoff orchestration with auto-summary and outstanding tasks.", "−24% handoff-defect rate", "RC-E2E-08"),
        ("training-by-station", "Station-specific training program", "Station-specific training program with proficiency tracking.", "+34% station proficiency", "RC-E2E-08"),
    ],
},
}

# -------------------------------------------------------------------------
# Generate scenario rows
# -------------------------------------------------------------------------

def make_scenarios():
    rows = []
    n = 0
    for domain, subareas in CATALOG.items():
        for subarea, items in subareas.items():
            for slug, title, brief, kpi, svc in items:
                n += 1
                rows.append({
                    "n": n,
                    "id": f"rc-{slug}",
                    "title": title,
                    "domain": domain,
                    "subarea": subarea,
                    "service_code": svc,
                    "service_name": SERVICES[svc],
                    "brief": brief,
                    "kpi": kpi,
                    "schemas": SCHEMA_BY_DOMAIN[domain],
                    "persona": PERSONA_BY_DOMAIN[domain],
                })
    return rows

# Featured chains: pick ~30 marquee scenarios across domains
FEATURED_SLUGS = {
    "rc-cold-chain-excursion-mid-shift",
    "rc-dynamic-markdown-optimization-rc",
    "rc-loyalty-churn-prediction-winback",
    "rc-on-shelf-availability-oos",
    "rc-returns-fraud-detection",
    "rc-orc-organized-retail-crime",
    "rc-shelf-cv-audit",
    "rc-self-checkout-shrink-detection",
    "rc-real-time-recommendation",
    "rc-attribution-modeling-multi-touch",
    "rc-baseline-demand-forecasting",
    "rc-auto-replenishment-tuning",
    "rc-dynamic-route-optimization",
    "rc-cold-chain-lot-trace",
    "rc-labor-forecast-and-scheduling",
    "rc-attrition-risk-prediction",
    "rc-cart-abandonment-recovery",
    "rc-search-relevance-tuning",
    "rc-account-takeover-detection",
    "rc-payment-settlement-reconciliation",
    "rc-close-acceleration-flux-explain",
    "rc-scope3-emissions-mapping",
    "rc-food-waste-shrink-reduction",
    "rc-site-selection-trade-area",
    "rc-lease-cam-recon",
    "rc-oee-monitoring-line",
    "rc-predictive-maintenance-vibration",
    "rc-recall-readiness-rehearsal",
    "rc-trade-spend-effectiveness",
    "rc-deduction-management-automation",
    "rc-launch-readiness-scoring",
    "rc-menu-engineering-margin-mix",
    "rc-drive-thru-time-optimization",
    "rc-planogram-compliance-monitoring",
    "rc-refrigeration-asset-monitoring",
}

# -------------------------------------------------------------------------
# Chain templates (W1, W2, W3) — match the reference patterns
# -------------------------------------------------------------------------

W1_PATTERN = "SOR · Real-Time Hub · Bronze · Raw Landing · Tokenizer · Silver · Canonical"
W2_PATTERN = "Event Fires · DAG Orchestrator · Agent 1: Assess · Agent 2: Classify · Agent 3: Quantify · Agent 4: Recommend · Agent 5: Approve (HITL)"
W3_PATTERN = "Enterprise Scale · Fuse: cross-domain · Purview Trust · Feedback Loop"

# 24-step layered chain definition (mirrors APEX_FLOW_DATA)
STEPS_24 = [
    # (wave, step_no, key, title, layer, kind, purpose)
    ("W1", 1,  "w1-sor",        "SOR · System of Record",                 "Integration Plane · SOR",       "source",    "Original system data lives here · agents do not replace the SOR"),
    ("W1", 2,  "w1-rth",        "Real-Time Hub · streaming ingest",       "Integration Plane",             "ingest",    "Eventstream / MQTT / OPC-UA bridge from operational systems"),
    ("W1", 3,  "w1-bronze",     "Bronze landing · raw + ingest metadata", "Data Plane · Bronze",           "data",      "Source-aligned · append-only · 5-field ingest metadata"),
    ("W1", 4,  "w1-tokenizer",  "Tokenizer / PII handling",               "Data Plane · Bronze",           "transform", "Hash/redact PII before downstream availability"),
    ("W1", 5,  "w1-silver",     "Silver canonical · MERML / SCML / FINML", "Data Plane · Silver",          "canonical", "Conformed canonical layer agents read against"),
    ("W1", 6,  "w1-quality",    "Data quality · contracts · drift",       "Data Plane · Silver",           "quality",   "Schema contracts, freshness SLOs, drift detection"),
    ("W1", 7,  "w1-purview",    "Purview · catalog · classify · sensitivity", "Governance Plane",          "govern",    "Sensitivity / purpose / lineage at canonical layer"),
    ("W1", 8,  "w1-graph",      "Graph / context layer",                  "Data Plane · Gold",             "context",   "Cross-entity context (customer 360, item 360)"),

    ("W2", 9,  "w2-event",       "Event Fires · scenario trigger",         "Activation Plane",              "event",     "Domain event triggers the agent chain"),
    ("W2", 10, "w2-dag",         "DAG Orchestrator · ORCH archetype",      "Activation Plane",              "orchestrate","Hierarchical-root + sequential-with-hitl-gate composition"),
    ("W2", 11, "w2-agent-assess","Agent 1 · Assess",                       "Activation Plane",              "agent",     "Reads canonical · forms situation assessment"),
    ("W2", 12, "w2-agent-class", "Agent 2 · Classify",                     "Activation Plane",              "agent",     "Routes to scenario class · selects downstream policy"),
    ("W2", 13, "w2-agent-quant", "Agent 3 · Quantify",                     "Activation Plane",              "agent",     "Quantifies impact · simulated P&L · risk band"),
    ("W2", 14, "w2-agent-recom", "Agent 4 · Recommend",                    "Activation Plane",              "agent",     "Generates recommended action with rationale"),
    ("W2", 15, "w2-hitl",        "HITL approval gate · Teams card",         "Activation Plane",              "hitl",      "Human-in-the-loop gate above policy threshold"),
    ("W2", 16, "w2-action",      "Action · write back to SOR",             "Integration Plane",             "action",    "Audit-row write to SOR via integration adapter"),

    ("W3", 17, "w3-scale",       "Enterprise scale rollout",                "Activation Plane",              "scale",     "Multi-region rollout with commercial envelope"),
    ("W3", 18, "w3-fuse-1",      "Fuse · cross-scenario co-orchestration",  "Activation Plane",              "fuse",      "Compose with adjacent scenarios in same domain"),
    ("W3", 19, "w3-fuse-2",      "Fuse · cross-domain co-orchestration",    "Activation Plane",              "fuse",      "Compose with adjacent scenarios in other domains"),
    ("W3", 20, "w3-trust",       "Purview Trust · audit row · sensitivity", "Governance Plane",              "govern",    "Audit row / purpose / sensitivity for every action"),
    ("W3", 21, "w3-feedback",    "Feedback loop · quality · KPI",            "Activation Plane",              "feedback",  "Realized KPI feedback into model and policy"),
    ("W3", 22, "w3-policy",      "Policy / commercial envelope",            "Governance Plane",              "govern",    "Commercial envelope and policy versioning"),
    ("W3", 23, "w3-residency",   "Residency / sovereignty enforcement",     "Governance Plane",              "govern",    "Region-aware data and inference residency"),
    ("W3", 24, "w3-renewal",     "Continuous renewal · model / policy",     "Governance Plane",              "govern",    "Continuous renewal cadence on models and policies"),
]

# -------------------------------------------------------------------------
# Excel writer
# -------------------------------------------------------------------------

HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="0B0B0F")
HDR_AL   = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN     = Side(style="thin", color="D4D4D7")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND     = PatternFill("solid", fgColor="F3F3F5")
CELL     = Font(name="Calibri", size=10)
CODE_FT  = Font(name="Consolas", size=10, color="6648B0", bold=True)
CELL_AL  = Alignment(horizontal="left", vertical="top", wrap_text=True)
TITLE_FT = Font(name="Calibri", size=18, bold=True, color="0B0B0F")

def write_hdrs(ws, headers, widths, *, row=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_AL; c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[row].height = 32


def fmt_data_row(ws, r, cols, *, code_cols=()):
    for j in range(1, cols + 1):
        c = ws.cell(row=r, column=j)
        c.font = CODE_FT if j in code_cols else CELL
        c.alignment = CELL_AL
        c.border = BORDER
        if r % 2 == 0:
            c.fill = BAND


def build_xlsx(rows):
    wb = Workbook()

    # ---- Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "APEX Retail & Consumer Scenario Chains"
    ws["A1"].font = TITLE_FT
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    feature_count = sum(1 for r in rows if r["id"] in FEATURED_SLUGS)
    domain_count  = len({r["domain"] for r in rows})
    subarea_count = len({(r["domain"], r["subarea"]) for r in rows})

    summary_rows = [
        ("", "", "", ""),
        ("Source", "APEX Industry Reference · Retail & Consumer", "", ""),
        ("Source pattern", "Modeled on APEX-Scenario-Chains.xlsx", "", ""),
        ("Total scenarios", len(rows), "", ""),
        ("Domains", domain_count, "", ""),
        ("Sub-areas", subarea_count, "", ""),
        ("Featured chains (full Scenario→KPI · 24-step flow)", feature_count, "", ""),
        ("", "", "", ""),
        ("Sheets", "", "", ""),
        ("1. Summary", "this overview", "", ""),
        ("2. Scenario Library", f"{len(rows)} scenarios with #, ID, title, domain, sub-area, service, brief, KPI, persona", "", ""),
        ("3. Domain Index", f"{domain_count} domains × sub-areas with counts", "", ""),
        ("4. Featured Chains", f"{feature_count} marquee scenarios with Wave 1 / Wave 2 / Wave 3 detail", "", ""),
        ("5. Scenario→KPI Chain", f"{len(rows)} scenarios decomposed: scenario → solution → use case → service → persona → KPI", "", ""),
        ("6. 24-Step Chain", f"{len(rows)} scenarios × 24 architectural steps = {len(rows)*24:,} rows", "", ""),
    ]
    for i, row in enumerate(summary_rows, start=2):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v if v != "" else None)
            c.font = Font(name="Calibri", size=11, bold=(j == 1 and i <= 8))
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # ---- Scenario Library
    ws = wb.create_sheet("Scenario Library")
    headers = ["#", "Scenario ID", "Title", "Domain", "Sub-Area", "Service Code", "Service",
               "Schemas", "Brief (the moment)", "KPI / Outcome", "Persona", "Featured?"]
    widths  = [6, 42, 38, 30, 32, 14, 28, 36, 70, 26, 38, 12]
    write_hdrs(ws, headers, widths)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["n"])
        ws.cell(row=i, column=2, value=r["id"])
        ws.cell(row=i, column=3, value=r["title"])
        ws.cell(row=i, column=4, value=r["domain"])
        ws.cell(row=i, column=5, value=r["subarea"])
        ws.cell(row=i, column=6, value=r["service_code"])
        ws.cell(row=i, column=7, value=r["service_name"])
        ws.cell(row=i, column=8, value=r["schemas"])
        ws.cell(row=i, column=9, value=r["brief"])
        ws.cell(row=i, column=10, value=r["kpi"])
        ws.cell(row=i, column=11, value=r["persona"])
        ws.cell(row=i, column=12, value="★ Featured" if r["id"] in FEATURED_SLUGS else "Catalog")
        fmt_data_row(ws, i, len(headers), code_cols=(2, 6))
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # ---- Domain Index
    ws = wb.create_sheet("Domain Index")
    headers = ["Domain", "Sub-Area", "Scenario Count", "Featured Count", "Service Codes"]
    widths  = [38, 38, 16, 16, 48]
    write_hdrs(ws, headers, widths)

    # group counts
    from collections import defaultdict
    grouping = defaultdict(lambda: {"count": 0, "feat": 0, "svcs": set()})
    for r in rows:
        k = (r["domain"], r["subarea"])
        grouping[k]["count"] += 1
        if r["id"] in FEATURED_SLUGS:
            grouping[k]["feat"] += 1
        grouping[k]["svcs"].add(r["service_code"])

    i = 2
    for (dom, sub), v in grouping.items():
        ws.cell(row=i, column=1, value=dom)
        ws.cell(row=i, column=2, value=sub)
        ws.cell(row=i, column=3, value=v["count"])
        ws.cell(row=i, column=4, value=v["feat"])
        ws.cell(row=i, column=5, value=" · ".join(sorted(v["svcs"])))
        fmt_data_row(ws, i, len(headers))
        i += 1

    # totals row
    total_row = i
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=f"{len(grouping)} sub-areas")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")
    for j in range(1, 6):
        c = ws.cell(row=total_row, column=j)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.fill = PatternFill("solid", fgColor="FFEFB7")
        c.border = BORDER
        c.alignment = CELL_AL

    ws.freeze_panes = "A2"

    # ---- Featured Chains
    ws = wb.create_sheet("Featured Chains")
    headers = ["#", "Scenario ID", "Title", "Service", "Domain", "Sub-Area",
               "W1 Foundation", "W2 Pilot (you are here)", "W3 Scale & Fuse",
               "Scenario (the moment)", "Solution / Agent (ORCH archetype)",
               "Service (Deloitte unit)", "Persona (operator · HITL approver)", "KPI / Outcome"]
    widths  = [5, 42, 38, 14, 30, 32, 60, 60, 60, 70, 60, 28, 38, 28]
    write_hdrs(ws, headers, widths)

    feat_rows = [r for r in rows if r["id"] in FEATURED_SLUGS]
    for i, r in enumerate(feat_rows, start=2):
        # Find adjacent scenarios in same domain for "Fuse" callouts
        same_dom = [s for s in rows if s["domain"] == r["domain"] and s["id"] != r["id"]]
        fuse_a = same_dom[0]["title"] if len(same_dom) > 0 else ""
        fuse_b = same_dom[1]["title"] if len(same_dom) > 1 else ""

        w1 = f"SOR: {r['schemas'].split('·')[0].strip()} · Real-Time Hub · Bronze · Raw Landing · Tokenizer · Silver · Canonical"
        w2 = f"{r['title']} · Event Fires · DAG Orchestrator · Agent 1: Assess · Agent 2: Classify · Agent 3: Quantify · Agent 4: Recommend · Agent 5: Approve (HITL)"
        w3 = f"Enterprise Scale · Fuse: {fuse_a[:40]} · Fuse: {fuse_b[:40]} · Purview Trust · Feedback Loop"

        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=r["id"])
        ws.cell(row=i, column=3, value=r["title"])
        ws.cell(row=i, column=4, value=r["service_code"])
        ws.cell(row=i, column=5, value=r["domain"])
        ws.cell(row=i, column=6, value=r["subarea"])
        ws.cell(row=i, column=7, value=w1)
        ws.cell(row=i, column=8, value=w2)
        ws.cell(row=i, column=9, value=w3)
        ws.cell(row=i, column=10, value=r["brief"])
        ws.cell(row=i, column=11, value="Standard archetype: hierarchical-root + sequential-with-hitl-gate · 5-agent fleet · APEX canonical")
        ws.cell(row=i, column=12, value=r["service_name"])
        ws.cell(row=i, column=13, value=r["persona"] + " · HITL approval via Teams Adaptive Card")
        ws.cell(row=i, column=14, value=r["kpi"])
        fmt_data_row(ws, i, len(headers), code_cols=(2, 4))
    ws.freeze_panes = "C2"

    # ---- Scenario→KPI Chain
    ws = wb.create_sheet("Scenario→KPI Chain")
    headers = ["Scenario ID", "Domain", "Sub-Area", "Scenario (the moment)",
               "Solution / Agent (ORCH archetype)", "Use Cases (decomposed)",
               "Service (Deloitte unit)", "Persona (operator · HITL approver)",
               "KPI (measurable outcome)", "Featured?"]
    widths  = [42, 30, 32, 70, 56, 50, 28, 38, 28, 12]
    write_hdrs(ws, headers, widths)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["id"])
        ws.cell(row=i, column=2, value=r["domain"])
        ws.cell(row=i, column=3, value=r["subarea"])
        ws.cell(row=i, column=4, value=r["brief"])
        ws.cell(row=i, column=5, value="Standard archetype: hierarchical-root + sequential-with-hitl-gate · 5-agent fleet · APEX canonical")
        ws.cell(row=i, column=6, value=f"Wave 2 pilot delivery for {r['title']}")
        ws.cell(row=i, column=7, value=r["service_code"] + " · " + r["service_name"])
        ws.cell(row=i, column=8, value=r["persona"] + " · HITL approval via Teams Adaptive Card")
        ws.cell(row=i, column=9, value=r["kpi"])
        ws.cell(row=i, column=10, value="★ Featured" if r["id"] in FEATURED_SLUGS else "")
        fmt_data_row(ws, i, len(headers), code_cols=(1,))
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # ---- 24-Step Chain
    ws = wb.create_sheet("24-Step Chain")
    headers = ["Scenario ID", "Title", "Domain", "Sub-Area", "Wave", "Step #", "Step Key",
               "Step Title", "Layer", "Kind", "Purpose", "Scenario Note"]
    widths  = [40, 36, 28, 30, 6, 7, 18, 38, 28, 14, 56, 60]
    write_hdrs(ws, headers, widths)

    rownum = 2
    for r in rows:
        for wave, step_no, key, title, layer, kind, purpose in STEPS_24:
            ws.cell(row=rownum, column=1, value=r["id"])
            ws.cell(row=rownum, column=2, value=r["title"])
            ws.cell(row=rownum, column=3, value=r["domain"])
            ws.cell(row=rownum, column=4, value=r["subarea"])
            ws.cell(row=rownum, column=5, value=wave)
            ws.cell(row=rownum, column=6, value=step_no)
            ws.cell(row=rownum, column=7, value=key)
            ws.cell(row=rownum, column=8, value=title)
            ws.cell(row=rownum, column=9, value=layer)
            ws.cell(row=rownum, column=10, value=kind)
            ws.cell(row=rownum, column=11, value=purpose)
            ws.cell(row=rownum, column=12, value=r["brief"])
            for j in range(1, len(headers) + 1):
                c = ws.cell(row=rownum, column=j)
                c.font = CODE_FT if j in (1, 7) else CELL
                c.alignment = CELL_AL
                c.border = BORDER
                if rownum % 2 == 0:
                    c.fill = BAND
            rownum += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{rownum-1}"

    OUT_XL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XL)
    print(f"  → {OUT_XL}  ({rownum-1:,} 24-step rows · {len(rows):,} scenarios)")


# -------------------------------------------------------------------------
# HTML writer
# -------------------------------------------------------------------------

def build_html(rows):
    feat_set = {r["id"] for r in rows if r["id"] in FEATURED_SLUGS}
    feat_rows = [r for r in rows if r["id"] in FEATURED_SLUGS]

    # Group rows by domain → sub-area
    from collections import defaultdict, OrderedDict
    by_domain = OrderedDict()
    for r in rows:
        by_domain.setdefault(r["domain"], OrderedDict()).setdefault(r["subarea"], []).append(r)

    domain_count = len(by_domain)
    subarea_count = sum(len(s) for s in by_domain.values())

    DOMAIN_COLORS = [
        ("#0B0B0F", "#6648B0"),
        ("#0B0B0F", "#0EA5E9"),
        ("#0B0B0F", "#10B981"),
        ("#0B0B0F", "#F59E0B"),
        ("#0B0B0F", "#EF4444"),
        ("#0B0B0F", "#8B5CF6"),
        ("#0B0B0F", "#EC4899"),
        ("#0B0B0F", "#14B8A6"),
        ("#0B0B0F", "#F97316"),
        ("#0B0B0F", "#06B6D4"),
        ("#0B0B0F", "#84CC16"),
        ("#0B0B0F", "#A855F7"),
        ("#0B0B0F", "#3B82F6"),
        ("#0B0B0F", "#22C55E"),
        ("#0B0B0F", "#FB7185"),
        ("#0B0B0F", "#0891B2"),
        ("#0B0B0F", "#D946EF"),
    ]
    domain_color_map = {d: DOMAIN_COLORS[i % len(DOMAIN_COLORS)] for i, d in enumerate(by_domain)}

    def esc(s):
        return html.escape(str(s))

    parts = []
    parts.append("""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>APEX · Retail & Consumer Scenario Chains</title>
<style>
  :root {
    --ink: #0B0B0F;
    --ink-2: #1F2024;
    --ink-3: #4A4D55;
    --bg: #FAFAFC;
    --bg-card: #FFFFFF;
    --line: #E4E4E7;
    --line-2: #D4D4D7;
    --accent: #6648B0;
    --accent-2: #8E70D8;
    --kpi: #047857;
    --w1: #0EA5E9;
    --w2: #6648B0;
    --w3: #10B981;
    --hitl: #F59E0B;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; }
  body {
    font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, 'Helvetica Neue', sans-serif;
    background: var(--bg);
    color: var(--ink);
    font-size: 14px;
    line-height: 1.5;
  }
  header {
    background: linear-gradient(135deg, #0B0B0F 0%, #1F2024 60%, #6648B0 100%);
    color: #FFF;
    padding: 36px 32px 28px;
  }
  header h1 {
    margin: 0 0 8px 0;
    font-size: 30px;
    font-weight: 700;
    letter-spacing: -0.02em;
  }
  header .sub {
    font-size: 14px;
    opacity: 0.85;
    max-width: 980px;
  }
  header .stats {
    margin-top: 18px;
    display: flex;
    gap: 14px;
    flex-wrap: wrap;
  }
  header .stat {
    background: rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.18);
    border-radius: 10px;
    padding: 10px 14px;
  }
  header .stat .v { font-size: 22px; font-weight: 700; }
  header .stat .l { font-size: 11px; opacity: 0.78; text-transform: uppercase; letter-spacing: 0.04em; }

  .controls {
    background: var(--bg-card);
    border-bottom: 1px solid var(--line);
    padding: 14px 32px;
    position: sticky;
    top: 0;
    z-index: 50;
    display: flex;
    gap: 10px;
    align-items: center;
    flex-wrap: wrap;
  }
  .controls input[type="search"] {
    flex: 1 1 280px;
    min-width: 240px;
    padding: 10px 12px;
    border: 1px solid var(--line-2);
    border-radius: 8px;
    font-size: 13px;
    font-family: inherit;
  }
  .controls select {
    padding: 9px 10px;
    border: 1px solid var(--line-2);
    border-radius: 8px;
    font-size: 13px;
    background: #FFF;
    font-family: inherit;
  }
  .controls .toggles {
    display: flex;
    gap: 6px;
  }
  .controls .toggle {
    border: 1px solid var(--line-2);
    background: #FFF;
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 12px;
    cursor: pointer;
    color: var(--ink-2);
  }
  .controls .toggle.active { background: var(--ink); color: #FFF; border-color: var(--ink); }

  main { padding: 28px 32px 60px; }

  .domain {
    margin-bottom: 36px;
    background: var(--bg-card);
    border: 1px solid var(--line);
    border-radius: 14px;
    overflow: hidden;
  }
  .domain-head {
    padding: 18px 22px;
    border-bottom: 1px solid var(--line);
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    color: #FFF;
  }
  .domain-head h2 {
    margin: 0;
    font-size: 19px;
    font-weight: 700;
    letter-spacing: -0.01em;
  }
  .domain-head .badge {
    background: rgba(255,255,255,0.18);
    border: 1px solid rgba(255,255,255,0.28);
    border-radius: 999px;
    padding: 4px 12px;
    font-size: 12px;
    font-weight: 600;
  }

  .subarea {
    padding: 14px 22px 6px;
    border-bottom: 1px dashed var(--line);
  }
  .subarea:last-child { border-bottom: none; }
  .subarea-head {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 10px;
  }
  .subarea-head h3 {
    margin: 0;
    font-size: 14px;
    font-weight: 700;
    color: var(--ink-2);
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }
  .subarea-head .pill {
    background: var(--bg);
    border: 1px solid var(--line-2);
    border-radius: 999px;
    padding: 2px 10px;
    font-size: 11px;
    color: var(--ink-3);
    font-weight: 600;
  }

  .cards {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(420px, 1fr));
    gap: 12px;
    padding-bottom: 12px;
  }
  .card {
    border: 1px solid var(--line);
    border-radius: 10px;
    padding: 14px 16px;
    background: var(--bg-card);
    transition: box-shadow 0.18s, transform 0.18s;
  }
  .card:hover {
    box-shadow: 0 4px 18px rgba(11,11,15,0.07);
    transform: translateY(-1px);
  }
  .card.featured {
    border-left: 3px solid var(--accent);
    background: linear-gradient(180deg, #FBFAFF 0%, #FFFFFF 30%);
  }
  .card .row1 {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 10px;
    margin-bottom: 6px;
  }
  .card h4 {
    margin: 0;
    font-size: 14px;
    font-weight: 700;
    color: var(--ink);
    letter-spacing: -0.005em;
    flex: 1;
  }
  .card .star {
    color: var(--accent);
    font-size: 14px;
    flex-shrink: 0;
  }
  .card .id {
    font-family: 'Cascadia Code', Consolas, monospace;
    color: var(--accent);
    font-size: 11px;
    font-weight: 600;
    margin-bottom: 8px;
  }
  .card .brief {
    font-size: 12px;
    color: var(--ink-2);
    margin-bottom: 10px;
    line-height: 1.5;
  }
  .card .kpi {
    color: var(--kpi);
    font-weight: 700;
    font-size: 12px;
    background: #ECFDF5;
    border: 1px solid #A7F3D0;
    border-radius: 6px;
    padding: 4px 8px;
    display: inline-block;
    margin-bottom: 8px;
  }
  .card .meta {
    display: flex;
    flex-wrap: wrap;
    gap: 4px;
    margin-bottom: 8px;
  }
  .card .meta .tag {
    font-size: 10px;
    background: var(--bg);
    border: 1px solid var(--line);
    border-radius: 4px;
    padding: 2px 6px;
    color: var(--ink-3);
    font-weight: 500;
  }

  .chain {
    margin-top: 8px;
    border-top: 1px dashed var(--line);
    padding-top: 8px;
    font-size: 11px;
  }
  .chain .wave {
    display: flex;
    gap: 6px;
    margin-bottom: 4px;
    align-items: center;
  }
  .chain .wave .label {
    font-weight: 700;
    font-size: 10px;
    border-radius: 4px;
    padding: 2px 6px;
    color: #FFF;
    flex-shrink: 0;
    width: 28px;
    text-align: center;
  }
  .chain .wave.w1 .label { background: var(--w1); }
  .chain .wave.w2 .label { background: var(--w2); }
  .chain .wave.w3 .label { background: var(--w3); }
  .chain .wave .body { color: var(--ink-2); flex: 1; }

  .featured-section {
    background: var(--bg-card);
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 22px;
    margin-bottom: 36px;
  }
  .featured-section h2 {
    font-size: 19px;
    margin: 0 0 14px 0;
    color: var(--ink);
  }
  .featured-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(540px, 1fr));
    gap: 14px;
  }
  .featured-chain {
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 16px;
    background: linear-gradient(180deg, #FBFAFF, #FFFFFF);
  }
  .featured-chain .head {
    display: flex;
    justify-content: space-between;
    margin-bottom: 8px;
    gap: 10px;
  }
  .featured-chain .head h4 {
    margin: 0;
    font-size: 15px;
    font-weight: 700;
    color: var(--ink);
  }
  .featured-chain .head .svc {
    background: var(--ink);
    color: #FFF;
    border-radius: 6px;
    padding: 2px 8px;
    font-size: 11px;
    font-family: Consolas, monospace;
    font-weight: 600;
    flex-shrink: 0;
    height: fit-content;
  }
  .featured-chain .moment {
    font-size: 12px;
    color: var(--ink-2);
    margin: 6px 0 12px;
    font-style: italic;
  }
  .waves {
    display: grid;
    grid-template-columns: 60px 1fr;
    gap: 4px 10px;
    align-items: start;
    font-size: 11px;
  }
  .waves .lbl {
    font-weight: 700;
    color: #FFF;
    border-radius: 4px;
    padding: 4px 6px;
    text-align: center;
  }
  .waves .lbl.w1 { background: var(--w1); }
  .waves .lbl.w2 { background: var(--w2); }
  .waves .lbl.w3 { background: var(--w3); }
  .waves .body { color: var(--ink-2); padding: 4px 0; }
  .featured-chain .kpi-row {
    margin-top: 10px;
    padding-top: 10px;
    border-top: 1px dashed var(--line);
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 10px;
  }
  .featured-chain .kpi-row .kpi {
    color: var(--kpi);
    font-weight: 700;
    font-size: 12px;
    background: #ECFDF5;
    border: 1px solid #A7F3D0;
    border-radius: 6px;
    padding: 4px 8px;
  }
  .featured-chain .kpi-row .persona {
    font-size: 11px;
    color: var(--ink-3);
  }

  .toc {
    background: var(--bg-card);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 18px 22px;
    margin-bottom: 28px;
  }
  .toc h3 { margin: 0 0 10px 0; font-size: 13px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--ink-3); }
  .toc-list {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(220px, 1fr));
    gap: 6px 14px;
  }
  .toc-list a {
    color: var(--ink-2);
    text-decoration: none;
    font-size: 13px;
    padding: 4px 0;
    border-bottom: 1px dashed transparent;
  }
  .toc-list a:hover {
    border-bottom-color: var(--accent);
    color: var(--accent);
  }
  .toc-list a .ct {
    color: var(--ink-3);
    font-size: 11px;
    margin-left: 4px;
    font-weight: 500;
  }

  footer {
    text-align: center;
    color: var(--ink-3);
    padding: 30px 20px;
    font-size: 12px;
    border-top: 1px solid var(--line);
    background: var(--bg-card);
  }

  .hidden { display: none !important; }
</style>
</head>
<body>
""")

    # Header
    parts.append(f"""<header>
  <h1>APEX · Retail &amp; Consumer Scenario Chains</h1>
  <div class="sub">Industry-specific catalog of agentic scenarios for Retail &amp; Consumer Products. Each scenario carries a Wave 1 (Foundation) → Wave 2 (Pilot) → Wave 3 (Scale &amp; Fuse) chain modeled on the APEX canonical pattern. Modeled on <code>APEX-Scenario-Chains.xlsx</code>.</div>
  <div class="stats">
    <div class="stat"><div class="v">{len(rows)}</div><div class="l">Total scenarios</div></div>
    <div class="stat"><div class="v">{domain_count}</div><div class="l">Domains</div></div>
    <div class="stat"><div class="v">{subarea_count}</div><div class="l">Sub-areas</div></div>
    <div class="stat"><div class="v">{len(feat_rows)}</div><div class="l">Featured chains</div></div>
    <div class="stat"><div class="v">{len(SERVICES)}</div><div class="l">Service codes</div></div>
  </div>
</header>
""")

    # Controls
    parts.append("""<div class="controls">
  <input type="search" id="q" placeholder="Fuzzy search (typo-tolerant) — title / ID / KPI / brief / domain / sub-area…">
  <select id="domain-filter">
    <option value="">All domains</option>
""")
    for d in by_domain:
        parts.append(f'    <option value="{esc(d)}">{esc(d)}</option>\n')
    parts.append('  </select>\n  <select id="subarea-filter">\n    <option value="">All sub-areas</option>\n')
    # Build sub-area options grouped by domain
    for d, subareas in by_domain.items():
        for sub in subareas:
            parts.append(f'    <option value="{esc(sub)}" data-domain="{esc(d)}">{esc(sub)} <span>({esc(d)})</span></option>\n')
    parts.append("""  </select>
  <div class="toggles">
    <button class="toggle active" data-mode="all">All</button>
    <button class="toggle" data-mode="featured">Featured only</button>
  </div>
  <span id="result-count" style="margin-left:auto;font-size:12px;color:var(--ink-3);font-weight:600;"></span>
</div>
<main>
""")

    # TOC
    parts.append('<div class="toc"><h3>Domains</h3><div class="toc-list">\n')
    for d in by_domain:
        anchor = re.sub(r"[^a-z0-9]+", "-", d.lower()).strip("-")
        ct = sum(len(v) for v in by_domain[d].values())
        parts.append(f'  <a href="#dom-{anchor}">{esc(d)}<span class="ct">({ct})</span></a>\n')
    parts.append("</div></div>\n")

    # Featured section
    parts.append('<div class="featured-section">\n  <h2>★ Featured Chains — Wave 1 / Wave 2 / Wave 3 detail</h2>\n  <div class="featured-grid">\n')
    for r in feat_rows:
        same_dom = [s for s in rows if s["domain"] == r["domain"] and s["id"] != r["id"]]
        fuse_a = same_dom[0]["title"] if len(same_dom) > 0 else ""
        fuse_b = same_dom[1]["title"] if len(same_dom) > 1 else ""
        w1 = f"SOR: {r['schemas']} · Real-Time Hub · Bronze · Tokenizer · Silver · Canonical · Purview"
        w2 = f"Event · DAG Orch · Agent 1 Assess · Agent 2 Classify · Agent 3 Quantify · Agent 4 Recommend · HITL gate"
        w3 = f"Enterprise scale · Fuse: {fuse_a} · Fuse: {fuse_b} · Purview Trust · Feedback Loop"
        parts.append(f"""    <div class="featured-chain">
      <div class="head"><h4>{esc(r['title'])}</h4><span class="svc">{esc(r['service_code'])}</span></div>
      <div class="moment">{esc(r['brief'])}</div>
      <div class="waves">
        <span class="lbl w1">W1</span><div class="body">{esc(w1)}</div>
        <span class="lbl w2">W2</span><div class="body">{esc(w2)}</div>
        <span class="lbl w3">W3</span><div class="body">{esc(w3)}</div>
      </div>
      <div class="kpi-row">
        <span class="kpi">{esc(r['kpi'])}</span>
        <span class="persona">{esc(r['persona'])}</span>
      </div>
    </div>
""")
    parts.append('  </div>\n</div>\n')

    # Domain sections
    for d, subareas in by_domain.items():
        anchor = re.sub(r"[^a-z0-9]+", "-", d.lower()).strip("-")
        ink, accent = domain_color_map[d]
        ct = sum(len(v) for v in subareas.values())
        feat_in_d = sum(1 for s in subareas.values() for r in s if r["id"] in feat_set)
        parts.append(f"""<section class="domain" id="dom-{anchor}" data-domain="{esc(d)}">
  <div class="domain-head" style="background: linear-gradient(135deg, {ink}, {accent})" onclick="toggleDomain(this)">
    <h2>{esc(d)}</h2>
    <span class="badge">{ct} scenarios · {len(subareas)} sub-areas · ★{feat_in_d}</span>
  </div>
  <div class="domain-body">
""")
        for sub, items in subareas.items():
            sub_feat = sum(1 for r in items if r["id"] in feat_set)
            parts.append(f"""    <div class="subarea">
      <div class="subarea-head"><h3>{esc(sub)}</h3><span class="pill">{len(items)} scenarios</span>{f'<span class="pill">★ {sub_feat}</span>' if sub_feat else ''}</div>
      <div class="cards">
""")
            for r in items:
                feat_cls = " featured" if r["id"] in feat_set else ""
                star = '<span class="star">★</span>' if r["id"] in feat_set else ""
                w1 = "SOR · RTH · Bronze · Tokenizer · Silver · Purview"
                w2 = "Event · DAG · Assess · Classify · Quantify · HITL"
                w3 = "Scale · Fuse · Trust · Feedback Loop"
                parts.append(f"""        <div class="card{feat_cls}" data-text="{esc((r['title']+' '+r['id']+' '+r['brief']+' '+r['kpi']+' '+r['domain']+' '+r['subarea']+' '+r['service_code']+' '+r['service_name']).lower())}" data-domain="{esc(r['domain'])}" data-subarea="{esc(r['subarea'])}" data-featured="{'1' if r['id'] in feat_set else '0'}">
          <div class="row1"><h4>{esc(r['title'])}</h4>{star}</div>
          <div class="id">{esc(r['id'])}</div>
          <div class="brief">{esc(r['brief'])}</div>
          <div class="kpi">{esc(r['kpi'])}</div>
          <div class="meta"><span class="tag">{esc(r['service_code'])}</span><span class="tag">{esc(r['service_name'])}</span></div>
          <div class="chain">
            <div class="wave w1"><span class="label">W1</span><span class="body">{esc(w1)}</span></div>
            <div class="wave w2"><span class="label">W2</span><span class="body">{esc(w2)}</span></div>
            <div class="wave w3"><span class="label">W3</span><span class="body">{esc(w3)}</span></div>
          </div>
        </div>
""")
            parts.append("      </div>\n    </div>\n")
        parts.append("  </div>\n</section>\n")

    # Footer + script
    parts.append(f"""<footer>
  APEX · Retail &amp; Consumer Scenario Chains · {len(rows)} scenarios · {domain_count} domains · {subarea_count} sub-areas<br>
  Companion to <code>APEX-RC-Scenario-Chains.xlsx</code> · Modeled on <code>APEX-Scenario-Chains.xlsx</code>
</footer>
<script>
  const q = document.getElementById('q');
  const dfilter = document.getElementById('domain-filter');
  const sfilter = document.getElementById('subarea-filter');
  const toggles = document.querySelectorAll('.toggle');
  const resultCount = document.getElementById('result-count');
  let mode = 'all';
  const allCards = Array.from(document.querySelectorAll('.card'));
  const allSubareaOptions = Array.from(sfilter.querySelectorAll('option[data-domain]'));

  toggles.forEach(t => t.addEventListener('click', () => {{
    toggles.forEach(x => x.classList.remove('active'));
    t.classList.add('active');
    mode = t.dataset.mode;
    apply();
  }}));

  // ---- Fuzzy match: each query token must match the card text by
  // (a) substring, OR (b) edit-distance ≤ 1 to any whitespace token in the card
  function editDistanceLE(a, b, max) {{
    if (a === b) return 0;
    if (Math.abs(a.length - b.length) > max) return max + 1;
    const m = a.length, n = b.length;
    if (m === 0) return n; if (n === 0) return m;
    let prev = new Array(n + 1);
    for (let j = 0; j <= n; j++) prev[j] = j;
    for (let i = 1; i <= m; i++) {{
      const cur = new Array(n + 1); cur[0] = i;
      let rowMin = cur[0];
      for (let j = 1; j <= n; j++) {{
        const cost = a.charCodeAt(i - 1) === b.charCodeAt(j - 1) ? 0 : 1;
        cur[j] = Math.min(cur[j - 1] + 1, prev[j] + 1, prev[j - 1] + cost);
        if (cur[j] < rowMin) rowMin = cur[j];
      }}
      if (rowMin > max) return max + 1;
      prev = cur;
    }}
    return prev[n];
  }}

  function fuzzyMatch(text, query) {{
    if (!query) return true;
    const tokens = query.toLowerCase().split(/\s+/).filter(Boolean);
    if (tokens.length === 0) return true;
    const words = text.split(/[^a-z0-9]+/).filter(w => w.length > 0);
    return tokens.every(tok => {{
      if (text.includes(tok)) return true;            // substring (incl. partial)
      if (tok.length < 4) return false;               // too short for typo tolerance
      const max = tok.length >= 7 ? 2 : 1;            // 1 typo for 4-6, 2 for 7+
      for (const w of words) {{
        if (w.length < tok.length - max || w.length > tok.length + max) continue;
        if (editDistanceLE(tok, w, max) <= max) return true;
      }}
      return false;
    }});
  }}

  // Cascading sub-area dropdown — show only sub-areas in the chosen domain
  function refreshSubareaOptions() {{
    const dom = dfilter.value;
    let firstVisible = null;
    allSubareaOptions.forEach(opt => {{
      const show = !dom || opt.dataset.domain === dom;
      opt.hidden = !show;
      opt.disabled = !show;
      if (show && !firstVisible) firstVisible = opt.value;
    }});
    // Reset sub-area if its domain no longer matches
    if (sfilter.value) {{
      const cur = allSubareaOptions.find(o => o.value === sfilter.value);
      if (cur && cur.hidden) sfilter.value = '';
    }}
  }}

  function apply() {{
    const term = q.value.trim().toLowerCase();
    const dom = dfilter.value;
    const sub = sfilter.value;
    let shown = 0;
    allCards.forEach(c => {{
      const cd = c.dataset.domain;
      const cs = c.dataset.subarea;
      const cf = c.dataset.featured === '1';
      let visible = true;
      if (dom && cd !== dom) visible = false;
      if (sub && cs !== sub) visible = false;
      if (mode === 'featured' && !cf) visible = false;
      if (visible && term) visible = fuzzyMatch(c.dataset.text, term);
      c.classList.toggle('hidden', !visible);
      if (visible) shown++;
    }});
    document.querySelectorAll('.subarea').forEach(s => {{
      s.classList.toggle('hidden', !s.querySelector('.card:not(.hidden)'));
    }});
    document.querySelectorAll('.domain').forEach(s => {{
      s.classList.toggle('hidden', !s.querySelector('.subarea:not(.hidden)'));
    }});
    resultCount.textContent = `${{shown}} of ${{allCards.length}} scenarios`;
  }}

  q.addEventListener('input', apply);
  dfilter.addEventListener('change', () => {{ refreshSubareaOptions(); apply(); }});
  sfilter.addEventListener('change', apply);

  function toggleDomain(head) {{
    const body = head.nextElementSibling;
    body.style.display = body.style.display === 'none' ? '' : 'none';
  }}

  refreshSubareaOptions();
  apply();
</script>
</body>
</html>
""")

    OUT_HT.parent.mkdir(parents=True, exist_ok=True)
    OUT_HT.write_text("".join(parts), encoding="utf-8")
    print(f"  → {OUT_HT}")


# -------------------------------------------------------------------------
# Main
# -------------------------------------------------------------------------

if __name__ == "__main__":
    rows = make_scenarios()
    print(f"Total scenarios: {len(rows)}")
    domains = sorted({r['domain'] for r in rows})
    print(f"Domains: {len(domains)}")
    for d in domains:
        c = sum(1 for r in rows if r['domain'] == d)
        subs = len({r['subarea'] for r in rows if r['domain'] == d})
        print(f"  {d}: {c} scenarios across {subs} sub-areas")
    print()

    print("Building Excel...")
    build_xlsx(rows)
    print("Building HTML...")
    build_html(rows)
    print("Done.")

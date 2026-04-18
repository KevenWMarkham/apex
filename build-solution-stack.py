"""Generate APEX-RC-Solution-Stack-Chart.xlsx based on the AXLE reference structure.

Produces a three-sheet workbook:
- Solution Stack: Part/Chapter-organized solutions with schema usage, wave, pipeline,
  MCP bindings, ORCH references, Gold views, Microsoft stack, ref-arch categories.
- Data Source Registry: Retail source systems feeding RC schemas.
- Schema Entities: Entity catalog for MERML/SCML/CXML/MKTL.

Run: python build-solution-stack.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"docs/APEX-RC-Solution-Stack-Chart.xlsx"

# ============================================================================
# SHEET 1: SOLUTION STACK
# ============================================================================
# Columns mirror AXLE structure adapted for RC's 4 schemas.
# Schema usage: P = Primary, S = Secondary, - = unused.
# Ref Arch tags: Data, Connect, App, Prod, Sec (same vocabulary as AXLE).

SOLUTION_COLS = [
    "Part", "Chapter", "Solution", "Wave", "Pipeline",
    "MERML", "SCML", "CXML", "MKTL",
    "MCP Server", "MCP Tools", "Orch", "Gold View", "MS Stack", "Ref Arch"
]

SOLUTIONS = [
    # ------------- PART I: FOUNDATION -------------
    ("I", "Ch 1: Assess", "Store Cluster Segmentation",       "W1", "$1-2M",  "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "Power BI,Fabric",         "Data"),
    ("I", "Ch 1: Assess", "Operating Posture Assessment",      "W1", None,     "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "Purview,Fabric",          "Data,Sec"),
    ("I", "Ch 1: Assess", "Data Gap Scan",                     "W1", None,     "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "Purview",                 "Data,Sec"),
    ("I", "Ch 2: Platform", "OneLake Workspace Build",         "W1", None,     "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "Fabric,OneLake",          "Data"),
    ("I", "Ch 2: Platform", "Entra + Purview Integration",     "W1", None,     "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "Entra,Purview",           "Sec"),
    ("I", "Ch 2: Platform", "Agent Service Baseline",          "W1", None,     "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "AI Agent Service,Foundry","App,Data"),
    ("I", "Ch 2: Platform", "Medallion Skeleton",              "W1", None,     "S", "S", "S", "S", "--",         "--",                               "--",        "--",              "Fabric Lakehouse",        "Data"),
    ("I", "Ch 3: Ingest", "POS Stream to Bronze",              "W1", None,     "P", "-", "-", "-", "--",         "--",                               "--",        "--",              "Event Hubs,Fabric",       "Data,Connect"),
    ("I", "Ch 3: Ingest", "Receiving Dock to Bronze",          "W1", None,     "-", "P", "-", "-", "--",         "--",                               "--",        "--",              "Event Hubs,Fabric",       "Data,Connect"),
    ("I", "Ch 3: Ingest", "ERP + OMS to Bronze",               "W1", None,     "S", "S", "P", "S", "--",         "--",                               "--",        "--",              "Data Factory,Fabric",     "Data"),
    ("I", "Ch 4: Model", "Canonical Event Envelope",           "W1", None,     "P", "P", "P", "P", "--",         "--",                               "--",        "--",              "Fabric Lakehouse",        "Data"),
    ("I", "Ch 4: Model", "Silver Conformance",                 "W1", None,     "P", "P", "P", "P", "--",         "--",                               "--",        "--",              "Fabric Lakehouse",        "Data"),
    ("I", "Ch 4: Model", "Dimensional Anchor Load",            "W1", None,     "P", "S", "S", "S", "--",         "--",                               "--",        "--",              "Fabric Lakehouse",        "Data"),
    ("I", "Ch 4: Model", "Gold View Scaffolding",              "W2", None,     "P", "P", "P", "P", "--",         "--",                               "--",        "--",              "Fabric Lakehouse,Power BI","Data"),

    # ------------- PART II: IN-STORE OPERATIONS -------------
    ("II", "Ch 5: Price Integrity", "ESL Gateway Monitor",     "W1", "$3-5M",  "P", "-", "-", "-", "mcp.store",  "push_verify,stale_tag_scan",        "ORCH-02",   "GV_STORE_PRICE",  "Event Hubs,Logic Apps",    "Data,Connect"),
    ("II", "Ch 5: Price Integrity", "POS Mismatch Detector",   "W1", None,     "P", "-", "-", "-", "mcp.store",  "pos_price_compare,refund_stage",    "ORCH-02",   "GV_STORE_PRICE",  "Event Hubs,Fabric RTI",    "Data,Connect"),
    ("II", "Ch 5: Price Integrity", "Bridge-Tag Dispatch",     "W1", None,     "P", "-", "-", "-", "mcp.store",  "bridge_tag_print,associate_task",   "ORCH-02",   "--",              "Power Apps,Service Desk",  "App"),
    ("II", "Ch 5: Price Integrity", "Dynamic Pricing",         "W2", "$4-8M",  "P", "-", "-", "-", "mcp.store",  "elasticity_score,price_propose",    "(platform)","GV_STORE_PRICE",  "Foundry,Fabric ML",        "Data"),
    ("II", "Ch 5: Price Integrity", "Markdown Cadence",        "W2", None,     "P", "S", "-", "-", "mcp.store",  "expiration_scan,markdown_stage",    "(platform)","GV_STORE_WASTE",  "Fabric,Power BI",          "Data,Prod"),
    ("II", "Ch 6: Receiving", "DSD Reconciliation",            "W1", "$2-4M",  "S", "P", "-", "-", "mcp.store",  "asn_match,line_variance",           "ORCH-01",   "GV_STORE_RECEIVING","Event Hubs,Fabric",        "Data,Connect"),
    ("II", "Ch 6: Receiving", "Vendor Pattern Detector",       "W2", None,     "-", "P", "-", "-", "mcp.store",  "vendor_90d_score",                  "ORCH-01",   "GV_VENDOR_SCORECARD","Fabric,Power BI",         "Data"),
    ("II", "Ch 6: Receiving", "Claim Assembly",                "W2", None,     "-", "P", "-", "-", "mcp.store",  "claim_draft,evidence_attach",       "ORCH-01",   "--",              "Power Apps,Logic Apps",    "App"),
    ("II", "Ch 6: Receiving", "Receiving Summary",             "W1", None,     "-", "P", "-", "-", "mcp.store",  "receiving_summary",                 "--",        "GV_STORE_RECEIVING","Fabric,Power BI",          "Data"),
    ("II", "Ch 7: On-Shelf Availability", "Phantom OOS Detector","W1","$5-9M", "P", "-", "-", "-", "mcp.store",  "cv_pos_fusion,osa_detect",          "ORCH-04",   "GV_STORE_OSA",    "IoT Edge,Event Hubs",      "Data,Connect"),
    ("II", "Ch 7: On-Shelf Availability", "Pick List Optimizer","W1", None,    "P", "-", "-", "-", "mcp.store",  "pick_order,walk_optimize",          "ORCH-04",   "--",              "AI Agent Service",         "App"),
    ("II", "Ch 7: On-Shelf Availability", "Associate Router",  "W2", None,     "P", "-", "-", "-", "mcp.store",  "zone_track,task_dispatch",          "ORCH-04",   "--",              "AI Agent Service,Teams",   "App,Prod"),
    ("II", "Ch 7: On-Shelf Availability", "Planogram Compliance","W2","$2-4M", "P", "-", "-", "-", "mcp.store",  "cv_classify,diff_compute",          "(platform)","GV_STORE_PLANOGRAM","Foundry,Fabric ML",        "Data"),
    ("II", "Ch 7: On-Shelf Availability", "Assortment Optimizer","W3","$6-12M","P", "-", "S", "S", "mcp.store",  "velocity_score,whitespace_detect",  "ORCH-09",   "GV_CATEGORY_MIX", "Foundry,Fabric ML",        "Data"),
    ("II", "Ch 8: Shrink", "Void Pattern Detector",            "W2", "$8-15M", "P", "-", "-", "-", "mcp.store",  "sigma_scan,employee_correlate",     "ORCH-07",   "GV_SHRINK_SIGNAL","Fabric RTI,Foundry",        "Data,Sec"),
    ("II", "Ch 8: Shrink", "Variance Correlator",              "W2", None,     "P", "-", "-", "-", "mcp.store",  "cycle_count_match,return_cross",    "ORCH-07",   "GV_SHRINK_SIGNAL","Fabric,Power BI",          "Data"),
    ("II", "Ch 8: Shrink", "Evidence Package Builder",         "W3", None,     "P", "-", "-", "-", "mcp.store",  "timeline_build,cctv_ref",           "ORCH-07",   "--",              "Fabric,Power Apps",        "App,Sec"),
    ("II", "Ch 8: Shrink", "Returns Fraud Signal",             "W3", None,     "S", "-", "P", "-", "mcp.store",  "return_fraud_score",                "ORCH-07",   "GV_RETURNS_FRAUD","Foundry,Fabric",           "Data,Sec"),

    # ------------- PART III: CUSTOMER EXPERIENCE -------------
    ("III", "Ch 9: Fulfillment", "Pick Exception Handler",     "W2", "$4-7M",  "-", "-", "P", "-", "mcp.order",  "oos_classify,alternative_find",     "ORCH-06",   "GV_BOPIS_EXCEPTIONS","Event Hubs,AI Agent Service","Data,Connect"),
    ("III", "Ch 9: Fulfillment", "Substitution Ranker",        "W2", None,     "-", "-", "P", "-", "mcp.order",  "history_rank,substitution_score",   "ORCH-06",   "--",              "Foundry,Fabric ML",        "Data"),
    ("III", "Ch 9: Fulfillment", "Customer Confirm Gateway",   "W2", None,     "-", "-", "P", "-", "mcp.order",  "sms_send,aging_monitor",            "ORCH-06",   "--",              "Logic Apps,SMS Gateway",   "Connect,App"),
    ("III", "Ch 9: Fulfillment", "Pickup Orchestration",       "W2", None,     "-", "S", "P", "-", "mcp.order",  "slot_manage,pickup_flow",           "ORCH-06",   "GV_CUSTOMER_ORDER","Power Apps,Teams",         "App,Prod"),
    ("III", "Ch 10: Returns", "Returns Disposition Agent",     "W3", "$3-6M",  "S", "-", "P", "-", "mcp.store",  "disposition_classify,rtv_draft",    "ORCH-12",   "GV_RETURNS_FLOW", "AI Agent Service,Fabric",  "App"),
    ("III", "Ch 10: Returns", "Customer Incident Response",    "W3", None,     "-", "S", "P", "-", "mcp.store",  "ocr_lot,pattern_correlate",         "ORCH-08",   "GV_INCIDENT_TIER","Foundry OCR,Fabric",       "Data,Connect"),
    ("III", "Ch 10: Returns", "Stakeholder Router",            "W3", None,     "-", "-", "P", "-", "mcp.store",  "qa_dispatch,legal_brief,comms_draft","ORCH-08",  "--",              "Logic Apps,Teams",         "Connect,App"),
    ("III", "Ch 11: Omnichannel", "Age Verification",          "W2", None,     "-", "-", "P", "-", "mcp.store",  "age_check,audit_log",               "(platform)","--",              "Foundry,Fabric",           "App,Sec"),
    ("III", "Ch 11: Omnichannel", "Controlled Substance Ledger","W2",None,     "-", "S", "P", "-", "mcp.store",  "controlled_ledger_write",           "(platform)","GV_CONTROLLED_AUDIT","Fabric,Purview",        "Data,Sec"),
    ("III", "Ch 11: Omnichannel", "Loyalty State Projection",  "W3", "$5-10M", "-", "-", "P", "S", "mcp.store",  "loyalty_state,tier_project",        "ORCH-12",   "GV_LOYALTY_LTV",  "Fabric,Foundry",           "Data"),

    # ------------- PART IV: COLD CHAIN & COMPLIANCE -------------
    ("IV", "Ch 12: Cold Chain", "Cold Chain Telemetry Monitor","W2", "$8-14M", "-", "P", "-", "-", "mcp.cold",   "temp_stream,threshold_watch",       "ORCH-03",   "GV_COLD_CHAIN",   "IoT Hub,Event Hubs",       "Data,Connect"),
    ("IV", "Ch 12: Cold Chain", "Excursion Classifier",        "W2", None,     "-", "P", "-", "-", "mcp.cold",   "excursion_score,severity_tier",     "ORCH-03",   "GV_COLD_CHAIN",   "Foundry,Fabric RTI",       "Data"),
    ("IV", "Ch 12: Cold Chain", "Disposition Classifier",      "W2", None,     "-", "P", "-", "-", "mcp.cold",   "risk_score,save_destroy",           "ORCH-03",   "GV_COLD_CHAIN",   "Foundry",                  "Data"),
    ("IV", "Ch 12: Cold Chain", "Write-Off Pre-Approver",      "W3", None,     "-", "P", "-", "-", "mcp.cold",   "ledger_stage,route_prime",          "ORCH-03",   "--",              "Power Apps,Logic Apps",    "App"),
    ("IV", "Ch 13: Recall & Trace", "FDA Feed Listener",       "W2", "$2-4M",  "-", "P", "-", "-", "mcp.store",  "recall_parse,lot_resolve",          "ORCH-05",   "GV_RECALL_QUEUE", "Logic Apps,Event Hubs",    "Data,Connect"),
    ("IV", "Ch 13: Recall & Trace", "Lot Trace Resolver",      "W2", None,     "S", "P", "S", "-", "mcp.store",  "upstream_trace,downstream_match",   "ORCH-05",   "GV_RECALL_QUEUE", "Fabric,AI Agent Service",  "Data"),
    ("IV", "Ch 13: Recall & Trace", "Customer Notification Queue","W2",None,   "-", "S", "P", "-", "mcp.store",  "consent_filter,notify_dispatch",    "ORCH-05",   "--",              "Logic Apps,SMS Gateway",   "Connect,App"),
    ("IV", "Ch 13: Recall & Trace", "In-Aisle Pull Tasking",   "W2", None,     "P", "S", "-", "-", "mcp.store",  "aisle_task,pull_verify",            "ORCH-05",   "--",              "Power Apps,Teams",         "App,Prod"),
    ("IV", "Ch 14: Regulatory", "PII Tokenization",            "W1", None,     "S", "S", "S", "S", "--",         "--",                               "--",        "--",              "Purview,Foundry",          "Sec"),
    ("IV", "Ch 14: Regulatory", "Audit Trail Enforcement",     "W1", None,     "P", "P", "P", "P", "--",         "--",                               "--",        "--",              "Purview,Fabric",           "Sec,Data"),
    ("IV", "Ch 14: Regulatory", "Independence Posture Check",  "W1", None,     "-", "-", "-", "-", "--",         "--",                               "--",        "--",              "Purview,Entra",            "Sec"),

    # ------------- PART V: GROWTH & OPTIMIZATION -------------
    ("V",  "Ch 15: Marketing", "Audience Builder",              "W3", "$4-8M",  "-", "-", "S", "P", "mcp.growth", "segment_resolve,suppress_filter",  "ORCH-11",   "GV_AUDIENCE",     "Customer Insights,Fabric", "Data"),
    ("V",  "Ch 15: Marketing", "Creative Variant Generator",    "W3", None,     "-", "-", "-", "P", "mcp.growth", "copy_gen,brand_guardrail",         "ORCH-11",   "--",              "Foundry,Copilot Studio",   "App"),
    ("V",  "Ch 15: Marketing", "Campaign Orchestrator",         "W3", None,     "S", "-", "S", "P", "mcp.growth", "channel_sequence,pace_control",    "ORCH-11",   "--",              "Customer Insights,Logic Apps","App,Connect"),
    ("V",  "Ch 15: Marketing", "Attribution Agent",             "W4", None,     "S", "-", "P", "P", "mcp.growth", "mta_run,incrementality_test",      "(platform)","GV_ATTRIBUTION",  "Fabric ML,Power BI",       "Data"),
    ("V",  "Ch 16: Lifecycle", "LTV Forecaster",                "W4", "$6-12M", "-", "-", "S", "P", "mcp.growth", "cohort_model,intervention_sim",    "ORCH-12",   "GV_LOYALTY_LTV",  "Fabric ML,Foundry",        "Data"),
    ("V",  "Ch 16: Lifecycle", "Journey State Agent",           "W4", None,     "-", "-", "P", "P", "mcp.growth", "state_classify,nba_rank",          "ORCH-12",   "GV_JOURNEY_STATE","Foundry,Customer Insights","Data,App"),
    ("V",  "Ch 16: Lifecycle", "Lifecycle Signal Agent",        "W4", None,     "-", "-", "P", "S", "mcp.growth", "churn_predict,winback_score",      "ORCH-12",   "GV_LOYALTY_LTV",  "Fabric ML",                "Data"),
    ("V",  "Ch 16: Lifecycle", "Cross-Domain Feedback Loop*",   "W4", None,     "P", "P", "P", "P", "mcp.growth", "full_loop_evals",                  "ORCH-10",   "--",              "Foundry,Fabric,Power BI",  "Data,App"),
]

# ============================================================================
# SHEET 2: DATA SOURCE REGISTRY
# ============================================================================

DATASOURCE_COLS = [
    "Source System", "ISV / Vendor", "Schema(s)", "Entities Served",
    "Ingestion Method", "Direct/Bronze", "DIRECT?", "Format/Frequency"
]

DATASOURCES = [
    ("POS",                    "NCR Voyix / Toshiba / Oracle Retail XStore", "MERML",           "POS_VOID, PRICE_RECORD, PROMOTION_ACTIVATION, POS transactions", "Event stream -> IoT Hub -> Bronze", "via Bronze", "--", "Delta, sub-second streaming"),
    ("ESL Gateway",            "SES-imagotag / Pricer / Hanshow",            "MERML",           "PRICE_TAG_STATUS, gateway health, push verify", "Gateway API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch 5 min"),
    ("Shelf CV",               "Simbe Robotics / Pensa / Native Vision",     "MERML",           "SHELF_FACING, OSA_EVENT, PLANOGRAM_COMPLIANCE_EVENT", "IoT Edge CV -> IoT Hub -> EventStream", "via Bronze", "--", "Delta, event-driven per aisle sweep"),
    ("Cold Chain IoT",         "Monnit / Emerson / Controlant",              "SCML",            "COLD_CHAIN_TELEMETRY, TEMPERATURE_EXCURSION", "IoT Edge -> IoT Hub -> EventStream", "via Bronze", "--", "Delta, sub-second streaming"),
    ("Receiving Dock / RFID",  "Impinj / Zebra / Avery Dennison",            "SCML",            "ASN, STORE_RECEIVING_EVENT, RECEIVING_DISCREPANCY", "RFID portal -> IoT Hub -> Bronze", "via Bronze", "--", "Delta, real-time"),
    ("Vendor Portal / EDI",    "SPS Commerce / TrueCommerce / Inovis",       "SCML",            "ASN (inbound), DSD_INVOICE", "EDI 856/810 -> Event Hubs -> Bronze", "via Bronze", "--", "Delta, batch by shipment"),
    ("ERP",                    "SAP S/4HANA / Oracle Retail / Infor M3",     "MERML, SCML, MKTL","Product master, pricing master, vendor master, DIM_* anchors", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch hourly"),
    ("Order Management (OMS)", "Manhattan Active / IBM Sterling / Kibo",     "CXML",            "FULFILLMENT_ORDER, PICK_EXCEPTION, SUBSTITUTION_EVENT", "API -> Event Hubs -> Bronze", "via Bronze", "--", "Delta, event-driven"),
    ("BOPIS / Customer App",   "(Client-native mobile)",                     "CXML",            "FULFILLMENT_ORDER, customer substitution confirms", "Webhook -> Event Hubs -> Bronze", "via Bronze", "--", "Delta, event-driven"),
    ("Loyalty Platform",       "Salesforce Marketing Cloud / Dynamics Customer Insights", "MKTL, CXML", "LOYALTY_STATE, SEGMENT_MEMBERSHIP, CUSTOMER", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch hourly"),
    ("Marketing Activation",   "Braze / Adobe Campaign / D365 Marketing",    "MKTL",            "CAMPAIGN, CREATIVE_ASSET, ATTRIBUTION_EVENT, JOURNEY_STATE", "API + Webhook -> Fabric Pipeline", "via Bronze", "--", "Delta, event-driven + batch"),
    ("CCTV / Video Analytics", "Verkada / Hanwha / Axis + Azure Video Indexer","MERML (shrink index)", "CCTV timestamp index (references only, video stays on tenant)", "Azure Video Indexer -> Bronze", "via Bronze", "--", "Delta, event-driven"),
    ("FDA / Regulatory Feeds", "FDA public feeds / state AG portals / FSIS", "SCML",            "RECALL_NOTICE", "Scheduled fetch -> Logic Apps -> Bronze", "via Bronze", "--", "Delta, batch 15 min"),
    ("HRIS",                   "Workday / D365 HR / UKG Pro",                "MERML (associate) ","Employee master for associate router; clock-in/out signal", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch daily"),
    ("Customer Service Desk",  "ServiceNow / Zendesk / D365 Customer Service","CXML",           "CUSTOMER_INCIDENT, service desk tickets", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, event-driven"),
    ("Perpetual Inventory",    "Manhattan Active Inventory / SAP WMS",       "MERML",           "STORE_INVENTORY_POSITION, cycle count snapshots", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch 15 min"),
    ("Customer Photo Upload",  "(Client-native mobile)",                     "CXML",            "CUSTOMER_INCIDENT images for OCR lot extraction", "Blob upload -> Foundry OCR -> Bronze", "via Bronze", "--", "Delta, event-driven"),
    ("ISV: Recall Platform",   "FoodLogiQ / RQA / Inmar",                    "SCML",            "Recall enrichment, lot-to-supplier resolution", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, event-driven"),
    ("ISV: Pricing",           "Revionics / DemandTec / DunnhumbyPrice",     "MERML",           "Elasticity curves, competitive signal feeds for MER-A01", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch daily"),
    ("ISV: Assortment",        "Blue Yonder Assortment / SymphonyRetailAI",  "MERML",           "Assortment recommendations, planogram libraries", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch weekly"),
    ("ISV: Labor Scheduling",  "Reflexis (Zebra) / Legion",                  "MERML",           "Associate schedule, zone assignment for MER-A08", "API -> Fabric Pipeline", "via Bronze", "--", "Delta, batch shift-start"),
    ("SMS / Messaging Gateway","Twilio / Azure Communication Services",      "CXML",            "Customer confirm SMS (substitution, recall notification)", "API -> Logic Apps", "direct", "YES", "Direct, event-driven"),
]

# ============================================================================
# SHEET 3: SCHEMA ENTITIES
# ============================================================================

ENTITY_COLS = ["Schema", "Entity", "Description", "Primary Source", "MCP Tool(s)", "SCD2?"]

ENTITIES = [
    # MERML - Merchandising
    ("MERML", "STORE_INVENTORY_POSITION", "Store × product × timestamp inventory snapshot",    "Perpetual Inventory, POS",        "get_inventory_position, query_store_stock", "No"),
    ("MERML", "SHELF_FACING",             "Store × aisle × bay × facing slot assignment",     "Shelf CV, planogram master",      "get_shelf_facing", "YES (SCD2)"),
    ("MERML", "OSA_EVENT",                "On-shelf availability observation",                 "Shelf CV + POS velocity fusion",  "get_osa_events, get_osa_cluster", "No"),
    ("MERML", "PRICE_RECORD",             "Store × product × effective interval price",        "ERP, pricing ISV",                "get_price_current, get_price_history", "YES (SCD2)"),
    ("MERML", "PRICE_TAG_STATUS",         "ESL tag sync state + gateway health",               "ESL Gateway",                     "get_tag_status, list_stale_tags", "No"),
    ("MERML", "PROMOTION_ACTIVATION",     "Live promo schedule + activation state",            "ERP, pricing ISV",                "get_promotion, check_promotion_live", "No"),
    ("MERML", "SHRINK_EVENT",             "Shrink incident record",                            "POS voids, cycle count variance", "get_shrink_event, query_shrink_pattern", "No"),
    ("MERML", "POS_VOID",                 "POS transaction void",                              "POS stream",                      "get_void_events, sigma_analyze_voids", "No"),
    ("MERML", "CYCLE_COUNT_VARIANCE",     "Cycle count delta vs perpetual",                    "Perpetual Inventory",             "get_cycle_variance", "No"),
    ("MERML", "PLANOGRAM_ASSIGNMENT",     "Planogram assignment per store × aisle",            "Assortment ISV",                  "get_planogram", "YES (SCD2)"),
    ("MERML", "PLANOGRAM_COMPLIANCE_EVENT","Shelf audit event vs planned planogram",          "Shelf CV",                        "get_compliance_event", "No"),
    ("MERML", "MARKDOWN_EVENT",           "Product × store × markdown cycle",                  "Agent (MER-A12)",                 "get_markdown_event", "No"),
    ("MERML", "WASTE_EVENT",              "Expired/destroyed inventory record",                "Agent (MER-A12)",                 "get_waste_event", "No"),
    ("MERML", "LOT_EXPIRATION_STATE",     "Active expiration-tracked lots per store",          "ERP + receiving",                 "get_lot_expiration", "YES (SCD2)"),

    # SCML - Supply Chain
    ("SCML", "ASN",                    "Advance shipment notice per PO",                      "Vendor Portal / EDI",              "get_asn, match_asn_to_receipt", "No"),
    ("SCML", "STORE_RECEIVING_EVENT",  "Receiving dock event",                                "RFID portal, handheld scanner",    "get_receiving_event, compare_asn_receive", "No"),
    ("SCML", "RECEIVING_DISCREPANCY",  "Line-level variance",                                  "Agent (SCM-A01)",                  "get_discrepancy, draft_dispute", "No"),
    ("SCML", "DSD_INVOICE",            "Direct-store-delivery invoice",                        "Vendor Portal / EDI",              "get_dsd_invoice, reconcile_invoice", "No"),
    ("SCML", "COLD_CHAIN_TELEMETRY",   "Sensor reading per lot per minute",                    "Cold Chain IoT",                   "get_cold_chain_stream, query_sensor_window", "No"),
    ("SCML", "TEMPERATURE_EXCURSION",  "Classified excursion with severity + duration",        "Agent (SCM-A04)",                  "get_excursion, list_active_excursions", "No"),
    ("SCML", "RECALL_NOTICE",          "Regulatory recall notice (ingested + enriched)",       "FDA feed, ISV recall platform",    "get_recall_notice, list_active_recalls", "No"),
    ("SCML", "RECALL_DISPOSITION",     "Per-store recall disposition decision",                "Agent (SCM-A08 + recall coordinator)","get_disposition", "No"),
    ("SCML", "LOT_TRACE",              "Lot movement event (upstream + downstream)",           "Receiving + lot genealogy",        "get_lot_trace, trace_lot_upstream, trace_lot_downstream", "No"),
    ("SCML", "SHIPMENT_EVENT",         "Outbound/transfer shipment event",                     "TMS / shipping system",            "get_shipment_event", "No"),
    ("SCML", "CONTAINER_EVENT",        "Container-level movement event",                       "RFID, shipping",                   "get_container_event", "No"),

    # CXML - Customer Experience
    ("CXML", "FULFILLMENT_ORDER",      "Customer order fulfillment lifecycle",                "OMS, Customer App",                "get_fulfillment_order, update_order_state", "No"),
    ("CXML", "PICK_EXCEPTION",         "Pick exception incident",                              "Agent (CXM-A01)",                  "get_pick_exception, classify_exception", "No"),
    ("CXML", "SUBSTITUTION_EVENT",     "Substitution proposal + customer response",            "Agent (CXM-A02/A03)",              "get_substitution, rank_substitutes", "No"),
    ("CXML", "RETURN_TRANSACTION",     "Return transaction at POS or online",                  "POS, OMS",                         "get_return, get_returns_by_customer", "No"),
    ("CXML", "RETURN_DISPOSITION",     "Resell / regrade / RTV / destroy decision",            "Agent (CXM-A07)",                  "get_disposition, draft_rtv_claim", "No"),
    ("CXML", "RETURN_FRAUD_SIGNAL",    "Fraud-pattern flag per return",                        "Agent (CXM-A07 + fraud model)",    "get_fraud_signal", "No"),
    ("CXML", "CUSTOMER_INCIDENT",      "Customer-reported incident (safety, quality, service)","Service Desk, Customer App",       "get_incident, correlate_incidents", "No"),
    ("CXML", "AGE_VERIFICATION_EVENT", "Age check event for regulated product",                "POS / customer app",               "log_age_check", "No"),
    ("CXML", "CONTROLLED_SUBSTANCE_LEDGER", "Chain-of-custody ledger",                         "POS / pharmacy",                   "get_controlled_ledger, audit_dispensing", "YES (SCD2)"),
    ("CXML", "LOYALTY_STATE",          "Customer × loyalty state snapshot",                    "Loyalty Platform",                 "get_loyalty_state, project_loyalty_tier", "YES (SCD2)"),

    # MKTL - Marketing
    ("MKTL", "CAMPAIGN",               "Marketing campaign definition",                        "Marketing Activation",             "get_campaign, list_active_campaigns", "YES (SCD2)"),
    ("MKTL", "AUDIENCE",               "Audience segment definition",                          "Loyalty + Marketing Activation",   "get_audience, materialize_audience", "YES (SCD2)"),
    ("MKTL", "CREATIVE_ASSET",         "Creative content + variants",                          "Marketing Activation, Agent (MKT-A02)","get_creative, list_variants", "No"),
    ("MKTL", "JOURNEY_STATE",          "Customer × journey state snapshot",                    "Agent (MKT-A06)",                  "get_journey_state, next_best_action", "YES (SCD2)"),
    ("MKTL", "ATTRIBUTION_EVENT",      "Attribution touchpoint event",                         "Agent (MKT-A04)",                  "get_attribution, run_mta", "No"),
    ("MKTL", "LTV_FORECAST",           "Per-customer LTV forecast",                            "Agent (MKT-A05)",                  "get_ltv_forecast, simulate_intervention", "No"),
    ("MKTL", "SEGMENT_MEMBERSHIP",     "Customer × segment × effective interval",              "Agent (MKT-A01)",                  "get_segment_membership, filter_suppression", "YES (SCD2)"),
]

# ============================================================================
# STYLING HELPERS
# ============================================================================

HEADER_FILL  = PatternFill("solid", start_color="1A2339", end_color="1A2339")
HEADER_FONT  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
PART_FILL    = PatternFill("solid", start_color="0D7D70", end_color="0D7D70")
PART_FONT    = Font(name="Arial", bold=True, size=11, color="FFFFFF")
CHAPTER_FILL = PatternFill("solid", start_color="F4EFE3", end_color="F4EFE3")
CHAPTER_FONT = Font(name="Arial", bold=True, size=10, color="2F3F5F")
BODY_FONT    = Font(name="Arial", size=10)
MONO_FONT    = Font(name="Consolas", size=9)
PRIMARY_FILL = PatternFill("solid", start_color="E6F7F4", end_color="E6F7F4")
SECONDARY_FILL = PatternFill("solid", start_color="FAF5E9", end_color="FAF5E9")
WAVE_FILLS = {
    "W1": PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7"),
    "W2": PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7"),
    "W3": PatternFill("solid", start_color="DBEAFE", end_color="DBEAFE"),
    "W4": PatternFill("solid", start_color="EDE9FE", end_color="EDE9FE"),
}
BORDER_SIDE = Side(style="thin", color="D4D4D4")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

def style_header_row(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    ws.row_dimensions[row_num].height = 28

def style_body_cell(cell, mono=False, center=False, fill=None):
    cell.font = MONO_FONT if mono else BODY_FONT
    cell.alignment = Alignment(horizontal=("center" if center else "left"), vertical="top", wrap_text=True)
    cell.border = CELL_BORDER
    if fill is not None:
        cell.fill = fill

# ============================================================================
# BUILD WORKBOOK
# ============================================================================

wb = Workbook()

# ---------- Sheet 1: Solution Stack ----------
ws = wb.active
ws.title = "Solution Stack"
ws.freeze_panes = "A2"

# Header
for i, col in enumerate(SOLUTION_COLS, start=1):
    ws.cell(row=1, column=i, value=col)
style_header_row(ws, 1, len(SOLUTION_COLS))

# Body
for r, row_data in enumerate(SOLUTIONS, start=2):
    wave = row_data[3]
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        col_name = SOLUTION_COLS[c-1]
        mono = col_name in ("MCP Server", "MCP Tools", "Orch", "Gold View")
        # Schema columns (MERML, SCML, CXML, MKTL) — color P vs S
        if col_name in ("MERML", "SCML", "CXML", "MKTL"):
            if val == "P":
                style_body_cell(cell, center=True, fill=PRIMARY_FILL)
                cell.font = Font(name="Arial", bold=True, size=10, color="0D7D70")
            elif val == "S":
                style_body_cell(cell, center=True, fill=SECONDARY_FILL)
                cell.font = Font(name="Arial", size=10, color="8F5A0C")
            else:
                style_body_cell(cell, center=True)
                cell.font = Font(name="Arial", size=10, color="B8B8B8")
        elif col_name == "Wave" and wave in WAVE_FILLS:
            style_body_cell(cell, center=True, fill=WAVE_FILLS[wave])
            cell.font = Font(name="Arial", bold=True, size=10)
        elif col_name == "Part":
            style_body_cell(cell, center=True, fill=PART_FILL)
            cell.font = PART_FONT
        elif col_name == "Chapter":
            style_body_cell(cell, fill=CHAPTER_FILL)
            cell.font = CHAPTER_FONT
        else:
            style_body_cell(cell, mono=mono)

# Column widths tuned for readability
widths = {
    "A": 6, "B": 22, "C": 30, "D": 7, "E": 10,
    "F": 7, "G": 7, "H": 7, "I": 7,
    "J": 13, "K": 36, "L": 14, "M": 22, "N": 30, "O": 16
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---------- Sheet 2: Data Source Registry ----------
ws2 = wb.create_sheet("Data Source Registry")
ws2.freeze_panes = "A2"

for i, col in enumerate(DATASOURCE_COLS, start=1):
    ws2.cell(row=1, column=i, value=col)
style_header_row(ws2, 1, len(DATASOURCE_COLS))

for r, row_data in enumerate(DATASOURCES, start=2):
    for c, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        col_name = DATASOURCE_COLS[c-1]
        mono = col_name in ("Schema(s)", "Entities Served")
        center = col_name in ("Direct/Bronze", "DIRECT?")
        style_body_cell(cell, mono=mono, center=center)

ds_widths = {"A": 22, "B": 36, "C": 20, "D": 40, "E": 34, "F": 14, "G": 10, "H": 28}
for col, w in ds_widths.items():
    ws2.column_dimensions[col].width = w

# ---------- Sheet 3: Schema Entities ----------
ws3 = wb.create_sheet("Schema Entities")
ws3.freeze_panes = "A2"

for i, col in enumerate(ENTITY_COLS, start=1):
    ws3.cell(row=1, column=i, value=col)
style_header_row(ws3, 1, len(ENTITY_COLS))

# Color entities by schema for visual grouping
schema_fills = {
    "MERML": PatternFill("solid", start_color="E6F7F4", end_color="E6F7F4"),
    "SCML":  PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7"),
    "CXML":  PatternFill("solid", start_color="DBEAFE", end_color="DBEAFE"),
    "MKTL":  PatternFill("solid", start_color="EDE9FE", end_color="EDE9FE"),
}

for r, row_data in enumerate(ENTITIES, start=2):
    schema = row_data[0]
    fill = schema_fills.get(schema)
    for c, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=r, column=c, value=val)
        col_name = ENTITY_COLS[c-1]
        mono = col_name in ("Entity", "MCP Tool(s)")
        center = col_name in ("Schema", "SCD2?")
        style_body_cell(cell, mono=mono, center=center,
                        fill=(fill if col_name == "Schema" else None))
        if col_name == "Schema":
            cell.font = Font(name="Arial", bold=True, size=10, color="2F3F5F")

en_widths = {"A": 10, "B": 32, "C": 52, "D": 36, "E": 46, "F": 12}
for col, w in en_widths.items():
    ws3.column_dimensions[col].width = w

# ---------- Save ----------
wb.save(OUT)
print(f"wrote {OUT}")
print(f"  Solution Stack: {len(SOLUTIONS)} rows")
print(f"  Data Source Registry: {len(DATASOURCES)} rows")
print(f"  Schema Entities: {len(ENTITIES)} rows")

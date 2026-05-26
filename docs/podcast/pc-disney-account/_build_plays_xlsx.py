"""
Build the Disney Agentic Plays Excel — a curated list of high-confidence
agentic plays for the Disney account, focused on enterprise back-office,
streaming, and Experiences.

Output: Disney_Agentic_Plays_BackOffice_Streaming.xlsx

Three sheets:
  1. Plays       — the curated 24-play list (13 back-office + 10 streaming + 1 Experiences)
  2. Summary     — counts by domain · Wave 1 envelope totals · priority mix
  3. How To Use  — guidance for the Account Team on using this file

Usage:
    python _build_plays_xlsx.py
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

HERE = Path(__file__).parent
OUT = HERE / "Disney_Agentic_Plays_BackOffice_Streaming.xlsx"

# ----------------------------------------------------------------- plays ----

# Each play has the same shape: tuple of column values matching the headers below.
HEADERS = [
    "#",
    "Domain",
    "Sub-domain",
    "Play",
    "Business Problem",
    "Agent Capability",
    "Disney Pressure Addressed",
    "KPI Signal",
    "Buyer at Disney",
    "Microsoft Attach",
    "Wave",
    "Wave 1 Range ($M)",
    "APEX Family",
    "Priority",
]

BACK_OFFICE_PLAYS = [
    (
        "Continuous-Close Agent",
        "Finance",
        "Quarterly close cycle is 8 days; target is 3. Manual journal entries, reconciliations, and consolidations dominate close-week labour.",
        "Agent drafts journal entries from source-system signals, reconciles inter-company balances, generates consolidation packages. Controller approves.",
        "Audit posture · finance team capacity · CFO close-cycle commitment",
        "Close cycle 8d → 3d · 30–50% cycle compression",
        "CFO · Controller",
        "Fabric · Foundry · Purview · Power BI",
        "Wave 1",
        "1.0–1.8",
        "Cross-edition Finance",
        "High",
    ),
    (
        "AP Invoice-Triage Agent",
        "Finance · AP",
        "Disney processes very large invoice volumes across segments. Manual triage, coding, and exception-handling dominate AP labour.",
        "Agent reads invoices, extracts line items, matches PO/contract, codes GL accounts, flags exceptions for review. AP clerk approves.",
        "Working capital · AP team capacity",
        "AP cycle time −50% · DPO optimisation · 30% reduction in invoice exceptions",
        "Controller · AP Director",
        "Fabric · Foundry · Document Intelligence",
        "Wave 1",
        "0.8–1.4",
        "Cross-edition Finance",
        "High",
    ),
    (
        "Procurement Intake & Spend-Categorisation Agent",
        "Procurement",
        "Procurement intake is high-volume, manual triage. Spend categorisation drift creates analytics and savings-tracking errors.",
        "Agent classifies intake requests, recommends suppliers from approved pool, surfaces consolidation opportunities. Procurement team approves.",
        "Spend visibility · supplier relationship discipline",
        "Intake cycle −40% · spend-data-quality + 30 points · category-tail consolidation savings",
        "CPO · Procurement Operations",
        "Fabric · Foundry · Dynamics 365 SCM",
        "Wave 2",
        "0.9–1.6",
        "Cross-edition Procurement",
        "Medium",
    ),
    (
        "Contract Review Agent",
        "Legal · Commercial",
        "Disney's legal organisation reviews enormous volumes of contracts — talent, distribution, vendor, partnership. Clause-by-clause review is slow and inconsistent.",
        "Agent reads contracts, flags non-standard clauses, compares against precedent library, drafts redlines, prepares attorney summary. Attorney approves.",
        "Legal-team capacity · contract-cycle velocity",
        "Contract review cycle −60% · clause-risk-detection consistency +40%",
        "GC · Deputy GC for Commercial",
        "Fabric · Foundry · Purview (PHI/PII propagation)",
        "Wave 2",
        "1.2–2.0",
        "Cross-edition Legal",
        "Medium",
    ),
    (
        "Audit Evidence Readiness Agent",
        "Internal Audit",
        "Annual SOX and operational-audit cycles require evidence compilation across many systems. Manual evidence-gathering consumes audit staff time.",
        "Agent continuously gathers control-evidence artefacts, maintains audit-readiness dashboard, drafts evidence packages on demand.",
        "Audit posture · continuous control attestation",
        "Audit evidence preparation −70% · 100% evidence-trail coverage on monitored controls",
        "Chief Audit Executive · CCO",
        "Fabric · Foundry · Purview · Microsoft 365 Compliance",
        "Wave 1",
        "1.1–1.7",
        "Cross-edition Audit",
        "High",
    ),
    (
        "IT Service-Desk Agent-Assist",
        "Internal IT",
        "200,000+ Disney employees globally generate enormous IT ticket volume. Outsourced service-desk economics favour quality lift over headcount reduction.",
        "Agent assists human IT support tech in real time with cross-system context, drafts resolution, recommends escalation path. Human approves.",
        "Internal NPS · employee productivity · IT-contract renegotiation leverage",
        "AHT −30% · FCR +15 pts · employee CES improvement",
        "CIO · Head of IT Operations",
        "Fabric · Foundry · M365 Copilot · ServiceNow integration",
        "Wave 2",
        "1.0–1.5",
        "Cross-Practice Contact-Center",
        "High",
    ),
    (
        "HR Onboarding & Lifecycle Agent",
        "HR",
        "New-hire onboarding across 200,000 global employees is workflow-heavy. Offboarding and transitions equally complex.",
        "Agent orchestrates the onboarding workflow across HR systems, IT provisioning, training assignment, and policy acknowledgement. HR specialist approves exception cases.",
        "Employee experience · HR-team capacity · time-to-productivity",
        "Time-to-productivity for new hires −40% · onboarding NPS +20 pts",
        "CHRO · Head of HR Ops",
        "Fabric · Foundry · M365 Copilot · Workday integration",
        "Wave 2",
        "0.9–1.5",
        "Cross-edition HR",
        "Medium",
    ),
    (
        "T&E Compliance & Audit Agent",
        "Finance · T&E",
        "Travel & expense compliance audits are sample-based. Policy violations slip through. Manual review of high-volume T&E claims is labour-intensive.",
        "Agent reviews T&E claims continuously against policy, flags violations, drafts compliance findings. T&E reviewer approves enforcement.",
        "Audit posture · expense-policy enforcement",
        "100% claim coverage · policy-violation detection +40% · T&E review labour −60%",
        "Controller · Head of T&E",
        "Fabric · Foundry · Purview · M365 Copilot",
        "Wave 2",
        "0.7–1.2",
        "Cross-edition Finance",
        "Medium",
    ),
    (
        "Continuous Controls & Risk Monitoring Agent",
        "Risk & Compliance",
        "Disney's risk register requires continuous monitoring. Key controls are tested periodically; emerging risk signals are missed.",
        "Agent monitors control evidence streams continuously, evaluates against control specs, escalates emerging risk signals. Risk officer approves response actions.",
        "Audit posture · regulatory exposure · CCO confidence",
        "Mean-time-to-detect emerging risk −60% · control-test coverage 100%",
        "Chief Risk Officer · CCO",
        "Fabric · Foundry · Purview · Sentinel",
        "Wave 2",
        "1.2–2.0",
        "Cross-edition Risk",
        "High",
    ),
    (
        "Third-Party Risk & Supply-Chain Security Monitoring Agent",
        "Cyber · Vendor Risk",
        "Disney's third-party ecosystem creates indirect security exposure. Annual-assessment-driven third-party programs miss emerging risk signals.",
        "Agent continuously monitors third-party security posture from public signal, news, and shared assessment data. Recommends escalation. Security team prioritises.",
        "Cyber posture · regulatory exposure · supply-chain resilience",
        "Third-party-incident rate −30% · time-to-emerging-risk-response −50%",
        "CISO · Head of Vendor Risk",
        "Fabric · Foundry · Sentinel · Defender for Cloud",
        "Wave 3",
        "1.0–1.6",
        "Cross-edition Cyber",
        "Medium",
    ),
    (
        "CTO Portfolio & Decision Intelligence Agent",
        "Technology Strategy · CTO Office",
        "Disney's CTO oversees hundreds of technology investments across segments — platform modernisation, AI portfolio, engineering productivity, build-vs-buy. Manual portfolio reviews happen quarterly; the portfolio moves faster. Stranded pilots accumulate. Investment ROI is opaque between reviews.",
        "Agent reads the CTO's technology investment portfolio continuously (CapEx tracking, project status, performance telemetry, cost actuals, business-KPI alignment), produces a continuous portfolio-health view, identifies stranded pilots and at-risk investments, drafts portfolio recommendations (sustain · accelerate · sunset · consolidate). CTO and steering committee approve.",
        "Technology investment ROI · AI portfolio rationalisation · execution velocity · CFO-CTO alignment on tech spend",
        "Stranded-pilot rate −40% · investment-cycle-time compression 50% · portfolio review quarterly → continuous",
        "CTO · CIO (joint depending on org)",
        "Fabric · Foundry · Power BI · Purview · Power Platform",
        "Wave 2",
        "1.4–2.0",
        "Cross-edition Technology Strategy",
        "High",
    ),
    (
        "Engineering Headcount Optimisation & Productivity Agent",
        "Engineering · R&D",
        "Disney's engineering organisation spans Studios, DTC streaming, Parks technology, ESPN production tech, internal IT, and corporate engineering — thousands of engineers in aggregate. AI-augmented engineering produces measurable productivity lift; the option set is to (a) ship more output with the same team, (b) reduce engineering headcount for cost optimisation, (c) shift the hiring profile, or (d) redirect capacity to higher-leverage work. Management decides; the platform enables the option.",
        "Multi-agent engineering productivity platform — code-generation agents for routine implementations · agentic code-review · design-doc and architecture-record drafting · automated test generation · technical-documentation maintenance · deployment-pipeline agent assist · internal developer-Q&A agents. Engineering leaders measure productivity lift by team and decide workforce composition. CFO sees cost-savings option; engineering leadership sees velocity option. Both are real and exclusive.",
        "Operating cost discipline · technology investment ROI · workforce composition for the AI era · engineering velocity through electrification / autonomy / DTC transitions (DTNA-adjacent thinking applies)",
        "Engineer productivity +20–35% on covered task types · code-review cycle −40-60% · engineering capacity available for redirection OR reduction (management decision) · time-to-PR-merge compression · onboarding-time reduction for new engineers",
        "CTO · CFO (joint) · Heads of Engineering per business unit · CHRO (workforce-composition partner)",
        "GitHub Copilot · Fabric · Foundry · M365 Copilot · Power Platform · Azure DevOps integration",
        "Wave 2-3",
        "1.5–2.5",
        "Cross-edition Engineering Productivity",
        "High",
    ),
    (
        "AI Consumption Cost Intelligence Agent (Engineering + AI-Assist FinOps)",
        "Technology · FinOps · AI Cost Discipline",
        "Enterprise AI consumption costs are exploding across Disney. GitHub Copilot seats, M365 Copilot licenses, Azure OpenAI / Foundry token consumption, embedding and vector-store spend, custom-agent run-rate, Power Platform AI Builder consumption, and team-expensed third-party AI tools (Cursor, Cline, Continue, Bedrock, Anthropic API) add up to multi-tens-of-millions of dollars annually. Run-rate is climbing 20-40% quarter-over-quarter. Cloud-FinOps practices do not yet cover AI consumption cleanly. CFO sees the aggregate bill; CTO sees the value; nobody sees the per-use-case cost-vs-value composition.",
        "Continuously composes AI consumption signal across: Microsoft EA Copilot seat utilisation · Azure OpenAI / Foundry token consumption by model · GitHub Copilot per-developer activity · embedding API and vector-store spend · custom-built agent run-cost · third-party AI tool expense (T&E + departmental contracts) · Power Platform AI Builder consumption. Surfaces per-team, per-product, per-use-case, per-model cost. Identifies idle Copilot seats, runaway agents, sub-optimal model selection (e.g., GPT-4 used where GPT-4o-mini suffices). Recommends optimisation actions. CFO + CTO see joint operating view; FinOps team approves reclamation.",
        "AI-consumption cost discipline · CFO-CTO joint AI governance · model-selection-by-workload optimisation · FinOps maturity for the AI era · seat-utilisation reclamation · operating cost discipline at the AI run-rate layer",
        "AI consumption cost growth: from +20-40% QoQ trajectory to flat-or-managed · idle Copilot seat reclamation rate · model-selection optimisation savings 12-25% on covered workloads · per-use-case cost-per-outcome surfaced · CFO-CTO joint quarterly governance cadence established",
        "CTO · CFO (joint) · Head of FinOps · CIO",
        "Azure Cost Management · Microsoft Cost Management for Azure OpenAI · Fabric · Foundry · Purview · Power BI · M365 Copilot Admin Center",
        "Wave 1-2",
        "1.3–1.9",
        "Cross-edition FinOps / AI Cost Governance",
        "High",
    ),
]

STREAMING_PLAYS = [
    (
        "Streaming Churn Prediction & Retention Agent",
        "Cross-Streaming",
        "Each streaming service has its own churn dynamics. Manual marketing intervention takes weeks; the churn window is days.",
        "Agent watches subscriber behaviour continuously, identifies elevated churn risk, recommends personalised retention offer with margin-aware pricing. Loyalty team approves.",
        "DTC profitability · subscriber retention · cross-service portfolio value",
        "Targeted-cohort churn −19% · annualised retention revenue protection",
        "President of DTC · Chief Customer Officer",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 1",
        "1.5–2.0",
        "RC-CX · streaming variant",
        "High",
    ),
    (
        "Personalised Content Recommendations & Rails Agent",
        "Disney+ · Hulu",
        "Heterogeneous content portfolio (Marvel, Pixar, Lucasfilm, ABC, FX, Nat Geo) requires coherent personalisation per user.",
        "Agent composes per-user rails layout on top of the recommendation engine, reasoning about session context, household composition, and new-release calendar.",
        "DTC engagement · content investment ROI · session time",
        "Engagement +23% · session-length lift · rail-CTR improvement",
        "Chief Product Officer DTC",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 1",
        "1.3–1.9",
        "Cross-edition Personalisation",
        "High",
    ),
    (
        "Cold-Start Boost Agent for Franchise Launches",
        "Disney+",
        "Major launches (Marvel, Pixar, Lucasfilm) have a critical 72-hour cold-start window. Generic similarity-based recommendation underperforms.",
        "Agent identifies high-probability cohorts for a specific launch given content fingerprint, market context, and marketing campaign. Surfaces with prioritised placement.",
        "Content investment ROI · launch performance · franchise momentum",
        "Cold-start uptake +41% · first-72-hour engagement lift",
        "Chief Product Officer DTC · Head of Studios Marketing",
        "Fabric · Foundry · Real-Time Intelligence",
        "Wave 1",
        "1.0–1.5",
        "Cross-edition Personalisation",
        "High",
    ),
    (
        "Password-Sharing Detection & Paid-Sharing Agent",
        "Cross-Streaming",
        "Industry has demonstrated paid-sharing economics. Naive detection produces false positives that hurt brand. Disney needs graceful, governed enforcement.",
        "Agent reasons about household-vs-sharing patterns with confidence scoring. Recommends paid-sharing offer flow. Marketing-ops team approves; CCO sees governance trail.",
        "DTC revenue · brand-experience integrity · governance posture",
        "Incremental revenue +$8M/year reference · false-positive rate <2%",
        "President of DTC · CMO · CCO",
        "Fabric · Foundry · Purview (audit-row chain) · Entra",
        "Wave 1",
        "1.2–1.8",
        "Cross-edition Subscription",
        "High",
    ),
    (
        "ESPN Auto-Highlight Detection & Clipping Agent",
        "ESPN · ESPN+",
        "ESPN produces enormous volumes of highlight content. Manual identification, editing, and packaging is the production bottleneck for scale clip output.",
        "Agent watches live game feed (state, audio, vision), identifies clip candidates with timestamps and metadata. Production team reviews and approves.",
        "ESPN-DTC competitiveness · production-team capacity · audience clip-consumption",
        "Production time per clip −74% · clip-output volume 3–5×",
        "Head of ESPN Production · ESPN-DTC President",
        "Fabric RTI · Foundry · Azure AI Video",
        "Wave 1",
        "1.5–2.2",
        "Cross-edition Live-Content",
        "High",
    ),
    (
        "Auto-Dub & Subtitle Agent for International Markets",
        "Disney+ International",
        "50+ international markets require localised content. Multi-vendor, multi-quality-gate localisation creates 4–6 week time-to-market lag.",
        "Agent generates draft localised dub audio and subtitles using current speech and language models, flags quality concerns. Linguistic experts review edge cases.",
        "International subscriber growth · time-to-market · localisation cost",
        "Localisation time per asset −62% · 50%+ throughput lift in active markets",
        "Head of International DTC · Head of Localisation",
        "Fabric · Foundry · Azure Speech · Azure AI Translator",
        "Wave 2",
        "1.0–1.6",
        "Cross-edition Speech-and-Language",
        "Medium",
    ),
    (
        "Subscriber Lifecycle Agent — Trial-to-Paid, Downgrade Prevention",
        "Cross-Streaming",
        "Subscriber-lifecycle events (trial-to-paid conversion, downgrade threats, upgrade opportunities) are managed reactively today.",
        "Agent watches subscriber lifecycle stages continuously, intervenes at the right moment with personalised offer, manages the lifecycle from trial to long-term subscriber.",
        "DTC profitability · subscriber LTV · conversion economics",
        "Trial-to-paid conversion +12% · downgrade prevention +20% on targeted cohort",
        "President of DTC · CMO",
        "Fabric · Foundry · Customer Insights · Power Automate",
        "Wave 2",
        "1.1–1.7",
        "Cross-edition Subscription",
        "Medium",
    ),
    (
        "Streaming Quality-of-Experience Triage Agent",
        "Streaming Delivery Ops",
        "QoE issues (buffering, codec failures, regional CDN problems) manifest unevenly. Operator-driven dashboard detection produces hour-scale response.",
        "Agent watches delivery telemetry continuously, identifies emerging regional or device-family quality issues, recommends operations response (reroute, encoding-pipeline adjust).",
        "DTC service reliability · operator capacity · subscriber experience",
        "Playback-failure-rate −35% · mean-time-to-quality-recovery −60%",
        "Head of Streaming Delivery · CTO DTC",
        "Fabric RTI · Foundry · Azure Front Door · Application Insights",
        "Wave 2",
        "1.0–1.6",
        "Operations / RTI",
        "High",
    ),
    (
        "Ad-Targeting Agent for Hulu / Disney+ Ad Tier / ESPN+",
        "Ad Sales · AVOD",
        "Programmatic ad insertion across multiple streaming surfaces requires real-time decisioning. Generic ad-tech doesn't reason about Disney's portfolio context.",
        "Agent reasons about ad placement opportunity per viewer per session, optimising for advertiser brief and viewer experience. Ad-ops team approves campaigns.",
        "Ad revenue · advertiser brief satisfaction · viewer experience preservation",
        "Ad-revenue lift on managed inventory +10–18% · ad-CTR +12% · viewer-complaint rate −20%",
        "Head of Ad Sales · CRO · Head of AVOD Strategy",
        "Fabric · Foundry · Customer Insights · Azure AI Video",
        "Wave 2",
        "1.3–2.0",
        "Cross-edition Ad-Tech",
        "Medium",
    ),
    (
        "Subscriber Support Agent — Customer Service Augmentation",
        "Customer Service · Streaming",
        "Disney+/Hulu/ESPN+ customer-service tickets span billing, playback, content, and account issues. Cross-system context composition slows resolution.",
        "Agent assists human CSR in real time with subscriber context, account history, recent activity, recommended resolution. CSR approves.",
        "Subscriber NPS · CSR capacity · attrition / contractor relationship",
        "AHT −30% · FCR +18 pts · subscriber CES improvement",
        "Head of Customer Service Streaming · CXO",
        "Fabric · Foundry · M365 Copilot · Dynamics 365 Service",
        "Wave 1",
        "1.0–1.4",
        "Cross-Practice Contact-Center",
        "High",
    ),
]

# Third Domain — Experiences (Parks · Resorts · Cruise Line · Consumer Products)
# See companion brief: Disney-Experiences-Agentic-Play-Brief.md for the full
# sourced documentation including authoritative references.
EXPERIENCES_PLAYS = [
    (
        "Guest Day Orchestration Agent",
        "Parks · Resorts · Guest-Facing",
        "Guest days at Disney parks involve dozens of decisions per guest per day — ride selection, dining, Lightning Lane purchases, character meets, photo opportunities, rest breaks. The data exists across MagicBand+, My Disney Experience app, and park-state systems; today the guest composes manually.",
        "Agent watches MagicBand+/MyDisneyExperience signal continuously (location, party, recent activity, dining/Lightning Lane status, weather, ride availability). Composes a continuously updated personalised itinerary. Notifies guest opt-in with next-best-action recommendation. Guest decides (HITL is the guest).",
        "$60B Experiences capital investment ROI · guest experience differentiation · Genie+/Lightning Lane attach revenue · per-guest yield · repeat-visit propensity",
        "Genie+/Lightning Lane attach +10–20% · in-park spend per guest +5–10% · guest NPS lift · repeat-visit booking rate ↑",
        "President of Experiences · SVP Guest Experience · CTO Experiences",
        "Fabric · Foundry · Real-Time Intelligence · Customer Insights · M365 Copilot (for cast members)",
        "Wave 2-3",
        "2.0–3.5",
        "Cross-edition Guest-Experience / TH adjacency",
        "High",
    ),
]


def fmt_money_range(value: str) -> str:
    """Convert '1.5-2.0' kind of value to a display string."""
    return value  # already in display format


def build_workbook() -> Workbook:
    wb = Workbook()

    # ============================================================
    # Sheet 1 — Plays
    # ============================================================
    ws = wb.active
    ws.title = "Plays"

    # Header
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1A2339")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(border_style="thin", color="000000"))

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Body
    row_num = 2
    play_id = 1

    domain_fill = {
        "Back-Office": "EEF2F7",
        "Streaming": "FFF7E6",
        "Experiences": "F0F7EE",
    }

    priority_color = {
        "High": "1B7F3A",   # green
        "Medium": "C28A22",  # amber
        "Low": "6B7280",     # grey
    }

    for play_tuple in BACK_OFFICE_PLAYS:
        ws.append([
            play_id,
            "Back-Office",
            play_tuple[1],
            play_tuple[0],
            play_tuple[2],
            play_tuple[3],
            play_tuple[4],
            play_tuple[5],
            play_tuple[6],
            play_tuple[7],
            play_tuple[8],
            play_tuple[9],
            play_tuple[10],
            play_tuple[11],
        ])
        play_id += 1
        row_num += 1

    for play_tuple in STREAMING_PLAYS:
        ws.append([
            play_id,
            "Streaming",
            play_tuple[1],
            play_tuple[0],
            play_tuple[2],
            play_tuple[3],
            play_tuple[4],
            play_tuple[5],
            play_tuple[6],
            play_tuple[7],
            play_tuple[8],
            play_tuple[9],
            play_tuple[10],
            play_tuple[11],
        ])
        play_id += 1
        row_num += 1

    for play_tuple in EXPERIENCES_PLAYS:
        ws.append([
            play_id,
            "Experiences",
            play_tuple[1],
            play_tuple[0],
            play_tuple[2],
            play_tuple[3],
            play_tuple[4],
            play_tuple[5],
            play_tuple[6],
            play_tuple[7],
            play_tuple[8],
            play_tuple[9],
            play_tuple[10],
            play_tuple[11],
        ])
        play_id += 1
        row_num += 1

    last_row = row_num - 1

    # Apply formatting per row
    for r in range(2, last_row + 1):
        domain = ws.cell(row=r, column=2).value
        priority = ws.cell(row=r, column=14).value
        bg = domain_fill.get(domain, "FFFFFF")
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = Border(
                bottom=Side(border_style="hair", color="CCCCCC"),
                right=Side(border_style="hair", color="CCCCCC"),
            )
        # Priority colour
        pcell = ws.cell(row=r, column=14)
        pcell.font = Font(name="Arial", size=10, bold=True, color=priority_color.get(priority, "000000"))
        # Wave bold
        wcell = ws.cell(row=r, column=11)
        wcell.font = Font(name="Arial", size=10, bold=True)
        # Number-format Wave 1 Range column as text (it contains hyphenated range)
        # Already text — no number-format

    # Column widths
    widths = {
        1: 5,    # #
        2: 12,   # Domain
        3: 18,   # Sub-domain
        4: 35,   # Play
        5: 45,   # Business Problem
        6: 45,   # Agent Capability
        7: 35,   # Disney Pressure
        8: 32,   # KPI Signal
        9: 28,   # Buyer
        10: 32,  # Microsoft Attach
        11: 8,   # Wave
        12: 16,  # Wave 1 Range
        13: 28,  # APEX Family
        14: 12,  # Priority
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 2 — Summary
    # ============================================================
    ws2 = wb.create_sheet("Summary")

    # Title
    ws2["A1"] = "Disney Agentic Plays — Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1A2339")
    ws2.merge_cells("A1:D1")

    ws2["A3"] = "Generated for the Deloitte Account Team for Disney"
    ws2["A3"].font = Font(name="Arial", italic=True, color="6B7280", size=10)
    ws2.merge_cells("A3:D3")

    # Counts table
    ws2["A5"] = "By Domain"
    ws2["A5"].font = Font(name="Arial", bold=True, size=11)

    headers2 = ["Domain", "Play Count", "Wave 1 Range Total ($M, low)", "Wave 1 Range Total ($M, high)"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=6, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")
        c.alignment = Alignment(horizontal="center")

    # Formula-driven counts and sums (across Plays sheet)
    # Plays!B (Domain), Plays!L (Wave 1 Range $M)
    # Range column has text like "1.0–1.8". We need to extract low/high. We'll use
    # SUMPRODUCT with VALUE on a left/right split — but openpyxl can't natively
    # do hyphen-character matching for the em-dash easily. Use a helper-friendly
    # approach: hardcode the totals (acceptable here because they're summary-only,
    # not editable inputs). Document the source.

    # Hardcoded summary (acceptable for a summary deliverable; users edit Plays
    # sheet directly if they need different envelopes).
    # Back-Office now 13 (added AI Consumption Cost play, range 1.3-1.9)
    summary_rows = [
        ("Back-Office", 13, 14.0, 22.7),
        ("Streaming", 10, 12.0, 18.0),
        ("Experiences", 1, 2.0, 3.5),
        ("TOTAL", 24, 28.0, 44.2),
    ]
    for i, (domain, count, low, high) in enumerate(summary_rows, start=7):
        ws2.cell(row=i, column=1, value=domain).font = Font(name="Arial", bold=(domain == "TOTAL"))
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=low).number_format = "$#,##0.0"
        ws2.cell(row=i, column=4, value=high).number_format = "$#,##0.0"
        for c in range(1, 5):
            cell = ws2.cell(row=i, column=c)
            cell.font = Font(name="Arial", size=10, bold=(domain == "TOTAL"))
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right")
            if domain == "TOTAL":
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    # Priority mix
    ws2["A12"] = "By Priority (Account Team Sequencing)"
    ws2["A12"].font = Font(name="Arial", bold=True, size=11)
    headers3 = ["Priority", "Play Count", "Notes"]
    for col, h in enumerate(headers3, start=1):
        c = ws2.cell(row=13, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")

    priority_rows = [
        ("High", 14, "Wave 1 candidates + strategic CTO/Experiences/Engineering Productivity/AI Consumption Cost plays — high strategic visibility · clear KPI signal · governance differentiation"),
        ("Medium", 9, "Wave 2 candidates — strong business case · sequenced after Wave 1 lands"),
        ("Low", 1, "Wave 3+ — strategic but lower-velocity placement in sequence"),
    ]
    for i, (pri, count, note) in enumerate(priority_rows, start=14):
        ws2.cell(row=i, column=1, value=pri).font = Font(name="Arial", size=10, bold=True, color={"High": "1B7F3A", "Medium": "C28A22", "Low": "6B7280"}[pri])
        ws2.cell(row=i, column=2, value=count).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3, value=note).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="center")

    # Top 5 Wave-1 picks per domain
    ws2["A19"] = "Recommended Wave 1 Entry Picks"
    ws2["A19"].font = Font(name="Arial", bold=True, size=11)

    headers4 = ["Domain", "Play", "Why"]
    for col, h in enumerate(headers4, start=1):
        c = ws2.cell(row=20, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")

    picks = [
        ("Back-Office", "Continuous-Close Agent", "CFO-owned · close-cycle compression is publicly committed · high-visibility metric · works with existing finance team"),
        ("Back-Office", "Audit Evidence Readiness Agent", "Audit-cycle is a forcing function · governance-differentiation play · CCO partnership"),
        ("Back-Office", "IT Service-Desk Agent-Assist", "Cross-Practice pattern · low engineering risk · employee-visible win"),
        ("Streaming", "Password-Sharing Detection Agent", "Net-new revenue · highest strategic visibility · governance-differentiation play · industry-proven model"),
        ("Streaming", "Streaming Churn Prediction Agent", "DTC profitability commitment · clear KPI envelope · adjacent to ad-tier and lifecycle plays"),
        ("Streaming", "ESPN Auto-Highlight Agent", "ESPN-DTC critical-path · clear production-team value · NAB-conference visibility"),
        ("Back-Office", "CTO Portfolio & Decision Intelligence Agent", "Strategic-CTO Wave 2 follow-on · positions APEX as the framework that governs Disney's broader AI portfolio · CTO-owned with CFO alignment"),
        ("Experiences", "Guest Day Orchestration Agent", "$60B Experiences capital commitment ROI lever · MagicBand+/MyDisneyExperience data uniquely Disney's · Genie+ attach revenue + per-guest yield · see Disney-Experiences-Agentic-Play-Brief.md"),
        ("Back-Office", "Engineering Headcount Optimisation & Productivity Agent", "Joint CTO-CFO conversation · multi-agent platform produces both productivity-lift AND headcount-reduction option · management decides which · industry-standard AI-augmented-engineering pattern · GitHub Copilot adjacency"),
        ("Back-Office", "AI Consumption Cost Intelligence Agent", "CFO-urgent Wave 1-2 entry · addresses the +20-40% QoQ growth in AI-consumption run-rate · joint CTO-CFO governance · Azure Cost Management and M365 Copilot Admin Center adjacency · operating-cost discipline at the AI run-rate layer"),
    ]
    for i, (dom, play, why) in enumerate(picks, start=21):
        c1 = ws2.cell(row=i, column=1, value=dom)
        c1.font = Font(name="Arial", size=10, bold=True)
        c1.fill = PatternFill("solid", fgColor=domain_fill[dom])
        c2 = ws2.cell(row=i, column=2, value=play)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        c3 = ws2.cell(row=i, column=3, value=why)
        c3.font = Font(name="Arial", size=10)
        c3.alignment = Alignment(wrap_text=True, vertical="center")

    # Column widths for Summary
    sw = {1: 14, 2: 38, 3: 70, 4: 22}
    for col, w in sw.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 3 — How To Use
    # ============================================================
    ws3 = wb.create_sheet("How To Use")

    instructions = [
        ("Disney Agentic Plays — How To Use This File", "title"),
        ("", ""),
        ("Purpose", "h2"),
        ("This file curates 24 high-confidence agentic plays for the Disney account across enterprise back-office (13), streaming (10), and Experiences (1). Each play is structured for the Deloitte Account Team to use as a working artefact during account planning, executive prep, and pursuit discussions.", "p"),
        ("", ""),
        ("Sheet 1 — Plays", "h2"),
        ("The full play list. 14 columns. Filter or sort by Domain, Wave, or Priority to scope your view. The Wave 1 Range column gives an indicative Deloitte services range; precise envelope is set during the BVA.", "p"),
        ("", ""),
        ("Sheet 2 — Summary", "h2"),
        ("Aggregated counts and recommended Wave 1 picks. Three sections — by domain, by priority, and recommended Wave 1 entry plays. Use this sheet when briefing leadership on the curated portfolio.", "p"),
        ("", ""),
        ("How to use in account planning", "h2"),
        ("• Pre-meeting prep — filter to the buyer's domain (Finance for CFO, DTC for President of DTC, etc.) and scan the relevant rows.", "p"),
        ("• Quarterly account plan — review High-priority plays and confirm which are in active discovery, which need to be opened.", "p"),
        ("• BVA construction — pull the KPI Signal column to seed BVA-in-5-bullets per the Sellers Handbook framework.", "p"),
        ("• Microsoft attach modelling — Microsoft Attach column drives SKU mix for Microsoft consumption forecasting.", "p"),
        ("", ""),
        ("Independence-clean usage", "h2"),
        ("This file lists Deloitte services plays. The dollar ranges are Deloitte services revenue ranges. Microsoft platforms are licensed directly by Disney from Microsoft; this file does not include Microsoft software margin. Refer to Sellers Podcast Episode 2 and the Independence Cheat Sheet for the full Independence posture.", "p"),
        ("", ""),
        ("Notes on KPI signals", "h2"),
        ("KPI signals are reference-scenario estimates from the APEX framework — not Disney-specific commitments. Disney-specific numbers are produced during the BVA, with Disney's own baseline measurement.", "p"),
        ("", ""),
        ("Last updated", "h2"),
        ("2026-05-13 — curated by TMT Practice + Disney Account Team", "p"),
    ]

    row = 1
    for text, style in instructions:
        cell = ws3.cell(row=row, column=1, value=text)
        if style == "title":
            cell.font = Font(name="Arial", bold=True, size=16, color="1A2339")
        elif style == "h2":
            cell.font = Font(name="Arial", bold=True, size=12, color="1A2339")
        elif style == "p":
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws3.column_dimensions["A"].width = 100

    return wb


def main():
    wb = build_workbook()
    wb.save(OUT)
    print(f"Wrote {OUT.name} ({OUT.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()

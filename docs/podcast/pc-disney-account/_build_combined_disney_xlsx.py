"""
Build the combined Disney Agentic Plays Excel — one workbook covering all
39 curated agentic plays for The Walt Disney Company across:

  Back-Office (13) — finance, procurement, legal, audit, IT, HR, T&E, risk,
                     cyber, CTO portfolio, engineering productivity,
                     AI consumption cost intelligence
  Streaming   (10) — churn, personalisation, cold-start, paid-sharing, ESPN
                     highlights, auto-dub, lifecycle, QoE, ad-targeting, CSR
  Experiences (1)  — Guest Day Orchestration Agent
  Studios     (15) — content-development & greenlight (4), production (5),
                     marketing/distribution/rights (6)

This file is the canonical Disney company-wide play book. It supersedes the
two segment-specific files (Disney_Agentic_Plays_BackOffice_Streaming.xlsx
and Disney_Studios_Agentic_Plays.xlsx) for the Account Team's primary
working view — those remain available for segment-specific working sessions.

Output: Disney_Agentic_Plays_Combined.xlsx

Four sheets:
  1. Plays       — all 39 plays with Segment + Domain columns for filtering
  2. Summary     — counts by segment · Wave 1 envelope · priority mix · top picks
  3. By Segment  — same data filtered/grouped by segment for quick scan
  4. How To Use  — guidance for the Account Team

Usage:
    python _build_combined_disney_xlsx.py
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "Disney_Agentic_Plays_Combined.xlsx"

# ---------------------------------------------------------------- headers --

HEADERS = [
    "#",
    "Segment",
    "Sub-domain",
    "Play",
    "Business Problem",
    "Agent Capability",
    "Disney Pressure Addressed",
    "KPI Signal",
    "Buyer at Disney",
    "Microsoft Attach",
    "Wave",
    "Wave 1 Range ($M)",
    "APEX Family",
    "Priority",
]

# Each play tuple shape (consistent with both source files):
#   (Play, Sub-domain, Business Problem, Agent Capability, Pressure, KPI,
#    Buyer, MSFT Attach, Wave, Range, APEX Family, Priority)

# ---------------------------------------------------------------- back-office ---

BACK_OFFICE_PLAYS = [
    (
        "Continuous-Close Agent",
        "Finance",
        "Quarterly close cycle is 8 days; target is 3. Manual journal entries, reconciliations, and consolidations dominate close-week labour.",
        "Agent drafts journal entries from source-system signals, reconciles inter-company balances, generates consolidation packages. Controller approves.",
        "Audit posture · finance team capacity · CFO close-cycle commitment",
        "Close cycle 8d → 3d · 30–50% cycle compression",
        "CFO · Controller",
        "Fabric · Foundry · Purview · Power BI",
        "Wave 1", "1.0–1.8", "Cross-edition Finance", "High",
    ),
    (
        "AP Invoice-Triage Agent",
        "Finance · AP",
        "Disney processes very large invoice volumes across segments. Manual triage, coding, and exception-handling dominate AP labour.",
        "Agent reads invoices, extracts line items, matches PO/contract, codes GL accounts, flags exceptions for review. AP clerk approves.",
        "Working capital · AP team capacity",
        "AP cycle time −50% · DPO optimisation · 30% reduction in invoice exceptions",
        "Controller · AP Director",
        "Fabric · Foundry · Document Intelligence",
        "Wave 1", "0.8–1.4", "Cross-edition Finance", "High",
    ),
    (
        "Procurement Intake & Spend-Categorisation Agent",
        "Procurement",
        "Procurement intake is high-volume, manual triage. Spend categorisation drift creates analytics and savings-tracking errors.",
        "Agent classifies intake requests, recommends suppliers from approved pool, surfaces consolidation opportunities. Procurement team approves.",
        "Spend visibility · supplier relationship discipline",
        "Intake cycle −40% · spend-data-quality + 30 points · category-tail consolidation savings",
        "CPO · Procurement Operations",
        "Fabric · Foundry · Dynamics 365 SCM",
        "Wave 2", "0.9–1.6", "Cross-edition Procurement", "Medium",
    ),
    (
        "Contract Review Agent",
        "Legal · Commercial",
        "Disney's legal organisation reviews enormous volumes of contracts — talent, distribution, vendor, partnership. Clause-by-clause review is slow and inconsistent.",
        "Agent reads contracts, flags non-standard clauses, compares against precedent library, drafts redlines, prepares attorney summary. Attorney approves.",
        "Legal-team capacity · contract-cycle velocity",
        "Contract review cycle −60% · clause-risk-detection consistency +40%",
        "GC · Deputy GC for Commercial",
        "Fabric · Foundry · Purview (PHI/PII propagation)",
        "Wave 2", "1.2–2.0", "Cross-edition Legal", "Medium",
    ),
    (
        "Audit Evidence Readiness Agent",
        "Internal Audit",
        "Annual SOX and operational-audit cycles require evidence compilation across many systems. Manual evidence-gathering consumes audit staff time.",
        "Agent continuously gathers control-evidence artefacts, maintains audit-readiness dashboard, drafts evidence packages on demand.",
        "Audit posture · continuous control attestation",
        "Audit evidence preparation −70% · 100% evidence-trail coverage on monitored controls",
        "Chief Audit Executive · CCO",
        "Fabric · Foundry · Purview · Microsoft 365 Compliance",
        "Wave 1", "1.1–1.7", "Cross-edition Audit", "High",
    ),
    (
        "IT Service-Desk Agent-Assist",
        "Internal IT",
        "200,000+ Disney employees globally generate enormous IT ticket volume. Outsourced service-desk economics favour quality lift over headcount reduction.",
        "Agent assists human IT support tech in real time with cross-system context, drafts resolution, recommends escalation path. Human approves.",
        "Internal NPS · employee productivity · IT-contract renegotiation leverage",
        "AHT −30% · FCR +15 pts · employee CES improvement",
        "CIO · Head of IT Operations",
        "Fabric · Foundry · M365 Copilot · ServiceNow integration",
        "Wave 2", "1.0–1.5", "Cross-Practice Contact-Center", "High",
    ),
    (
        "HR Onboarding & Lifecycle Agent",
        "HR",
        "New-hire onboarding across 200,000 global employees is workflow-heavy. Offboarding and transitions equally complex.",
        "Agent orchestrates the onboarding workflow across HR systems, IT provisioning, training assignment, and policy acknowledgement. HR specialist approves exception cases.",
        "Employee experience · HR-team capacity · time-to-productivity",
        "Time-to-productivity for new hires −40% · onboarding NPS +20 pts",
        "CHRO · Head of HR Ops",
        "Fabric · Foundry · M365 Copilot · Workday integration",
        "Wave 2", "0.9–1.5", "Cross-edition HR", "Medium",
    ),
    (
        "T&E Compliance & Audit Agent",
        "Finance · T&E",
        "Travel & expense compliance audits are sample-based. Policy violations slip through. Manual review of high-volume T&E claims is labour-intensive.",
        "Agent reviews T&E claims continuously against policy, flags violations, drafts compliance findings. T&E reviewer approves enforcement.",
        "Audit posture · expense-policy enforcement",
        "100% claim coverage · policy-violation detection +40% · T&E review labour −60%",
        "Controller · Head of T&E",
        "Fabric · Foundry · Purview · M365 Copilot",
        "Wave 2", "0.7–1.2", "Cross-edition Finance", "Medium",
    ),
    (
        "Continuous Controls & Risk Monitoring Agent",
        "Risk & Compliance",
        "Disney's risk register requires continuous monitoring. Key controls are tested periodically; emerging risk signals are missed.",
        "Agent monitors control evidence streams continuously, evaluates against control specs, escalates emerging risk signals. Risk officer approves response actions.",
        "Audit posture · regulatory exposure · CCO confidence",
        "Mean-time-to-detect emerging risk −60% · control-test coverage 100%",
        "Chief Risk Officer · CCO",
        "Fabric · Foundry · Purview · Sentinel",
        "Wave 2", "1.2–2.0", "Cross-edition Risk", "High",
    ),
    (
        "Third-Party Risk & Supply-Chain Security Monitoring Agent",
        "Cyber · Vendor Risk",
        "Disney's third-party ecosystem creates indirect security exposure. Annual-assessment-driven third-party programs miss emerging risk signals.",
        "Agent continuously monitors third-party security posture from public signal, news, and shared assessment data. Recommends escalation. Security team prioritises.",
        "Cyber posture · regulatory exposure · supply-chain resilience",
        "Third-party-incident rate −30% · time-to-emerging-risk-response −50%",
        "CISO · Head of Vendor Risk",
        "Fabric · Foundry · Sentinel · Defender for Cloud",
        "Wave 3", "1.0–1.6", "Cross-edition Cyber", "Medium",
    ),
    (
        "CTO Portfolio & Decision Intelligence Agent",
        "Technology Strategy · CTO Office",
        "Disney's CTO oversees hundreds of technology investments across segments. Manual portfolio reviews happen quarterly; the portfolio moves faster. Stranded pilots accumulate. Investment ROI is opaque between reviews.",
        "Agent reads the CTO's technology investment portfolio continuously (CapEx tracking, project status, performance telemetry, cost actuals, business-KPI alignment), produces a continuous portfolio-health view, identifies stranded pilots and at-risk investments, drafts portfolio recommendations.",
        "Technology investment ROI · AI portfolio rationalisation · execution velocity · CFO-CTO alignment on tech spend",
        "Stranded-pilot rate −40% · investment-cycle-time compression 50% · portfolio review quarterly → continuous",
        "CTO · CIO (joint depending on org)",
        "Fabric · Foundry · Power BI · Purview · Power Platform",
        "Wave 2", "1.4–2.0", "Cross-edition Technology Strategy", "High",
    ),
    (
        "Engineering Headcount Optimisation & Productivity Agent",
        "Engineering · R&D",
        "Disney's engineering organisation spans Studios, DTC streaming, Parks technology, ESPN production tech, internal IT, and corporate engineering. AI-augmented engineering produces measurable productivity lift; the option set is (a) more output same team, (b) reduce headcount, (c) shift hiring profile, (d) redirect capacity. Management decides; platform enables.",
        "Multi-agent engineering productivity platform — code-generation · agentic code-review · design-doc drafting · automated test generation · technical-documentation maintenance · deployment-pipeline assist · internal developer-Q&A. Engineering leaders measure lift by team and decide composition.",
        "Operating cost discipline · technology investment ROI · workforce composition for the AI era · engineering velocity through transitions",
        "Engineer productivity +20–35% on covered task types · code-review cycle −40-60% · capacity available for redirection OR reduction (management decision) · onboarding-time reduction",
        "CTO · CFO (joint) · Heads of Engineering per BU · CHRO (workforce-composition partner)",
        "GitHub Copilot · Fabric · Foundry · M365 Copilot · Power Platform · Azure DevOps integration",
        "Wave 2-3", "1.5–2.5", "Cross-edition Engineering Productivity", "High",
    ),
    (
        "AI Consumption Cost Intelligence Agent (Engineering + AI-Assist FinOps)",
        "Technology · FinOps · AI Cost Discipline",
        "Enterprise AI consumption costs are exploding across Disney. GitHub Copilot seats, M365 Copilot licenses, Azure OpenAI / Foundry token consumption, embedding and vector-store spend, custom-agent run-rate, Power Platform AI Builder consumption, team-expensed third-party AI tools (Cursor, Cline, Continue, Bedrock, Anthropic) add up to multi-tens-of-millions of dollars annually. Run-rate is climbing 20-40% quarter-over-quarter. Cloud-FinOps practices do not yet cover AI consumption cleanly.",
        "Continuously composes AI consumption signal: Microsoft EA Copilot seat utilisation · Azure OpenAI/Foundry token consumption by model · GitHub Copilot per-developer activity · embedding API and vector-store spend · custom-built agent run-cost · third-party AI tool expense · Power Platform AI Builder consumption. Surfaces per-team, per-product, per-use-case, per-model cost. Identifies idle seats, runaway agents, sub-optimal model selection. CFO + CTO see joint operating view; FinOps team approves reclamation.",
        "AI-consumption cost discipline · CFO-CTO joint AI governance · model-selection-by-workload optimisation · FinOps maturity for the AI era · operating cost discipline at the AI run-rate layer",
        "AI consumption cost growth: from +20-40% QoQ trajectory to flat-or-managed · idle Copilot seat reclamation rate · model-selection optimisation savings 12-25% on covered workloads · per-use-case cost-per-outcome surfaced · CFO-CTO joint quarterly governance cadence",
        "CTO · CFO (joint) · Head of FinOps · CIO",
        "Azure Cost Management · Microsoft Cost Management for Azure OpenAI · Fabric · Foundry · Purview · Power BI · M365 Copilot Admin Center",
        "Wave 1-2", "1.3–1.9", "Cross-edition FinOps / AI Cost Governance", "High",
    ),
]

# ---------------------------------------------------------------- streaming ---

STREAMING_PLAYS = [
    (
        "Streaming Churn Prediction & Retention Agent",
        "Cross-Streaming",
        "Each streaming service has its own churn dynamics. Manual marketing intervention takes weeks; the churn window is days.",
        "Agent watches subscriber behaviour continuously, identifies elevated churn risk, recommends personalised retention offer with margin-aware pricing. Loyalty team approves.",
        "DTC profitability · subscriber retention · cross-service portfolio value",
        "Targeted-cohort churn −19% · annualised retention revenue protection",
        "President of DTC · Chief Customer Officer",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 1", "1.5–2.0", "RC-CX · streaming variant", "High",
    ),
    (
        "Personalised Content Recommendations & Rails Agent",
        "Disney+ · Hulu",
        "Heterogeneous content portfolio (Marvel, Pixar, Lucasfilm, ABC, FX, Nat Geo) requires coherent personalisation per user.",
        "Agent composes per-user rails layout on top of the recommendation engine, reasoning about session context, household composition, and new-release calendar.",
        "DTC engagement · content investment ROI · session time",
        "Engagement +23% · session-length lift · rail-CTR improvement",
        "Chief Product Officer DTC",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 1", "1.3–1.9", "Cross-edition Personalisation", "High",
    ),
    (
        "Cold-Start Boost Agent for Franchise Launches",
        "Disney+",
        "Major launches (Marvel, Pixar, Lucasfilm) have a critical 72-hour cold-start window. Generic similarity-based recommendation underperforms.",
        "Agent identifies high-probability cohorts for a specific launch given content fingerprint, market context, and marketing campaign. Surfaces with prioritised placement.",
        "Content investment ROI · launch performance · franchise momentum",
        "Cold-start uptake +41% · first-72-hour engagement lift",
        "Chief Product Officer DTC · Head of Studios Marketing",
        "Fabric · Foundry · Real-Time Intelligence",
        "Wave 1", "1.0–1.5", "Cross-edition Personalisation", "High",
    ),
    (
        "Password-Sharing Detection & Paid-Sharing Agent",
        "Cross-Streaming",
        "Industry has demonstrated paid-sharing economics. Naive detection produces false positives that hurt brand. Disney needs graceful, governed enforcement.",
        "Agent reasons about household-vs-sharing patterns with confidence scoring. Recommends paid-sharing offer flow. Marketing-ops team approves; CCO sees governance trail.",
        "DTC revenue · brand-experience integrity · governance posture",
        "Incremental revenue +$8M/year reference · false-positive rate <2%",
        "President of DTC · CMO · CCO",
        "Fabric · Foundry · Purview (audit-row chain) · Entra",
        "Wave 1", "1.2–1.8", "Cross-edition Subscription", "High",
    ),
    (
        "ESPN Auto-Highlight Detection & Clipping Agent",
        "ESPN · ESPN+",
        "ESPN produces enormous volumes of highlight content. Manual identification, editing, and packaging is the production bottleneck for scale clip output.",
        "Agent watches live game feed (state, audio, vision), identifies clip candidates with timestamps and metadata. Production team reviews and approves.",
        "ESPN-DTC competitiveness · production-team capacity · audience clip-consumption",
        "Production time per clip −74% · clip-output volume 3–5×",
        "Head of ESPN Production · ESPN-DTC President",
        "Fabric RTI · Foundry · Azure AI Video",
        "Wave 1", "1.5–2.2", "Cross-edition Live-Content", "High",
    ),
    (
        "Auto-Dub & Subtitle Agent for International Markets",
        "Disney+ International",
        "50+ international markets require localised content. Multi-vendor, multi-quality-gate localisation creates 4–6 week time-to-market lag.",
        "Agent generates draft localised dub audio and subtitles using current speech and language models, flags quality concerns. Linguistic experts review edge cases.",
        "International subscriber growth · time-to-market · localisation cost",
        "Localisation time per asset −62% · 50%+ throughput lift in active markets",
        "Head of International DTC · Head of Localisation",
        "Fabric · Foundry · Azure Speech · Azure AI Translator",
        "Wave 2", "1.0–1.6", "Cross-edition Speech-and-Language", "Medium",
    ),
    (
        "Subscriber Lifecycle Agent — Trial-to-Paid, Downgrade Prevention",
        "Cross-Streaming",
        "Subscriber-lifecycle events (trial-to-paid conversion, downgrade threats, upgrade opportunities) are managed reactively today.",
        "Agent watches subscriber lifecycle stages continuously, intervenes at the right moment with personalised offer, manages the lifecycle from trial to long-term subscriber.",
        "DTC profitability · subscriber LTV · conversion economics",
        "Trial-to-paid conversion +12% · downgrade prevention +20% on targeted cohort",
        "President of DTC · CMO",
        "Fabric · Foundry · Customer Insights · Power Automate",
        "Wave 2", "1.1–1.7", "Cross-edition Subscription", "Medium",
    ),
    (
        "Streaming Quality-of-Experience Triage Agent",
        "Streaming Delivery Ops",
        "QoE issues (buffering, codec failures, regional CDN problems) manifest unevenly. Operator-driven dashboard detection produces hour-scale response.",
        "Agent watches delivery telemetry continuously, identifies emerging regional or device-family quality issues, recommends operations response (reroute, encoding-pipeline adjust).",
        "DTC service reliability · operator capacity · subscriber experience",
        "Playback-failure-rate −35% · mean-time-to-quality-recovery −60%",
        "Head of Streaming Delivery · CTO DTC",
        "Fabric RTI · Foundry · Azure Front Door · Application Insights",
        "Wave 2", "1.0–1.6", "Operations / RTI", "High",
    ),
    (
        "Ad-Targeting Agent for Hulu / Disney+ Ad Tier / ESPN+",
        "Ad Sales · AVOD",
        "Programmatic ad insertion across multiple streaming surfaces requires real-time decisioning. Generic ad-tech doesn't reason about Disney's portfolio context.",
        "Agent reasons about ad placement opportunity per viewer per session, optimising for advertiser brief and viewer experience. Ad-ops team approves campaigns.",
        "Ad revenue · advertiser brief satisfaction · viewer experience preservation",
        "Ad-revenue lift on managed inventory +10–18% · ad-CTR +12% · viewer-complaint rate −20%",
        "Head of Ad Sales · CRO · Head of AVOD Strategy",
        "Fabric · Foundry · Customer Insights · Azure AI Video",
        "Wave 2", "1.3–2.0", "Cross-edition Ad-Tech", "Medium",
    ),
    (
        "Subscriber Support Agent — Customer Service Augmentation",
        "Customer Service · Streaming",
        "Disney+/Hulu/ESPN+ customer-service tickets span billing, playback, content, and account issues. Cross-system context composition slows resolution.",
        "Agent assists human CSR in real time with subscriber context, account history, recent activity, recommended resolution. CSR approves.",
        "Subscriber NPS · CSR capacity · attrition / contractor relationship",
        "AHT −30% · FCR +18 pts · subscriber CES improvement",
        "Head of Customer Service Streaming · CXO",
        "Fabric · Foundry · M365 Copilot · Dynamics 365 Service",
        "Wave 1", "1.0–1.4", "Cross-Practice Contact-Center", "High",
    ),
]

# ---------------------------------------------------------------- experiences ---

EXPERIENCES_PLAYS = [
    (
        "Guest Day Orchestration Agent",
        "Parks · Resorts · Guest-Facing",
        "Guest days at Disney parks involve dozens of decisions per guest per day — ride selection, dining, Lightning Lane purchases, character meets, photo opportunities, rest breaks. The data exists across MagicBand+, My Disney Experience app, and park-state systems; today the guest composes manually.",
        "Agent watches MagicBand+/MyDisneyExperience signal continuously. Composes continuously updated personalised itinerary. Notifies guest opt-in with next-best-action recommendation. Guest decides (HITL is the guest).",
        "$60B Experiences capital investment ROI · guest experience differentiation · Genie+/Lightning Lane attach revenue · per-guest yield · repeat-visit propensity",
        "Genie+/Lightning Lane attach +10–20% · in-park spend per guest +5–10% · guest NPS lift · repeat-visit booking rate ↑",
        "President of Experiences · SVP Guest Experience · CTO Experiences",
        "Fabric · Foundry · Real-Time Intelligence · Customer Insights · M365 Copilot",
        "Wave 2-3", "2.0–3.5", "Cross-edition Guest-Experience", "High",
    ),
]

# ---------------------------------------------------------------- studios ---

STUDIOS_CONTENT_DEV = [
    (
        "Greenlight Decision Support Agent",
        "Studios · Development & Greenlight",
        "Greenlight at tentpole level requires alignment across 12+ stakeholders. Cycle runs weeks-to-months. Decision-quality variance is high because debate is partly about reconciling differing data interpretations.",
        "Agent composes per-project view across comparable-title performance, talent-availability, production-cost benchmark, marketing-spend comparable, awards-positioning, genre-trends, international market readiness. Executive committee sees same composed view. Agent does NOT recommend greenlight; committee decides.",
        "Content cost inflation · hit-and-miss volatility · per-title ROI",
        "Greenlight cycle time -30 to -50% on tentpole projects · decision-quality variance reduction · post-greenlight surprise reduction",
        "Studio President · Chief Creative Officer · Head of Business Affairs (multi-buyer)",
        "Fabric · Foundry · Power BI · Purview",
        "Wave 1", "1.5-2.2", "TMT Content / Decision Support", "High",
    ),
    (
        "Audience Science & Sentiment Agent",
        "Studios · Audience Research & Marketing",
        "Audience-science data arrives in batches — preview screenings (12hr), trailer tests (48hr), social sentiment (days). Decisions on cut and marketing have to happen on the same compressed cadence. Composition of multiple streams is manual.",
        "Agent reads audience-research streams continuously, composes per-project sentiment-and-response view, recommends focal areas for creative-team review. Creative team decides what to change.",
        "Hit-and-miss volatility · marketing campaign effectiveness · cut-and-creative decision velocity",
        "Time-to-composed-audience-insight -60 to -70% · decision-velocity on cut/marketing adjustments · creative-team feedback cycle accelerated",
        "Chief Marketing Officer · Head of Audience Science · Studio President",
        "Fabric · Foundry · Real-Time Intelligence · Customer Insights",
        "Wave 1", "1.2-1.8", "TMT Audience / Real-Time Intelligence", "High",
    ),
    (
        "IP & Franchise Opportunity Scoring Agent",
        "Studios · IP Strategy & Development",
        "Disney's IP portfolio is enormous (Marvel, Star Wars, Pixar, Walt Disney library, 20th Century library, Searchlight catalog). Identifying highest-opportunity IP is largely judgement-based; portfolio view is not continuously composed; opportunities surface episodically.",
        "Agent reads continuously across IP-performance signal, audience-affinity, social signal, competitive landscape, talent-attachment availability. Surfaces emerging opportunities; development leadership reviews and prioritises.",
        "Franchise-extension revenue · long-tail IP monetisation · slate-portfolio coverage",
        "Underutilised-IP-identification rate · time-to-opportunity-surfaced · development-pipeline coverage of high-opportunity IP",
        "Chief Creative Officer · Head of IP Strategy · Studio President",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 2", "1.0-1.6", "TMT IP / Portfolio Analytics", "Medium",
    ),
    (
        "Talent & Casting Intelligence Agent",
        "Studios · Casting & Business Affairs",
        "Composing implications of a casting choice is manual today. Comparable-performance, audience-affinity, awards track-record, compensation comparables, calendar availability, existing commitments — all exist but composed by head of casting plus studio executive in conversation.",
        "Agent composes dimensions for candidate talent or ensemble. Surfaces composite-fit view. Casting director, director, studio executive review and decide. Agent does NOT generate or simulate performances, does NOT use likenesses without explicit consent, does NOT displace casting directors. Operational composition only.",
        "Talent-and-AI policy compliance · casting-decision velocity · production economics",
        "Casting decision cycle compression · coverage of full candidate-set · casting-decision stakeholder alignment time reduction",
        "Head of Casting · Studio President · Director (per project)",
        "Fabric · Foundry · Purview (talent-rights propagation)",
        "Wave 2", "0.9-1.4", "TMT Talent / Decision Support", "Medium",
    ),
]

STUDIOS_PRODUCTION = [
    (
        "Production Schedule & Budget Intelligence Agent",
        "Studios · Physical Production",
        "Tentpole production budgets run $150-300M. Schedule variance drives budget variance — every over-schedule day cascades. Production accounting and schedule systems are separate. Composition into early-warning signal is weekly-cadence human work.",
        "Agent reads production accounting, daily call-sheet completion, scene-coverage progress, vendor-invoice signal, post-production dependency signal. Continuously composes trajectory. Identifies emerging variance patterns. Decisions stay human.",
        "Per-title ROI · content cost inflation · production economics integrity",
        "Time-to-detect emerging variance: weekly -> daily · variance-recovery rate improvement · budget-and-schedule integrity across annual slate",
        "President of Physical Production · Studio CFO · Senior Line Producer",
        "Fabric · Foundry · Power BI · Real-Time Intelligence",
        "Wave 1", "1.4-2.0", "TMT Production / Real-Time Intelligence", "High",
    ),
    (
        "VFX Pipeline Optimisation Agent",
        "Studios · VFX & Post-Production",
        "VFX is the largest production-cost line item on tentpoles. Productions carry 2000-3000 VFX shots across multiple vendors globally. Vendor coordination is high-variance and high-stakes. Status flows through vendor tracking + emails + weekly meetings.",
        "Agent reads continuously across vendor tracking, shot-version repos, review-feedback, delivery schedules. Identifies emerging risk patterns. Surfaces to VFX supervisor and post-production producer.",
        "Per-title ROI · VFX cost discipline · Marvel/Lucasfilm pipeline integrity · ILM internal operations",
        "VFX-delivery-risk detection compression · final-delivery slippage rate reduction · VFX-budget-variance reduction",
        "VFX Supervisor · VFX Producer · Post-Production Producer · ILM Operations (Lucasfilm)",
        "Fabric · Foundry · Real-Time Intelligence · Power BI",
        "Wave 1", "1.5-2.2", "TMT VFX / Cross-Vendor Coordination", "High",
    ),
    (
        "Pixar Animation Pipeline Agent",
        "Studios · Long-Cycle Animation",
        "Pixar production cycle is 4-5 years per feature. Story revisions cascade into design revisions cascade into animation reworks. Render-farm capacity is finite; render time per shot compounds. Long-cycle visibility is structural work.",
        "Agent reads continuously across story-revision tracking, design-asset versioning, animation-shot status, render-farm queue, sound-pipeline status. Composes multi-year pipeline view. Agent does NOT replace creative judgement; operational substrate only.",
        "Pixar production economics · animation-pipeline cycle integrity · story-cascade risk · render-capacity optimisation",
        "Pipeline-cycle-time variance reduction · render-capacity utilisation · earlier surface of story-revision cascade risk",
        "Pixar President · Chief Creative Officer (Pixar) · Head of Production (Pixar)",
        "Fabric · Foundry · Real-Time Intelligence",
        "Wave 2-3", "1.3-1.9", "TMT Animation / Long-Cycle Pipeline", "Medium",
    ),
    (
        "Post-Production Workflow Agent",
        "Studios · Post-Production & Delivery",
        "Post-production is editing, sound, color, music, VFX integration, delivery prep. Dependency chain is long; schedule tight. Late delivery of one element cascades. Visibility across the dependency chain is partially manual.",
        "Agent reads continuously across editing milestones, sound milestones, VFX delivery, color-and-finishing milestones, delivery preparation. Identifies dependency-cascade risk. Surfaces to post-production supervisor with timeline-adjustment recommendations.",
        "Post-production schedule integrity · final-delivery on-schedule rate · production economics",
        "Dependency-cascade detection time compression · post-production schedule integrity · final-delivery-on-schedule rate improvement",
        "Post-Production Supervisor · Post-Production Producer",
        "Fabric · Foundry · Power BI",
        "Wave 2", "0.9-1.4", "TMT Post / Schedule Coordination", "Medium",
    ),
    (
        "Production Safety & Wellness Agent",
        "Studios · Production Operations & Safety",
        "Safety-incident-pattern detection across multiple concurrent productions is partial today. Incidents at one production may indicate systemic patterns (vendor practices, scheduling pressure, stunt/equipment types) not surfaced until pattern matures.",
        "Agent reads safety-incident reporting, crew-sentiment signals (opt-in), set-incident logs, protocol-compliance signal. Surfaces emerging patterns. Anonymised pattern signal at production-level and protocol-level. Agent does NOT surveil individuals.",
        "Insurance economics · talent comfort · regulatory compliance · production-environment quality",
        "Time-to-pattern-detection across productions · safety-incident-rate trends · crew-wellness signal · production-environment-quality metric",
        "Studio President · President of Physical Production · Head of Safety",
        "Fabric · Foundry · Purview (strict PII governance) · Sentinel",
        "Wave 3", "0.8-1.3", "TMT Production / Safety Telemetry", "Low",
    ),
]

STUDIOS_MARKETING_RIGHTS = [
    (
        "Trailer Performance & Audience Testing Agent",
        "Studios · Marketing & Audience Testing",
        "Trailer-test results arrive in 48-hour batches. By the time results are reviewed, trailer-house has often started next-cut iteration. Composition with social-signal, competitive-trailer-performance, and audience-segment context is human work.",
        "Agent composes test-vendor results, social-signal post-launch, comparable-trailer benchmarks, competitive-landscape positioning. Surfaces composed view within hours. Marketing leadership decides on cut iteration and placement.",
        "Marketing campaign effectiveness · opening-weekend correlation · trailer-decision velocity",
        "Time-to-composed-trailer-insight compression 70-80% · decision-velocity on cut iteration · trailer-effectiveness improvement",
        "Chief Marketing Officer · Head of Worldwide Marketing · Studio Marketing President",
        "Fabric · Foundry · Real-Time Intelligence · Customer Insights",
        "Wave 1", "1.1-1.7", "TMT Marketing / Real-Time Intelligence", "High",
    ),
    (
        "Marketing Campaign Optimisation Agent",
        "Studios · Marketing & Cross-Channel",
        "Theatrical-tentpole campaigns run across broadcast, digital, social, programmatic, OOH, influencer, partnership, talent press, premiere, in-cinema. Per-channel attribution is partial; cross-channel optimisation more partial.",
        "Agent reads continuously across channel-performance, audience-segment response, competitive landscape, creative-effectiveness data. Recommends within-campaign reallocations. Marketing-operations team reviews and approves spend shifts.",
        "Marketing ROI · campaign-spend optimisation · marketing-team capacity",
        "Within-campaign reallocation velocity · marketing ROI improvement 8-15% on covered campaigns · cross-channel attribution maturity",
        "Chief Marketing Officer · Head of Media Strategy · CMO Operations",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 1", "1.2-1.8", "TMT Marketing / Cross-Channel Optimisation", "High",
    ),
    (
        "Theatrical-vs-Streaming Windowing Decision Agent",
        "Studios · Distribution Strategy",
        "Windowing decisions are strategic, not procedural. Theatrical-only vs theatrical-plus-streaming-30-days vs streaming-direct reshape revenue, marketing economics, talent relationships. Composing factors across title, market, strategic, talent-contract dimensions is largely manual.",
        "Agent composes the factors across title and market dimensions. Surfaces composite-decision view to windowing committee. Committee decides; agent does NOT recommend the decision.",
        "Per-title revenue capture · DTC subscriber-acquisition strategy · talent compensation provisions · executive-committee decision velocity",
        "Windowing-decision cycle compression · decision-quality variance reduction · post-decision execution quality",
        "Disney Entertainment Chairperson · Studio Presidents · Head of Worldwide Distribution · CFO Entertainment (executive committee)",
        "Fabric · Foundry · Power BI · Purview",
        "Wave 3", "1.6-2.4", "TMT Distribution / Strategic Decision Support", "Medium",
    ),
    (
        "Awards Campaign Management Agent",
        "Studios · Awards & Specialty Marketing",
        "Awards campaign coordination is year-round, peaking during Oscar and Emmy seasons. Searchlight's specialty-theatrical business depends on awards-season performance. Campaigns coordinate academy screenings, talent press, FYC advertising, industry-relationship management, festival positioning, eligibility tracking.",
        "Agent reads campaign-event calendar, eligibility-tracking, talent-availability, FYC-advertising performance, social-and-industry sentiment, competitive-campaign signal. Composes campaign-status view continuously. Awards-team approves campaign-decision recommendations.",
        "Searchlight brand viability · awards-driven theatrical extension · talent-relationship management",
        "Awards-campaign coordination velocity · nomination-and-win conversion rate · specialty-theatrical viability for Searchlight",
        "President of Searchlight · Awards Campaign Director · CMO Searchlight (Searchlight is natural lead)",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 2", "1.1-1.6", "TMT Awards / Specialty Marketing", "High",
    ),
    (
        "Rights Compliance & Content Reuse Agent",
        "Studios · Rights & Business Affairs",
        "Disney's content library is enormous. Reusing archive content requires rights clearance — music, talent, footage, image rights. Each reuse is a clearance project. Work is labour-intensive and slow; rights status across the library is partial.",
        "Agent reads continuously across rights-management systems, talent-contract repositories, music-licensing systems, footage-rights databases. Maintains continuous rights-status view. Surfaces rights-availability for candidate reuse projects.",
        "Long-tail IP monetisation · library-reuse velocity · compliance-risk reduction · business-affairs team capacity",
        "Rights-clearance cycle compression 50-70% · reuse-opportunity surfaced rate · compliance-risk reduction",
        "General Counsel Studios · Head of Business Affairs · Chief Rights Officer",
        "Fabric · Foundry · Purview (heavy PII governance) · M365 Compliance",
        "Wave 2", "1.0-1.5", "TMT Rights / Library Governance", "Medium",
    ),
    (
        "Music Synchronisation Clearance Agent",
        "Studios · Rights & Music",
        "Music sync clearance is high-volume — soundtrack assembly, trailer scoring, marketing-content scoring all require sync clearance. Each musical cue requires clearance from music publisher AND record label. Cycle from 'we want this song' to 'cleared and licensed' can run weeks.",
        "Agent reads continuously across music-publisher and record-label clearance systems, internal-licensing trackers, clearance-history databases. Surfaces clearance-availability and indicative pricing. Music supervisor decides; business-affairs approves.",
        "Music-clearance cycle compression · trailer-and-marketing schedule risk · music-supervisor productivity",
        "Music-clearance cycle compression · music-supervisor productivity · reduced clearance-related schedule risk on tight-deadline projects",
        "Music Supervisor · Head of Music Business Affairs · GC Studios",
        "Fabric · Foundry · Purview · M365 Compliance",
        "Wave 2", "0.7-1.2", "TMT Rights / Music Clearance", "Medium",
    ),
]


def build_workbook() -> Workbook:
    wb = Workbook()

    # ============================================================
    # Sheet 1 - Plays
    # ============================================================
    ws = wb.active
    ws.title = "Plays"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1A2339")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(border_style="thin", color="000000"))

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    segment_fill = {
        "Back-Office": "EEF2F7",
        "Streaming":   "FFF7E6",
        "Experiences": "F0F7EE",
        "Studios":     "F6F0FA",  # soft lavender — distinct from Disney Account segments
    }

    priority_color = {
        "High": "1B7F3A",
        "Medium": "C28A22",
        "Low": "6B7280",
    }

    row_num = 2
    play_id = 1

    def append_plays(plays, segment_label):
        nonlocal row_num, play_id
        for t in plays:
            ws.append([
                play_id, segment_label,
                t[1], t[0], t[2], t[3], t[4], t[5], t[6], t[7], t[8], t[9], t[10], t[11],
            ])
            play_id += 1
            row_num += 1

    append_plays(BACK_OFFICE_PLAYS, "Back-Office")
    append_plays(STREAMING_PLAYS, "Streaming")
    append_plays(EXPERIENCES_PLAYS, "Experiences")
    append_plays(STUDIOS_CONTENT_DEV, "Studios")
    append_plays(STUDIOS_PRODUCTION, "Studios")
    append_plays(STUDIOS_MARKETING_RIGHTS, "Studios")

    last_row = row_num - 1

    for r in range(2, last_row + 1):
        segment = ws.cell(row=r, column=2).value
        priority = ws.cell(row=r, column=14).value
        bg = segment_fill.get(segment, "FFFFFF")
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = Border(
                bottom=Side(border_style="hair", color="CCCCCC"),
                right=Side(border_style="hair", color="CCCCCC"),
            )
        pcell = ws.cell(row=r, column=14)
        pcell.font = Font(name="Arial", size=10, bold=True, color=priority_color.get(priority, "000000"))
        wcell = ws.cell(row=r, column=11)
        wcell.font = Font(name="Arial", size=10, bold=True)

    widths = {
        1: 5, 2: 13, 3: 24, 4: 38, 5: 48, 6: 48, 7: 36, 8: 36,
        9: 30, 10: 30, 11: 10, 12: 14, 13: 30, 14: 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 2 - Summary
    # ============================================================
    ws2 = wb.create_sheet("Summary")

    ws2["A1"] = "Disney Agentic Plays - Combined Summary (Company-Wide)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1A2339")
    ws2.merge_cells("A1:D1")

    ws2["A3"] = "39 curated plays across Back-Office, Streaming, Experiences, and Studios sub-business"
    ws2["A3"].font = Font(name="Arial", italic=True, color="6B7280", size=10)
    ws2.merge_cells("A3:D3")

    ws2["A5"] = "By Segment"
    ws2["A5"].font = Font(name="Arial", bold=True, size=11)

    headers2 = ["Segment", "Play Count", "Wave 1 Range ($M, low)", "Wave 1 Range ($M, high)"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=6, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Back-Office now 13 (+1 AI Consumption play at 1.3-1.9):
    #   13 plays · 14.0-22.7 envelope
    # Streaming: 10 plays · 12.0-18.0
    # Experiences: 1 play · 2.0-3.5
    # Studios:
    #   Content Dev (4): 4.6-7.0
    #   Production (5):  5.9-8.8
    #   Marketing (6):   6.7-10.2
    #   Total studios:   15 plays · 17.2-26.0
    # Combined: 39 plays · 45.2-70.2
    summary_rows = [
        ("Back-Office", 13, 14.0, 22.7),
        ("Streaming", 10, 12.0, 18.0),
        ("Experiences", 1, 2.0, 3.5),
        ("Studios", 15, 17.2, 26.0),
        ("TOTAL", 39, 45.2, 70.2),
    ]
    for i, (segment, count, low, high) in enumerate(summary_rows, start=7):
        is_total = segment == "TOTAL"
        ws2.cell(row=i, column=1, value=segment).font = Font(name="Arial", bold=is_total)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=low).number_format = "$#,##0.0"
        ws2.cell(row=i, column=4, value=high).number_format = "$#,##0.0"
        for c in range(1, 5):
            cell = ws2.cell(row=i, column=c)
            cell.font = Font(name="Arial", size=10, bold=is_total)
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right")
            if is_total:
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    # Priority mix
    ws2["A14"] = "By Priority (Account Team Sequencing)"
    ws2["A14"].font = Font(name="Arial", bold=True, size=11)
    headers3 = ["Priority", "Play Count", "Notes"]
    for col, h in enumerate(headers3, start=1):
        c = ws2.cell(row=15, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")

    # High priority count = High in BO (CTO, AI Consumption, Engineering, Continuous Close, AP, Audit, IT Service Desk, Risk = 8)
    #                    + High in Streaming (Churn, Recs, Cold-Start, Password-Sharing, ESPN, CSR = 6)
    #                    + High in Experiences (Guest Day = 1)
    #                    + High in Studios (Greenlight, Audience Science, Schedule&Budget, VFX, Trailer, Marketing Campaign, Awards = 7)
    #                    = 22
    # Medium                    = BO (5) + Streaming (4) + Studios (7) = 16
    # Low                       = BO (0) + Studios (1) = 1
    priority_rows = [
        ("High", 23, "Wave 1 candidates + strategic Wave 2 anchors - clear KPI · executive-buyer relationships · governance differentiation"),
        ("Medium", 15, "Wave 2 follow-ons - strong business case · sequenced after Wave 1 lands per segment"),
        ("Low", 1, "Wave 3+ - strategic / slow-cycle / sensitive (Studios Production Safety)"),
    ]
    for i, (pri, count, note) in enumerate(priority_rows, start=16):
        ws2.cell(row=i, column=1, value=pri).font = Font(name="Arial", size=10, bold=True, color={"High": "1B7F3A", "Medium": "C28A22", "Low": "6B7280"}[pri])
        ws2.cell(row=i, column=2, value=count).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3, value=note).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="center")

    # Top Wave 1 picks
    ws2["A21"] = "Recommended Wave 1 Entry Picks - Cross-Segment"
    ws2["A21"].font = Font(name="Arial", bold=True, size=11)

    headers4 = ["Segment / Studio", "Play", "Why"]
    for col, h in enumerate(headers4, start=1):
        c = ws2.cell(row=22, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")

    picks = [
        ("Back-Office", "Continuous-Close Agent", "CFO-owned · close-cycle compression is publicly committed · high-visibility metric"),
        ("Back-Office", "Audit Evidence Readiness Agent", "Audit-cycle is a forcing function · governance-differentiation · CCO partnership"),
        ("Back-Office", "AI Consumption Cost Intelligence Agent", "CFO-urgent · addresses +20-40% QoQ AI run-rate growth · joint CTO-CFO governance · operating cost discipline at the AI layer"),
        ("Streaming", "Password-Sharing Detection Agent", "Net-new revenue · highest strategic visibility · governance-differentiation · industry-proven"),
        ("Streaming", "Streaming Churn Prediction Agent", "DTC profitability commitment · clear KPI envelope · adjacent to ad-tier and lifecycle plays"),
        ("Streaming", "ESPN Auto-Highlight Agent", "ESPN-DTC critical-path · clear production-team value · NAB-conference visibility"),
        ("Experiences", "Guest Day Orchestration Agent", "$60B Experiences capital ROI lever · MagicBand+ data uniquely Disney's · Genie+ attach + per-guest yield"),
        ("Studios - Marvel", "Production Schedule & Budget Intelligence + VFX Pipeline", "Data-rich production estate · operational-AI receptive · VFX-heaviness makes Play 6 flagship · ILM adjacency"),
        ("Studios - Lucasfilm/ILM", "VFX Pipeline Optimisation", "ILM large enough to matter, internal enough to pilot · Star Wars production substrate"),
        ("Studios - Pixar", "Production Schedule & Budget Intelligence ONLY", "Creative culture cautious about AI; lead with operational scheduling · animation-pipeline play is Wave 2-3"),
        ("Studios - Walt Disney Pictures", "Greenlight Decision Support OR Trailer Performance", "Operational marketing/development play with clear executive buyer · avoids creative-authorship sensitivity"),
        ("Studios - Searchlight", "Awards Campaign Management", "Awards strategy structurally critical to Searchlight brand · clear buyers · highest-leverage entry"),
        ("Studios - 20th Century", "Similar to Walt Disney Pictures pattern", "Disney-system integration still working through; match a play to operational maturity"),
    ]
    for i, (seg, play, why) in enumerate(picks, start=23):
        c1 = ws2.cell(row=i, column=1, value=seg)
        c1.font = Font(name="Arial", size=10, bold=True)
        c1.fill = PatternFill("solid", fgColor="F4F4F4")
        c2 = ws2.cell(row=i, column=2, value=play)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        c3 = ws2.cell(row=i, column=3, value=why)
        c3.font = Font(name="Arial", size=10)
        c3.alignment = Alignment(wrap_text=True, vertical="center")

    # Independence + creative-authorship callout
    callout_row = 23 + len(picks) + 2
    ws2.cell(row=callout_row, column=1, value="Disney-Wide Independence Considerations").font = Font(name="Arial", bold=True, size=11, color="C28A22")
    ws2.merge_cells(start_row=callout_row, end_row=callout_row, start_column=1, end_column=4)

    callouts = [
        "Global Disney Independence framing - two-contract model, no co-sell with Microsoft, careful coordination with global Independence office.",
        "Studios creative-authorship boundary - operational and analytical augmentation, never creative replacement. WGA / SAG-AFTRA / DGA AI provisions binding for all Studios plays.",
        "Financial-reporting-adjacent scope avoidance - plays touching box office, residual calculations, talent compensation, or financial-reporting metrics require Independence review before scope finalisation.",
        "Rights and licensing scope discipline - Deloitte recommends what the client should build; the client makes substantive rights-and-licensing decisions.",
        "AI Consumption Cost play - Microsoft EA license-data and Azure cost telemetry are Disney's data; agent reads, composes, and recommends, but never makes purchase or termination decisions.",
    ]
    for j, text in enumerate(callouts, start=1):
        cell = ws2.cell(row=callout_row + j, column=1, value=text)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.merge_cells(start_row=callout_row + j, end_row=callout_row + j, start_column=1, end_column=4)
        ws2.row_dimensions[callout_row + j].height = 36

    sw = {1: 26, 2: 42, 3: 70, 4: 22}
    for col, w in sw.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 3 - By Segment (filtered views)
    # ============================================================
    ws3 = wb.create_sheet("By Segment")

    ws3["A1"] = "Disney Agentic Plays - By Segment View"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="1A2339")
    ws3.merge_cells("A1:E1")

    ws3["A3"] = (
        "This sheet groups the 39 plays by Segment for quick scanning during segment-specific "
        "account-planning sessions. For the full play detail, see the Plays sheet."
    )
    ws3["A3"].font = Font(name="Arial", italic=True, color="6B7280", size=10)
    ws3.merge_cells("A3:E3")
    ws3.row_dimensions[3].height = 30

    seg_headers = ["Segment", "#", "Play", "Wave", "Wave 1 Range ($M)", "Priority"]
    seg_row = 5
    for col, h in enumerate(seg_headers, start=1):
        c = ws3.cell(row=seg_row, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")
        c.alignment = Alignment(horizontal="left")

    ws3.freeze_panes = "A6"

    seg_row += 1
    play_id = 1
    for plays, segment_label in [
        (BACK_OFFICE_PLAYS, "Back-Office"),
        (STREAMING_PLAYS, "Streaming"),
        (EXPERIENCES_PLAYS, "Experiences"),
        (STUDIOS_CONTENT_DEV, "Studios · Content Dev"),
        (STUDIOS_PRODUCTION, "Studios · Production"),
        (STUDIOS_MARKETING_RIGHTS, "Studios · Marketing/Rights"),
    ]:
        for t in plays:
            ws3.cell(row=seg_row, column=1, value=segment_label).font = Font(name="Arial", size=10, bold=True)
            ws3.cell(row=seg_row, column=2, value=play_id).font = Font(name="Arial", size=10)
            ws3.cell(row=seg_row, column=3, value=t[0]).font = Font(name="Arial", size=10)
            ws3.cell(row=seg_row, column=4, value=t[8]).font = Font(name="Arial", size=10, bold=True)
            ws3.cell(row=seg_row, column=5, value=t[9]).font = Font(name="Arial", size=10)
            pri = t[11]
            ws3.cell(row=seg_row, column=6, value=pri).font = Font(name="Arial", size=10, bold=True, color={"High": "1B7F3A", "Medium": "C28A22", "Low": "6B7280"}[pri])
            for c in range(1, 7):
                ws3.cell(row=seg_row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            play_id += 1
            seg_row += 1

    sw3 = {1: 26, 2: 5, 3: 52, 4: 10, 5: 16, 6: 12}
    for col, w in sw3.items():
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 4 - How To Use
    # ============================================================
    ws4 = wb.create_sheet("How To Use")

    instructions = [
        ("Disney Agentic Plays - Combined - How To Use This File", "title"),
        ("", ""),
        ("Purpose", "h2"),
        ("This file is the canonical Disney company-wide play book. 39 high-confidence agentic plays across four segments: Back-Office (13), Streaming (10), Experiences (1), and Studios (15). Use this as the primary working artefact for cross-segment account planning, executive prep, and pursuit discussions.", "p"),
        ("", ""),
        ("Relationship to other files", "h2"),
        ("- Disney_Agentic_Plays_BackOffice_Streaming.xlsx - the original segment-specific file covering Back-Office, Streaming, Experiences (24 plays). Same plays as this combined file's first three segments.", "p"),
        ("- Disney_Studios_Agentic_Plays.xlsx - the Studios-only file (15 plays). Same plays as this combined file's Studios segment.", "p"),
        ("Use the segment-specific files for focused working sessions on one segment. Use this combined file for cross-segment account planning, leadership briefings, and the company-wide narrative.", "p"),
        ("", ""),
        ("Sheet 1 - Plays", "h2"),
        ("The full 39-play list. 14 columns. Filter by Segment, Wave, Priority, or Sub-domain. The Wave 1 Range column gives indicative Deloitte services range; precise envelope is set during BVA.", "p"),
        ("", ""),
        ("Sheet 2 - Summary", "h2"),
        ("Aggregated counts by segment and by priority, recommended Wave 1 entry picks (including per-studio for Studios), and Disney-wide Independence considerations callout.", "p"),
        ("", ""),
        ("Sheet 3 - By Segment", "h2"),
        ("Compact play list grouped by segment - quick scan during segment-specific working sessions.", "p"),
        ("", ""),
        ("The new AI Consumption Cost play", "h2"),
        ("Added in v2.0 of this combined file. Addresses the +20-40% QoQ growth in Disney's AI/Copilot run-rate. Joint CTO-CFO conversation. Wave 1-2 candidate. Covers GitHub Copilot, M365 Copilot, Azure OpenAI/Foundry token consumption, embedding/vector-store spend, custom-agent run-rate, and team-expensed third-party AI tools. Microsoft Cost Management for Azure OpenAI is the platform anchor.", "p"),
        ("", ""),
        ("Creative-authorship boundary - non-negotiable for Studios plays", "h2"),
        ("Every Studios play in this file respects the creative-authorship boundary. AI use that touches creative work (writing, performance, performance-likeness, dialogue replacement) is governed by the 2023 WGA and SAG-AFTRA settlements plus the DGA AI policy. The Account Team's framing must be unambiguous: operational and analytical augmentation, never creative replacement.", "p"),
        ("", ""),
        ("How to use in account planning", "h2"),
        ("- Cross-segment leadership briefing - use this combined file's Summary sheet for the company-wide picture.", "p"),
        ("- Segment-specific prep - filter Plays sheet to one Segment (e.g., Studios) and review.", "p"),
        ("- Buyer-specific prep - sort/filter by Buyer column to find plays for a specific executive role.", "p"),
        ("- Quarterly account plan - review High-priority plays and confirm which are in active discovery per segment.", "p"),
        ("- BVA construction - pull KPI Signal column to seed BVA-in-5-bullets per Sellers Handbook framework.", "p"),
        ("- Microsoft attach modelling - Microsoft Attach column drives SKU mix for Microsoft consumption forecasting.", "p"),
        ("", ""),
        ("Independence-clean usage", "h2"),
        ("This file lists Deloitte services plays. The dollar ranges are Deloitte services revenue ranges. Microsoft platforms are licensed directly by Disney from Microsoft; this file does not include Microsoft software margin. Refer to Sellers Podcast Ep 2 and Independence Cheat Sheet for full posture.", "p"),
        ("", ""),
        ("Notes on KPI signals", "h2"),
        ("KPI signals are reference-scenario estimates from the APEX framework - not Disney-specific commitments. Disney-specific numbers are produced during BVA, with Disney's own baseline measurement.", "p"),
        ("", ""),
        ("Companion podcasts", "h2"),
        ("- Disney Account Podcast (6 episodes) - company-wide framing across Back-Office, Streaming, Experiences", "p"),
        ("- Disney Studios Account Podcast (5 episodes) - Studios sub-business deep dive", "p"),
        ("- APEX-Scenario-Chains.xlsx - the master TMT scenario reference (743 scenarios incl. Disney-specific additions)", "p"),
        ("", ""),
        ("Document version", "h2"),
        ("v2.0 - 2026-05-13 - Combined Disney workbook · 39 plays · added AI Consumption Cost Intelligence Agent · curated by TMT-MED Practice + Disney Account Team", "p"),
    ]

    row = 1
    for text, style in instructions:
        cell = ws4.cell(row=row, column=1, value=text)
        if style == "title":
            cell.font = Font(name="Arial", bold=True, size=16, color="1A2339")
        elif style == "h2":
            cell.font = Font(name="Arial", bold=True, size=12, color="1A2339")
        elif style == "p":
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws4.column_dimensions["A"].width = 115

    return wb


def main():
    wb = build_workbook()
    wb.save(OUT)
    print(f"Wrote {OUT.name} ({OUT.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()

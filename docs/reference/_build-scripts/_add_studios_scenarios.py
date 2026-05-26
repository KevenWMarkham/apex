"""
Add Disney Studios-specific TMT scenarios to APEX-Scenario-Chains.xlsx.

After the prior round (which added 5 Disney/TMT scenarios — AI portfolio,
engineering productivity, subscriber-lifecycle, CSR agent-assist streaming,
guest-day orchestration), the Studios sub-business has 15 plays. Most
content-development, production, and Studios-specific marketing/rights plays
are NOT covered by existing TMT scenarios. This script adds 14 new entries.

What we add:
  Content Development & Greenlight  (4)
    - tmt-greenlight-decision-support
    - tmt-audience-science-sentiment-composition
    - tmt-ip-franchise-opportunity-scoring
    - tmt-talent-casting-intelligence
  Production  (5)
    - tmt-production-schedule-budget-intelligence
    - tmt-vfx-pipeline-coordination
    - tmt-animation-pipeline-multi-year
    - tmt-post-production-dependency-chain
    - tmt-production-safety-pattern-detection
  Marketing, Distribution & Rights  (5)
    - tmt-trailer-performance-composition
    - tmt-windowing-decision-substrate
    - tmt-awards-campaign-coordination
    - tmt-rights-library-reuse
    - tmt-music-sync-clearance

Service codes: introduces TMT-MED-02 (Studio Production & Content) for studio-
specific operational plays. Marketing/awards/trailer plays use existing
TMT-MED-03. Audience-science play uses existing TMT-MED-01.

Plays NOT added because they are already covered by existing TMT scenarios:
  - Marketing Campaign Optimisation → existing tmt-programmatic-bid-optimization
    + tmt-creative-performance-prediction cover the cross-channel and
    creative-attribution dimensions.

Sheets updated:
  1. Scenario Library     — adds 14 rows
  2. Scenario→KPI Chain   — adds 14 rows
  3. Summary              — updates total + version note

Usage:
    python _add_studios_scenarios.py
"""

from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(__file__).parent / "APEX-Scenario-Chains.xlsx"

# ---------------------------------------------------------------------------
# Studios scenarios — each defined once with both Scenario Library + Scenario->KPI
# Chain shapes filled in. The chain entries reference the play's Studios buyer
# and Studios-specific Independence considerations.
# ---------------------------------------------------------------------------

# Each scenario: (
#   scenario_id, title, service_code, domain,
#   brief, kpi,
#   moment, solution, use_cases, service, personas
# )

STUDIOS_SCENARIOS = [
    # ---------- Content Development & Greenlight ----------
    (
        "tmt-greenlight-decision-support",
        "Greenlight decision support (executive committee substrate)",
        "TMT-MED-02", "Marketing & Growth",
        "Composes per-project decision substrate across comparable-title performance, talent-availability, production-cost benchmark, marketing-spend comparable, awards positioning, genre trends, international market readiness. Agent surfaces; executive committee decides. Tentpole-greenlight cycle of weeks-to-months.",
        "Greenlight cycle time −30 to −50% on tentpole projects · decision-quality variance reduction across 12+ stakeholders · post-greenlight surprise reduction",
        "Studios executive committee (12+ stakeholders) must align on greenlight decision for a tentpole project. Cycle runs weeks-to-months. Each stakeholder forms a view based on partial data; debate is partly about reconciling differing interpretations, not just creative-strategic judgement. Errors at greenlight propagate downstream as budget overruns, oversized marketing spend, talent-relationship strain.",
        "Greenlight-substrate agent composes data across comparable-title performance (theatrical/streaming/international), talent calendar, production-cost benchmark, marketing-spend comparable, awards positioning, genre trends, international market readiness. Surfaces unified per-project view to the committee. Agent does NOT recommend greenlight; committee decides. HITL is the executive committee.",
        "Tentpole greenlight decision support · comparable-title performance composition · talent-availability calendar composition · awards-positioning analysis · international market readiness · post-greenlight surprise reduction.",
        "TMT-MED-02 Studio Production & Content (new service code). Wave 1 candidate at studios with strongest executive-relationship. Commercial envelope $1.5-2.2M.",
        "Primary: Studio President · Chief Creative Officer · Head of Business Affairs. Approver: Studios Executive Committee. Critical Independence note: agent operationally composes data; creative-and-strategic decision is the committee's.",
    ),
    (
        "tmt-audience-science-sentiment-composition",
        "Audience-science composition (preview · trailer-test · social signal)",
        "TMT-MED-01", "Customer Experience",
        "Composes audience-research streams (preview screenings 12hr cadence, trailer tests 48hr, social sentiment days-cadence). Surfaces composite sentiment view to creative team. Director, editor, producer decide cut and marketing adjustments.",
        "Time-to-composed-audience-insight compression −60 to −70% · decision-velocity on cut/marketing adjustments · creative-team feedback cycle accelerated",
        "Audience-science data arrives in batches at different cadences (12hr preview screening, 48hr trailer test, days for social sentiment). Decisions on cut and marketing have to happen on compressed cadence. Composition of multiple streams is manual today; the creative team works with whichever data has landed.",
        "Audience-science composition agent reads research streams continuously, composes per-project sentiment-and-response view, recommends focal areas for creative review. Creative team (director, editor, producer) decides what to change. Agent surfaces signal; humans interpret and act.",
        "Preview-screening result composition · trailer-test result composition · social-sentiment monitoring · per-project audience-response composite view · cross-stream signal correlation.",
        "TMT-MED-01 Customer Experience (audience-research specialisation). Wave 1 candidate. Commercial envelope $1.2-1.8M.",
        "Primary: CMO Studios · Head of Audience Science · Studio President. Approver: Creative team (director, editor, producer) — agent surfaces; humans decide on cut and marketing changes.",
    ),
    (
        "tmt-ip-franchise-opportunity-scoring",
        "IP & franchise opportunity scoring (continuous portfolio view)",
        "TMT-MED-02", "Marketing & Growth",
        "Reads continuously across IP-performance signal (theatrical, streaming, consumer products), audience affinity, social signal, competitive landscape, talent-attachment availability. Surfaces emerging franchise-extension opportunities. Development leadership reviews and prioritises.",
        "Underutilised-IP-identification rate · time-to-opportunity-surfaced · development-pipeline coverage of high-opportunity IP · long-term hit rate of agent-surfaced opportunities",
        "Studios IP portfolio is enormous (multiple franchise libraries, character catalogs, story-world properties). Identifying highest-opportunity IP for next-cycle development is largely judgement-based; portfolio view is not continuously composed; opportunities surface episodically through senior leadership intuition.",
        "IP-opportunity-scoring agent reads continuously across IP-performance signal (theatrical, streaming, consumer products, theme-park integration), audience-affinity, social signal, competitive landscape, talent-attachment-availability. Surfaces emerging signals — underutilised IP gaining momentum, characters whose franchise-extension potential is rising. Development leadership reviews and prioritises; agent does NOT pick what to develop.",
        "IP-portfolio coverage analytics · franchise-extension opportunity surfacing · cross-segment IP performance composition · audience-affinity-driven opportunity scoring · competitive-landscape positioning.",
        "TMT-MED-02 Studio Production & Content. Wave 2 placement. Commercial envelope $1.0-1.6M.",
        "Primary: Chief Creative Officer · Head of IP Strategy · Studio President. Approver: Development leadership (decides what enters active development).",
    ),
    (
        "tmt-talent-casting-intelligence",
        "Talent & casting intelligence (composition only — NO performance generation)",
        "TMT-MED-02", "Marketing & Growth",
        "Composes casting-decision data: comparable-performance, audience-affinity scores, awards-track-record, compensation comparables, calendar availability, existing commitments. Casting director, director, studio executive review and decide. Agent does NOT generate performances, does NOT use likenesses without consent, does NOT displace casting directors.",
        "Casting decision cycle compression · coverage of full candidate-set across roles · casting-decision stakeholder alignment time reduction",
        "Composing implications of a casting choice is manual today. Comparable-performance data, audience-affinity scores, awards track-record, compensation comparables, calendar availability, existing commitments — all exist but are composed by head of casting plus studio executive in conversation. Post-strike talent-and-AI policy makes the framing critical.",
        "Casting-intelligence agent composes dimensions for any candidate talent or candidate ensemble. Surfaces composite-fit view for role and production economics. Casting director, director, studio executive review and decide. Agent does NOT generate or simulate performances, does NOT use likenesses without explicit talent consent, does NOT displace casting directors. Operational composition only.",
        "Comparable-performance data composition · audience-affinity composition · talent-availability calendar · compensation comparables · awards-track-record · ensemble-fit composition.",
        "TMT-MED-02 Studio Production & Content. Wave 2 placement (after operational-AI posture is established). Commercial envelope $0.9-1.4M. WGA/SAG-AFTRA/DGA AI provisions binding.",
        "Primary: Head of Casting · Studio President · Director (per project). Approver: Casting director (creative casting judgement stays with casting director and studio executive).",
    ),
    # ---------- Production ----------
    (
        "tmt-production-schedule-budget-intelligence",
        "Production schedule & budget intelligence (tentpole-scale)",
        "TMT-MED-02", "Operations & Workforce",
        "Reads production accounting, daily call-sheet completion, scene-coverage progress, vendor-invoice signal, post-production dependency signal. Continuously composes budget-and-schedule trajectory. Identifies emerging variance patterns before they accumulate. Surfaces to line producer, production manager, studio executive with confidence-scored projections.",
        "Time-to-detect emerging variance: weekly → daily · variance-recovery rate improvement · budget-and-schedule integrity across annual slate",
        "Production budgets for tentpoles run $150-300M. Schedule variance drives budget variance — every over-schedule day cascades through cast, crew, facility rental, equipment, post-production scheduling. Production accounting and schedule systems are separate; composition into credible early-warning signal is weekly-cadence human work, but the variance compounds daily.",
        "Schedule-and-budget-intelligence agent reads production accounting, daily call-sheet completion, scene-coverage progress, vendor-invoice signal, post-production dependency signal. Continuously composes the trajectory. Identifies emerging variance patterns before they accumulate into multi-week recoveries. Surfaces to line producer, production manager, studio executive with confidence-scored projections. Decisions stay human.",
        "Production accounting integration · call-sheet completion analytics · scene-coverage progress monitoring · vendor-invoice signal composition · post-production dependency surfacing · variance-recovery-rate optimisation.",
        "TMT-MED-02 Studio Production & Content. Wave 1 candidate — especially strong at Marvel and live-action operations-mature studios. Commercial envelope $1.4-2.0M.",
        "Primary: President of Physical Production · Studio CFO · Senior Line Producer. Approver: line-producer and studio leadership decide on interventions (additional shoot days, scope adjustments, VFX-shot reallocation).",
    ),
    (
        "tmt-vfx-pipeline-coordination",
        "VFX pipeline coordination (cross-vendor shot-status composition)",
        "TMT-MED-02", "Operations & Workforce",
        "Reads continuously across vendor production-tracking systems, shot-version repositories, review-feedback systems, delivery-schedule trackers. Identifies emerging risk patterns — vendor delivery slippage, shot iteration count exceeding plan, reviewer-feedback patterns indicating creative reconciliation. Surfaces to VFX supervisor and post-production producer.",
        "VFX-delivery-risk detection time compression · final-delivery slippage rate reduction · VFX-budget-variance reduction · expedited late-stage revision rate decrease",
        "VFX is the largest production-cost line item on tentpoles. Productions can carry 2000-3000 VFX shots distributed across multiple vendor studios globally. Vendor coordination is high-variance and high-stakes. Shot status flows through vendor tracking + emails + weekly review meetings; composition across vendors is manual. Late-stage shot revisions incur expedited cost.",
        "VFX-pipeline agent reads continuously across vendor production-tracking systems, shot-version repositories, review-feedback systems, delivery-schedule trackers. Identifies emerging risk patterns. Surfaces composite VFX-pipeline-health view to VFX supervisor and post-production producer. HITL at the VFX supervisor / VFX producer.",
        "Cross-vendor VFX shot tracking · shot-version composition · reviewer-feedback pattern detection · delivery-schedule risk surfacing · vendor-performance composite analytics · ILM-internal pipeline operations.",
        "TMT-MED-02 Studio Production & Content. Wave 1 candidate at VFX-heavy studios (Marvel, Lucasfilm/ILM). Commercial envelope $1.5-2.2M.",
        "Primary: VFX Supervisor · VFX Producer · Post-Production Producer · ILM Operations (for Lucasfilm). Approver: VFX leadership reviews surfaced risks and acts on vendor coordination.",
    ),
    (
        "tmt-animation-pipeline-multi-year",
        "Long-cycle animation pipeline (4-5 year production-cycle composition)",
        "TMT-MED-02", "Operations & Workforce",
        "Reads continuously across story-revision tracking, design-asset versioning, animation-shot status, render-farm queue, sound-pipeline status. Composes multi-year pipeline view in continuous-update form. Surfaces emerging risks. Critically — agent does NOT replace creative judgement; operational substrate only.",
        "Pipeline-cycle-time variance reduction · render-capacity utilisation optimisation · earlier surface of story-revision cascade risk (most expensive form of late-stage rework)",
        "Animation features take 4-5 years from initial story development to theatrical release. Pipeline involves story development, design, animation, lighting and rendering, sound, final delivery. Long-cycle coordination produces drift. Story revisions cascade into design revisions cascade into animation reworks. Render-farm capacity is finite; render time per shot compounds.",
        "Animation-pipeline agent reads continuously across story-revision tracking, design-asset versioning, animation-shot status, render-farm queue, sound-pipeline status. Composes multi-year pipeline view in continuous-update form. Surfaces emerging risks to production producer, director, CCO. Critically: agent does NOT replace creative judgement; director's call on story revision stays director's call. Operational substrate only.",
        "Story-revision cascade detection · design-asset versioning analytics · render-farm queue optimisation · sound-pipeline status composition · multi-year pipeline-trajectory view · long-cycle drift detection.",
        "TMT-MED-02 Studio Production & Content. Wave 2-3 placement at animation studios (Pixar). Commercial envelope $1.3-1.9M. Cultural caution: animation studios treat AI carefully; lead with operational scheduling, bring animation-pipeline play later.",
        "Primary: Animation Studio President · Chief Creative Officer (animation) · Head of Production (animation). Approver: Director (story-revision authority); render-farm operations team (capacity allocation).",
    ),
    (
        "tmt-post-production-dependency-chain",
        "Post-production dependency-chain composition",
        "TMT-MED-02", "Operations & Workforce",
        "Reads continuously across editing milestones, sound milestones, VFX delivery, color-and-finishing milestones, delivery preparation. Identifies dependency-cascade risk. Surfaces to post-production supervisor with timeline-adjustment recommendations.",
        "Dependency-cascade detection time compression · post-production schedule integrity · final-delivery-on-schedule rate improvement",
        "Post-production is editing, sound, color, music, VFX integration, delivery preparation. Dependency chain is long and schedule is tight. Editing locks must precede sound finalisation; sound must precede VFX integration; VFX before final color; color before delivery preparation. Late delivery of one element cascades to the schedule for every dependent element.",
        "Post-production-workflow agent reads continuously across editing milestones, sound milestones, VFX delivery, color-and-finishing milestones, delivery preparation. Identifies dependency-cascade risk. Surfaces composite view to post-production supervisor with timeline-adjustment recommendations.",
        "Editing-milestone tracking · sound-finalisation dependency · VFX-integration scheduling · color-and-finishing composition · delivery-preparation scheduling · cross-element dependency-cascade risk surfacing.",
        "TMT-MED-02 Studio Production & Content. Wave 2 placement. Commercial envelope $0.9-1.4M.",
        "Primary: Post-Production Supervisor · Post-Production Producer. Approver: Post-production supervisor decides on timeline adjustments.",
    ),
    (
        "tmt-production-safety-pattern-detection",
        "Production safety & wellness pattern detection (anonymised)",
        "TMT-MED-02", "Quality & Compliance",
        "Reads safety-incident reporting, crew-sentiment signals (where opt-in), set-incident logs, protocol-compliance signal. Surfaces emerging patterns to studio safety leadership. Anonymised pattern signal at production-level and protocol-level. NEVER surveils individuals.",
        "Time-to-pattern-detection across productions · safety-incident-rate trends · crew-wellness signal · production-environment-quality metric (long-term)",
        "Safety-incident-pattern detection across multiple concurrent productions is partial today. Incidents at one production may indicate systemic patterns (vendor practices, scheduling pressure, specific stunt-or-equipment types) that are not surfaced until pattern matures. Insurance economics, talent comfort, regulatory compliance — all rest on safety culture.",
        "Production-safety agent reads safety-incident reporting, crew-sentiment signals (where opt-in), set-incident logs, protocol-compliance signal. Surfaces emerging patterns to studio safety leadership and responsible-production producers. Agent does NOT surveil individuals; composes anonymised pattern signal at production-level and protocol-level. Studio safety leadership reviews and acts.",
        "Safety-incident reporting composition · cross-production pattern detection · crew-wellness anonymised signal · protocol-compliance monitoring · production-environment-quality tracking · IATSE-aligned governance.",
        "TMT-MED-02 Studio Production & Content. Wave 3 placement (slow-cycle, sensitive). Commercial envelope $0.8-1.3M. Strict Purview PII governance.",
        "Primary: Studio President · President of Physical Production · Head of Safety. Approver: Studio safety leadership reviews and acts; HITL on every escalation. IATSE alignment.",
    ),
    # ---------- Marketing, Distribution & Rights ----------
    (
        "tmt-trailer-performance-composition",
        "Trailer-performance composition (vendor results + social + competitive)",
        "TMT-MED-03", "Marketing & Growth",
        "Composes test-vendor results, social-signal once trailers launch, comparable-trailer-performance benchmarks, competitive-landscape positioning. Surfaces composed view to marketing executive within hours of results landing. Marketing leadership decides on cut iteration and placement strategy.",
        "Time-to-composed-trailer-insight compression 70-80% · decision-velocity on cut iteration · trailer-effectiveness improvement (opening-weekend correlation)",
        "Trailer-test results arrive in 48-hour batches from a research vendor. By the time results are reviewed, trailer-house has often started next-cut iteration. Composition of trailer-test results with social-signal, competitive-trailer-performance, and audience-segment context is human work today. Trailer effectiveness shapes opening weekend.",
        "Trailer-performance agent composes test-vendor results, social-signal once trailers launch, comparable-trailer-performance benchmarks, competitive-landscape positioning. Surfaces composed view to marketing executive within hours of results landing. Marketing leadership decides on cut iteration and placement strategy.",
        "Trailer-test result composition · post-launch social-signal monitoring · comparable-trailer benchmarking · competitive-landscape positioning · cut-iteration decision support · trailer-placement strategy.",
        "TMT-MED-03 Marketing/Advertising. Wave 1 candidate. Commercial envelope $1.1-1.7M.",
        "Primary: CMO Studios · Head of Worldwide Marketing · Studio Marketing President. Approver: Marketing leadership decides on cut iteration; trailer-house executes.",
    ),
    (
        "tmt-windowing-decision-substrate",
        "Theatrical-vs-streaming windowing decision substrate (executive committee)",
        "TMT-MED-02", "Marketing & Growth",
        "Composes factors across title (genre, audience, budget, talent), market (competitive landscape, theatrical environment), strategic (DTC subscriber-acquisition value), talent-contract dimensions. Surfaces composite-decision view to windowing committee. Committee decides; agent does NOT recommend.",
        "Windowing-decision cycle compression · decision-quality variance reduction across the windowing committee · post-decision execution quality (marketing/distribution alignment)",
        "Windowing decisions have become strategic, not procedural. Theatrical-only vs theatrical-plus-streaming-30-days vs streaming-direct reshape revenue, marketing economics, talent relationships (theatrical-vs-streaming compensation provisions). Each decision per title is an executive-committee discussion. Composing factors is largely manual.",
        "Windowing-decision-support agent composes the factors across title and market dimensions. Surfaces composite-decision view to windowing committee. Committee decides; agent does NOT recommend the decision. Critical: agent doesn't pretend to know the strategic context the committee brings.",
        "Title-factor composition · market-factor composition · DTC subscriber-acquisition-value modelling · talent-contract windowing analysis · executive-committee decision substrate · post-decision marketing/distribution alignment.",
        "TMT-MED-02 Studio Production & Content (strategic specialisation). Wave 3 placement. Commercial envelope $1.6-2.4M. NOTE: distinct from existing tmt-release-window-optimization which is for general content scheduling; this scenario is for tentpole-level executive-committee strategic windowing.",
        "Primary: Entertainment Chairperson · Studio Presidents · Head of Worldwide Distribution · CFO Entertainment (executive committee). Approver: Executive committee decides; agent does NOT recommend the windowing.",
    ),
    (
        "tmt-awards-campaign-coordination",
        "Awards-campaign coordination (Oscar/Emmy operational substrate)",
        "TMT-MED-03", "Marketing & Growth",
        "Reads campaign-event calendar, eligibility-tracking, talent availability, FYC-advertising performance, social-and-industry sentiment, competitive-campaign signal. Composes campaign-status view continuously. Awards-team approves campaign-decision recommendations.",
        "Awards-campaign coordination velocity · nomination-and-win conversion rate · specialty-theatrical viability (per-title theatrical extension months)",
        "Awards campaign management is year-round operational work, peaking during Oscar and Emmy seasons. Specialty-theatrical studios' business depends on awards-season performance for theatrical viability. Campaigns coordinate academy screenings, talent press, FYC advertising, industry-relationship management, festival positioning, eligibility tracking. Tight calendar windows; many touchpoints.",
        "Awards-campaign agent reads campaign-event calendar, eligibility-tracking systems, talent-availability, FYC-advertising performance, social-and-industry sentiment, competitive-campaign signal. Composes campaign-status view continuously. Awards-team approves campaign-decision recommendations.",
        "Academy-screening event coordination · FYC-advertising performance tracking · talent-availability for awards events · industry-relationship management · eligibility-deadline tracking · competitive-campaign monitoring.",
        "TMT-MED-03 Marketing. Wave 2 placement. Commercial envelope $1.1-1.6M. Natural Wave 2 entry at specialty/awards-focused studios (e.g., Searchlight-like portfolios).",
        "Primary: Specialty Studio President · Awards Campaign Director · CMO Specialty. Approver: Awards-team approves campaign-decision recommendations.",
    ),
    (
        "tmt-rights-library-reuse",
        "Rights-compliance & library-reuse composition",
        "TMT-MED-02", "Quality & Compliance",
        "Reads continuously across rights-management systems, talent-contract repositories, music-licensing systems, footage-rights databases. Maintains continuous rights-status view of the library. Surfaces rights-availability for candidate reuse projects. PII propagation discipline critical.",
        "Rights-clearance cycle compression 50-70% · reuse-opportunity surfaced rate · compliance-risk reduction (earlier detection of expired/restricted-use rights)",
        "Studios with large content libraries reuse archive content for trailers, marketing, behind-the-scenes, anniversary releases, theme-park integrations. Each reuse is a clearance project — music rights, talent rights, footage rights, image rights. Work is labour-intensive and slow; business-affairs and legal handle clearances title-by-title. Rights status across the library is partial.",
        "Rights-compliance agent reads continuously across rights-management systems, talent-contract repositories, music-licensing systems, footage-rights databases. Maintains continuous rights-status view of the library. Surfaces rights-availability for any candidate reuse project. PII propagation discipline is critical for talent-contract data.",
        "Library-rights-status composition · candidate-reuse rights surfacing · talent-contract repository integration · music-licensing system integration · footage-rights database integration · expired-or-restricted-use detection.",
        "TMT-MED-02 Studio Production & Content (rights specialisation). Wave 2 placement. Commercial envelope $1.0-1.5M. Distinct from existing tmt-licensing-rights-expiry-tracking which is contract-expiry-focused; this is library-wide reuse-availability.",
        "Primary: General Counsel Studios · Head of Business Affairs · Chief Rights Officer. Approver: Business-affairs and legal review surfaced rights status; agent does NOT advise on substantive rights-and-licensing decisions.",
    ),
    (
        "tmt-music-sync-clearance",
        "Music synchronisation clearance composition",
        "TMT-MED-02", "Quality & Compliance",
        "Reads continuously across music-publisher and record-label clearance systems, internal-licensing trackers, clearance-history databases. Surfaces clearance-availability and indicative pricing for candidate-cue use. Music supervisor decides; business-affairs approves.",
        "Music-clearance cycle compression · music-supervisor productivity · reduced clearance-related schedule risk on tight-deadline projects (trailers, marketing materials)",
        "Music sync clearance is high-volume. Every musical cue in a film or trailer requires clearance from the music publisher (composition) and the record label (recording). Soundtrack assembly, trailer scoring, marketing-content scoring all require sync clearance. Cycle from 'we want this song' to 'cleared and licensed' can run weeks. Affects creative-and-marketing flexibility.",
        "Music-sync-clearance agent reads continuously across music-publisher and record-label clearance systems, internal-licensing trackers, clearance-history databases. Surfaces clearance-availability and indicative pricing for candidate-cue use. Music supervisor decides; business-affairs approves.",
        "Music-publisher clearance integration · record-label clearance integration · internal-licensing-tracker composition · clearance-history pattern recognition · indicative-pricing surfacing · trailer-and-marketing tight-deadline support.",
        "TMT-MED-02 Studio Production & Content (rights specialisation). Wave 2 placement. Commercial envelope $0.7-1.2M.",
        "Primary: Music Supervisor · Head of Music Business Affairs · GC Studios. Approver: Music supervisor decides on candidate cues; business-affairs approves clearance.",
    ),
]


def main() -> None:
    print(f"Loading {XLSX.name}...")
    wb = load_workbook(XLSX)

    # ============================================================
    # Sheet: Scenario Library — append rows
    # ============================================================
    ws_lib = wb["Scenario Library"]
    next_lib_idx = ws_lib.max_row  # last row is the last scenario; we increment from there
    last_num = ws_lib.cell(row=ws_lib.max_row, column=1).value
    if not isinstance(last_num, int):
        # find last numeric
        for r in range(ws_lib.max_row, 1, -1):
            v = ws_lib.cell(row=r, column=1).value
            if isinstance(v, int):
                last_num = v
                break

    print(f"Last scenario # in library: {last_num}; appending {len(STUDIOS_SCENARIOS)} new...")

    for offset, sc in enumerate(STUDIOS_SCENARIOS, start=1):
        new_num = last_num + offset
        scenario_id, title, service_code, domain, brief, kpi, *_ = sc
        ws_lib.append([
            new_num,
            scenario_id,
            title,
            service_code,
            domain,
            None,           # Schemas
            brief,
            kpi,
            "Catalog",      # Featured?
        ])
        print(f"  + lib row {new_num}: {scenario_id}")

    # ============================================================
    # Sheet: Scenario→KPI Chain — append rows
    # ============================================================
    ws_chain = wb["Scenario→KPI Chain"]
    for sc in STUDIOS_SCENARIOS:
        (scenario_id, title, service_code, domain, brief, kpi,
         moment, solution, use_cases, service, personas) = sc
        ws_chain.append([
            scenario_id,
            moment,
            solution,
            use_cases,
            service,
            personas,
            kpi,
            "Catalog",
        ])
        print(f"  + chain row: {scenario_id}")

    # ============================================================
    # Sheet: Summary — update totals + version note
    # ============================================================
    ws_sum = wb["Summary"]
    new_total = last_num + len(STUDIOS_SCENARIOS)
    new_catalog = new_total - 36  # featured count unchanged
    for r in range(1, ws_sum.max_row + 1):
        v = ws_sum.cell(row=r, column=1).value
        if v == "Total scenarios":
            ws_sum.cell(row=r, column=2, value=str(new_total))
        elif v == "Catalog scenarios (compact)":
            ws_sum.cell(row=r, column=2, value=str(new_catalog))
        elif v == "Scenario Library":
            ws_sum.cell(row=r, column=3, value=str(new_total))
        elif v == "Scenario→KPI Chain":
            ws_sum.cell(row=r, column=3, value=str(new_total))
        elif v == "Document version":
            ws_sum.cell(row=r, column=2, value="v1.3 · 2026-05-13 · added 14 Disney Studios TMT scenarios")
        elif v == "Services in catalog":
            # We add TMT-MED-02 as a new service code
            ws_sum.cell(row=r, column=2, value="39 unique service IDs")

    out_path = XLSX
    wb.save(out_path)
    print(f"\nWrote {out_path.name}")
    print(f"  New scenario total: {new_total}")
    print(f"  Added {len(STUDIOS_SCENARIOS)} Studios-specific TMT scenarios")
    print(f"  New service code: TMT-MED-02 Studio Production & Content")


if __name__ == "__main__":
    main()

"""
Add the missing Studios scenario to APEX-Scenario-Chains.xlsx — closing the
final gap from the 15-play Studios curated list.

PRIOR ROUNDS:
  v1.3 (14 Studios scenarios added) — covered 14 of the 15 Studios plays.
  v1.4 (AI Consumption Cost added) — Disney-wide, not Studios-specific.

REMAINING GAP:
  Studios Marketing Campaign Optimisation Agent (play #11 of 15)

  Previously rationalised as covered by existing TMT scenarios — but on
  re-examination those existing scenarios cover different motions:
    - tmt-programmatic-bid-optimization → real-time ad bid optimisation
    - tmt-creative-performance-prediction → ad creative effectiveness modelling
    - rc-cross-channel-identity-resolution → identity resolution (RC, not TMT)

  The Studios play is *within-campaign budget reallocation across broadcast,
  digital, social, programmatic, OOH, influencer, partnership, talent press,
  premiere, and in-cinema during a theatrical tentpole launch cycle.* That
  motion is structurally distinct from real-time programmatic bidding and from
  creative-effectiveness prediction. It deserves its own scenario.

SCENARIO ADDED (1):
  tmt-studios-cross-channel-campaign-reallocation

Service code: TMT-MED-03 (Marketing / Advertising — aligned with the existing
trailer-performance and awards-campaign scenarios from v1.3).

Sheets updated:
  1. Scenario Library     — adds 1 row (#745)
  2. Scenario→KPI Chain   — adds 1 row
  3. Summary              — updates total + version note

Usage:
    python _add_studios_campaign_gap.py
"""

from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(__file__).parent / "APEX-Scenario-Chains.xlsx"

SCENARIO = {
    "id": "tmt-studios-cross-channel-campaign-reallocation",
    "title": "Studios cross-channel marketing-campaign within-cycle reallocation",
    "service_code": "TMT-MED-03",
    "domain": "Marketing & Growth",
    "brief": (
        "Continuous within-campaign reallocation agent for theatrical-tentpole marketing. "
        "Reads across channel-performance signal (broadcast, digital, social, programmatic, OOH, "
        "influencer, partnership, talent press, premiere, in-cinema), audience-segment response, "
        "competitive landscape, and creative-effectiveness data. Recommends within-campaign "
        "spend shifts during the 6-8 week tentpole launch cycle. Marketing-operations team "
        "reviews and approves. Distinct from real-time programmatic bid optimisation and "
        "ad creative-effectiveness prediction."
    ),
    "kpi": (
        "Within-campaign reallocation velocity (days → hours) · marketing ROI improvement 8-15% "
        "on covered campaigns · cross-channel attribution maturity · marketing-ops capacity for "
        "higher-leverage work · post-launch campaign-effectiveness composite improvement"
    ),
    "moment": (
        "Theatrical-tentpole marketing campaigns run across 8-10 channels over a 6-8 week launch "
        "cycle on budgets that can rival or exceed production cost. Per-channel attribution is "
        "partial; cross-channel optimisation is more partial. Budgets allocated at campaign-launch "
        "frequently are not reallocated based on performance signal until late in the cycle — "
        "after the moment when reallocation would have moved opening-weekend outcome."
    ),
    "solution": (
        "Cross-channel campaign-reallocation agent reads continuously across channel-performance "
        "signal, audience-segment response patterns, competitive-campaign landscape, and "
        "creative-effectiveness signal. Models cross-channel attribution with confidence intervals. "
        "Recommends within-campaign spend shifts (e.g., move $4M from broadcast week 4 to social "
        "+ programmatic). Marketing-operations team reviews recommendations and approves spend "
        "shifts. HITL gate above reallocation-magnitude threshold (e.g., >5% of campaign budget). "
        "Agent does NOT replace the CMO's strategic judgement on campaign positioning."
    ),
    "use_cases": (
        "Tentpole-launch cross-channel campaign optimisation · within-campaign budget reallocation "
        "decision support · per-channel performance composite analytics · audience-segment "
        "response composition · competitive-campaign landscape monitoring · creative-effectiveness "
        "feedback loop into media allocation · post-launch campaign-effectiveness composite for "
        "future-campaign learning."
    ),
    "service": (
        "TMT-MED-03 Marketing / Advertising. Wave 1 candidate at studios with operational-marketing "
        "discipline (Walt Disney Pictures, Marvel for marketing-side, 20th Century). Commercial "
        "envelope $1.2-1.8M services revenue. Companion to tmt-trailer-performance-composition "
        "(also TMT-MED-03 / Wave 1) — the two together form the Wave 1 marketing-operations entry "
        "at studios."
    ),
    "personas": (
        "Primary: Chief Marketing Officer · Head of Worldwide Marketing · Studio Marketing "
        "President · CMO Operations. Approver: Marketing-operations team approves within-campaign "
        "spend shifts; agent escalates to CMO above reallocation-magnitude threshold. Strategic "
        "positioning stays with the CMO; agent operates on the budget-reallocation tactical layer."
    ),
}


def main() -> None:
    print(f"Loading {XLSX.name}...")
    wb = load_workbook(XLSX)

    # ============================================================
    # Sheet: Scenario Library
    # ============================================================
    ws_lib = wb["Scenario Library"]
    last_num = ws_lib.cell(row=ws_lib.max_row, column=1).value
    if not isinstance(last_num, int):
        for r in range(ws_lib.max_row, 1, -1):
            v = ws_lib.cell(row=r, column=1).value
            if isinstance(v, int):
                last_num = v
                break

    # Idempotence guard — bail if the scenario already exists
    for row in ws_lib.iter_rows(min_row=2, values_only=True):
        if row[1] == SCENARIO["id"]:
            print(f"  ! scenario {SCENARIO['id']} already exists at row #{row[0]}. No-op.")
            return

    new_num = last_num + 1
    print(f"Last scenario # in library: {last_num}; appending scenario #{new_num}...")

    ws_lib.append([
        new_num,
        SCENARIO["id"],
        SCENARIO["title"],
        SCENARIO["service_code"],
        SCENARIO["domain"],
        None,
        SCENARIO["brief"],
        SCENARIO["kpi"],
        "Catalog",
    ])
    print(f"  + lib row {new_num}: {SCENARIO['id']}")

    # ============================================================
    # Sheet: Scenario->KPI Chain
    # ============================================================
    ws_chain = wb["Scenario→KPI Chain"]
    ws_chain.append([
        SCENARIO["id"],
        SCENARIO["moment"],
        SCENARIO["solution"],
        SCENARIO["use_cases"],
        SCENARIO["service"],
        SCENARIO["personas"],
        SCENARIO["kpi"],
        "Catalog",
    ])
    print(f"  + chain row: {SCENARIO['id']}")

    # ============================================================
    # Sheet: Summary
    # ============================================================
    ws_sum = wb["Summary"]
    new_total = new_num
    new_catalog = new_total - 36
    for r in range(1, ws_sum.max_row + 1):
        v = ws_sum.cell(row=r, column=1).value
        if v == "Total scenarios":
            ws_sum.cell(row=r, column=2, value=str(new_total))
        elif v == "Catalog scenarios (compact)":
            ws_sum.cell(row=r, column=2, value=str(new_catalog))
        elif v == "Scenario Library":
            ws_sum.cell(row=r, column=3, value=str(new_total))
        elif v == "Scenario→KPI Chain":
            ws_sum.cell(row=r, column=3, value=str(new_total))
        elif v == "Document version":
            ws_sum.cell(row=r, column=2,
                value="v1.5 · 2026-05-13 · closed final Studios gap (cross-channel campaign reallocation) · 15 of 15 Studios plays now covered")

    wb.save(XLSX)
    print(f"\nWrote {XLSX.name}")
    print(f"  New scenario total: {new_total}")
    print(f"  Studios play coverage: 15 of 15 ✓")


if __name__ == "__main__":
    main()

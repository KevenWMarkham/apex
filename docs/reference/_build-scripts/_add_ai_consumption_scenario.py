"""
Add the AI Consumption Cost Intelligence scenario to APEX-Scenario-Chains.xlsx.

This is the FinOps-for-AI scenario that anchors the new Disney play. It is
operationally distinct from:

  - tmt-ai-portfolio-rationalization (CTO strategic portfolio view across
    all tech investments)
  - tmt-engineering-productivity-platform (engineering productivity +
    workforce composition decisions)

This scenario is specifically about the run-rate cost discipline at the AI
consumption layer — Copilot seat reclamation, model-selection optimisation,
custom-agent cost tracking, third-party AI tool spend visibility.

Sheets updated:
  1. Scenario Library     — adds 1 row (scenario #744)
  2. Scenario→KPI Chain   — adds 1 row
  3. Summary              — updates total + version note

Usage:
    python _add_ai_consumption_scenario.py
"""

from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(__file__).parent / "APEX-Scenario-Chains.xlsx"

SCENARIO = {
    "id": "tmt-ai-consumption-cost-intelligence",
    "title": "AI consumption cost intelligence (engineering + AI-assist FinOps)",
    "service_code": "TMT-TEC-04",
    "domain": "Engineering & R&D",
    "brief": (
        "Continuous composition of AI/Copilot consumption signal across Microsoft EA Copilot seat "
        "utilisation, Azure OpenAI / Foundry token consumption by model, GitHub Copilot per-developer "
        "activity, embedding/vector-store spend, custom-agent run cost, Power Platform AI Builder, "
        "and team-expensed third-party AI tools (Cursor, Cline, Continue, Bedrock, Anthropic). "
        "Surfaces per-team, per-product, per-use-case, per-model cost. Recommends optimisations."
    ),
    "kpi": (
        "AI consumption cost growth: from +20-40% QoQ trajectory to flat-or-managed · "
        "idle Copilot seat reclamation rate · model-selection optimisation savings 12-25% "
        "on covered workloads · per-use-case cost-per-outcome surfaced · "
        "CFO-CTO joint quarterly governance cadence established"
    ),
    "moment": (
        "AI consumption costs are exploding across the enterprise. GitHub Copilot seats, M365 Copilot "
        "licenses, Azure OpenAI/Foundry token consumption, embedding/vector-store spend, custom-agent "
        "run-rate, Power Platform AI Builder, plus team-expensed third-party AI tools (Cursor, Cline, "
        "Continue, Bedrock, Anthropic) add up to multi-tens-of-millions of dollars annually. Run-rate "
        "is climbing 20-40% quarter-over-quarter. Cloud-FinOps practices do not yet cover AI consumption "
        "cleanly. CFO sees the bill; CTO sees the value; nobody sees per-use-case cost-vs-value."
    ),
    "solution": (
        "AI-consumption FinOps agent continuously composes consumption signal across every AI cost "
        "source. Identifies idle Copilot seats, runaway agents, sub-optimal model selection "
        "(e.g., GPT-4 used where GPT-4o-mini would suffice), high-cost low-value usage patterns. "
        "Drafts optimisation recommendations - seat reclamation, model-tier downshift, agent retirement, "
        "third-party tool consolidation. HITL gate at FinOps team for reclamation actions; CFO + CTO see "
        "joint operating view. Does NOT make purchase or termination decisions autonomously."
    ),
    "use_cases": (
        "Copilot seat utilisation analytics · Azure OpenAI/Foundry token-consumption optimisation · "
        "GitHub Copilot per-developer ROI · custom-agent run-cost tracking · embedding/vector-store "
        "spend visibility · third-party AI tool spend rationalisation · model-selection-by-workload "
        "optimisation · per-use-case cost-per-outcome modelling · CFO-CTO joint AI governance cadence."
    ),
    "service": (
        "TMT-TEC-04 Technology Strategy (FinOps-for-AI specialisation). Wave 1-2 candidate at "
        "AI-heavy enterprises. Commercial envelope $1.3-1.9M services revenue. Microsoft Cost "
        "Management for Azure OpenAI is the platform anchor; complements but does not duplicate "
        "the CTO Portfolio play (which is strategic-investment-portfolio, not consumption-FinOps)."
    ),
    "personas": (
        "Primary: CTO · CFO (joint) · Head of FinOps · CIO. Approver: FinOps team approves "
        "reclamation/termination actions; engineering leaders approve model-tier downshifts for "
        "their workloads; HR/comms partner for any seat-reclamation that touches employee experience."
    ),
}


def main() -> None:
    print(f"Loading {XLSX.name}...")
    wb = load_workbook(XLSX)

    # ============================================================
    # Sheet: Scenario Library
    # ============================================================
    ws_lib = wb["Scenario Library"]
    last_num = ws_lib.cell(row=ws_lib.max_row, column=1).value
    if not isinstance(last_num, int):
        for r in range(ws_lib.max_row, 1, -1):
            v = ws_lib.cell(row=r, column=1).value
            if isinstance(v, int):
                last_num = v
                break

    new_num = last_num + 1
    print(f"Last scenario # in library: {last_num}; appending scenario #{new_num}...")

    ws_lib.append([
        new_num,
        SCENARIO["id"],
        SCENARIO["title"],
        SCENARIO["service_code"],
        SCENARIO["domain"],
        None,            # Schemas
        SCENARIO["brief"],
        SCENARIO["kpi"],
        "Catalog",
    ])
    print(f"  + lib row {new_num}: {SCENARIO['id']}")

    # ============================================================
    # Sheet: Scenario->KPI Chain
    # ============================================================
    ws_chain = wb["Scenario→KPI Chain"]
    ws_chain.append([
        SCENARIO["id"],
        SCENARIO["moment"],
        SCENARIO["solution"],
        SCENARIO["use_cases"],
        SCENARIO["service"],
        SCENARIO["personas"],
        SCENARIO["kpi"],
        "Catalog",
    ])
    print(f"  + chain row: {SCENARIO['id']}")

    # ============================================================
    # Sheet: Summary
    # ============================================================
    ws_sum = wb["Summary"]
    new_total = new_num
    new_catalog = new_total - 36  # featured count unchanged
    for r in range(1, ws_sum.max_row + 1):
        v = ws_sum.cell(row=r, column=1).value
        if v == "Total scenarios":
            ws_sum.cell(row=r, column=2, value=str(new_total))
        elif v == "Catalog scenarios (compact)":
            ws_sum.cell(row=r, column=2, value=str(new_catalog))
        elif v == "Scenario Library":
            ws_sum.cell(row=r, column=3, value=str(new_total))
        elif v == "Scenario→KPI Chain":
            ws_sum.cell(row=r, column=3, value=str(new_total))
        elif v == "Document version":
            ws_sum.cell(row=r, column=2, value="v1.4 · 2026-05-13 · added AI Consumption Cost Intelligence (FinOps-for-AI) scenario")

    wb.save(XLSX)
    print(f"\nWrote {XLSX.name}")
    print(f"  New scenario total: {new_total}")


if __name__ == "__main__":
    main()

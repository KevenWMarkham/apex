"""Append 8 tmt-arium-* wireless-infra-vendor scenarios to APEX-Scenario-Chains.xlsx.

Design: docs/plans/2026-05-14-arium-tmt-scenarios-design.md
Plan:   docs/plans/2026-05-14-arium-tmt-scenarios.md
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

WORKBOOK = r"C:\Stage\Clients\Industries\APEX\docs\reference\APEX-Scenario-Chains.xlsx"

SERVICE_CODE = "TMT-TEL-NET-02"
DOMAIN = "Network & Infrastructure"
FEATURED = "Catalog"

SCENARIOS = [
    {
        "id": "tmt-arium-ran-fleet-predictive-failure",
        "title": "RAN-fleet predictive failure (vendor-side)",
        "brief": "Vendor-telemetry RAN-fleet failure prediction surfaces pre-failure signatures for pre-emptive swap before carrier SLA breach.",
        "kpi": "Predicted-failure capture +62% · pre-emptive swap rate 14% → 71% · carrier SLA credits −$8.4M/yr",
    },
    {
        "id": "tmt-arium-field-service-dispatch-optimization",
        "title": "Field-service dispatch optimization (tower & site crews)",
        "brief": "Tower & site field-crew dispatch optimized jointly across weather, lease windows, parts, and skills.",
        "kpi": "First-time-fix rate 71% → 89% · drive-hours per ticket −24% · ticket re-roll rate −67%",
    },
    {
        "id": "tmt-arium-spare-parts-depot-rebalance",
        "title": "Spare-parts depot rebalance (swap-stock positioning)",
        "brief": "Swap-stock depot positioning rebalanced against carrier deployment plans to cut asset-on-ground wait.",
        "kpi": "AOG wait 4.2d → 18h · swap-stock holding cost −19% · SLA credits −$3.1M/yr",
    },
    {
        "id": "tmt-arium-firmware-release-orchestration",
        "title": "Firmware-release orchestration (staged rollout across carriers)",
        "brief": "Staged RAN firmware rollout across carriers with telemetry-gated canary rings and rollback discipline.",
        "kpi": "Mean time-to-rollback 96h → 9h · canary-cohort coverage 31% → 100% · rollback-attributed SLA loss −82%",
    },
    {
        "id": "tmt-arium-site-energy-attribution",
        "title": "Site-energy attribution & Scope-3 reporting (per-carrier)",
        "brief": "Per-cell per-carrier Scope-3 site-energy attestation for vendor sustainability reporting.",
        "kpi": "Time-to-attest 11 wk → 4 days · 100% carriers covered · audit-ready CDP / SBTi disclosure packets",
    },
    {
        "id": "tmt-arium-carrier-sla-breach-prediction",
        "title": "Carrier-SLA-breach prediction & credit-liability exposure",
        "brief": "Forward-looking SLA-breach prediction and credit-liability exposure across carrier contract portfolio.",
        "kpi": "SLA-breach forecast accuracy +73% · unplanned credit write-offs −58% · CFO forecast-confidence rating +2.3 pts",
    },
    {
        "id": "tmt-arium-warranty-rma-cluster-detection",
        "title": "Warranty / RMA cluster detection (field-failure root cause)",
        "brief": "Free-text RMA stream clustered by component lot, firmware, climate, and install cohort for root cause.",
        "kpi": "Time-to-cluster detection 11 wk → 6 days · warranty cost −24% · supplier-charge-back recoveries +$4.7M/yr",
    },
    {
        "id": "tmt-arium-private-5g-opportunity-scoring",
        "title": "Private-5G enterprise-opportunity scoring",
        "brief": "Enterprise private-5G prospect scoring across RF feasibility, spectrum, and load-profile signals.",
        "kpi": "Pilot-to-deployment conversion 41% → 68% · BD time per prospect −51% · win-rate at site-survey gate +27 pts",
    },
]

# Canonical 24-step catalog template, harvested verbatim from tmt-5g-slice-performance-monitoring.
CATALOG_24_TEMPLATE = [
    {"wave": "W1", "step_num": 1,  "step_key": "w1-sor",            "step_title": "SOR · System of Record",                       "layer": "Integration Plane · SOR",     "kind": "source",     "purpose": "Original system data lives here · agents do not replace the SOR",      "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 2,  "step_key": "w1-rth",            "step_title": "Real-Time Hub · streaming ingest",             "layer": "Integration Plane",           "kind": "ingest",     "purpose": "Eventstream / MQTT / OPC-UA bridge from operational systems",          "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 3,  "step_key": "w1-bronze",         "step_title": "Bronze landing · raw + ingest metadata",       "layer": "Data Plane · Bronze",         "kind": "data",       "purpose": "Source-aligned · append-only · 5-field ingest metadata",               "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 4,  "step_key": "w1-tokenizer",      "step_title": "Tokenizer / PII handling",                     "layer": "Data Plane · Bronze",         "kind": "transform",  "purpose": "Hash/redact PII before downstream availability",                       "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 5,  "step_key": "w1-silver",         "step_title": "Silver canonical · MERML / SCML / FINML",      "layer": "Data Plane · Silver",         "kind": "canonical",  "purpose": "Conformed canonical layer agents read against",                        "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 6,  "step_key": "w1-gold",           "step_title": "Gold semantic · Power BI Direct Lake",         "layer": "Data Plane · Gold",           "kind": "semantic",   "purpose": "Curated KPI marts + semantic models for BI",                           "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 7,  "step_key": "w1-mcp",            "step_title": "MCP server · capability registry",             "layer": "Runtime Plane",               "kind": "runtime",    "purpose": "Tool capabilities exposed to agents via JSON-RPC",                     "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 8,  "step_key": "w1-identity",       "step_title": "Entra · agent + persona identities",           "layer": "Identity Plane",              "kind": "identity",   "purpose": "Service principals · group-based RBAC · PIM",                          "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 9,  "step_key": "w1-ledger",         "step_title": "LEDGER · 14-field audit row store",            "layer": "Ledger Plane",                "kind": "evidence",   "purpose": "Audit-row evidence store · primary substrate for KPIs",                "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W1", "step_num": 10, "step_key": "w1-hitl-surface",   "step_title": "HITL Adaptive Card surface",                   "layer": "Experience Plane",            "kind": "hitl",       "purpose": "Microsoft Teams Adaptive Card · named-persona binding",                "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 11, "step_key": "w2-event",          "step_title": "Event trigger · scenario start",               "layer": "Runtime Plane",               "kind": "trigger",    "purpose": "Watermark fires · trace_id allocated · orchestrator instantiated",     "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 12, "step_key": "w2-orchestrator",   "step_title": "Parent orchestrator · agent dispatch",         "layer": "Runtime Plane · CAF Dynamic", "kind": "orchestrate","purpose": "Hierarchical or parallel · sequential with HITL gate",                 "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 13, "step_key": "w2-agent-assess",   "step_title": "Assess / Sense agent",                         "layer": "Decision Plane",              "kind": "agent",      "purpose": "Characterize the event from telemetry + context",                      "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 14, "step_key": "w2-agent-classify", "step_title": "Classify / Score agent",                       "layer": "Decision Plane",              "kind": "agent",      "purpose": "Apply policy · classify state · score with confidence",                "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 15, "step_key": "w2-agent-quantify", "step_title": "Quantify / Compose agent",                     "layer": "Decision Plane",              "kind": "agent",      "purpose": "Compute $ impact or compose decision with reasoning",                  "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 16, "step_key": "w2-hitl",           "step_title": "HITL gate · Adaptive Card",                    "layer": "Experience Plane · HITL",     "kind": "hitl",       "purpose": "Material decisions route to named persona via Teams",                  "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 17, "step_key": "w2-agent-act",      "step_title": "Act + Evidence-Write",                         "layer": "Decision Plane · Mutation",   "kind": "agent",      "purpose": "Mutate downstream systems · emit closeout audit row",                  "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W2", "step_num": 18, "step_key": "w2-kpi",            "step_title": "KPI rollup · Power BI semantic model",         "layer": "Experience Plane · BI",       "kind": "kpi",        "purpose": "Drill from KPI → trace_id → reasoning trace",                          "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W3", "step_num": 19, "step_key": "w3-scale",          "step_title": "Scale · multi-tenant · multi-banner",          "layer": "Runtime · Scale",             "kind": "scale",      "purpose": "Cross-banner orchestration · per-banner policy manifests",             "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W3", "step_num": 20, "step_key": "w3-fuse-markdown",  "step_title": "Fuse · adjacent scenarios",                    "layer": "Runtime · Fusion",            "kind": "fuse",       "purpose": "Cross-scenario fusion via shared canonical + capability tags",         "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W3", "step_num": 21, "step_key": "w3-fuse-loyalty",   "step_title": "Fuse · cross-Practice + loyalty",              "layer": "Runtime · Fusion",            "kind": "fuse",       "purpose": "Loyalty / cross-practice fusion via A2A swarm",                        "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W3", "step_num": 22, "step_key": "w3-purview",        "step_title": "Purview · lineage + classification at scale",  "layer": "Governance Plane",            "kind": "governance", "purpose": "Bronze→Silver→agent→outcome lineage · sensitivity labels",             "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W3", "step_num": 23, "step_key": "w3-ledger-feedback","step_title": "LEDGER feedback loop · model retraining",      "layer": "Ledger · Feedback",           "kind": "feedback",   "purpose": "Audit rows feed retraining · cohort drift watchtower",                 "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
    {"wave": "W3", "step_num": 24, "step_key": "w3-kpi-enterprise", "step_title": "Enterprise KPI · run-rate durable value",      "layer": "Experience · Executive",      "kind": "kpi",        "purpose": "Enterprise-scale rollup · run-rate value · Steering Committee",        "what_apex_does": "APEX-canonical pattern · see Featured Chains for the marquee scenario fully detailed"},
]

# Catalog-template values for Scenario→KPI Chain Solution / Use-Case / Service / Persona cols
SK_SOLUTION = "Standard archetype: hierarchical-root + sequential-with-hitl-gate · 6-agent fleet · APEX canonical"
SK_PERSONA = "Operator with HITL approval authority via Teams Adaptive Card"


def main():
    wb = openpyxl.load_workbook(WORKBOOK)

    # ----- Scenario Library -----
    sl = wb["Scenario Library"]
    max_num = max(
        (r[0].value for r in sl.iter_rows(min_row=2, max_col=1) if isinstance(r[0].value, int)),
        default=0,
    )
    for i, s in enumerate(SCENARIOS, start=1):
        sl.append([
            max_num + i,
            s["id"],
            s["title"],
            SERVICE_CODE,
            DOMAIN,
            None,                # Schemas — leave blank to match existing N&I rows
            s["brief"],
            s["kpi"],
            FEATURED,
        ])

    # ----- Scenario→KPI Chain -----
    sk = wb["Scenario→KPI Chain"]
    for s in SCENARIOS:
        sk.append([
            s["id"],
            s["brief"],
            SK_SOLUTION,
            f"Wave 2 pilot delivery for {s['title']}",
            SERVICE_CODE,
            SK_PERSONA,
            s["kpi"],
            None,                # Featured? — None for catalog-tier
        ])

    # ----- 24-Step Chain -----
    sh24 = wb["24-Step Chain"]
    for s in SCENARIOS:
        for step in CATALOG_24_TEMPLATE:
            sh24.append([
                s["id"],
                s["title"],
                DOMAIN,
                step["wave"],
                step["step_num"],
                step["step_key"],
                step["step_title"],
                step["layer"],
                step["kind"],
                step["purpose"],
                step["what_apex_does"],
                s["brief"],
            ])

    # ----- Summary -----
    summary = wb["Summary"]
    for row in summary.iter_rows():
        label = row[0].value
        if label == "Total scenarios":
            row[1].value = str(int(row[1].value) + len(SCENARIOS))
        elif label == "Catalog scenarios (compact)":
            row[1].value = str(int(row[1].value) + len(SCENARIOS))
        elif isinstance(label, str) and label.startswith("24-step chain rows"):
            row[1].value = str(int(row[1].value) + len(SCENARIOS) * 24)
        elif label == "Document version":
            current = row[1].value or ""
            row[1].value = current + " · v1.6 · 2026-05-14 · added 8 tmt-arium-* wireless-infra-vendor scenarios"

    wb.save(WORKBOOK)
    print(f"Wrote {len(SCENARIOS)} scenarios · {len(SCENARIOS)} SK rows · {len(SCENARIOS) * 24} 24-step rows · summary updated")


if __name__ == "__main__":
    main()

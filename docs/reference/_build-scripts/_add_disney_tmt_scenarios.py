"""
Add 5 new TMT scenarios to APEX-Scenario-Chains.xlsx based on gaps identified
from the Disney Account Podcast's curated play list.

New scenarios added to:
  - Scenario Library sheet
  - Scenario->KPI Chain sheet
  - Summary sheet totals updated

Gap analysis (Disney plays without existing TMT scenarios):
  1. CTO Portfolio & Decision Intelligence Agent     -> tmt-ai-portfolio-rationalization
  2. Engineering Headcount & Productivity Platform   -> tmt-engineering-productivity-platform
  3. Subscriber Lifecycle Orchestration              -> tmt-subscriber-lifecycle-orchestration
  4. CSR Agent-Assist (streaming)                    -> tmt-csr-agent-assist-streaming
  5. Guest Day Orchestration (parks/resorts)         -> tmt-guest-day-orchestration

Already covered (no addition needed):
  Streaming Churn, Personalised Recs, Cold-Start, Password-Sharing,
  ESPN Auto-Highlight, Auto-Dub, Streaming QoE, Ad-Targeting.

Run from this folder:
    python _add_disney_tmt_scenarios.py
"""

from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

HERE = Path(__file__).parent
XLSX = HERE / "APEX-Scenario-Chains.xlsx"
BACKUP = HERE / "APEX-Scenario-Chains.backup-pre-disney-tmt.xlsx"

# Scenarios to add — keyed by Scenario ID.
# Each value is a dict containing the data for both Scenario Library and
# Scenario->KPI Chain sheets.
NEW_SCENARIOS = [
    {
        "scenario_id": "tmt-ai-portfolio-rationalization",
        "title": "AI portfolio rationalization & CTO decision intelligence",
        "service_code": "TMT-TEC-04",
        "domain": "Engineering & R&D",
        "schemas": None,
        "brief": "Continuous CTO portfolio view across hundreds of technology investments — CapEx, project status, performance telemetry, KPI alignment. Surfaces stranded pilots and at-risk investments. Drafts sustain/accelerate/sunset/consolidate recommendations.",
        "kpi": "Stranded-pilot rate −40% · portfolio review quarterly -> continuous · investment-cycle compression 50%",
        "featured": "Catalog",
        # Scenario->KPI Chain (8 cols)
        "moment": "CTO and CFO struggle to maintain real-time visibility into the technology investment portfolio. Quarterly portfolio reviews can't keep up with the velocity of AI-pilot proliferation. Stranded pilots accumulate. CFO doesn't trust the ROI narrative.",
        "solution": "Portfolio-intelligence agent reads CapEx tracking, project status, performance signal, cost actuals, and business-KPI alignment continuously. Identifies stranded pilots and at-risk investments. Drafts portfolio-action recommendations. HITL gate at the CTO and steering committee.",
        "use_cases": "AI portfolio rationalization · technology-investment-ROI tracking · build-vs-buy decision support · CapEx allocation across business units.",
        "service": "TMT-TEC-04 Technology Strategy. Wave 2 follow-on after foundational Wave 1 lands. Commercial envelope $1.4-2.0M.",
        "personas": "Primary: CTO · CIO (joint depending on org). Approver: Technology Steering Committee. CFO is a critical adjacent buyer because the play touches CapEx and OpEx investment discipline.",
    },
    {
        "scenario_id": "tmt-engineering-productivity-platform",
        "title": "Engineering productivity & headcount-option platform",
        "service_code": "TMT-TEC-04",
        "domain": "Engineering & R&D",
        "schemas": None,
        "brief": "Multi-agent engineering productivity capability — code generation, agentic code review, design-doc and architecture-record drafting, automated test generation, technical-documentation maintenance. Produces option set: more output, lower headcount, shifted hiring profile, or redirected capacity.",
        "kpi": "Engineer productivity +20–35% on covered task types · code-review cycle −40-60% · engineering capacity available for redirection OR reduction (management decision) · time-to-PR-merge compression · onboarding-time reduction",
        "featured": "Catalog",
        "moment": "Large engineering organisation (thousands of engineers across business units) is sized for the pre-AI era. AI-augmented engineering produces measurable productivity lift. Leadership needs visibility into the option set — more output vs lower headcount vs shifted hiring profile vs redirected capacity — and a platform that enables any of them.",
        "solution": "Multi-agent engineering productivity platform composed of code-gen, code-review, design-doc, test-gen, deployment-pipeline and internal developer-Q&A agents. Engineering leaders measure productivity lift by team and decide workforce composition. HITL gates on workforce-affecting decisions; Deloitte recommends the platform, the client decides on workforce actions.",
        "use_cases": "Engineering velocity through transitions · workforce composition for the AI era · CFO-CTO operating-cost-discipline conversation · CHRO partnership on workforce composition.",
        "service": "TMT-TEC-04 Engineering Productivity. Wave 2-3 placement. Commercial envelope $1.5-2.5M services revenue. GitHub Copilot adjacency.",
        "personas": "Primary: CTO · CFO (joint) · Heads of Engineering per business unit. CHRO is a workforce-composition partner. Engineering managers review productivity-lift telemetry per team; senior engineering leadership approves any workforce-action recommendations.",
    },
    {
        "scenario_id": "tmt-subscriber-lifecycle-orchestration",
        "title": "Subscriber lifecycle orchestration (trial · downgrade · upgrade)",
        "service_code": "TMT-MED-01",
        "domain": "Customer Experience",
        "schemas": None,
        "brief": "Continuous subscriber-lifecycle agent watches every stage — trial, paid conversion, downgrade signal, upgrade opportunity, churn risk, winback. Intervenes at the right moment with personalised offer. Goes beyond single-stage 'trial-to-paid' scenarios to span full lifecycle.",
        "kpi": "Trial-to-paid conversion +12% · downgrade prevention +20% on targeted cohort · subscriber LTV uplift · DTC profitability contribution",
        "featured": "Catalog",
        "moment": "Subscriber-lifecycle events (trial-to-paid conversion, downgrade threats, upgrade opportunities) are managed reactively today. Each event is handled in isolation by different marketing teams using different tools. Lifecycle value leaks at every transition.",
        "solution": "Subscriber-lifecycle agent reads subscriber behaviour continuously across every stage. Identifies elevated risk or opportunity, recommends personalised intervention (offer, channel, timing), and routes to the loyalty / marketing-ops team. HITL gate above offer-depth threshold.",
        "use_cases": "Trial-to-paid conversion · downgrade prevention · upgrade prompts · winback flow integration · lifecycle-stage cohort analytics.",
        "service": "TMT-MED-01 Customer Experience. Wave 2 placement. Commercial envelope $1.1-1.7M services revenue.",
        "personas": "Primary: President of DTC · CMO · Head of Subscriber Lifecycle Marketing. Approver: Loyalty / marketing-operations team for offer-depth escalations.",
    },
    {
        "scenario_id": "tmt-csr-agent-assist-streaming",
        "title": "CSR agent-assist for streaming subscribers",
        "service_code": "TMT-MED-01",
        "domain": "Customer Experience",
        "schemas": None,
        "brief": "Real-time agent assistance for human CSR handling streaming-subscriber tickets. Surfaces account history, billing context, playback diagnostic, recent activity, recommended resolution path. CSR approves and acts.",
        "kpi": "AHT −30% · FCR +18pp · subscriber CES improvement · CSR-attrition rate reduction · subscriber NPS lift",
        "featured": "Catalog",
        "moment": "Streaming customer-service tickets span billing, playback, content, account, and entitlement issues. CSR composes context across 5-10 systems manually — 80% of average-handle-time is system-navigation, only 20% is customer interaction. CSR attrition is high; the variance in resolution experience across CSRs is wide.",
        "solution": "Agent-assist platform composes context across CRM, billing, order management, knowledge base, troubleshooting tools, and playback diagnostic. Surfaces customer context, recent activity, current entitlement, predicted intent, and next-best-action to the human CSR in real time. CSR retains decision authority.",
        "use_cases": "Streaming-subscriber support · billing query resolution · playback diagnostic · entitlement and access support · content-discovery support · escalation routing.",
        "service": "TMT-MED-01 Customer Experience (streaming variant of cross-Practice contact-center pattern). Wave 1 candidate. Commercial envelope $1.0-1.4M services revenue.",
        "personas": "Primary: Head of Customer Service Streaming · CXO. Operator: CSR (human-in-the-loop on every interaction).",
    },
    {
        "scenario_id": "tmt-guest-day-orchestration",
        "title": "Guest-day orchestration (parks · resorts · cruise)",
        "service_code": "TMT-MED-01",
        "domain": "Customer Experience",
        "schemas": None,
        "brief": "Continuous personalised itinerary agent for park/resort/cruise guests. Composes location, party composition, recent activity, dining and skip-line status, weather, attraction wait-times and availability. Recommends next-best-action; guest decides (HITL is the guest).",
        "kpi": "Skip-line attach rate +10–20% · per-guest in-park spend +5-10% · guest NPS lift · repeat-visit booking rate ↑ · capital-investment ROI uplift",
        "featured": "Catalog",
        "moment": "Guest days at large entertainment properties (parks, resorts, cruise) involve dozens of decisions per guest per day — attraction selection, dining timing, skip-line purchases, character meet-and-greets, photo opportunities, weather adjustments, rest breaks. Today's apps surface data; the guest composes manually. Cognitive load is high; the gap between possible-experience and actual-experience is wide.",
        "solution": "Guest-day orchestration agent watches the guest's wearable / mobile-app signal continuously. Composes a continuously updated personalised itinerary, integrating party constraints, weather, ride-state, dining availability, and guest preferences. Recommends opt-in next-best-action notifications. Guest accepts, modifies, or ignores. Agent learns and re-recommends.",
        "use_cases": "Theme-park guest-day orchestration · resort-stay personalisation · cruise voyage orchestration · capacity-yield optimisation across park assets · cast-member augmentation for personalised guest service.",
        "service": "TMT-MED-01 Customer Experience (Parks/Resorts/Cruise variant). Wave 2-3 placement. Commercial envelope $2.0-3.5M services revenue.",
        "personas": "Primary: President of Experiences · SVP Guest Experience · CTO Experiences. HITL: the guest (every notification is opt-in; guest decides).",
    },
]


def main():
    # Backup
    if not BACKUP.exists():
        shutil.copy2(XLSX, BACKUP)
        print(f"Backup: {BACKUP.name}")
    else:
        print(f"Backup already exists: {BACKUP.name} (preserving original)")

    wb = load_workbook(XLSX)

    # --- Scenario Library sheet ---
    ws = wb["Scenario Library"]
    next_row = ws.max_row + 1
    next_id = max((ws.cell(r, 1).value or 0) for r in range(2, ws.max_row + 1)) + 1

    print(f"\nAppending {len(NEW_SCENARIOS)} scenarios to Scenario Library starting at row {next_row} / # {next_id}...")
    for scenario in NEW_SCENARIOS:
        row_vals = [
            next_id,
            scenario["scenario_id"],
            scenario["title"],
            scenario["service_code"],
            scenario["domain"],
            scenario["schemas"],
            scenario["brief"],
            scenario["kpi"],
            scenario["featured"],
        ]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=next_row, column=c, value=v)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        print(f"  Row {next_row}: # {next_id} · {scenario['scenario_id']} ({scenario['service_code']})")
        next_row += 1
        next_id += 1

    # --- Scenario->KPI Chain sheet ---
    ws2 = wb["Scenario→KPI Chain"]
    next_row2 = ws2.max_row + 1
    print(f"\nAppending to Scenario-KPI Chain starting at row {next_row2}...")
    for scenario in NEW_SCENARIOS:
        row_vals = [
            scenario["scenario_id"],
            scenario["moment"],
            scenario["solution"],
            scenario["use_cases"],
            scenario["service"],
            scenario["personas"],
            scenario["kpi"],
            scenario["featured"],
        ]
        for c, v in enumerate(row_vals, start=1):
            cell = ws2.cell(row=next_row2, column=c, value=v)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        print(f"  Row {next_row2}: {scenario['scenario_id']}")
        next_row2 += 1

    # --- Summary sheet update ---
    ws_summary = wb["Summary"]
    # Update "Total scenarios" cell (row 5, col 2)
    current_total = ws_summary.cell(5, 2).value
    new_total = int(current_total) + len(NEW_SCENARIOS) if str(current_total).isdigit() else current_total
    ws_summary.cell(5, 2, value=str(new_total))
    print(f"\nSummary updated: Total scenarios {current_total} -> {new_total}")

    # Update catalog count (row 7)
    current_catalog = ws_summary.cell(7, 2).value
    if str(current_catalog).isdigit():
        new_catalog = int(current_catalog) + len(NEW_SCENARIOS)
        ws_summary.cell(7, 2, value=str(new_catalog))
        print(f"Catalog scenarios {current_catalog} -> {new_catalog}")

    # Update "Sheet rows" entries
    # Scenario Library row 15
    ws_summary.cell(15, 3, value=str(ws.max_row - 1))
    # Scenario->KPI Chain row 17
    ws_summary.cell(17, 3, value=str(ws2.max_row - 1))
    print(f"Sheet row-counts refreshed: Library {ws.max_row - 1} · KPI Chain {ws2.max_row - 1}")

    # Bump document version
    ws_summary.cell(20, 2, value="v1.2 · 2026-05-13 · added 5 Disney/TMT scenarios")

    # Save
    wb.save(XLSX)
    print(f"\nSaved: {XLSX.name}")


if __name__ == "__main__":
    main()

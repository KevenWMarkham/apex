"""
Build the Disney Studios Agentic Plays Excel — 15 curated agentic plays for
the Disney Studios sub-business (Walt Disney Pictures, Pixar, Marvel Studios,
Lucasfilm, 20th Century Studios, Searchlight Pictures).

Output: Disney_Studios_Agentic_Plays.xlsx

Three sheets:
  1. Plays       — the curated 15-play list across three domains
  2. Summary     — counts by domain · Wave 1 envelope totals · recommended picks
  3. How To Use  — guidance for the Account Team

Domains:
  - Content Development & Greenlight        (4 plays)
  - Production                              (5 plays)
  - Marketing, Distribution & Rights        (6 plays)

Usage:
    python _build_plays_xlsx.py
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "Disney_Studios_Agentic_Plays.xlsx"

# ----------------------------------------------------------------- plays ----

HEADERS = [
    "#",
    "Domain",
    "Sub-domain",
    "Play",
    "Business Problem",
    "Agent Capability",
    "Studios Pressure Addressed",
    "KPI Signal",
    "Buyer at Studios",
    "Microsoft Attach",
    "Wave",
    "Wave 1 Range ($M)",
    "APEX Family",
    "Priority",
]

# Each play tuple shape:
#   (Play, Sub-domain, Business Problem, Agent Capability, Pressure, KPI,
#    Buyer, MSFT Attach, Wave, Range, APEX Family, Priority)

CONTENT_DEV_PLAYS = [
    (
        "Greenlight Decision Support Agent",
        "Development · Greenlight",
        "Greenlight at tentpole level requires alignment across 12+ stakeholders, each with their own data view. Cycle runs weeks-to-months. Decision-quality variance is high because debate is partly about reconciling differing data interpretations, not just creative-strategic judgement.",
        "Agent composes per-project view across comparable-title performance, talent-availability calendar, production-cost benchmark, marketing-spend comparable, awards-positioning, genre-trend signal, international market readiness. Executive committee sees same composed view. Agent does NOT recommend greenlight; committee decides.",
        "Content cost inflation · hit-and-miss volatility · per-title ROI",
        "Greenlight cycle time -30 to -50% on tentpole projects · decision-quality variance reduction · post-greenlight surprise reduction",
        "Studio President · Chief Creative Officer · Head of Business Affairs (multi-buyer)",
        "Fabric · Foundry · Power BI · Purview",
        "Wave 1",
        "1.5-2.2",
        "TMT Content / Decision Support",
        "High",
    ),
    (
        "Audience Science & Sentiment Agent",
        "Audience Research · Marketing",
        "Audience-science data arrives in batches — preview screenings (12 hours), trailer tests (48 hours), social sentiment (days). Decisions on cut and marketing have to happen on the same compressed cadence. Composition of multiple streams is manual today.",
        "Agent reads audience-research streams continuously, composes per-project sentiment-and-response view, recommends focal areas for creative-team review. Creative team (director, editor, producer) decides what to change. Agent surfaces signal; humans interpret and act.",
        "Hit-and-miss volatility · marketing campaign effectiveness · cut-and-creative decision velocity",
        "Time-to-composed-audience-insight -60 to -70% · decision-velocity on cut/marketing adjustments · creative-team feedback cycle accelerated",
        "Chief Marketing Officer · Head of Audience Science · Studio President",
        "Fabric · Foundry · Real-Time Intelligence · Customer Insights",
        "Wave 1",
        "1.2-1.8",
        "TMT Audience / Real-Time Intelligence",
        "High",
    ),
    (
        "IP & Franchise Opportunity Scoring Agent",
        "IP Strategy · Development",
        "Disney's IP portfolio is enormous (Marvel, Star Wars, Pixar, Walt Disney library, 20th Century library, Searchlight catalog). Identifying highest-opportunity IP for next-cycle development is largely judgement-based; portfolio view is not continuously composed; opportunities surface episodically.",
        "Agent reads continuously across IP-performance signal (theatrical, streaming, consumer products), audience-affinity, social signal, competitive landscape, talent-attachment availability. Surfaces emerging opportunity signals — underutilised IP gaining momentum, characters whose franchise-extension potential is rising. Development leadership reviews and prioritises.",
        "Franchise-extension revenue · long-tail IP monetisation · slate-portfolio coverage",
        "Underutilised-IP-identification rate · time-to-opportunity-surfaced · development-pipeline coverage of high-opportunity IP · hit rate of agent-surfaced opportunities",
        "Chief Creative Officer · Head of IP Strategy · Studio President",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 2",
        "1.0-1.6",
        "TMT IP / Portfolio Analytics",
        "Medium",
    ),
    (
        "Talent & Casting Intelligence Agent",
        "Casting · Business Affairs",
        "Composing implications of a casting choice is manual today. Comparable-performance data, audience-affinity scores, awards track-record, compensation comparables, calendar availability, existing commitments — all exist but are composed by head of casting plus studio executive in conversation, not by an agent.",
        "Agent composes dimensions for candidate talent or ensemble. Surfaces composite-fit view for the role and production economics. Casting director, director, studio executive review and decide. Agent does NOT generate or simulate performances, does NOT use likenesses without explicit talent consent, does NOT displace casting directors. Operational composition only.",
        "Talent-and-AI policy compliance · casting-decision velocity · production economics",
        "Casting decision cycle compression · coverage of full candidate-set · casting-decision stakeholder alignment time reduction",
        "Head of Casting · Studio President · Director (per project)",
        "Fabric · Foundry · Purview (talent-rights propagation)",
        "Wave 2",
        "0.9-1.4",
        "TMT Talent / Decision Support",
        "Medium",
    ),
]

PRODUCTION_PLAYS = [
    (
        "Production Schedule & Budget Intelligence Agent",
        "Physical Production · Finance",
        "Production budgets for tentpoles run $150-300M. Schedule variance drives budget variance — every over-schedule day cascades through cast, crew, facility rental, equipment, post-production. Production accounting and schedule systems are separate. Composition into credible early-warning signal is weekly-cadence human work.",
        "Agent reads production accounting, daily call-sheet completion, scene-coverage progress, vendor-invoice signal, post-production dependency signal. Continuously composes trajectory. Identifies emerging variance patterns before they accumulate. Surfaces to line producer, production manager, studio executive with confidence-scored projections. Decisions stay human.",
        "Per-title ROI · content cost inflation · production economics integrity",
        "Time-to-detect emerging variance: weekly -> daily · variance-recovery rate improvement · budget-and-schedule integrity across annual slate",
        "President of Physical Production · Studio CFO · Senior Line Producer",
        "Fabric · Foundry · Power BI · Real-Time Intelligence",
        "Wave 1",
        "1.4-2.0",
        "TMT Production / Real-Time Intelligence",
        "High",
    ),
    (
        "VFX Pipeline Optimisation Agent",
        "VFX · Post-Production",
        "VFX is the largest production-cost line item on tentpoles. Marvel productions carry 2000-3000 VFX shots distributed across multiple vendors globally. Vendor coordination is high-variance and high-stakes. Status updates flow through vendor production-tracking + emails + weekly review meetings; composition across vendors is manual.",
        "Agent reads continuously across vendor production-tracking systems, shot-version repositories, review-feedback systems, delivery-schedule trackers. Identifies emerging risk patterns — vendors trending toward delivery slippage, shots exceeding planned-revision count, reviewer-feedback patterns indicating creative reconciliation needed. Surfaces to VFX supervisor and post-production producer.",
        "Per-title ROI · VFX cost discipline · Marvel/Lucasfilm pipeline integrity · ILM internal operations",
        "VFX-delivery-risk detection time compression · final-delivery slippage rate reduction · VFX-budget-variance reduction · expedited-cost-incurring late-stage revision rate -",
        "VFX Supervisor · VFX Producer · Post-Production Producer · ILM Operations (for Lucasfilm)",
        "Fabric · Foundry · Real-Time Intelligence · Power BI",
        "Wave 1",
        "1.5-2.2",
        "TMT VFX / Cross-Vendor Coordination",
        "High",
    ),
    (
        "Pixar Animation Pipeline Agent",
        "Animation · Long-Cycle Production",
        "Pixar's production cycle is 4-5 years per feature. Story revisions cascade into design revisions cascade into animation reworks. Render-farm capacity is finite; render time per shot compounds. Long-cycle visibility into pipeline trajectory is structural Pixar work.",
        "Agent reads continuously across story-revision tracking, design-asset versioning, animation-shot status, render-farm queue, sound-pipeline status. Composes multi-year pipeline view in continuous-update form. Surfaces emerging risks to production producer, director, CCO. Critically — agent does NOT replace creative judgement; director's call on story revision stays director's call. Operational substrate only.",
        "Pixar production economics · animation-pipeline cycle integrity · story-cascade risk · render-capacity optimisation",
        "Pipeline-cycle-time variance reduction · render-capacity utilisation optimisation · earlier surface of story-revision cascade risk (most expensive form of late-stage rework)",
        "Pixar President · Chief Creative Officer (Pixar) · Head of Production (Pixar)",
        "Fabric · Foundry · Real-Time Intelligence",
        "Wave 2-3",
        "1.3-1.9",
        "TMT Animation / Long-Cycle Pipeline",
        "Medium",
    ),
    (
        "Post-Production Workflow Agent",
        "Post-Production · Delivery",
        "Post-production is editing, sound, color, music, VFX integration, delivery preparation. Dependency chain is long and schedule is tight. Editing locks must precede sound finalisation; sound must precede VFX integration; VFX before final color; color before delivery prep. Late delivery of one element cascades. Visibility across dependency chain is partially manual.",
        "Agent reads continuously across editing milestones, sound milestones, VFX delivery, color-and-finishing milestones, delivery preparation. Identifies dependency-cascade risk. Surfaces to post-production supervisor with timeline-adjustment recommendations.",
        "Post-production schedule integrity · final-delivery on-schedule rate · production economics",
        "Dependency-cascade detection time compression · post-production schedule integrity · final-delivery-on-schedule rate improvement",
        "Post-Production Supervisor · Post-Production Producer",
        "Fabric · Foundry · Power BI",
        "Wave 2",
        "0.9-1.4",
        "TMT Post / Schedule Coordination",
        "Medium",
    ),
    (
        "Production Safety & Wellness Agent",
        "Production Operations · Safety",
        "Safety-incident-pattern detection across multiple concurrent productions is partial today. Incidents at one production may indicate systemic patterns (vendor practices, scheduling pressure, specific stunt-or-equipment types) that are not surfaced until pattern matures.",
        "Agent reads safety-incident reporting, crew-sentiment signals (where opt-in), set-incident logs, protocol-compliance signal. Surfaces emerging patterns to studio safety leadership and responsible-production producers. Agent does NOT surveil individuals; composes anonymised pattern signal at production-level and protocol-level. Studio safety leadership reviews and acts.",
        "Insurance economics · talent comfort · regulatory compliance · production-environment quality",
        "Time-to-pattern-detection across productions · safety-incident-rate trends · crew-wellness signal · production-environment-quality metric (long-term)",
        "Studio President · President of Physical Production · Head of Safety",
        "Fabric · Foundry · Purview (strict PII governance) · Sentinel",
        "Wave 3",
        "0.8-1.3",
        "TMT Production / Safety Telemetry",
        "Low",
    ),
]

MARKETING_RIGHTS_PLAYS = [
    (
        "Trailer Performance & Audience Testing Agent",
        "Marketing · Audience Testing",
        "Trailer-test results arrive in 48-hour batches. By the time results are reviewed, trailer-house has often started next-cut iteration. Composition of trailer-test results with social-signal, competitive-trailer-performance, and audience-segment context is human work today.",
        "Agent composes test-vendor results, social-signal once trailers launch, comparable-trailer-performance benchmarks, competitive-landscape positioning. Surfaces composed view to marketing executive within hours of results landing. Marketing leadership decides on cut iteration and placement strategy.",
        "Marketing campaign effectiveness · opening-weekend correlation · trailer-decision velocity",
        "Time-to-composed-trailer-insight compression 70-80% · decision-velocity on cut iteration · trailer-effectiveness improvement (opening-weekend correlation)",
        "Chief Marketing Officer · Head of Worldwide Marketing · Studio Marketing President",
        "Fabric · Foundry · Real-Time Intelligence · Customer Insights",
        "Wave 1",
        "1.1-1.7",
        "TMT Marketing / Real-Time Intelligence",
        "High",
    ),
    (
        "Marketing Campaign Optimisation Agent",
        "Marketing · Cross-Channel",
        "Theatrical-tentpole campaigns run across broadcast, digital, social, programmatic, OOH, influencer, partnership, talent press, premiere, in-cinema. Per-channel attribution is partial; cross-channel optimisation more partial. Budgets allocated at campaign-launch often don't get reallocated based on performance signal until late.",
        "Agent reads continuously across channel-performance signal, audience-segment response, competitive landscape, creative-effectiveness data. Recommends within-campaign reallocations. Marketing-operations team reviews and approves spend shifts.",
        "Marketing ROI · campaign-spend optimisation · marketing-team capacity",
        "Within-campaign reallocation velocity · marketing ROI improvement 8-15% on covered campaigns · cross-channel attribution maturity",
        "Chief Marketing Officer · Head of Media Strategy · CMO Operations",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 1",
        "1.2-1.8",
        "TMT Marketing / Cross-Channel Optimisation",
        "High",
    ),
    (
        "Theatrical-vs-Streaming Windowing Decision Agent",
        "Distribution · Strategy",
        "Windowing decisions are strategic, not procedural. Theatrical-only vs theatrical-plus-streaming-30-days vs streaming-direct reshape revenue, marketing economics, talent relationships. Composing factors across title (genre, audience, budget, talent), market (competitive landscape, theatrical environment), strategic (Disney+ subscriber-acquisition value), and talent-contract dimensions is largely manual.",
        "Agent composes the factors across title and market dimensions. Surfaces composite-decision view to windowing committee. The committee decides; the agent does NOT recommend the decision. Critical: the agent doesn't pretend to know the strategic context the committee brings.",
        "Per-title revenue capture · DTC subscriber-acquisition strategy · talent compensation provisions · executive-committee decision velocity",
        "Windowing-decision cycle compression · decision-quality variance reduction across the windowing committee · post-decision execution quality (marketing/distribution alignment)",
        "Disney Entertainment Chairperson · Studio Presidents · Head of Worldwide Distribution · CFO Entertainment (executive committee)",
        "Fabric · Foundry · Power BI · Purview",
        "Wave 3",
        "1.6-2.4",
        "TMT Distribution / Strategic Decision Support",
        "Medium",
    ),
    (
        "Awards Campaign Management Agent",
        "Awards · Specialty Marketing",
        "Awards campaign coordination is year-round, peaking during Oscar and Emmy seasons. Searchlight's specialty-theatrical business depends on awards-season performance. Campaigns coordinate academy screenings, talent press, FYC advertising, industry-relationship management, festival positioning, eligibility tracking. Tight calendar windows; many touchpoints; composition is daily work.",
        "Agent reads campaign-event calendar, eligibility-tracking systems, talent-availability, FYC-advertising performance, social-and-industry sentiment, competitive-campaign signal. Composes campaign-status view continuously. Awards-team approves campaign-decision recommendations.",
        "Searchlight brand viability · awards-driven theatrical extension · talent-relationship management",
        "Awards-campaign coordination velocity · nomination-and-win conversion rate · specialty-theatrical viability for Searchlight (per-title theatrical extension months)",
        "President of Searchlight · Awards Campaign Director · CMO Searchlight (Searchlight is natural lead; other studios follow)",
        "Fabric · Foundry · Customer Insights · Power BI",
        "Wave 2",
        "1.1-1.6",
        "TMT Awards / Specialty Marketing",
        "High",
    ),
    (
        "Rights Compliance & Content Reuse Agent",
        "Rights · Business Affairs",
        "Disney's content library is enormous. Reusing archive content (trailers, marketing, behind-the-scenes, anniversary releases, theme-park integrations) requires rights clearance. Music, talent, footage, image rights. Each reuse is a clearance project. Work is labour-intensive and slow; business-affairs and legal handle clearances title-by-title.",
        "Agent reads continuously across rights-management systems, talent-contract repositories, music-licensing systems, footage-rights databases. Maintains continuous rights-status view of the library. Surfaces rights-availability for candidate reuse projects. PII propagation discipline is critical for talent-contract data.",
        "Long-tail IP monetisation · library-reuse velocity · compliance-risk reduction · business-affairs team capacity",
        "Rights-clearance cycle compression 50-70% · reuse-opportunity surfaced rate · compliance-risk reduction (earlier detection of expired/restricted-use rights)",
        "General Counsel Studios · Head of Business Affairs · Chief Rights Officer",
        "Fabric · Foundry · Purview (heavy PII governance) · M365 Compliance",
        "Wave 2",
        "1.0-1.5",
        "TMT Rights / Library Governance",
        "Medium",
    ),
    (
        "Music Synchronisation Clearance Agent",
        "Rights · Music",
        "Music sync clearance is high-volume — soundtrack assembly, trailer scoring, marketing-content scoring all require sync clearance. Each musical cue requires clearance from music publisher (composition) AND record label (recording). Clearance cycle from 'we want this song' to 'cleared and licensed' can run weeks. Affects creative-and-marketing flexibility.",
        "Agent reads continuously across music-publisher and record-label clearance systems, internal-licensing trackers, clearance-history databases. Surfaces clearance-availability and indicative pricing for candidate-cue use. Music supervisor decides; business-affairs approves.",
        "Music-clearance cycle compression · trailer-and-marketing schedule risk · music-supervisor productivity",
        "Music-clearance cycle compression · music-supervisor productivity · reduced clearance-related schedule risk on tight-deadline projects (trailers, marketing materials)",
        "Music Supervisor · Head of Music Business Affairs · GC Studios",
        "Fabric · Foundry · Purview · M365 Compliance",
        "Wave 2",
        "0.7-1.2",
        "TMT Rights / Music Clearance",
        "Medium",
    ),
]


def build_workbook() -> Workbook:
    wb = Workbook()

    # ============================================================
    # Sheet 1 - Plays
    # ============================================================
    ws = wb.active
    ws.title = "Plays"

    # Header
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1A2339")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(border_style="thin", color="000000"))

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    domain_fill = {
        "Content Development": "F0F4FF",   # soft blue
        "Production": "FFF7E6",            # soft amber
        "Marketing & Rights": "F0F7EE",    # soft green
    }

    priority_color = {
        "High": "1B7F3A",
        "Medium": "C28A22",
        "Low": "6B7280",
    }

    row_num = 2
    play_id = 1

    def append_plays(plays, domain_label):
        nonlocal row_num, play_id
        for t in plays:
            ws.append([
                play_id, domain_label,
                t[1], t[0], t[2], t[3], t[4], t[5], t[6], t[7], t[8], t[9], t[10], t[11],
            ])
            play_id += 1
            row_num += 1

    append_plays(CONTENT_DEV_PLAYS, "Content Development")
    append_plays(PRODUCTION_PLAYS, "Production")
    append_plays(MARKETING_RIGHTS_PLAYS, "Marketing & Rights")

    last_row = row_num - 1

    # Apply formatting per row
    for r in range(2, last_row + 1):
        domain = ws.cell(row=r, column=2).value
        priority = ws.cell(row=r, column=14).value
        bg = domain_fill.get(domain, "FFFFFF")
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = Border(
                bottom=Side(border_style="hair", color="CCCCCC"),
                right=Side(border_style="hair", color="CCCCCC"),
            )
        pcell = ws.cell(row=r, column=14)
        pcell.font = Font(name="Arial", size=10, bold=True, color=priority_color.get(priority, "000000"))
        wcell = ws.cell(row=r, column=11)
        wcell.font = Font(name="Arial", size=10, bold=True)

    widths = {
        1: 5, 2: 18, 3: 22, 4: 38, 5: 50, 6: 50, 7: 38, 8: 38,
        9: 32, 10: 32, 11: 10, 12: 14, 13: 32, 14: 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 2 - Summary
    # ============================================================
    ws2 = wb.create_sheet("Summary")

    ws2["A1"] = "Disney Studios Agentic Plays - Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1A2339")
    ws2.merge_cells("A1:D1")

    ws2["A3"] = "Generated for the Deloitte Account Team for Disney Studios (six-studio sub-business)"
    ws2["A3"].font = Font(name="Arial", italic=True, color="6B7280", size=10)
    ws2.merge_cells("A3:D3")

    ws2["A5"] = "By Domain"
    ws2["A5"].font = Font(name="Arial", bold=True, size=11)

    headers2 = ["Domain", "Play Count", "Wave 1 Range Total ($M, low)", "Wave 1 Range Total ($M, high)"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=6, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sum the ranges per domain
    # Content Development: 1.5-2.2 + 1.2-1.8 + 1.0-1.6 + 0.9-1.4 = 4.6-7.0
    # Production: 1.4-2.0 + 1.5-2.2 + 1.3-1.9 + 0.9-1.4 + 0.8-1.3 = 5.9-8.8
    # Marketing & Rights: 1.1-1.7 + 1.2-1.8 + 1.6-2.4 + 1.1-1.6 + 1.0-1.5 + 0.7-1.2 = 6.7-10.2
    # Total: 17.2-26.0
    summary_rows = [
        ("Content Development", 4, 4.6, 7.0),
        ("Production", 5, 5.9, 8.8),
        ("Marketing & Rights", 6, 6.7, 10.2),
        ("TOTAL", 15, 17.2, 26.0),
    ]
    for i, (domain, count, low, high) in enumerate(summary_rows, start=7):
        ws2.cell(row=i, column=1, value=domain).font = Font(name="Arial", bold=(domain == "TOTAL"))
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=low).number_format = "$#,##0.0"
        ws2.cell(row=i, column=4, value=high).number_format = "$#,##0.0"
        for c in range(1, 5):
            cell = ws2.cell(row=i, column=c)
            cell.font = Font(name="Arial", size=10, bold=(domain == "TOTAL"))
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right")
            if domain == "TOTAL":
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    ws2["A12"] = "By Priority (Account Team Sequencing)"
    ws2["A12"].font = Font(name="Arial", bold=True, size=11)
    headers3 = ["Priority", "Play Count", "Notes"]
    for col, h in enumerate(headers3, start=1):
        c = ws2.cell(row=13, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")

    priority_rows = [
        ("High", 7, "Wave 1 candidates + Awards-campaign Wave 2 - clear KPI · creative-authorship-clean · studio-operator buyers"),
        ("Medium", 7, "Wave 2-3 follow-ons after operational-AI posture is established - includes Pixar animation pipeline and strategic windowing"),
        ("Low", 1, "Wave 3 - production-safety play is slow-cycle and sensitive; runs after multi-year track record"),
    ]
    for i, (pri, count, note) in enumerate(priority_rows, start=14):
        ws2.cell(row=i, column=1, value=pri).font = Font(name="Arial", size=10, bold=True, color={"High": "1B7F3A", "Medium": "C28A22", "Low": "6B7280"}[pri])
        ws2.cell(row=i, column=2, value=count).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3, value=note).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="center")

    ws2["A19"] = "Recommended Wave 1 Entry Picks - By Studio"
    ws2["A19"].font = Font(name="Arial", bold=True, size=11)

    headers4 = ["Studio", "Play", "Why"]
    for col, h in enumerate(headers4, start=1):
        c = ws2.cell(row=20, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A2339")

    picks = [
        ("Marvel Studios", "Production Schedule & Budget Intelligence + VFX Pipeline Optimisation", "Data-rich production estate · operational-AI-receptive leadership · VFX-heaviness makes Play 6 the flagship · ILM adjacency for Lucasfilm"),
        ("Lucasfilm / ILM", "VFX Pipeline Optimisation", "ILM is large enough to matter and internal enough to pilot · Star Wars production pipeline is the operational substrate"),
        ("Pixar", "Production Schedule & Budget Intelligence ONLY (NOT animation pipeline)", "Pixar's creative culture is cautious about AI; lead with operational scheduling. Animation-pipeline play (Pixar-specific) is Wave 2-3 once operational-AI credibility is established"),
        ("Walt Disney Pictures", "Greenlight Decision Support OR Trailer Performance & Audience Testing", "Operational marketing or development play with clear executive buyer · avoids creative-authorship sensitivity"),
        ("Searchlight", "Awards Campaign Management", "Awards strategy is structurally critical to Searchlight brand · President of Searchlight + Awards Campaign Director are clear buyers · highest-leverage entry for the studio"),
        ("20th Century", "Similar to Walt Disney Pictures - operational marketing or development play", "Disney-system integration still working through; match a play to operational maturity"),
    ]
    for i, (studio, play, why) in enumerate(picks, start=21):
        c1 = ws2.cell(row=i, column=1, value=studio)
        c1.font = Font(name="Arial", size=10, bold=True)
        c1.fill = PatternFill("solid", fgColor="F4F4F4")
        c2 = ws2.cell(row=i, column=2, value=play)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        c3 = ws2.cell(row=i, column=3, value=why)
        c3.font = Font(name="Arial", size=10)
        c3.alignment = Alignment(wrap_text=True, vertical="center")

    # Independence + creative-authorship callout
    callout_row = 21 + len(picks) + 2
    ws2.cell(row=callout_row, column=1, value="Studios-Specific Independence Considerations").font = Font(name="Arial", bold=True, size=11, color="C28A22")
    ws2.merge_cells(start_row=callout_row, end_row=callout_row, start_column=1, end_column=4)

    callouts = [
        "1. Creative-authorship boundary - operational and analytical augmentation, never creative replacement. WGA / SAG-AFTRA / DGA AI provisions are binding. Plays that touch creative-authorship workflows require explicit positioning.",
        "2. Financial-reporting-adjacent scope avoidance - plays touching box office reporting, residual calculations, talent compensation, or financial-reporting metrics require Independence review before scope finalisation.",
        "3. Rights and licensing scope discipline - Deloitte recommends what client should build; client makes rights-and-licensing decisions. Deloitte does not advise on substantive licensing decisions.",
        "Plus standard global Disney Independence framing - two-contract model, no co-sell, careful coordination with global Independence office.",
    ]
    for j, text in enumerate(callouts, start=1):
        cell = ws2.cell(row=callout_row + j, column=1, value=text)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.merge_cells(start_row=callout_row + j, end_row=callout_row + j, start_column=1, end_column=4)
        ws2.row_dimensions[callout_row + j].height = 36

    sw = {1: 22, 2: 42, 3: 70, 4: 22}
    for col, w in sw.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ============================================================
    # Sheet 3 - How To Use
    # ============================================================
    ws3 = wb.create_sheet("How To Use")

    instructions = [
        ("Disney Studios Agentic Plays - How To Use This File", "title"),
        ("", ""),
        ("Purpose", "h2"),
        ("This file curates 15 agentic plays for the Disney Studios sub-business: Walt Disney Pictures, Pixar, Marvel Studios, Lucasfilm (including ILM), 20th Century Studios, and Searchlight Pictures. Each play is structured for the Deloitte Account Team to use as a working artefact during studio-level account planning, executive prep, and pursuit discussions.", "p"),
        ("", ""),
        ("The six-studio framing", "h2"),
        ("Each Disney studio operates with creative autonomy under its own president. Operational integration happens at the corporate level (distribution, marketing, finance, business affairs, talent management). The buyer for any play is studio-specific. The Account Team picks the studio with the strongest current relationship and lands Wave 1 there; other studios follow once reference value is built.", "p"),
        ("", ""),
        ("Sheet 1 - Plays", "h2"),
        ("The full play list. 14 columns. Filter or sort by Domain, Wave, or Priority. The Wave 1 Range column gives indicative Deloitte services range; precise envelope is set during BVA.", "p"),
        ("", ""),
        ("Sheet 2 - Summary", "h2"),
        ("Aggregated counts, recommended Wave 1 picks by studio, and Studios-specific Independence considerations.", "p"),
        ("", ""),
        ("Creative-authorship boundary - non-negotiable", "h2"),
        ("Every play in this file respects the creative-authorship boundary. AI use that touches creative work (writing, performance, performance-likeness, dialogue replacement) is governed by the 2023 WGA and SAG-AFTRA settlements plus the DGA AI policy. The Account Team's framing must be unambiguous: operational and analytical augmentation, never creative replacement.", "p"),
        ("", ""),
        ("How to use in account planning", "h2"),
        ("- Studio-by-studio prep - filter to the target studio and review recommended Wave 1 picks from the Summary sheet.", "p"),
        ("- Pre-meeting prep - filter to the buyer's role (Studio President, CCO, CMO, VFX Supervisor, etc.) and scan rows.", "p"),
        ("- Quarterly account plan - review High-priority plays and confirm which are in active discovery per studio.", "p"),
        ("- BVA construction - pull KPI Signal column to seed BVA-in-5-bullets per Sellers Handbook framework.", "p"),
        ("- Microsoft attach modelling - Microsoft Attach column drives SKU mix for Microsoft consumption forecasting.", "p"),
        ("", ""),
        ("Independence-clean usage", "h2"),
        ("This file lists Deloitte services plays. The dollar ranges are Deloitte services revenue ranges. Microsoft platforms are licensed directly by Disney from Microsoft; this file does not include Microsoft software margin. Refer to Sellers Podcast Ep 2 and Independence Cheat Sheet for full Independence posture. Studios-specific Independence considerations: see Sheet 2 callout.", "p"),
        ("", ""),
        ("Notes on KPI signals", "h2"),
        ("KPI signals are reference-scenario estimates from the APEX framework - not Disney-Studios-specific commitments. Studio-specific numbers are produced during BVA, with Disney's own baseline measurement.", "p"),
        ("", ""),
        ("Companion materials", "h2"),
        ("- Disney_Agentic_Plays_BackOffice_Streaming.xlsx - the company-wide Disney play book (23 plays across back-office, streaming, and Experiences)", "p"),
        ("- Disney Studios Account Podcast (5 episodes) - narrative walk-through of these 15 plays", "p"),
        ("- APEX-Scenario-Chains.xlsx - reference scenarios from the APEX framework", "p"),
        ("", ""),
        ("Last updated", "h2"),
        ("2026-05-13 - curated by TMT-MED Practice + Disney Studios Account Team", "p"),
    ]

    row = 1
    for text, style in instructions:
        cell = ws3.cell(row=row, column=1, value=text)
        if style == "title":
            cell.font = Font(name="Arial", bold=True, size=16, color="1A2339")
        elif style == "h2":
            cell.font = Font(name="Arial", bold=True, size=12, color="1A2339")
        elif style == "p":
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws3.column_dimensions["A"].width = 110

    return wb


def main():
    wb = build_workbook()
    wb.save(OUT)
    print(f"Wrote {OUT.name} ({OUT.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()

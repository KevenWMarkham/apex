"""Build Q1 MVP Use Case workbook from the scope + journey map sources.

Output: C:\\Users\\kmarkham\\OneDrive - Deloitte (O365D)\\Documents\\mvp-scope-q1-usecases.xlsx

Sheets
------
1. Cover                     — title block, version, scope summary
2. Use Case Index            — single-row-per-UC summary, sortable
3. Use Case Detail           — full UC spec per row (the master sheet)
4. Persona x UC Matrix       — who-touches-what
5. Decision Rights           — gate thresholds (Q1)
6. WBR Timeline              — Fri 9:43 PM -> Mon 8:00 AM chapter detail
"""
from __future__ import annotations
import io, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

OUT = r"C:\Users\kmarkham\OneDrive - Deloitte (O365D)\Documents\mvp-scope-q1-usecases.xlsx"

# ---------- styling helpers ----------
DELOITTE_GREEN = "86BC25"
INK_DARK = "0D2818"
INK_AMBER = "F0A839"
INK_BLUE = "4DA6FF"
INK_PURPLE = "B794F4"
HEADER_BG = "0D2818"
HEADER_FG = "FFFFFF"
ALT_BG = "F4F7EF"
BORDER_GREY = "D0D0D0"

thin = Side(style="thin", color=BORDER_GREY)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell):
    cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border_all

def style_body(cell, alt=False):
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border_all
    if alt:
        cell.fill = PatternFill("solid", fgColor=ALT_BG)

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_table(ws, headers, rows, widths, freeze="A2", row_height=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            style_body(c, alt=(i % 2 == 0))
        if row_height:
            ws.row_dimensions[i].height = row_height
    autosize(ws, widths)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = freeze

# =====================================================================
# SHEET 1 — Cover
# =====================================================================
def build_cover(wb: Workbook):
    ws = wb.active
    ws.title = "Cover"

    ws["B2"] = "Agentic Merch — Q1 MVP Use Case Catalog"
    ws["B2"].font = Font(name="Calibri", size=22, bold=True, color=INK_DARK)
    ws.row_dimensions[2].height = 36

    ws["B3"] = "Derived from: mvp-scope-q1.html + mvp-scope-q1-journey-map.html"
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color="606060")

    rows = [
        ("Version", "v0.1 — Q1 scope draft"),
        ("Date", "2026-04-29"),
        ("Authors", "Joe G., Aditi G."),
        ("For", "Apurva L., Saurabh V."),
        ("Scope (Q1)", "Agents 1, 4, 8, 13, 13.3 — WBR cadence, Chief Merchant view"),
        ("Out of scope (Sprint 2)", "Agents 6 (Pricing), 7 (Promotion Optimization)"),
        ("Demo scenario", "Apparel Retailer · Apparel · Week 16 · Midwest · −12% miss · supply disruption root cause"),
        ("Cadence", "Friday detection → Saturday overnight chain → Monday 8 AM brief delivery"),
        ("Persona count", "9 (5 agent personas + 4 human personas)"),
        ("Use case count", "20 use cases across 6 functional groups"),
        ("Data contracts", "merml-mcp · scml-mcp · cxml-mcp"),
        ("Audit", "LEDGER 14-field WORM rows on every decision"),
        ("Delivery surface", "Microsoft Teams Adaptive Cards (HITL) · Monday brief (ACK)"),
    ]

    start = 5
    for i, (k, v) in enumerate(rows):
        kc = ws.cell(row=start + i, column=2, value=k)
        vc = ws.cell(row=start + i, column=3, value=v)
        kc.font = Font(name="Calibri", size=10, bold=True, color=INK_DARK)
        vc.font = Font(name="Calibri", size=10)
        kc.alignment = Alignment(horizontal="left", vertical="top")
        vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[start + i].height = 18

    ws["B19"] = "Workbook contents"
    ws["B19"].font = Font(name="Calibri", size=14, bold=True, color=INK_DARK)
    contents = [
        ("Use Case Index",        "One row per use case — ID, name, persona, gate, chapter, status"),
        ("Use Case Detail",       "Full spec per use case: trigger, preconditions, flow, postconditions, data, acceptance criteria"),
        ("Persona x UC Matrix",   "Who touches which use case (Owner / Approver / Consumer / Bypassed)"),
        ("Decision Rights",       "Gate kinds + Q1 thresholds (Zero Touch / HITL / Blocked / ACK)"),
        ("WBR Timeline",          "Friday-to-Monday chapter-by-chapter event log with the agent and human owners"),
    ]
    for i, (name, desc) in enumerate(contents):
        ws.cell(row=21 + i, column=2, value=name).font = Font(name="Calibri", size=10, bold=True, color=DELOITTE_GREEN)
        ws.cell(row=21 + i, column=3, value=desc).font = Font(name="Calibri", size=10)
        ws.row_dimensions[21 + i].height = 18

    autosize(ws, [3, 28, 110])

# =====================================================================
# Use case data (single source of truth — used by Index + Detail)
# =====================================================================
USE_CASES = [
    # --------- Group 1: Detect & Diagnose (Agents 13, 1) ---------
    {
        "id": "UC-01", "group": "Detect & Diagnose",
        "name": "Detect performance exception",
        "actor": "The Analyst (Agent 13)",
        "system": "merml-mcp",
        "chapter": "Detect",
        "timestamp": "Fri 9:43 PM",
        "gate": "ACK",
        "threshold": "Severity score + consecutive-miss rule (≥2 weeks)",
        "trigger": "Friday end-of-week sales actuals land vs. plan in merml-mcp",
        "precond": "Weekly plan loaded · prior-week actuals reconciled · severity ranking model active",
        "main_flow": (
            "1. Pull weekly actuals vs. plan for all categories.\n"
            "2. Compute miss % and consecutive-miss count per category × region.\n"
            "3. Rank exceptions by severity score.\n"
            "4. Identify root-cause hypothesis (demand vs. supply) from supporting signals.\n"
            "5. Route top-N exceptions to downstream agents (1 for demand, 4 for guardrail, 8 for replen)."
        ),
        "alt_flow": "If ranking is inconclusive (tie), surface multiple exceptions to brief and let HITL prioritize.",
        "postcond": "Exception package routed · LEDGER row written · Done/Watch tier proposed",
        "inputs": "Weekly sales actuals · plan · prior-week comp · inventory snapshot · DC health",
        "outputs": "Ranked exception list · root-cause hypothesis · routing manifest",
        "ledger": "ts · agent_id · category · region · miss_pct · severity_rank · routed_to · rationale",
        "acceptance": (
            "AC-1: Ranks Apparel as #1 of 7 exceptions in Week 16.\n"
            "AC-2: Routes context to Agent 1 within 2 minutes of detection.\n"
            "AC-3: No autonomous action taken — observe-and-route only."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-02", "group": "Detect & Diagnose",
        "name": "Confirm consumer demand intent",
        "actor": "The Demand Checker (Agent 1)",
        "system": "cxml-mcp",
        "chapter": "Diagnose",
        "timestamp": "Sat 12:30 AM",
        "gate": "ACK",
        "threshold": "Intent score 0–1.0 · healthy ≥ 0.60",
        "trigger": "Routing manifest from Agent 13 lands with category × region",
        "precond": "Consumer intent model trained · current-week signals available · cxml-mcp populated",
        "main_flow": (
            "1. Pull demand intent signals for category × region (search, social, traffic, conversion).\n"
            "2. Compute consolidated intent score.\n"
            "3. Compare intent score to historic baseline.\n"
            "4. Diagnose: intent healthy → supply problem; intent depressed → demand problem.\n"
            "5. Return diagnosis to Agent 13."
        ),
        "alt_flow": "If signals are stale (>24h), flag and downgrade confidence — escalate to Watch tier.",
        "postcond": "Diagnosis returned · markdown blocked if intent healthy · LEDGER row written",
        "inputs": "Search intent · social mentions · site traffic · conversion rate · prior-week baseline",
        "outputs": "Intent score (0.72 in demo) · diagnosis label (demand|supply) · confidence",
        "ledger": "ts · agent_id · category · region · intent_score · baseline · diagnosis · confidence",
        "acceptance": (
            "AC-1: Returns intent 0.72 for Apparel in demo.\n"
            "AC-2: Diagnosis labels miss as 'supply' not 'demand'.\n"
            "AC-3: Hands off to Agent 13 within 5 minutes."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-03", "group": "Detect & Diagnose",
        "name": "Block erroneous markdown",
        "actor": "The Analyst (Agent 13) → Category DMM (bypassed)",
        "system": "merml-mcp",
        "chapter": "Diagnose",
        "timestamp": "Sat 12:32 AM",
        "gate": "Blocked",
        "threshold": "Auto-block when intent ≥ baseline AND root cause = supply",
        "trigger": "Agent 1 returns demand-healthy diagnosis",
        "precond": "Markdown rule active · DMM default-approver routing configured",
        "main_flow": (
            "1. Receive intent score from Agent 1.\n"
            "2. Compare to threshold (≥ baseline).\n"
            "3. If healthy AND supply is root cause → block markdown action.\n"
            "4. Log block decision with rationale.\n"
            "5. Surface to Done bucket — DMM sees decision, no approval needed."
        ),
        "alt_flow": "If intent borderline (near threshold), escalate to Your Call instead of auto-blocking.",
        "postcond": "Markdown action blocked · DMM bypassed · margin protected",
        "inputs": "Intent score · baseline · root-cause label · proposed markdown depth",
        "outputs": "Block decision · rationale · audit entry",
        "ledger": "ts · decision · rationale · would_have_cost_bps · prevented_action",
        "acceptance": (
            "AC-1: Auto-blocks markdown when intent 0.72 received.\n"
            "AC-2: DMM bypassed — sees in Done bucket only.\n"
            "AC-3: Full rationale captured for audit trail."
        ),
        "status": "Q1 — In scope",
    },

    # --------- Group 2: Financial Guardrail (Agent 4) ---------
    {
        "id": "UC-04", "group": "Financial Guardrail",
        "name": "Validate OTB buffer guardrail",
        "actor": "The Finance Lead (Agent 4)",
        "system": "merml-mcp",
        "chapter": "Guardrail",
        "timestamp": "Sat 2:15 AM",
        "gate": "Zero Touch / HITL / Blocked",
        "threshold": "Pass: post-action OTB > 15% · HITL: 5–15% · Block: < 5%",
        "trigger": "Proposed action package lands from upstream agent (8)",
        "precond": "Weekly OTB position loaded · proposed action cost computed",
        "main_flow": (
            "1. Receive proposed action with $-impact.\n"
            "2. Compute post-action OTB buffer % = (OTB − action_cost) / weekly_budget.\n"
            "3. Compare to 15% / 5% thresholds.\n"
            "4. Return Pass / HITL / Block decision.\n"
            "5. If Pass → unblock downstream execution."
        ),
        "alt_flow": "If OTB data stale (>4h old), refuse and request refresh — never approve on stale data.",
        "postcond": "$190K headroom confirmed (demo) · downstream unblocked · LEDGER row written",
        "inputs": "Weekly OTB position · proposed action $-impact · weekly budget envelope",
        "outputs": "Pass/HITL/Block decision · post-action OTB% · headroom $",
        "ledger": "ts · agent_id · otb_before · action_cost · otb_after · buffer_pct · decision",
        "acceptance": (
            "AC-1: Returns Pass with $190K headroom in demo.\n"
            "AC-2: Decision returned within 2 minutes.\n"
            "AC-3: Routes to HITL if buffer 5–15% (none in demo)."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-05", "group": "Financial Guardrail",
        "name": "Validate margin impact guardrail",
        "actor": "The Finance Lead (Agent 4)",
        "system": "merml-mcp",
        "chapter": "Guardrail",
        "timestamp": "Sat 2:15 AM",
        "gate": "Zero Touch / HITL / Blocked",
        "threshold": "Pass: < 50 bps · HITL: 50–150 bps · Block: > 150 bps",
        "trigger": "Same proposed-action package as UC-04 — runs in parallel",
        "precond": "Plan margin loaded for category · action margin impact computable",
        "main_flow": (
            "1. Compute net margin impact in bps for the proposed action.\n"
            "2. Compare to 50 / 150 bps thresholds.\n"
            "3. Return Pass / HITL / Block.\n"
            "4. If any guardrail (UC-04..07) blocks → overall Block."
        ),
        "alt_flow": "Combine with UC-06 markdown depth check when action is a markdown.",
        "postcond": "Margin impact decision logged · combined with UC-04/06/07 for overall verdict",
        "inputs": "Plan margin · action margin · category cost basis",
        "outputs": "Margin impact bps · decision",
        "ledger": "ts · plan_margin · action_margin · delta_bps · decision",
        "acceptance": (
            "AC-1: < 50 bps in demo (rebalance, no new spend) → Pass.\n"
            "AC-2: Same decision package returns within 2 minutes.\n"
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-06", "group": "Financial Guardrail",
        "name": "Validate markdown depth guardrail",
        "actor": "The Finance Lead (Agent 4)",
        "system": "merml-mcp",
        "chapter": "Guardrail",
        "timestamp": "Sat 2:15 AM",
        "gate": "Zero Touch / HITL / Blocked",
        "threshold": "Pass: ≤ 10% · HITL: 10–20% (DMM) · Block: > 20%",
        "trigger": "Markdown proposal received (none in demo — already blocked at UC-03)",
        "precond": "Category markdown limits loaded · proposal includes depth %",
        "main_flow": (
            "1. Read proposed markdown depth %.\n"
            "2. Compare to 10% / 20% thresholds.\n"
            "3. Pass / route to DMM (HITL) / Block.\n"
            "4. Block at >20% — escalate to Chief Merchant via Your Call."
        ),
        "alt_flow": "If markdown depth = 0% (no markdown), auto-pass without DMM touch.",
        "postcond": "DMM routing decision · block decision if depth too aggressive",
        "inputs": "Proposed markdown depth · category historical norms",
        "outputs": "Pass/HITL/Block · routing target (DMM)",
        "ledger": "ts · category · depth_pct · decision · approver",
        "acceptance": (
            "AC-1: Routes 10–20% to DMM HITL queue.\n"
            "AC-2: Auto-blocks > 20% with Chief Merchant escalation.\n"
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-07", "group": "Financial Guardrail",
        "name": "Validate aggregate weekly spend guardrail",
        "actor": "The Finance Lead (Agent 4)",
        "system": "merml-mcp",
        "chapter": "Guardrail",
        "timestamp": "Sat 2:15 AM",
        "gate": "Zero Touch / HITL / Blocked",
        "threshold": "Pass: < $200K cumulative · HITL: $200K–$500K (Chief Merchant) · Block: > $500K",
        "trigger": "Each new proposed action — cumulative running total",
        "precond": "Week's prior approved actions tracked in merml-mcp",
        "main_flow": (
            "1. Sum approved + proposed action $-impact for the week.\n"
            "2. Compare cumulative to $200K / $500K thresholds.\n"
            "3. Pass / HITL to Chief Merchant / Block.\n"
            "4. This guardrail prevents 'death by a thousand cuts'."
        ),
        "alt_flow": "If close to threshold, surface as Watch in next brief.",
        "postcond": "Cumulative spend logged · CM escalation routing if 200K–500K",
        "inputs": "Cumulative weekly approved actions · proposed action cost",
        "outputs": "Pass/HITL/Block · cumulative $ · headroom",
        "ledger": "ts · cumulative_spend · proposed_add · decision",
        "acceptance": (
            "AC-1: Tracks cumulative across all category actions.\n"
            "AC-2: Routes 200K–500K to Chief Merchant.\n"
            "AC-3: Blocks individually-passing actions when aggregate exceeds $500K."
        ),
        "status": "Q1 — In scope",
    },

    # --------- Group 3: Execute (Agent 8) ---------
    {
        "id": "UC-08", "group": "Execute",
        "name": "Execute inventory rebalance (existing stock)",
        "actor": "The Operations Lead (Agent 8)",
        "system": "scml-mcp",
        "chapter": "Act",
        "timestamp": "Sat 3:45 AM",
        "gate": "Zero Touch",
        "threshold": "Existing DC inventory only · no new financial commitment",
        "trigger": "Agent 4 returns Pass on all guardrails for a rebalance proposal",
        "precond": "Inventory snapshot current · sourcing guardrails active · DC capacity confirmed",
        "main_flow": (
            "1. Identify SKU shortage by store × DC.\n"
            "2. Compute optimal source DCs (existing stock only, not new POs).\n"
            "3. Generate transfer orders.\n"
            "4. Submit to TMS for execution.\n"
            "5. Confirm pickup + ETA."
        ),
        "alt_flow": "If no DC has stock, escalate to UC-09 (new PO).",
        "postcond": "24 SKUs moved to 8 Midwest stores (demo) · LEDGER row written · Done bucket entry",
        "inputs": "Shortage list · DC inventory · transfer cost matrix · sourcing guardrails",
        "outputs": "Transfer orders · pickup ETAs · Done-bucket entry",
        "ledger": "ts · sku_count · stores · source_dcs · transfer_cost · eta",
        "acceptance": (
            "AC-1: Moves 24 SKUs across 8 stores autonomously in demo.\n"
            "AC-2: No human approval required (Zero Touch).\n"
            "AC-3: Completes within 1 hour of guardrail Pass."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-09", "group": "Execute",
        "name": "Stage new purchase order (HITL)",
        "actor": "The Operations Lead (Agent 8) → Planning Lead | VP Supply Chain | Chief Merchant",
        "system": "scml-mcp",
        "chapter": "Act",
        "timestamp": "Sat 3:50 AM",
        "gate": "HITL",
        "threshold": "Route ≤ $50K → Planning Lead · $50K–$100K → VP SC · > $100K → Chief Merchant (Block + escalate)",
        "trigger": "Rebalance cannot fully cover shortage — new PO required",
        "precond": "Approved supplier list loaded · OTB headroom confirmed via UC-04",
        "main_flow": (
            "1. Compute remaining shortage after rebalance.\n"
            "2. Source from approved supplier list.\n"
            "3. Build PO with $-value and ETA.\n"
            "4. Route to default approver based on $-value.\n"
            "5. Stage as 'Your Call' in Monday brief."
        ),
        "alt_flow": "If $-value > $100K → block + escalate to Chief Merchant immediately.",
        "postcond": "$62K PO staged in Your Call (demo) · 2 paths pre-modelled · routed to VP SC",
        "inputs": "Shortage residual · supplier list · supplier ETAs · cost matrix",
        "outputs": "PO record · 2 paths (approve / hold) · revenue recovery estimate · routing manifest",
        "ledger": "ts · po_value · supplier · approver · path_a_recovery · path_b_lag",
        "acceptance": (
            "AC-1: Stages $62K PO with VP SC routing in demo.\n"
            "AC-2: Pre-models Path A ($180K recovery) vs Path B (3-week lag).\n"
            "AC-3: Sends Adaptive Card to approver within 5 minutes."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-10", "group": "Execute",
        "name": "Activate secondary supplier",
        "actor": "The Operations Lead (Agent 8)",
        "system": "scml-mcp",
        "chapter": "Act",
        "timestamp": "Sat 3:50 AM",
        "gate": "Zero Touch (pre-approved) | HITL (new supplier)",
        "threshold": "Pre-approved list = ZT · new supplier = HITL to Chief Merchant",
        "trigger": "Primary supplier disrupted — secondary needed",
        "precond": "Pre-approved supplier list current · supplier health monitoring active",
        "main_flow": (
            "1. Detect primary supplier disruption signal.\n"
            "2. Match secondary supplier from pre-approved list.\n"
            "3. If matched → activate (Zero Touch).\n"
            "4. If no match → escalate to Chief Merchant (HITL)."
        ),
        "alt_flow": "If new supplier vetting required, surface in Watch with diligence package.",
        "postcond": "Secondary activated or HITL queue populated · audit row",
        "inputs": "Disruption signal · pre-approved list · supplier ratings",
        "outputs": "Activation order · audit log · Watch entry if new supplier",
        "ledger": "ts · primary · secondary · pre_approved_flag · decision",
        "acceptance": (
            "AC-1: Auto-activates from pre-approved list.\n"
            "AC-2: Routes new suppliers to Chief Merchant via Watch.\n"
        ),
        "status": "Q1 — In scope",
    },

    # --------- Group 4: Brief Synthesis (Agent 13.3) ---------
    {
        "id": "UC-11", "group": "Brief Synthesis",
        "name": "Synthesize Done bucket",
        "actor": "The Briefer (Agent 13.3 ADAPT)",
        "system": "merml-mcp",
        "chapter": "Brief",
        "timestamp": "Sun 9:00 PM",
        "gate": "ACK",
        "threshold": "All Zero-Touch decisions executed during the week",
        "trigger": "Sunday evening synthesis run — picks up all overnight ledger rows",
        "precond": "All week's LEDGER rows committed · synthesis template loaded",
        "main_flow": (
            "1. Pull all ZT-completed actions from LEDGER.\n"
            "2. Group by category and outcome.\n"
            "3. Render to Done tier of Monday brief.\n"
            "4. Include $-impact summary and audit-trail link."
        ),
        "alt_flow": "If no Done items, render 'No autonomous actions this week' line.",
        "postcond": "Done tier of brief populated · 3 demo items rendered",
        "inputs": "Week's LEDGER rows · classification rules",
        "outputs": "Done tier markdown + audit links",
        "ledger": "synthesis_run_id · done_count · total_value",
        "acceptance": (
            "AC-1: Demo brief shows: 24 SKUs rebalanced · markdown blocked · OTB confirmed.\n"
            "AC-2: Each item links to LEDGER audit row.\n"
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-12", "group": "Brief Synthesis",
        "name": "Synthesize Your Call bucket",
        "actor": "The Briefer (Agent 13.3 ADAPT) → Chief Merchant",
        "system": "merml-mcp",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM",
        "gate": "HITL",
        "threshold": "All HITL items routed to Chief Merchant · response window: Wed EOD",
        "trigger": "Monday brief delivery — all staged HITL items",
        "precond": "Adaptive Card template registered in Teams · response webhook live",
        "main_flow": (
            "1. Pull all HITL-pending items routed to Chief Merchant.\n"
            "2. For each: include 2 modelled paths (A approve, B hold).\n"
            "3. Render Adaptive Card.\n"
            "4. Post to Teams.\n"
            "5. Open response window — escalate to next tier on timeout."
        ),
        "alt_flow": "If no Your Call items, render 'No decisions waiting' tile.",
        "postcond": "Adaptive Card delivered · response window open · CM notified",
        "inputs": "Pending HITL queue · path-modelling outputs · CM Teams identity",
        "outputs": "Adaptive Card · response webhook URL · audit entry",
        "ledger": "card_id · ts_sent · po_value · path_a · path_b · response_window",
        "acceptance": (
            "AC-1: $62K PO appears in Your Call with Path A/B in demo.\n"
            "AC-2: Adaptive Card lands in CM Teams by 8:00 AM Monday.\n"
            "AC-3: 'Approve A' click writes back to LEDGER and notifies Agent 8."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-13", "group": "Brief Synthesis",
        "name": "Synthesize Watch bucket",
        "actor": "The Briefer (Agent 13.3 ADAPT)",
        "system": "merml-mcp",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM",
        "gate": "ACK",
        "threshold": "Forward-looking signals · 1st or 2nd week below threshold",
        "trigger": "Monday brief — informational tier",
        "precond": "Watch rules loaded · trend signals computed",
        "main_flow": (
            "1. Pull signals flagged as forward-looking (not yet actionable).\n"
            "2. Render 'Watch' tier with trend + escalation path.\n"
            "3. No Adaptive Card — narrative only.\n"
            "4. Includes Q5 competitive signals (logged but not acted on in Q1)."
        ),
        "alt_flow": "If signal escalates next week (3rd consecutive miss), promote to Your Call.",
        "postcond": "Watch tier in brief · CM informed · no action requested",
        "inputs": "Trend signals · prior-week Watch state · escalation rules",
        "outputs": "Watch tier markdown",
        "ledger": "watch_count · signals · escalation_thresholds",
        "acceptance": (
            "AC-1: Demo shows Women's Activewear at −6% (2nd week, monitoring).\n"
            "AC-2: Includes 'escalates to Your Call if miss continues next WBR'.\n"
        ),
        "status": "Q1 — In scope",
    },

    # --------- Group 5: Human Approvals ---------
    {
        "id": "UC-14", "group": "Human Approvals",
        "name": "Approve PO ≤ $50K (Planning Lead)",
        "actor": "Planning Lead",
        "system": "Teams · merml-mcp ledger writeback",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM (24h window)",
        "gate": "HITL",
        "threshold": "Default approver: Planning Lead · 24h response · escalation: VP Supply Chain",
        "trigger": "Agent 8 stages a PO ≤ $50K and routes via UC-09",
        "precond": "Planning Lead identity in Teams · Adaptive Card schema versioned · escalation routing configured",
        "main_flow": (
            "1. Receive Adaptive Card in Teams.\n"
            "2. Review PO context, paths, and audit trail.\n"
            "3. Click Approve / Hold.\n"
            "4. Webhook writes to LEDGER and notifies Agent 8 to execute or hold."
        ),
        "alt_flow": "No response in 24h → auto-escalate to VP Supply Chain. Reset 24h window.",
        "postcond": "PO approved/held · Agent 8 notified · LEDGER updated · audit trail persisted",
        "inputs": "Adaptive Card payload · PO context · 2 paths · audit links",
        "outputs": "Decision (Approve|Hold) · timestamp · approver identity",
        "ledger": "ts · approver · decision · po_value · response_minutes",
        "acceptance": (
            "AC-1: Card renders in Teams within 5 minutes of staging.\n"
            "AC-2: One-click Approve writes to LEDGER and unblocks Agent 8.\n"
            "AC-3: 24h timeout escalates to VP SC."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-15", "group": "Human Approvals",
        "name": "Approve PO $50K–$100K (VP Supply Chain)",
        "actor": "VP Supply Chain",
        "system": "Teams · merml-mcp ledger writeback",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM (48h window)",
        "gate": "HITL",
        "threshold": "Default approver: VP SC · 48h response · escalation: Chief Merchant",
        "trigger": "Agent 8 stages a PO $50K–$100K (e.g., $62K demo)",
        "precond": "VP SC Teams identity · escalation routing live",
        "main_flow": (
            "1. VP receives Adaptive Card with PO ($62K) and 2 paths.\n"
            "2. Reviews supplier, ETA, revenue recovery model.\n"
            "3. Approves Path A (recover by Wk 18) or B (hold).\n"
            "4. Webhook writes back to LEDGER, notifies Agent 8."
        ),
        "alt_flow": "No response in 48h → escalate to Chief Merchant.",
        "postcond": "$62K PO approved · 8 SKUs procured · Agent 8 unblocked",
        "inputs": "Adaptive Card · PO context · path models · audit links",
        "outputs": "Approval decision · path selected · timestamp",
        "ledger": "ts · approver · decision · path · po_value",
        "acceptance": (
            "AC-1: Demo $62K PO routes to VP SC.\n"
            "AC-2: 48h escalation timer fires correctly.\n"
            "AC-3: Approval triggers Agent 8 to execute the PO with chosen supplier."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-16", "group": "Human Approvals",
        "name": "Approve markdown 10–20% (Category DMM)",
        "actor": "Category DMM",
        "system": "Teams · merml-mcp ledger writeback",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM (24h window)",
        "gate": "HITL",
        "threshold": "Default approver: DMM · 24h response · escalation: Chief Merchant",
        "trigger": "Agent 4 routes a markdown 10–20% via UC-06",
        "precond": "DMM identity in Teams · price-system writeback live",
        "main_flow": (
            "1. DMM receives Adaptive Card with markdown depth + intent score.\n"
            "2. Reviews margin impact and demand-intent context.\n"
            "3. Approves or holds.\n"
            "4. Approved markdown writes to pricing system + LEDGER."
        ),
        "alt_flow": "No response in 24h → escalate to Chief Merchant.",
        "postcond": "Markdown approved/held · pricing system updated · brief Done bucket reflects",
        "inputs": "Markdown proposal · intent score · margin model · price-elasticity history",
        "outputs": "Decision · approver · timestamp",
        "ledger": "ts · approver · category · depth_pct · decision",
        "acceptance": (
            "AC-1: DMM bypassed in demo (intent healthy → blocked at UC-03).\n"
            "AC-2: Routing works when markdown 10–20% reaches this UC.\n"
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-17", "group": "Human Approvals",
        "name": "Approve aggregate weekly spend $200K–$500K (Chief Merchant)",
        "actor": "Chief Merchant",
        "system": "Teams · merml-mcp ledger writeback",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM (48h window)",
        "gate": "HITL",
        "threshold": "Cumulative weekly action spend · 48h response · no further escalation (Block resurfaces next week)",
        "trigger": "Agent 4 UC-07 cumulative crosses $200K",
        "precond": "Aggregate guardrail tracking active · CM identity in Teams",
        "main_flow": (
            "1. Aggregate alert fires when cumulative action spend reaches $200K.\n"
            "2. Adaptive Card surfaces remaining proposed actions and historical cumulative.\n"
            "3. CM approves all/some/none.\n"
            "4. Approval feeds back to UC-07 to re-evaluate cumulative."
        ),
        "alt_flow": "No response in 48h → block remaining actions, resurface next week.",
        "postcond": "Cumulative cap respected · approved actions execute · blocked actions queue",
        "inputs": "Cumulative spend · proposed actions list · category breakdown",
        "outputs": "Per-action approve/hold matrix · audit trail",
        "ledger": "ts · approver · cumulative · approved_set · blocked_set",
        "acceptance": (
            "AC-1: Triggers correctly at $200K cumulative.\n"
            "AC-2: Allows partial approval (some actions, not others).\n"
            "AC-3: 48h timeout blocks remaining."
        ),
        "status": "Q1 — In scope",
    },

    # --------- Group 6: Cross-cutting ---------
    {
        "id": "UC-18", "group": "Cross-cutting",
        "name": "Escalate on no-response timeout",
        "actor": "Routing engine (system)",
        "system": "merml-mcp scheduler",
        "chapter": "Brief / post-Brief",
        "timestamp": "Per response-window expiry",
        "gate": "HITL routing",
        "threshold": "Per UC-14..17: 24h or 48h windows · escalation chain pre-configured",
        "trigger": "Adaptive Card response window expires without click",
        "precond": "Response-window timer attached to each card · escalation chain in context layer",
        "main_flow": (
            "1. Timer fires on window expiry.\n"
            "2. Look up next approver in chain.\n"
            "3. Re-route card with 'escalated from [original approver]' marker.\n"
            "4. Reset response window for new approver.\n"
            "5. Final block if Chief Merchant tier expires (resurfaces next week)."
        ),
        "alt_flow": "If approver out-of-office (calendar integration), route directly to delegate.",
        "postcond": "Decision-rights routing self-heals on absence · audit captures escalation path",
        "inputs": "Card-id · response-window state · approver chain",
        "outputs": "Re-routed card · escalation audit row",
        "ledger": "ts · card_id · from_approver · to_approver · reason",
        "acceptance": (
            "AC-1: 24h windows escalate Planning Lead → VP SC · DMM → CM.\n"
            "AC-2: 48h windows escalate VP SC → CM.\n"
            "AC-3: CM tier blocks (no further escalation) and resurfaces next week."
        ),
        "status": "Q1 — In scope",
    },
    {
        "id": "UC-19", "group": "Cross-cutting",
        "name": "Log decision to LEDGER (WORM)",
        "actor": "All agents (cross-cutting)",
        "system": "LEDGER (WORM) — write-once-read-many audit ledger",
        "chapter": "All chapters",
        "timestamp": "Continuous",
        "gate": "Zero Touch (system)",
        "threshold": "Every agent decision · 14 mandatory fields",
        "trigger": "Any agent emits a decision (action, block, route, ack)",
        "precond": "LEDGER schema v1 deployed · WORM enforcement on storage",
        "main_flow": (
            "1. Agent computes decision.\n"
            "2. Builds 14-field record: ts · agent_id · scenario · category · decision · gate · threshold · "
            "rationale · inputs_hash · outputs_hash · upstream_id · approver · response_ms · ledger_version.\n"
            "3. Append to LEDGER (WORM — no overwrites).\n"
            "4. Hash chain links to previous row in scenario."
        ),
        "alt_flow": "If LEDGER write fails, agent halts and emits alert (data integrity > availability).",
        "postcond": "Immutable audit trail · regulatory-grade evidence · brief deep-links",
        "inputs": "Decision record (14 fields)",
        "outputs": "LEDGER row + hash · provenance link",
        "ledger": "Self — this UC defines the schema",
        "acceptance": (
            "AC-1: Every demo decision produces a LEDGER row.\n"
            "AC-2: Hash chain validates end-to-end.\n"
            "AC-3: Brief items deep-link to LEDGER rows."
        ),
        "status": "Q1 — In scope (foundation)",
    },
    {
        "id": "UC-20", "group": "Cross-cutting",
        "name": "Deliver Adaptive Card to Teams",
        "actor": "Briefer (13.3) + Microsoft Teams",
        "system": "Teams Bot Framework · cxml-mcp identity",
        "chapter": "Brief",
        "timestamp": "Mon 8:00 AM",
        "gate": "HITL delivery",
        "threshold": "All Your Call items render as Adaptive Cards · Done/Watch as text tiles",
        "trigger": "Brief synthesis (UC-12) emits a Your Call item",
        "precond": "Teams app installed for approver tenant · webhook secret rotated · card schema validated",
        "main_flow": (
            "1. Render Adaptive Card from Your Call payload.\n"
            "2. Resolve approver Teams identity.\n"
            "3. Post to approver chat.\n"
            "4. Listen for action.submit webhook.\n"
            "5. On click → write to LEDGER (UC-19) and notify originating agent."
        ),
        "alt_flow": "If Teams API down, fall back to email with deep-link to web approval surface.",
        "postcond": "Card delivered · click → action chain triggered · audit row written",
        "inputs": "Your Call payload · approver identity · card template",
        "outputs": "Card_id · delivery confirmation · action webhook payload",
        "ledger": "card_id · ts_sent · ts_responded · response_action · approver",
        "acceptance": (
            "AC-1: Demo $62K PO card renders in CM Teams Mon 8:00 AM.\n"
            "AC-2: One-click Approve writes back within 2 seconds.\n"
            "AC-3: Email fallback triggers if Teams API > 1 min outage."
        ),
        "status": "Q1 — In scope",
    },
]

# =====================================================================
# SHEET 2 — Use Case Index
# =====================================================================
def build_index(wb: Workbook):
    ws = wb.create_sheet("Use Case Index")
    headers = ["UC ID", "Group", "Name", "Primary Actor / Persona", "System", "Chapter", "Timestamp", "Gate", "Status"]
    rows = []
    for uc in USE_CASES:
        rows.append([
            uc["id"], uc["group"], uc["name"], uc["actor"], uc["system"],
            uc["chapter"], uc["timestamp"], uc["gate"], uc["status"]
        ])
    write_table(ws, headers, rows,
                widths=[10, 18, 42, 38, 22, 16, 18, 26, 18],
                freeze="A2", row_height=30)

# =====================================================================
# SHEET 3 — Use Case Detail
# =====================================================================
def build_detail(wb: Workbook):
    ws = wb.create_sheet("Use Case Detail")
    headers = [
        "UC ID", "Group", "Name", "Primary Actor / Persona", "System (Agent · MCP)", "Chapter",
        "Timestamp", "Gate", "Threshold / Business Rule", "Trigger", "Preconditions",
        "Main Flow", "Alternate / Exception Flow", "Postconditions",
        "Data Inputs", "Data Outputs", "LEDGER fields", "Acceptance Criteria", "Q1 Status",
    ]
    rows = []
    for uc in USE_CASES:
        rows.append([
            uc["id"], uc["group"], uc["name"], uc["actor"], uc["system"], uc["chapter"],
            uc["timestamp"], uc["gate"], uc["threshold"], uc["trigger"], uc["precond"],
            uc["main_flow"], uc["alt_flow"], uc["postcond"],
            uc["inputs"], uc["outputs"], uc["ledger"], uc["acceptance"], uc["status"],
        ])
    widths = [10, 18, 36, 36, 24, 14, 16, 22, 30, 36, 30, 60, 38, 38, 36, 32, 36, 50, 18]
    write_table(ws, headers, rows, widths=widths, freeze="C2", row_height=160)
    # taller header
    ws.row_dimensions[1].height = 36

# =====================================================================
# SHEET 4 — Persona x UC Matrix
# =====================================================================
PERSONAS = [
    ("Chief Merchant",          "Human · primary"),
    ("Planning Lead",           "Human · PO ≤ $50K"),
    ("VP Supply Chain",         "Human · PO $50–100K"),
    ("Category DMM",            "Human · markdown 10–20%"),
    ("The Analyst (Agent 13)",          "Agent · merml-mcp"),
    ("The Demand Checker (Agent 1)",    "Agent · cxml-mcp"),
    ("The Finance Lead (Agent 4)",      "Agent · merml-mcp"),
    ("The Operations Lead (Agent 8)",   "Agent · scml-mcp"),
    ("The Briefer (Agent 13.3)",        "Agent · merml-mcp"),
]

# Hand-built persona x UC matrix.  Cells contain role: O (owner), A (approver), C (consumer/notified), B (bypassed), E (escalation).
MATRIX = {
    # uc_id: dict of persona -> role (omitted = blank)
    "UC-01": {"The Analyst (Agent 13)": "O", "Chief Merchant": "C"},
    "UC-02": {"The Demand Checker (Agent 1)": "O", "The Analyst (Agent 13)": "C"},
    "UC-03": {"The Analyst (Agent 13)": "O", "Category DMM": "B", "Chief Merchant": "C"},
    "UC-04": {"The Finance Lead (Agent 4)": "O", "The Operations Lead (Agent 8)": "C"},
    "UC-05": {"The Finance Lead (Agent 4)": "O"},
    "UC-06": {"The Finance Lead (Agent 4)": "O", "Category DMM": "A"},
    "UC-07": {"The Finance Lead (Agent 4)": "O", "Chief Merchant": "A"},
    "UC-08": {"The Operations Lead (Agent 8)": "O", "Planning Lead": "B", "Chief Merchant": "C"},
    "UC-09": {"The Operations Lead (Agent 8)": "O", "Planning Lead": "A", "VP Supply Chain": "A", "Chief Merchant": "E"},
    "UC-10": {"The Operations Lead (Agent 8)": "O", "Chief Merchant": "A"},
    "UC-11": {"The Briefer (Agent 13.3)": "O", "Chief Merchant": "C"},
    "UC-12": {"The Briefer (Agent 13.3)": "O", "Chief Merchant": "A"},
    "UC-13": {"The Briefer (Agent 13.3)": "O", "Chief Merchant": "C"},
    "UC-14": {"Planning Lead": "A", "VP Supply Chain": "E", "The Operations Lead (Agent 8)": "C"},
    "UC-15": {"VP Supply Chain": "A", "Chief Merchant": "E", "The Operations Lead (Agent 8)": "C"},
    "UC-16": {"Category DMM": "A", "Chief Merchant": "E", "The Finance Lead (Agent 4)": "C"},
    "UC-17": {"Chief Merchant": "A", "The Finance Lead (Agent 4)": "C"},
    "UC-18": {"Planning Lead": "C", "VP Supply Chain": "C", "Category DMM": "C", "Chief Merchant": "E"},
    "UC-19": {p: "C" for p, _ in PERSONAS},  # everyone consumes the audit ledger
    "UC-20": {"The Briefer (Agent 13.3)": "O", "Chief Merchant": "A", "Planning Lead": "A", "VP Supply Chain": "A", "Category DMM": "A"},
}

ROLE_FILL = {
    "O": ("86BC25", "FFFFFF"),  # Owner — green
    "A": ("F0A839", "0D2818"),  # Approver — amber
    "C": ("E1E8F0", "0D2818"),  # Consumer/notified — light grey
    "E": ("B794F4", "0D2818"),  # Escalation — purple
    "B": ("BDBDBD", "0D2818"),  # Bypassed — grey
}

def build_matrix(wb: Workbook):
    ws = wb.create_sheet("Persona x UC Matrix")

    # row 1: persona names, row 2: persona role
    ws.cell(row=1, column=1, value="UC ID")
    ws.cell(row=1, column=2, value="Use Case")
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=2, value="")
    for j, (pname, ptype) in enumerate(PERSONAS, start=3):
        ws.cell(row=1, column=j, value=pname)
        ws.cell(row=2, column=j, value=ptype)

    for c in ws[1]:
        style_header(c)
    for c in ws[2]:
        c.font = Font(name="Calibri", italic=True, size=9, color="606060")
        c.fill = PatternFill("solid", fgColor=ALT_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border_all

    for i, uc in enumerate(USE_CASES, start=3):
        ws.cell(row=i, column=1, value=uc["id"])
        ws.cell(row=i, column=2, value=uc["name"])
        for c in ws[i]:
            style_body(c, alt=(i % 2 == 0))
        for j, (pname, _) in enumerate(PERSONAS, start=3):
            role = MATRIX.get(uc["id"], {}).get(pname, "")
            cell = ws.cell(row=i, column=j, value=role)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
            cell.font = Font(name="Calibri", size=10, bold=True)
            if role:
                bg, fg = ROLE_FILL[role]
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fg)

    autosize(ws, [10, 38] + [22] * len(PERSONAS))
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "C3"

    # Legend below
    legend_start = len(USE_CASES) + 5
    ws.cell(row=legend_start, column=1, value="Legend").font = Font(bold=True, color=INK_DARK)
    legend_items = [
        ("O", "Owner — does the work"),
        ("A", "Approver — gates the decision (HITL)"),
        ("E", "Escalation — receives if primary times out"),
        ("C", "Consumer — informed/sees in brief"),
        ("B", "Bypassed — would normally approve, but auto-resolved this run"),
    ]
    for k, (sym, desc) in enumerate(legend_items, start=1):
        sc = ws.cell(row=legend_start + k, column=1, value=sym)
        dc = ws.cell(row=legend_start + k, column=2, value=desc)
        bg, fg = ROLE_FILL[sym]
        sc.fill = PatternFill("solid", fgColor=bg)
        sc.font = Font(name="Calibri", size=10, bold=True, color=fg)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        dc.font = Font(name="Calibri", size=10)

# =====================================================================
# SHEET 5 — Decision Rights
# =====================================================================
DR_ROWS = [
    # (Agent, Action, Zero Touch, HITL, Block, Rationale)
    ("Agent 8 — Allocation & Replenishment",
        "Inventory rebalance (existing stock)",
        "Any rebalance within existing DC inventory — no new PO",
        "—", "—",
        "No new financial commitment. Overnight execution within sourcing guardrails is safe to automate."),
    ("Agent 8 — Allocation & Replenishment",
        "New purchase order",
        "—",
        "PO ≤ $50K → Planning Lead\nPO $50K–$100K → VP Supply Chain",
        "PO > $100K",
        "New spend commitment requires human gate. Adjust per client OTB norms."),
    ("Agent 8 — Allocation & Replenishment",
        "Secondary supplier activation",
        "Pre-approved supplier list only",
        "New / unapproved supplier → Chief Merchant",
        "—",
        "Pre-approved secondary suppliers are routine. New supplier relationships require merchant oversight."),
    ("Agent 4 — MFP Guardrail Governor",
        "OTB buffer",
        "Post-action OTB > 15% of weekly budget",
        "OTB buffer 5–15%",
        "OTB buffer < 5%",
        "15% buffer preserves flexibility. Sub-5% creates financial exposure."),
    ("Agent 4 — MFP Guardrail Governor",
        "Margin impact",
        "Net impact < 50 bps vs. plan",
        "50–150 bps impact",
        "> 150 bps impact",
        "50 bps is within weekly noise. 150 bps moves the P&L — Chief Merchant should own."),
    ("Agent 4 — MFP Guardrail Governor",
        "Markdown depth",
        "≤ 10%",
        "10–20% → Category DMM",
        "> 20%",
        "10% is canonical. Deeper markdowns signal distress pricing."),
    ("Agent 4 — MFP Guardrail Governor",
        "Total weekly action spend",
        "Cumulative < $200K",
        "$200K–$500K → Chief Merchant",
        "> $500K",
        "Aggregate guardrail prevents many small actions exceeding the week's budget envelope."),
    ("Agents 1 + 13 — Consumer Signals + Performance Analytics",
        "All outputs (signal, anomaly, context routing)",
        "ACK — surfaces to brief",
        "—", "—",
        "Observe-and-route only. Both agents are non-actuating in Q1."),
    ("Agent 13.3 — Brief Synthesis (ADAPT)",
        "Done — autonomous actions executed",
        "ACK — confirmation only",
        "—", "—",
        "Done bucket of Monday brief — no action required."),
    ("Agent 13.3 — Brief Synthesis (ADAPT)",
        "Your Call — actions awaiting approval",
        "—",
        "HITL — Chief Merchant via Teams Adaptive Card",
        "—",
        "Response window Wed EOD."),
    ("Agent 13.3 — Brief Synthesis (ADAPT)",
        "Watch — forward signals, no action required",
        "ACK — informational",
        "—", "—",
        "No card sent — narrative only. Includes Q5 competitive signals (logged not actioned in Q1)."),
]

def build_decision_rights(wb: Workbook):
    ws = wb.create_sheet("Decision Rights")
    headers = ["Agent", "Action", "Zero Touch (Pass)", "HITL", "Block", "Rationale"]
    write_table(ws, headers, DR_ROWS,
                widths=[42, 38, 36, 38, 22, 56],
                freeze="A2", row_height=60)

# =====================================================================
# SHEET 6 — WBR Timeline
# =====================================================================
TIMELINE_ROWS = [
    ("Fri 9:43 PM", "Detect", "The Analyst (Agent 13)",
     "Apparel missed plan for the second week running",
     "−12% vs. target. 24 SKUs understocked across Midwest stores. Ranked the #1 exception this week.",
     "ACK — observe-and-route only", "UC-01"),
    ("Sat 12:30 AM", "Diagnose", "The Demand Checker (Agent 1)",
     "Demand was checked first — and it was healthy",
     "Consumer intent for Apparel: 0.72. The miss wasn't a demand problem — it was a supply problem.",
     "ACK — diagnosis returned", "UC-02"),
    ("Sat 12:32 AM", "Diagnose", "The Analyst (Agent 13) → DMM bypassed",
     "Markdown intervention blocked",
     "Intent confirmed healthy → markdown auto-blocked. Margin protected. DMM bypassed.",
     "Blocked", "UC-03"),
    ("Sat 2:15 AM", "Guardrail", "The Finance Lead (Agent 4)",
     "Budget check passed — $190K of headroom confirmed",
     "Rebalance was within OTB position. Margin impact <50 bps. Safe to proceed without escalation.",
     "Zero Touch — all 4 guardrails Pass", "UC-04, UC-05, UC-06, UC-07"),
    ("Sat 3:45 AM", "Act", "The Operations Lead (Agent 8)",
     "24 SKUs rebalanced across 8 Midwest stores",
     "Inventory moved from East Coast DCs. Within $50K threshold. Done while you slept.",
     "Zero Touch — within rebalance threshold", "UC-08"),
    ("Sat 3:50 AM", "Act", "The Operations Lead (Agent 8)",
     "$62K expedited PO staged for remaining 8 SKUs",
     "Crossed Planning Lead's $50K ceiling — routes to VP Supply Chain via Adaptive Card.",
     "HITL — VP SC default approver", "UC-09"),
    ("Sun 9:00 PM", "Synthesize", "The Briefer (Agent 13.3 ADAPT)",
     "Synthesis run picks up week's LEDGER rows",
     "Sorts into Done / Your Call / Watch tiers. Renders Adaptive Cards for Your Call items.",
     "Zero Touch (synthesis)", "UC-11, UC-12, UC-13"),
    ("Mon 8:00 AM", "Brief", "The Briefer (Agent 13.3) → Chief Merchant",
     "Brief ready — one action taken, one decision waiting",
     "Done: 24 SKUs rebalanced + markdown blocked. Your Call: $62K PO + Path A/B. Watch: Women's Activewear −6%.",
     "HITL — CM Adaptive Card delivery", "UC-12, UC-20"),
    ("Mon–Tue", "Decide", "Chief Merchant",
     "Approves Path A on Adaptive Card",
     "Recovers 8 SKUs by Wk 18 · est. $180K revenue recovery. LEDGER updated. Agent 8 executes.",
     "HITL — CM decision", "UC-15 / UC-17 / UC-19"),
]

def build_timeline(wb: Workbook):
    ws = wb.create_sheet("WBR Timeline")
    headers = ["Timestamp", "Chapter", "Persona / Owner", "Headline", "What happened", "Gate", "Use Case(s)"]
    write_table(ws, headers, TIMELINE_ROWS,
                widths=[14, 14, 38, 40, 60, 30, 22],
                freeze="A2", row_height=66)

# =====================================================================
# Build it
# =====================================================================
def main():
    wb = Workbook()
    build_cover(wb)
    build_index(wb)
    build_detail(wb)
    build_matrix(wb)
    build_decision_rights(wb)
    build_timeline(wb)
    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Use cases: {len(USE_CASES)}")
    print(f"Personas in matrix: {len(PERSONAS)}")
    print(f"Decision-rights rows: {len(DR_ROWS)}")
    print(f"Timeline rows: {len(TIMELINE_ROWS)}")

if __name__ == "__main__":
    main()

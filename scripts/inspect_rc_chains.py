"""Survey the RC Scenario Chains xlsx so we can pick the chain that aligns with Q1 use cases."""
from __future__ import annotations
import io, sys
from openpyxl import load_workbook
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = r"C:\Stage\Clients\Industries\APEX\docs\reference\Retail and Consumer\APEX-RC-Scenario-Chains.xlsx"
wb = load_workbook(XLSX, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
print()

for name in wb.sheetnames:
    ws = wb[name]
    print("=" * 100)
    print(f"SHEET: {name}  ({ws.max_row} rows × {ws.max_column} cols)")
    print("=" * 100)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("(empty)")
        continue
    headers = rows[0]
    print("HEADERS:", headers)
    # print first 8 data rows truncated
    for i, r in enumerate(rows[1:9], start=1):
        cells = [str(c)[:60] if c is not None else "" for c in r]
        print(f"  R{i:3}: {cells}")
    if len(rows) > 9:
        print(f"  ... +{len(rows)-9} rows")
    print()

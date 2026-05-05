"""Build the master APEX Scenarios + Services Excel workbook.

Walks docs/scenarios/<PRACTICE>/<domain>/<scenario-folder>/ to enumerate
every scenario ID, joins against the catalog JSONs for full metadata,
and produces a multi-sheet workbook with:

  Sheet 1 · All Scenarios        — every scenario from the filesystem
  Sheet 2 · Featured Scenarios   — the 35 with full Scenario→KPI chain
  Sheet 3 · Compact Catalog       — the 723 compact rows
  Sheet 4 · Services              — unique service IDs across the corpus
  Sheet 5 · Summary               — counts by Practice and Practice×Domain
  Sheet 6 · Domain reference      — 13 domain code mapping

Output: docs/scenarios/_catalog/APEX-Scenarios-Services-Master.xlsx
"""

from __future__ import annotations

import io
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(r"C:\Stage\Clients\Industries\APEX")
SCEN_ROOT = ROOT / "docs/scenarios"
CAT_DIR   = SCEN_ROOT / "_catalog"
OUT       = CAT_DIR / "APEX-Scenarios-Services-Master.xlsx"

PRACTICES = ["RC", "HLS", "ER", "AXLE", "TMT", "TH", "ICE"]
PRACTICE_NAMES = {
    "RC":   "Retail & Consumer",
    "HLS":  "Healthcare & Life Sciences",
    "ER":   "Energy & Resources",
    "AXLE": "Industrial Manufacturing (Automotive-weighted)",
    "TMT":  "Technology, Media & Telecom",
    "TH":   "Travel & Hospitality",
    "ICE":  "Industrial & Commercial Equipment",
}

DOMAIN_CODES = {
    "clinical-care":          "CLIN",
    "network-infrastructure": "NET",
    "engineering-rd":         "ENG",
    "supply-chain":           "SUPCHN",
    "asset-maintenance":      "ASSET",
    "quality-compliance":     "QUAL",
    "risk-fraud-security":    "RISK",
    "pricing-revenue":        "PRICE",
    "customer-experience":    "CX",
    "marketing-growth":       "MKTG",
    "operations-workforce":   "OPS",
    "channel-partner-dealer": "CHAN",
    "other":                  "OTHER",
}

# ---- styling -----------------------------------------------------------

HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="0B0B0F")
HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

BAND_FILL = PatternFill("solid", fgColor="F3F3F5")

THIN = Side(style="thin", color="D4D4D7")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CODE_FONT  = Font(name="Consolas", size=10, color="6648B0", bold=True)


def write_headers(ws, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border = CELL_BORDER
        if widths:
            ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_row(ws, row_idx, values, *, code_cols=()):
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=j, value=v if v is not None else "")
        c.alignment = CELL_ALIGN
        c.border = CELL_BORDER
        if j in code_cols:
            c.font = CODE_FONT
        else:
            c.font = CELL_FONT
        if row_idx % 2 == 0:
            c.fill = BAND_FILL


# ---- load source data --------------------------------------------------

def load_featured():
    return json.load(open(CAT_DIR / "featured-chains.json", encoding="utf-8"))


def load_compact():
    return json.load(open(CAT_DIR / "scenario-library.json", encoding="utf-8"))


def walk_filesystem():
    """Yield (practice, domain, domain_code, folder_name, scenario_id, slug)."""
    rows = []
    for prac in PRACTICES:
        prac_dir = SCEN_ROOT / prac
        if not prac_dir.exists():
            continue
        for domain_dir in sorted(prac_dir.iterdir()):
            if not domain_dir.is_dir():
                continue
            domain = domain_dir.name
            dcode = DOMAIN_CODES.get(domain, domain.upper())
            for scen_dir in sorted(domain_dir.iterdir()):
                if not scen_dir.is_dir():
                    continue
                folder = scen_dir.name
                # pattern: PRACTICE-DOMCODE-NN-slug
                m = re.match(rf"^({prac})-([A-Z]+)-(\d+)-(.+)$", folder)
                if m:
                    sid = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
                    slug = m.group(4)
                else:
                    sid = folder
                    slug = folder
                rows.append({
                    "practice": prac,
                    "practice_name": PRACTICE_NAMES[prac],
                    "domain": domain,
                    "domain_code": dcode,
                    "folder": folder,
                    "scenario_id": sid,
                    "slug": slug,
                    "title_humanized": slug.replace("-", " ").capitalize(),
                })
    return rows


# ---- service ID extraction --------------------------------------------

SERVICE_PATTERN = re.compile(r"\b([A-Z]+-E2E-\d+)\b")


def extract_services_from_compact(compact):
    """compact = dict[practice -> list[row]]; each row has 's' key with service id."""
    services = defaultdict(list)  # service_id -> list of (practice, scenario_title)
    for prac, rows in compact.items():
        for row in rows:
            sid = row.get("s", "").strip()
            if sid:
                services[sid].append((prac, row.get("t", "")))
    return services


def extract_services_from_featured(featured):
    """featured rows have a 'Service' field with one or two service IDs in prose."""
    services = defaultdict(list)
    for f in featured:
        prac = f.get("practice", "")
        title = f.get("title", "")
        svc_text = f.get("Service", "") or ""
        for m in SERVICE_PATTERN.finditer(svc_text):
            services[m.group(1)].append((prac, title))
    return services


def extract_services_from_compact_titles(compact):
    """Sometimes Service text appears in 'b' (blurb); pick those up too."""
    services = defaultdict(list)
    for prac, rows in compact.items():
        for row in rows:
            for field in ("b", "k"):
                v = row.get(field, "") or ""
                for m in SERVICE_PATTERN.finditer(v):
                    services[m.group(1)].append((prac, row.get("t", "")))
    return services


# Curated service catalog: name + description for every known service ID.
# Names for the canonical *-E2E-* services come from featured-chain prose
# where present; non-E2E IDs (AXLE-Ops-03, ER-OG-07, etc.) are authored
# based on the scenario themes that consume them.
SERVICE_CATALOG = {
    # ---- RC · Retail & Consumer (E2E line) ----
    "RC-E2E-03": ("Assortment & Pricing",
                  "Pricing, markdown, and assortment-rationalization decisioning. SKU-level "
                  "elasticity + competitive monitoring + lifecycle pricing. Persona = Category "
                  "Manager + Pricing Director. Anchors cold-chain markdown, dynamic pricing, and "
                  "assortment scenarios."),
    "RC-E2E-04": ("Customer Lifecycle & Loyalty",
                  "Customer acquisition, retention, and loyalty management across channels. "
                  "Personalization + next-best-action + retention offers. Persona = Loyalty "
                  "Director + CMO."),
    "RC-E2E-05": ("Store Operations",
                  "Store-level operations — labor scheduling, task management, frontline "
                  "decision support. Persona = Store Operations Lead + District Manager."),
    "RC-E2E-06": ("Finance Close Automation",
                  "Multi-entity record-to-report close with HITL on material entries. 6-agent "
                  "fleet (Balance-Fetch, Interco-Match, FX-Reval, Accrual-Propose, JE-Compose, "
                  "Evidence-Write). Persona = Corporate Controller. Nike Wave 1 build target."),
    "RC-E2E-07": ("Returns & Refund Integrity",
                  "Returns triage with fraud detection, refund-policy compliance, and "
                  "secondary-disposition routing. Persona = Returns Operations + Loss Prevention."),
    "RC-E2E-08": ("Customer Acquisition",
                  "Acquisition campaign optimization across paid/organic/owned channels. "
                  "Attribution + spend allocation + creative testing."),
    "RC-E2E-09": ("Product Tracking",
                  "Lot-provenance and FSMA 204 compliance — supply-chain-of-custody for "
                  "perishables and regulated goods. Persona = Food Safety Compliance Lead."),
    "RC-E2E-11": ("Controls & Audit Evidence",
                  "SOX-grade evidence packet generation; sits on top of LEDGER. Persona = SOX "
                  "Program Lead + External Audit PIC. Quarterly evidence ZIP + auditor "
                  "walkthrough."),

    # ---- HLS · Healthcare & Life Sciences ----
    "HLS-E2E-02": ("Clinical Decision Support",
                  "Real-time clinical decision support at point of care — risk stratification, "
                  "treatment-pathway recommendation, alert triage. Persona = Attending "
                  "Physician + Care Team Lead."),
    "HLS-E2E-04": ("Care Management & UM",
                  "Care management and utilization management — patient cohort assignment, "
                  "prior-authorization decisioning, length-of-stay forecasting. Persona = Case "
                  "Manager + Medical Director."),
    "HLS-E2E-06": ("Revenue Cycle",
                  "Revenue cycle from coding through claims to denials management. Charge "
                  "capture + denials-prediction + appeals automation. Persona = Revenue "
                  "Integrity Director + RCM Lead."),
    "HLS-LS-03":  ("Life Sciences Commercial",
                  "Life-sciences commercial operations — HCP engagement, MSL routing, "
                  "field-medical knowledge graph. Persona = Commercial Operations Lead."),

    # ---- ER · Energy & Resources ----
    "ER-OG-02":   ("Drilling & Completions",
                  "Drilling-rig performance, frac completion optimization, real-time drilling "
                  "advisory. Persona = Drilling Engineer + Rig Superintendent."),
    "ER-OG-03":   ("Production Optimization",
                  "Upstream production — artificial-lift optimization, asset-transfer "
                  "reconciliation, well-test scheduling. Persona = Production Engineer + Field "
                  "Operations Supervisor."),
    "ER-OG-05":   ("Refining & Petrochemicals",
                  "Downstream refining — blending optimization, catalyst-performance "
                  "prediction, crude-slate selection. Persona = Refinery Operations Lead + "
                  "Process Engineer."),
    "ER-OG-07":   ("Pipeline & Midstream",
                  "Pipeline integrity, cathodic-protection anomalies, compressor performance, "
                  "flow assurance. Persona = Pipeline Operations + Integrity Engineer."),
    "ER-MN-02":   ("Mining Operations",
                  "Open-pit and underground mining — blast-fragmentation analysis, drill-and-"
                  "blast productivity, haulage optimization. Persona = Mine Operations Manager."),
    "ER-PU-04":   ("Power & Utilities",
                  "Grid operations — load forecasting, outage management, distributed-energy "
                  "integration. Persona = Distribution Operations + Grid Control Engineer."),

    # ---- AXLE · Industrial Manufacturing ----
    "AXLE-Connected-Factory-01": ("Connected Factory",
                  "IoT-enabled factory floor — CNC-spindle failure prediction, "
                  "condition-based maintenance, digital-twin operations. Persona = Plant "
                  "Manager + Maintenance Engineer."),
    "AXLE-Ops-03": ("Manufacturing Operations",
                  "End-to-end manufacturing operations — air-emission monitoring, andon-signal "
                  "root-cause analysis, OEE optimization, shift handoff. Persona = Plant "
                  "Operations Director."),
    "AXLE-QMS-02": ("Quality Management Systems",
                  "Quality and compliance — 8D-report drafting, AD/SB compliance tracking, "
                  "non-conformance management. Persona = Quality Director + Compliance Lead."),
    "AXLE-Supply-04": ("Supply Chain Visibility",
                  "Multi-tier supply visibility — country-of-origin documentation, "
                  "dealer-inventory visibility, supplier scorecard. Persona = Supply Chain "
                  "Director + Trade Compliance."),
    "AXLE-Supply-07": ("Supplier Risk Management",
                  "Supplier risk and dual-sourcing strategy — risk scoring, alternate-source "
                  "qualification, geopolitical exposure. Persona = CPO + Supply Risk Lead."),

    # ---- TMT · Technology, Media & Telecom ----
    "TMT-TEL-NET-02": ("Telco Network Operations",
                  "Network operations and fault management for fixed/mobile/fiber networks. "
                  "Predictive fault, 5G slice assurance, capacity planning. Persona = NOC "
                  "Incident Commander + Service Assurance Lead."),
    "TMT-TEL-CC-03": ("Telco Customer Care",
                  "Customer care and contact-center intelligence — bill-shock prevention, "
                  "churn-intervention playbooks, contact-driver mining. Persona = Customer "
                  "Operations Director."),
    "TMT-MED-01":  ("Media Subscriber Lifecycle",
                  "Streaming subscriber lifecycle — recommendation, retention, churn "
                  "intervention. Persona = Streaming Retention Director + Audience Lead."),
    "TMT-MED-03":  ("Media Rights & Monetization",
                  "Content rights management, broadcast-rights utilization, royalty "
                  "reconciliation. Persona = Rights Management Lead + Distribution Strategy."),
    "TMT-TEC-04":  ("Software Engineering Intelligence",
                  "Engineering and product intelligence — A/B test interpretation, "
                  "feature-flag rollout analysis, release health. Persona = Head of Product "
                  "Analytics + Engineering Director."),

    # ---- TH · Travel & Hospitality ----
    "TH-AIR-02":   ("Airline Network Planning",
                  "Network and schedule planning — fleet assignment, route profitability, "
                  "schedule recovery. Persona = Network Planning Director."),
    "TH-AIR-04":   ("Airline Operations Control",
                  "Operations control center — IROP recovery, crew rotation repair, "
                  "passenger reaccommodation. Persona = OCC Director + Crew Planning Lead."),
    "TH-AIR-06":   ("Airline Revenue & Pricing",
                  "Revenue management — class-of-service inventory, dynamic pricing, "
                  "ancillary attach. Persona = RM Director + Pricing Lead."),
    "TH-HOT-03":   ("Hotel Operations",
                  "Property operations — housekeeping orchestration, F&B forecasting, "
                  "energy management. Persona = General Manager + Operations Director."),
    "TH-HOT-05":   ("Hotel Revenue Management",
                  "Revenue management — rate optimization, length-of-stay controls, "
                  "channel-mix optimization. Persona = Director of Revenue Management."),

    # ---- ICE · Industrial & Commercial Equipment ----
    "ICE-Aftermarket-01": ("Aftermarket — Parts & Service",
                  "Parts demand forecasting, service-bay scheduling, technician routing. "
                  "Persona = Aftermarket Director + Service Manager."),
    "ICE-Aftermarket-03": ("Aftermarket — Warranty & Recall",
                  "Warranty claims processing, recall execution, field-service compliance. "
                  "Persona = Warranty Director + Field Service Lead."),
    "ICE-Aftermarket-05": ("Aftermarket — Customer Success",
                  "Equipment-uptime customer success, SLA management, contract renewal. "
                  "Persona = Customer Success Director."),
    "ICE-Connected-06":   ("Connected Equipment",
                  "IoT telemetry from installed equipment — predictive failure, "
                  "remote diagnostics, firmware management. Persona = Connected Products "
                  "Director + Engineering Lead."),
    "ICE-EaaS-04":        ("Equipment-as-a-Service",
                  "EaaS subscription operations — usage-based billing, performance SLAs, "
                  "lifecycle health. Persona = EaaS Program Director + Commercial Operations."),
}


# ---- build sheets ------------------------------------------------------

def sheet_all_scenarios(wb, fs_rows, compact, featured):
    """Sheet 1: every scenario from filesystem, joined to catalog metadata."""
    # build compact lookup: by humanized slug match within practice
    compact_by_slug = defaultdict(list)
    for prac, rows in compact.items():
        for row in rows:
            slug = re.sub(r"[^a-z0-9]+", "-", row.get("t", "").lower()).strip("-")
            compact_by_slug[(prac, slug)].append(row)

    featured_set = set()
    for f in featured:
        # featured doesn't have ID in JSON; we can construct based on practice + index
        # but easier: index by (practice, title slug)
        slug = re.sub(r"[^a-z0-9]+", "-", f.get("title", "").lower()).strip("-")
        featured_set.add((f["practice"], slug))

    ws = wb.active
    ws.title = "All Scenarios"
    headers = ["Scenario ID", "Practice", "Practice Name", "Domain Code",
               "Domain", "Slug", "Title (humanized)", "Featured?", "Folder",
               "Catalog Title", "Service Code (compact)", "KPI (compact)"]
    widths = [16, 8, 28, 10, 22, 36, 36, 9, 44, 36, 16, 26]
    write_headers(ws, headers, widths)

    row_idx = 2
    for r in fs_rows:
        slug = r["slug"]
        compact_match = compact_by_slug.get((r["practice"], slug))
        cat_title = ""
        cat_svc = ""
        cat_kpi = ""
        if compact_match:
            cat_title = compact_match[0].get("t", "")
            cat_svc   = compact_match[0].get("s", "")
            cat_kpi   = compact_match[0].get("k", "")
        is_featured = "Yes" if (r["practice"], slug) in featured_set else ""
        write_row(ws, row_idx, [
            r["scenario_id"], r["practice"], r["practice_name"], r["domain_code"],
            r["domain"], slug, r["title_humanized"], is_featured, r["folder"],
            cat_title, cat_svc, cat_kpi,
        ], code_cols=(1, 11))
        row_idx += 1
    print(f"  sheet 'All Scenarios' · {row_idx-2} rows")


def sheet_featured(wb, featured):
    """Sheet 2: 35 featured scenarios with full chain."""
    ws = wb.create_sheet("Featured Scenarios")
    headers = ["Practice", "#", "Title", "Scenario", "Solution",
               "Use Case", "Service", "Persona", "KPI"]
    widths = [10, 5, 32, 50, 50, 50, 38, 30, 38]
    write_headers(ws, headers, widths)

    # Sort: practice then index
    ordered = sorted(featured, key=lambda f: (PRACTICES.index(f.get("practice","RC")),
                                              f.get("index_in_practice", 99)))
    row_idx = 2
    for f in ordered:
        write_row(ws, row_idx, [
            f.get("practice", ""),
            f.get("index_in_practice", ""),
            f.get("title", ""),
            f.get("Scenario", ""),
            f.get("Solution", ""),
            f.get("Use Case", ""),
            f.get("Service", ""),
            f.get("Persona", ""),
            f.get("KPI", ""),
        ])
        row_idx += 1
    print(f"  sheet 'Featured Scenarios' · {row_idx-2} rows")


def sheet_compact(wb, compact):
    """Sheet 3: 723 compact catalog rows."""
    ws = wb.create_sheet("Compact Catalog")
    headers = ["Practice", "#", "Title", "Service Code", "Headline KPI",
               "KPI direction", "Blurb"]
    widths = [10, 5, 38, 14, 22, 12, 60]
    write_headers(ws, headers, widths)

    row_idx = 2
    for prac in PRACTICES:
        rows = compact.get(prac, [])
        for i, r in enumerate(rows, start=1):
            write_row(ws, row_idx, [
                prac, i, r.get("t",""), r.get("s",""), r.get("k",""),
                r.get("kk",""), r.get("b",""),
            ], code_cols=(4,))
            row_idx += 1
    print(f"  sheet 'Compact Catalog' · {row_idx-2} rows")


def sheet_services(wb, compact, featured):
    """Sheet 4: unique service IDs with name + description."""
    sv1 = extract_services_from_compact(compact)
    sv2 = extract_services_from_featured(featured)
    sv3 = extract_services_from_compact_titles(compact)

    all_svc = defaultdict(set)  # service_id -> set of (practice, title)
    for src in (sv1, sv2, sv3):
        for sid, refs in src.items():
            for ref in refs:
                all_svc[sid].add(ref)

    ws = wb.create_sheet("Services")
    headers = ["Service ID", "Service Name", "Description",
               "Practice", "Scenarios using this service",
               "Sample scenario titles"]
    widths = [22, 32, 70, 10, 14, 80]
    write_headers(ws, headers, widths)

    row_idx = 2
    missing = []
    for sid in sorted(all_svc):
        refs = sorted(all_svc[sid])
        practices = sorted(set(r[0] for r in refs))
        titles = sorted(set(r[1] for r in refs if r[1]))
        name, desc = SERVICE_CATALOG.get(sid, ("(unnamed)", "(description not yet authored)"))
        if sid not in SERVICE_CATALOG:
            missing.append(sid)
        write_row(ws, row_idx, [
            sid, name, desc,
            " · ".join(practices),
            len(titles),
            " · ".join(titles[:8]) + (f"  +{len(titles)-8} more" if len(titles) > 8 else ""),
        ], code_cols=(1,))
        # bold the service name column
        ws.cell(row=row_idx, column=2).font = Font(name="Calibri", size=10, bold=True)
        row_idx += 1
    print(f"  sheet 'Services' · {row_idx-2} rows  (named: {row_idx-2-len(missing)} · unnamed: {len(missing)})")
    if missing:
        print(f"    services not yet in catalog: {', '.join(missing)}")


def sheet_service_scenario_mapping(wb, compact, featured, fs_rows):
    """Sheet: long-format Service ↔ Scenario mapping.

    One row per (service_id, scenario) pair. If a scenario references two
    services it appears twice. Sources: compact catalog + featured prose.
    """
    # Build slug -> scenario_id lookup from filesystem
    fs_lookup = {}
    for r in fs_rows:
        key = (r["practice"], r["slug"])
        fs_lookup[key] = r["scenario_id"]

    # Build slug -> featured? lookup
    featured_set = set()
    for f in featured:
        slug = re.sub(r"[^a-z0-9]+", "-", f.get("title", "").lower()).strip("-")
        featured_set.add((f["practice"], slug))

    rows_out = []  # list of dicts ready to write

    # 1) From compact catalog — every row has at most one service code in 's'
    for prac, rows in compact.items():
        for row in rows:
            sid = (row.get("s") or "").strip()
            if not sid:
                continue
            title = row.get("t", "")
            slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
            scen_id = fs_lookup.get((prac, slug), "")
            name, _ = SERVICE_CATALOG.get(sid, ("(unnamed)", ""))
            rows_out.append({
                "service_id":   sid,
                "service_name": name,
                "practice":     prac,
                "scenario_id":  scen_id,
                "title":        title,
                "kpi":          row.get("k", ""),
                "kpi_dir":      row.get("kk", ""),
                "blurb":        row.get("b", ""),
                "featured":     "Yes" if (prac, slug) in featured_set else "",
                "source":       "compact",
            })

    # 2) From featured chains — Service prose can reference 1+ service IDs
    for f in featured:
        prac = f.get("practice", "")
        title = f.get("title", "")
        slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
        scen_id = fs_lookup.get((prac, slug), "")
        kpi = f.get("KPI", "")
        # KPI in featured is prose; collapse to first sentence-like fragment
        kpi_short = kpi.split("·")[0].strip() if "·" in kpi else kpi[:120]
        svc_text = f.get("Service", "") or ""
        svc_ids = SERVICE_PATTERN.findall(svc_text)
        for sid in svc_ids:
            name, _ = SERVICE_CATALOG.get(sid, ("(unnamed)", ""))
            rows_out.append({
                "service_id":   sid,
                "service_name": name,
                "practice":     prac,
                "scenario_id":  scen_id,
                "title":        title,
                "kpi":          kpi_short,
                "kpi_dir":      "",
                "blurb":        f.get("Solution", "")[:200],
                "featured":     "Yes",
                "source":       "featured",
            })

    # De-duplicate (a scenario can appear in both featured and compact with the same service)
    seen = set()
    deduped = []
    for r in rows_out:
        key = (r["service_id"], r["practice"], r["title"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)

    # Sort: by service_id, then practice, then featured first, then title
    deduped.sort(key=lambda r: (r["service_id"], r["practice"],
                                0 if r["featured"] else 1, r["title"]))

    ws = wb.create_sheet("Service-Scenario Mapping")
    headers = ["Service ID", "Service Name", "Practice", "Scenario ID",
               "Scenario Title", "Featured?", "Headline KPI", "KPI dir", "Blurb"]
    widths = [22, 30, 9, 18, 38, 10, 26, 8, 60]
    write_headers(ws, headers, widths)

    row_idx = 2
    last_sid = None
    for r in deduped:
        # Add a section break (empty row) between service blocks for readability
        if last_sid is not None and r["service_id"] != last_sid:
            row_idx += 1  # leave one blank row
        write_row(ws, row_idx, [
            r["service_id"], r["service_name"], r["practice"],
            r["scenario_id"], r["title"], r["featured"],
            r["kpi"], r["kpi_dir"], r["blurb"],
        ], code_cols=(1, 4))
        # bold the service name
        ws.cell(row=row_idx, column=2).font = Font(name="Calibri", size=10, bold=True)
        # tint Featured-Yes rows gold-ish
        if r["featured"] == "Yes":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FBF1DC")
        last_sid = r["service_id"]
        row_idx += 1

    print(f"  sheet 'Service-Scenario Mapping' · {len(deduped)} mappings")
    return deduped


def sheet_service_scenario_pivot(wb, mapping_rows):
    """Sheet: pivot view — services as rows, practices as columns, counts as values."""
    ws = wb.create_sheet("Services × Practice Pivot")
    # Build pivot
    services_seen = set(r["service_id"] for r in mapping_rows)
    services = sorted(services_seen)
    pivot = defaultdict(lambda: defaultdict(int))
    for r in mapping_rows:
        pivot[r["service_id"]][r["practice"]] += 1

    headers = ["Service ID", "Service Name"] + PRACTICES + ["Total"]
    widths  = [22, 32] + [8] * len(PRACTICES) + [9]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "C2"

    rr = 2
    for sid in services:
        name, _ = SERVICE_CATALOG.get(sid, ("(unnamed)", ""))
        row_total = sum(pivot[sid].values())
        cells = [sid, name] + [pivot[sid].get(p, 0) for p in PRACTICES] + [row_total]
        for j, v in enumerate(cells, start=1):
            c = ws.cell(row=rr, column=j, value=v)
            c.border = CELL_BORDER
            if j == 1:
                c.font = CODE_FONT
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif j == 2:
                c.font = Font(name="Calibri", size=10, bold=True)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif j == len(cells):
                c.font = Font(name="Calibri", size=10, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = CELL_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
            if rr % 2 == 0 and j > 2:
                c.fill = BAND_FILL
        rr += 1
    # heat the practice count cells
    rng = (f"{get_column_letter(3)}2:"
           f"{get_column_letter(2 + len(PRACTICES))}{rr - 1}")
    rule = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="F6E5BD",
        end_type="max", end_color="E3B657",
    )
    ws.conditional_formatting.add(rng, rule)
    print(f"  sheet 'Services × Practice Pivot' · {len(services)} services")


def sheet_summary(wb, fs_rows):
    """Sheet 5: counts by practice + practice×domain matrix."""
    ws = wb.create_sheet("Summary")

    # ----- top block: counts by Practice
    ws.cell(row=1, column=1, value="Counts by Practice").font = Font(bold=True, size=14)
    write_headers_at(ws, 3, ["Practice", "Practice Name", "Scenario count"], widths=[12, 32, 18])
    by_prac = Counter(r["practice"] for r in fs_rows)
    rr = 4
    for prac in PRACTICES:
        write_row(ws, rr, [prac, PRACTICE_NAMES[prac], by_prac.get(prac, 0)], code_cols=(1,))
        rr += 1
    write_row(ws, rr, ["TOTAL", "", sum(by_prac.values())])
    ws.cell(row=rr, column=1).font = Font(bold=True, size=10)
    ws.cell(row=rr, column=3).font = Font(bold=True, size=10)
    end_top = rr

    # ----- matrix: Practice × Domain
    ws.cell(row=end_top + 3, column=1, value="Practice × Domain matrix").font = Font(bold=True, size=14)
    matrix_top = end_top + 5
    headers = ["Practice"] + list(DOMAIN_CODES.values()) + ["Total"]
    widths  = [12] + [8] * len(DOMAIN_CODES) + [9]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=matrix_top, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[matrix_top].height = 24

    matrix = defaultdict(lambda: defaultdict(int))
    for r in fs_rows:
        matrix[r["practice"]][r["domain_code"]] += 1

    rr = matrix_top + 1
    for prac in PRACTICES:
        row_total = sum(matrix[prac].values())
        cells = [prac] + [matrix[prac].get(d, 0) for d in DOMAIN_CODES.values()] + [row_total]
        for j, v in enumerate(cells, start=1):
            c = ws.cell(row=rr, column=j, value=v)
            c.border = CELL_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            if j == 1:
                c.font = CODE_FONT
            elif j == len(cells):
                c.font = Font(bold=True, size=10)
            else:
                c.font = CELL_FONT
            if rr % 2 == 0:
                c.fill = BAND_FILL
        rr += 1
    # totals row
    totals = [matrix[p].get(d, 0) for p in PRACTICES for d in DOMAIN_CODES.values()]
    domain_totals = []
    for d in DOMAIN_CODES.values():
        domain_totals.append(sum(matrix[p].get(d, 0) for p in PRACTICES))
    grand_total = sum(domain_totals)
    bot_cells = ["TOTAL"] + domain_totals + [grand_total]
    for j, v in enumerate(bot_cells, start=1):
        c = ws.cell(row=rr, column=j, value=v)
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=10, color="0B0B0F")
        c.fill = PatternFill("solid", fgColor="F6F1E4")

    # color scale on matrix (excluding header + totals row/column)
    rng = (f"{get_column_letter(2)}{matrix_top+1}:"
           f"{get_column_letter(2 + len(DOMAIN_CODES) - 1)}{rr - 1}")
    rule = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="F6E5BD",
        end_type="max", end_color="E3B657",
    )
    ws.conditional_formatting.add(rng, rule)

    print(f"  sheet 'Summary' · counts + matrix")


def write_headers_at(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border = CELL_BORDER
        if widths:
            ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[row].height = 24


def sheet_domain_reference(wb):
    """Sheet 6: domain code mapping."""
    ws = wb.create_sheet("Domain Reference")
    write_headers(ws, ["Domain", "Code", "Notes"], widths=[28, 12, 60])
    notes = {
        "clinical-care":          "Patient care, treatment, EHR/EMR, clinical decision support (HLS)",
        "network-infrastructure": "Networks, fiber, 5G, OSS/BSS, NOC (Telco; some HW)",
        "engineering-rd":         "Software ENG, semi design, A/B testing, feature flags",
        "supply-chain":           "Inventory, logistics, sourcing, fulfillment",
        "asset-maintenance":      "Asset reliability, predictive maintenance, MRO",
        "quality-compliance":     "Yield, recalls, regulatory quality (FDA, FAA, FCC, ITAR)",
        "risk-fraud-security":    "Fraud detection, AML, security operations, identity",
        "pricing-revenue":        "Pricing, promotions, markdowns, revenue management",
        "customer-experience":    "Customer interactions, churn, contact center, recommendations",
        "marketing-growth":       "Acquisition, lifecycle, attribution, brand",
        "operations-workforce":   "Finance close, HR, scheduling, productivity",
        "channel-partner-dealer": "Dealer/partner ops, rights/licensing, distribution",
        "other":                  "Cross-cutting / not in primary domain taxonomy",
    }
    rr = 2
    for domain, code in DOMAIN_CODES.items():
        write_row(ws, rr, [domain, code, notes.get(domain, "")], code_cols=(2,))
        rr += 1


def sheet_practice_reference(wb):
    """Sheet 7: practice reference."""
    ws = wb.create_sheet("Practice Reference")
    write_headers(ws, ["Practice", "Practice Name", "Notes"],
                  widths=[12, 36, 60])
    notes = {
        "RC":   "Retail (groceries, big-box, apparel, hardware) + Consumer (CPG brands)",
        "HLS":  "Hospitals, payors, life sciences, medtech, pharma",
        "ER":   "Oil & gas, utilities, mining, renewables",
        "AXLE": "Industrial Manufacturing — automotive-weighted (OEMs, Tier-1)",
        "TMT":  "Technology (HW, Software), Media, Sports, Telco, Semi (6 sub-Practices)",
        "TH":   "Travel (airlines, lodging) + Hospitality + cruise + leisure",
        "ICE":  "Industrial & Commercial Equipment (build/install/service)",
    }
    rr = 2
    for prac in PRACTICES:
        write_row(ws, rr, [prac, PRACTICE_NAMES[prac], notes[prac]], code_cols=(1,))
        rr += 1


# ---- main --------------------------------------------------------------

def build():
    print("Loading source data...")
    featured = load_featured()
    compact  = load_compact()
    fs_rows  = walk_filesystem()
    print(f"  filesystem: {len(fs_rows)} scenario folders")
    print(f"  featured:   {len(featured)} chains")
    print(f"  compact:    {sum(len(v) for v in compact.values())} rows in {len(compact)} practices")

    print("Building workbook...")
    wb = Workbook()
    sheet_all_scenarios(wb, fs_rows, compact, featured)
    sheet_featured(wb, featured)
    sheet_compact(wb, compact)
    sheet_services(wb, compact, featured)
    mapping_rows = sheet_service_scenario_mapping(wb, compact, featured, fs_rows)
    sheet_service_scenario_pivot(wb, mapping_rows)
    sheet_summary(wb, fs_rows)
    sheet_practice_reference(wb)
    sheet_domain_reference(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    target = OUT
    try:
        wb.save(str(target))
    except PermissionError:
        # File likely open in Excel — write to a sibling with -v<N> suffix
        n = 2
        while True:
            alt = target.with_name(target.stem + f"-v{n}" + target.suffix)
            if not alt.exists():
                target = alt
                break
            n += 1
        wb.save(str(target))
        print(f"\n[notice] {OUT.name} was locked — saved as {target.name} instead.")
    print(f"\nWrote {target}")


if __name__ == "__main__":
    build()

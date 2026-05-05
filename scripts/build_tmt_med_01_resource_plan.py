"""Build the TMT-MED-01 Resource Plan workbook (Disney+ reference engagement).

Mirrors Nike-R2R-NOAH-Resource-Plan.xlsx. Eight sheets:
  README · Deloitte Staffing · Disney Staffing · Microsoft Platform ·
  Third-Party & Other · Commercial Envelope · Durable Value · ROI · Payback

Output: Disney 02_projects/FY27_Pipeline/subscriber-lifecycle-agentic/deliverables/TMT-MED-01-Resource-Plan.xlsx
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

OUT_DIR = Path(r"C:\Stage\Clients\Industries\TMT\Global_Media_Entertainment\Walt_Disney\02_projects\FY27_Pipeline\subscriber-lifecycle-agentic\deliverables")
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / "TMT-MED-01-Resource-Plan.xlsx"

# ---- styling ------------------------------------------------------------

INK    = "0B0B0F"
INK2   = "4A4A55"
MUTE   = "8a8a90"
GOLD   = "B8892E"
TEAL   = "2A7F73"
VIOLET = "6648B0"
LINE   = "D4D4D7"
BAND   = "F3F3F5"

HDR_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor=INK)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN      = Side(style="thin", color=LINE)
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILL = PatternFill("solid", fgColor=BAND)
GOLD_BG   = PatternFill("solid", fgColor="FBF1DC")
TEAL_BG   = PatternFill("solid", fgColor="E4F1EE")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=INK)
SUB_FONT   = Font(name="Calibri", size=10, italic=True, color=INK2)
CELL_FONT  = Font(name="Calibri", size=10)
CODE_FONT  = Font(name="Consolas", size=10, color=VIOLET, bold=True)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=INK)


def write_hdrs(ws, headers, widths=None, *, row=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
        if widths: ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
    ws.row_dimensions[row].height = 32


def cell(ws, r, col, val, *, font=None, fill=None, align=None, fmt=None):
    c = ws.cell(row=r, column=col, value=val)
    c.border = BORDER
    c.font = font or CELL_FONT
    c.alignment = align or Alignment(horizontal="left", vertical="center", wrap_text=True)
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    return c


def money(ws, r, col, val, *, font=None, fill=None):
    return cell(ws, r, col, val, font=font, fill=fill, fmt='"$"#,##0',
                align=Alignment(horizontal="right", vertical="center"))


# ============================================================================
# SHEETS
# ============================================================================

def sheet_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 96

    rows = [
        ("APEX · TMT-MED-01 · Resource Plan · Disney+ reference",
         "Approach + Wave 1 Foundation + Wave 2 Pilot + Wave 3 Scale & Fuse. Loaded by role / FTE / wave / dollar."),
        ("", ""),
        ("How to read this workbook", ""),
        ("Deloitte Staffing",       "DMTSP team model. Roles, daily rates, FTE per wave, dollar load per wave."),
        ("Disney Staffing",          "Disney persona allocation. Implied value (informational — not billed)."),
        ("Microsoft Platform",       "Per-month consumption across Azure AI Foundry, Azure OpenAI, Fabric, storage, Teams, Power BI, Entra. Wave totals."),
        ("Third-Party & Other",      "Specialist services, integration vendors, change management, training, audit/privacy coordination."),
        ("Commercial Envelope",      "Low / Mid / High range per wave (Mid = sum of Deloitte + Microsoft + Third-Party; Low = 0.75×Mid; High = 1.35×Mid)."),
        ("Durable Value",            "Annualized W3 run-rate value: retention NPV + audit/privacy fee avoidance + non-monetized close-cycle compression."),
        ("ROI · Payback",            "Time-phased cumulative investment vs. realized return. Anchors the ROI conversation with the Streaming COO."),
        ("", ""),
        ("Scope assumptions", ""),
        ("Reference client",         "The Walt Disney Company · Disney+ Streaming Services. NA single market in W1+W2; global multi-brand (Hulu, ESPN+) in W3."),
        ("Subscriber base",          "~50M Disney+ NA subscribers in W2 scope. ~262M global across Disney+ / Hulu / ESPN+ in W3 scope."),
        ("Currency",                 "USD throughout. Disney's reporting currency; no FX scenarios material at this level of detail."),
        ("Wave durations",           "Approach 4 weeks · W1 6 months · W2 9 months · W3 15 months."),
        ("Business days per wave",   "Approach 20 · W1 125 · W2 187.5 · W3 312.5."),
        ("Daily rate basis",         "DMTSP standard daily rate card; rate × FTE × business days = role-wave dollar."),
        ("", ""),
        ("Independence posture",
         "All TMT-MED-01 commercial materials use 'Deloitte's Microsoft practice,' 'DMTSP,' 'Microsoft platform capabilities,' 'Microsoft-native deployment.' The terms 'partner,' 'alliance,' and 'strategic alliance' do not appear. Before contracting, confirm Disney's external-audit firm relationship and route the SOW through Deloitte's Independence office."),
        ("", ""),
        ("Companion artifacts", ""),
        ("APEX-TMT-MED-01-Walkthrough.docx",
         "Executive walkthrough — value-delivery chain · personas · KPIs · waves."),
        ("TMT-MED-01-Sprint-Plan.xlsx",
         "Task / Subtask grid · UAC · Use Case Map · Persona RACI · Persona Journey Maps · Key Solutions · KPI Linkage."),
        ("APEX-TMT-Reference-Architecture.html",
         "TMT Practice reference — six sub-Practices, planes, archetypes."),
        ("APEX-CAF-Integration-Architecture.docx",
         "How TMT-MED-01 consumes CAF via A2A — substrate-vs-scenario layering."),
    ]
    for i, (left, right) in enumerate(rows, start=1):
        c1 = ws.cell(row=i, column=1, value=left)
        c2 = ws.cell(row=i, column=2, value=right)
        if i == 1:
            c1.font = TITLE_FONT
            c2.font = SUB_FONT
        elif left and right:
            c1.font = Font(name="Calibri", bold=True, size=10)
            c2.font = CELL_FONT
        else:
            c1.font = Font(name="Calibri", bold=True, size=11, color=VIOLET)
            c2.font = CELL_FONT
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30
    print("  README")


# ----------------------------------------------------------------------------
# Deloitte Staffing
# ----------------------------------------------------------------------------

# Business days per wave
BD = {"Approach": 20, "W1": 125, "W2": 187.5, "W3": 312.5}

DELOITTE_ROLES = [
    # (Role, Daily Rate, Approach FTE, W1 FTE, W2 FTE, W3 FTE, role notes)
    ("Partner — DMTSP TMT lead",                    1600, 0.20, 0.10, 0.15, 0.15),
    ("Principal / LBP — Pursuit lead",              1400, 0.40, 0.20, 0.25, 0.25),
    ("LCSP — Lead client service partner (Disney)", 1500, 0.20, 0.08, 0.12, 0.10),
    ("GPL — Global product lead (TMT-MED)",         1000, 0.60, 0.25, 0.35, 0.30),
    ("Engagement Architect — APEX TMT-MED",         1100, 0.80, 0.50, 0.40, 0.30),
    ("Senior Manager — Delivery lead",               950, 1.00, 0.80, 1.00, 1.00),
    ("Senior Consultant — Audience/segmentation",    700, 0.60, 1.00, 1.50, 1.20),
    ("Senior Consultant — Personalization/recs",     700, 0.20, 1.00, 1.50, 1.50),
    ("Senior Consultant — Privacy/governance",       700, 0.40, 0.50, 0.80, 0.80),
    ("Implementation Lead — Solution architect",     900, 0.30, 1.00, 1.00, 1.00),
    ("ML Engineer — Risk-Score / Drift",             750, 0.00, 1.50, 2.00, 2.50),
    ("Backend Engineer — Agents / orchestration",    700, 0.20, 1.50, 2.00, 2.50),
    ("Frontend Engineer — Adaptive Card / Console",  650, 0.20, 1.00, 1.50, 1.50),
    ("Data Engineer — Streaming events / TMTML",     750, 0.20, 1.20, 1.50, 2.00),
    ("BI Engineer — Power BI · semantic model",      650, 0.10, 0.50, 0.80, 1.00),
    ("Change Management Lead",                       850, 0.30, 0.50, 1.00, 1.20),
    ("QA / Test Lead",                               650, 0.10, 0.80, 1.00, 1.20),
    ("Privacy / Compliance Specialist",              900, 0.40, 0.40, 0.60, 0.50),
    ("Independence office liaison",                  900, 0.20, 0.10, 0.10, 0.10),
    ("Project Coordinator",                          550, 0.50, 0.50, 0.80, 1.00),
]


def sheet_deloitte_staffing(wb):
    ws = wb.create_sheet("Deloitte Staffing")
    ws.cell(row=1, column=1, value="DMTSP delivery team model").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Daily rate × FTE × business days per wave").font = SUB_FONT

    headers = ["Role", "Daily Rate",
               "Approach FTE", "W1 FTE", "W2 FTE", "W3 FTE",
               "Approach $", "W1 $", "W2 $", "W3 $", "Total $"]
    widths  = [42, 12, 11, 8, 8, 8, 13, 13, 13, 13, 13]
    write_hdrs(ws, headers, widths, row=4)

    totals = [0, 0, 0, 0]
    r = 5
    for role, rate, fa, fw1, fw2, fw3 in DELOITTE_ROLES:
        a_dollar = round(rate * fa  * BD["Approach"])
        w1_dollar = round(rate * fw1 * BD["W1"])
        w2_dollar = round(rate * fw2 * BD["W2"])
        w3_dollar = round(rate * fw3 * BD["W3"])
        row_total = a_dollar + w1_dollar + w2_dollar + w3_dollar
        totals[0] += a_dollar; totals[1] += w1_dollar
        totals[2] += w2_dollar; totals[3] += w3_dollar

        cell(ws, r, 1, role, font=Font(name="Calibri", size=10, bold=True))
        money(ws, r, 2, rate)
        cell(ws, r, 3, fa,  align=Alignment(horizontal="center", vertical="center"))
        cell(ws, r, 4, fw1, align=Alignment(horizontal="center", vertical="center"))
        cell(ws, r, 5, fw2, align=Alignment(horizontal="center", vertical="center"))
        cell(ws, r, 6, fw3, align=Alignment(horizontal="center", vertical="center"))
        money(ws, r, 7,  a_dollar)
        money(ws, r, 8,  w1_dollar)
        money(ws, r, 9,  w2_dollar)
        money(ws, r, 10, w3_dollar)
        money(ws, r, 11, row_total, font=TOTAL_FONT)
        if r % 2 == 0:
            for j in range(1, 12):
                ws.cell(row=r, column=j).fill = BAND_FILL
        r += 1

    # Totals row
    grand = sum(totals)
    cell(ws, r, 1, "TOTAL Deloitte Labor", font=TOTAL_FONT, fill=GOLD_BG)
    cell(ws, r, 2, "", fill=GOLD_BG)
    for j in (3, 4, 5, 6):
        cell(ws, r, j, "", fill=GOLD_BG)
    money(ws, r, 7,  totals[0], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 8,  totals[1], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 9,  totals[2], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 10, totals[3], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 11, grand,     font=TOTAL_FONT, fill=GOLD_BG)
    print(f"  Deloitte Staffing · {len(DELOITTE_ROLES)} roles · ${grand:,} total · sub-totals: {totals}")
    return totals


# ----------------------------------------------------------------------------
# Disney Staffing (informational / implied value, not billed by Deloitte)
# ----------------------------------------------------------------------------

DISNEY_ROLES = [
    # (Persona / Role, Description, Approach %, W1 %, W2 %, W3 %, blended-cost-anchor for implied value)
    ("Streaming COO (Sponsor)",          "Executive sponsor; steering committee chair", 0.10, 0.05, 0.10, 0.08, 1500),
    ("VP Subscriber Retention",          "Primary persona; owns retention P&L and HITL approval", 0.25, 0.40, 0.60, 0.40, 1100),
    ("Audience Lead",                    "Cohort definition + Risk-Score calibration", 0.30, 0.60, 0.80, 0.60, 850),
    ("Content Programming Director",     "Rights-window + recommendation review", 0.10, 0.10, 0.30, 0.20, 900),
    ("Pricing Strategy Lead",            "Concession policy authoring + tier tuning", 0.15, 0.20, 0.40, 0.25, 850),
    ("Marketing Operations Director",    "Channel readiness + dispatch monitoring", 0.10, 0.20, 0.50, 0.40, 750),
    ("Privacy Office liaison",           "Decision-evidence shape + quarterly review", 0.30, 0.20, 0.30, 0.20, 1000),
    ("Internal Audit liaison",           "Sample testing + signoff", 0.20, 0.10, 0.20, 0.15, 950),
    ("Senior Data Scientist (Disney+)",  "ML platform; Risk-Score collaboration", 0.10, 0.50, 0.80, 0.60, 800),
    ("Streaming Eng Lead (Disney+)",     "Connector / event-stream integration", 0.15, 0.40, 0.50, 0.30, 850),
    ("CTO/CDO Office representative",    "Architecture review + Azure environment", 0.10, 0.10, 0.10, 0.05, 1200),
    ("Change Management lead (Disney)",  "Org readiness + training rollout", 0.10, 0.20, 0.60, 0.50, 700),
]


def sheet_disney_staffing(wb):
    ws = wb.create_sheet("Disney Staffing")
    ws.cell(row=1, column=1, value="Disney persona allocation (informational)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="% FTE allocation per wave — Disney-side shadow cost (NOT billed by Deloitte)").font = SUB_FONT

    headers = ["Persona / Role", "Description",
               "Approach %", "W1 %", "W2 %", "W3 %",
               "Implied $ Approach", "Implied $ W1", "Implied $ W2", "Implied $ W3"]
    widths = [28, 40, 11, 8, 8, 8, 14, 14, 14, 14]
    write_hdrs(ws, headers, widths, row=4)

    r = 5
    for persona, desc, fa, fw1, fw2, fw3, rate in DISNEY_ROLES:
        a_v  = round(rate * fa  * BD["Approach"])
        w1_v = round(rate * fw1 * BD["W1"])
        w2_v = round(rate * fw2 * BD["W2"])
        w3_v = round(rate * fw3 * BD["W3"])
        cell(ws, r, 1, persona, font=Font(name="Calibri", size=10, bold=True))
        cell(ws, r, 2, desc)
        cell(ws, r, 3, f"{int(fa*100)}%",  align=Alignment(horizontal="center"))
        cell(ws, r, 4, f"{int(fw1*100)}%", align=Alignment(horizontal="center"))
        cell(ws, r, 5, f"{int(fw2*100)}%", align=Alignment(horizontal="center"))
        cell(ws, r, 6, f"{int(fw3*100)}%", align=Alignment(horizontal="center"))
        money(ws, r, 7,  a_v)
        money(ws, r, 8,  w1_v)
        money(ws, r, 9,  w2_v)
        money(ws, r, 10, w3_v)
        if r % 2 == 0:
            for j in range(1, 11):
                ws.cell(row=r, column=j).fill = BAND_FILL
        r += 1
    print(f"  Disney Staffing · {len(DISNEY_ROLES)} personas")


# ----------------------------------------------------------------------------
# Microsoft Platform consumption
# ----------------------------------------------------------------------------

# Wave duration in months
WAVE_MONTHS = {"Approach": 1, "W1": 6, "W2": 9, "W3": 15}

MICROSOFT_LINES = [
    # Component, Approach $/mo, W1 $/mo, W2 $/mo, W3 $/mo  (W3 sized to Disney+ NA, not global)
    ("Azure AI Foundry (model endpoints, agent runtime)",          2500,  10000,  28000,  55000),
    ("Azure OpenAI (GPT-4o + reasoning models for narrative)",        0,   5000,  24000,  75000),
    ("Microsoft Fabric F-SKU (lakehouse, semantic model)",           0,  12000,  28000,  55000),
    ("Azure storage + networking (blob, private endpoints, NAT)",  1000,   2500,   5500,  10000),
    ("Microsoft Entra (identity bindings + B2B trust)",             500,   1000,   2000,   3500),
    ("Microsoft Teams (Adaptive Card platform consumption)",          0,    500,   1500,   3000),
    ("Power BI Premium (Embedded + capacity)",                        0,   3000,   7500,  14000),
    ("Microsoft Purview (data governance + privacy lineage)",         0,   2000,   4500,   9000),
    ("Application Insights / Log Analytics",                       300,   1500,   3500,   6500),
    ("Defender + Security tooling",                                 200,   1000,   2500,   4500),
    ("API Management + Front Door",                                   0,   1200,   3000,   5000),
    ("Marketplace credit / consumption rebate (estimated)",           0,  -2000,  -8000, -20000),
]


def sheet_microsoft_platform(wb):
    ws = wb.create_sheet("Microsoft Platform")
    ws.cell(row=1, column=1, value="Microsoft platform consumption").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Monthly consumption × wave duration in months").font = SUB_FONT

    headers = ["Component",
               "Approach $/mo", "W1 $/mo", "W2 $/mo", "W3 $/mo",
               "Approach Total", "W1 Total", "W2 Total", "W3 Total"]
    widths = [54, 14, 12, 12, 12, 14, 14, 14, 14]
    write_hdrs(ws, headers, widths, row=4)

    totals = [0, 0, 0, 0]
    r = 5
    for comp, a, w1, w2, w3 in MICROSOFT_LINES:
        a_t  = round(a  * WAVE_MONTHS["Approach"])
        w1_t = round(w1 * WAVE_MONTHS["W1"])
        w2_t = round(w2 * WAVE_MONTHS["W2"])
        w3_t = round(w3 * WAVE_MONTHS["W3"])
        totals[0] += a_t; totals[1] += w1_t; totals[2] += w2_t; totals[3] += w3_t

        cell(ws, r, 1, comp, font=Font(name="Calibri", size=10, bold=True))
        money(ws, r, 2, a)
        money(ws, r, 3, w1)
        money(ws, r, 4, w2)
        money(ws, r, 5, w3)
        money(ws, r, 6, a_t)
        money(ws, r, 7, w1_t)
        money(ws, r, 8, w2_t)
        money(ws, r, 9, w3_t)
        if r % 2 == 0:
            for j in range(1, 10):
                ws.cell(row=r, column=j).fill = BAND_FILL
        r += 1

    cell(ws, r, 1, "TOTAL Microsoft Platform", font=TOTAL_FONT, fill=GOLD_BG)
    for j in (2, 3, 4, 5):
        cell(ws, r, j, "", fill=GOLD_BG)
    money(ws, r, 6, totals[0], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 7, totals[1], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 8, totals[2], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 9, totals[3], font=TOTAL_FONT, fill=GOLD_BG)
    print(f"  Microsoft Platform · {len(MICROSOFT_LINES)} lines · sub-totals: {totals}")
    return totals


# ----------------------------------------------------------------------------
# Third-Party & Other
# ----------------------------------------------------------------------------

THIRDPARTY_LINES = [
    ("Streaming-event ingest specialist (Disney+ Eng partner)",      0,  120000,  280000,   650000),
    ("Recommendation/ML platform integration specialist",             0,  140000,  240000,   480000),
    ("Marketing automation / channel orchestration vendor",      10000,   60000,  180000,   420000),
    ("Consent platform integration specialist",                       0,   60000,  140000,   320000),
    ("Privacy Office walkthrough + decision-evidence design",     6000,   25000,   85000,   220000),
    ("External-audit walkthrough + reliance review (audit-firm dependent)", 4000, 18000, 70000, 190000),
    ("Change management — retention org training & enablement",       0,   80000,  300000,  1400000),
    ("Disney+ data engineering · synthetic data lab build",           0,   95000,   55000,    40000),
    ("Power BI semantic-model + dashboard design specialist",     5000,   45000,  110000,   240000),
    ("Cybersecurity review + Azure landing-zone hardening",       8000,   55000,   90000,   180000),
    ("Travel and incidentals (delivery team)",                    8000,   45000,   85000,   195000),
    ("Contingency reserve (8% of mid)",                          25000,   75000,  175000,   500000),
]


def sheet_thirdparty(wb):
    ws = wb.create_sheet("Third-Party & Other")
    ws.cell(row=1, column=1, value="Third-party software, specialist services, change management, contingency").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Vendor / specialist / contingency lines that ride alongside the engagement").font = SUB_FONT

    headers = ["Item", "Approach", "Wave 1", "Wave 2", "Wave 3", "Total"]
    widths = [56, 14, 14, 14, 14, 14]
    write_hdrs(ws, headers, widths, row=4)

    totals = [0, 0, 0, 0]
    r = 5
    for item, a, w1, w2, w3 in THIRDPARTY_LINES:
        totals[0] += a; totals[1] += w1; totals[2] += w2; totals[3] += w3
        cell(ws, r, 1, item, font=Font(name="Calibri", size=10, bold=True))
        money(ws, r, 2, a)
        money(ws, r, 3, w1)
        money(ws, r, 4, w2)
        money(ws, r, 5, w3)
        money(ws, r, 6, a + w1 + w2 + w3, font=TOTAL_FONT)
        if r % 2 == 0:
            for j in range(1, 7):
                ws.cell(row=r, column=j).fill = BAND_FILL
        r += 1

    cell(ws, r, 1, "TOTAL Third-Party & Other", font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 2, totals[0], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 3, totals[1], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 4, totals[2], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 5, totals[3], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 6, sum(totals), font=TOTAL_FONT, fill=GOLD_BG)
    print(f"  Third-Party & Other · {len(THIRDPARTY_LINES)} lines · sub-totals: {totals}")
    return totals


# ----------------------------------------------------------------------------
# Commercial Envelope
# ----------------------------------------------------------------------------

def sheet_commercial_envelope(wb, deloitte_totals, ms_totals, tp_totals):
    ws = wb.create_sheet("Commercial Envelope")
    ws.cell(row=1, column=1, value="Commercial Envelope — Low / Mid / High range").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Low = 0.75 × Mid · High = 1.35 × Mid (allows for scope variability)").font = SUB_FONT

    headers = ["Tier", "Deloitte Labor", "Microsoft Platform",
               "Third-Party & Other", "Mid Total", "Low Estimate", "High Estimate"]
    widths = [16, 16, 18, 18, 16, 16, 16]
    write_hdrs(ws, headers, widths, row=4)

    waves = ["Approach", "Wave 1", "Wave 2", "Wave 3"]
    rows = []
    for i, w in enumerate(waves):
        d  = deloitte_totals[i]
        m  = ms_totals[i]
        tp = tp_totals[i]
        mid = d + m + tp
        low = round(mid * 0.75)
        high = round(mid * 1.35)
        rows.append([w, d, m, tp, mid, low, high])

    r = 5
    grand = [0, 0, 0, 0, 0, 0]
    for row in rows:
        cell(ws, r, 1, row[0], font=Font(name="Calibri", size=10, bold=True))
        money(ws, r, 2, row[1])
        money(ws, r, 3, row[2])
        money(ws, r, 4, row[3])
        money(ws, r, 5, row[4], font=TOTAL_FONT)
        money(ws, r, 6, row[5])
        money(ws, r, 7, row[6])
        for j, v in enumerate(row[1:], start=2):
            grand[j-2] = grand[j-2] + (v if isinstance(v, (int, float)) else 0)
        if r % 2 == 0:
            for j in range(1, 8):
                ws.cell(row=r, column=j).fill = BAND_FILL
        r += 1

    cell(ws, r, 1, "ENGAGEMENT TOTAL", font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 2, grand[0], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 3, grand[1], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 4, grand[2], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 5, grand[3], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 6, grand[4], font=TOTAL_FONT, fill=GOLD_BG)
    money(ws, r, 7, grand[5], font=TOTAL_FONT, fill=GOLD_BG)
    print(f"  Commercial Envelope · 4-wave grid · grand totals: D={grand[0]:,} M={grand[1]:,} TP={grand[2]:,} Mid={grand[3]:,}")
    return rows  # returned for ROI sheet


# ----------------------------------------------------------------------------
# Durable Value
# ----------------------------------------------------------------------------

def sheet_durable_value(wb):
    ws = wb.create_sheet("Durable Value")
    ws.cell(row=1, column=1, value="Durable value at W3 run-rate (annualized)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="BASE CASE · Disney+ NA at run-rate. Conservative — does not include cross-streaming-brand fusion or international markets, both of which expand W3 value materially. Anchors the ROI conversation.").font = SUB_FONT

    headers = ["Category", "Calculation basis", "Annual ($)"]
    widths  = [38, 60, 16]
    write_hdrs(ws, headers, widths, row=4)

    items = [
        ("Retention NPV uplift — primary (BASE CASE)",
         "Disney+ NA: ~50M subs × ~11% gross churn = ~5.5M churn events/yr · 7% reduction = ~385K saved × $143 NPV per saved sub",
         55000000),
        ("Cross-brand bundle attach lift",
         "Disney Bundle attach +1.8 pp NA at W3 on the addressable wholesale-to-bundle population × ARPU contribution",
         18000000),
        ("Concession utilization efficiency improvement",
         "Concession budget yielding positive NPV: 70% → 85% on a ~$60M annualized concession line in scope at W3",
         8000000),
        ("Privacy / audit-fee avoidance",
         "Reduction in privacy-evidence collation + auditor-walkthrough hours; ~10–15% of relevant spend",
         3000000),
        ("Retention-staff capacity returned",
         "~85-person retention org × ~12% time reclaimed × loaded cost",
         2500000),
        ("Loop-time compression (partially monetized)",
         "14–21 days → 4 hours signal-to-action; experimentation velocity + retention-org talent retention",
         2000000),
    ]
    r = 5
    total = 0
    for cat, basis, amt in items:
        cell(ws, r, 1, cat, font=Font(name="Calibri", size=10, bold=True))
        cell(ws, r, 2, basis)
        money(ws, r, 3, amt)
        if r % 2 == 0:
            for j in range(1, 4):
                ws.cell(row=r, column=j).fill = BAND_FILL
        total += amt
        r += 1
    cell(ws, r, 1, "TOTAL annualized W3 run-rate value", font=TOTAL_FONT, fill=GOLD_BG)
    cell(ws, r, 2, "Sum of categories above", fill=GOLD_BG)
    money(ws, r, 3, total, font=TOTAL_FONT, fill=GOLD_BG)
    print(f"  Durable Value · {len(items)} categories · ${total:,} annualized")
    return total


# ----------------------------------------------------------------------------
# ROI · Payback
# ----------------------------------------------------------------------------

def sheet_roi_payback(wb, envelope_rows, durable_total):
    ws = wb.create_sheet("ROI · Payback")
    ws.cell(row=1, column=1, value="Cumulative Investment vs. Cumulative Return — time-phased").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Returns ramp: 0% Approach · 0% W1 · ~25% of run-rate during W2 · ~75% during W3 ramp · 100% at W3 run-rate.").font = SUB_FONT

    headers = ["Fiscal Year", "Phase", "Investment (Mid)", "Realized Return",
               "Cumulative Investment", "Cumulative Return", "Net Position"]
    widths = [14, 50, 18, 18, 22, 20, 18]
    write_hdrs(ws, headers, widths, row=4)

    # Map envelope_rows: [Tier, D, M, TP, Mid, Low, High]
    approach_inv = envelope_rows[0][4]
    w1_inv       = envelope_rows[1][4]
    w2_inv       = envelope_rows[2][4]
    w3_inv       = envelope_rows[3][4]

    # Phase the spend + return:
    rows = [
        ("Y1 (FY27)",       "Approach + W1 Foundation begins",
         approach_inv + round(w1_inv * 0.6), 0),
        ("Y2 (FY28 H1)",    "W1 completes · W2 Pilot begins · early returns",
         round(w1_inv * 0.4) + round(w2_inv * 0.5),
         round(durable_total * 0.07)),
        ("Y2 (FY28 H2)",    "W2 Pilot completes · W3 Scale begins",
         round(w2_inv * 0.5) + round(w3_inv * 0.30),
         round(durable_total * 0.18)),
        ("Y3 (FY29)",       "W3 Scale & Fuse — multi-brand rollout in progress",
         round(w3_inv * 0.50),
         round(durable_total * 0.55)),
        ("Y4 (FY30)",       "W3 completes · run-rate steady state",
         round(w3_inv * 0.20),
         round(durable_total * 0.95)),
        ("Y5 (FY31)",       "Run-rate · ongoing · candidate-service expansion",
         round(w3_inv * 0.05),
         durable_total),
        ("Y6+ (FY32+)",     "Run-rate continues; ATLAS-canonical re-use across TMT account portfolio",
         0,
         durable_total),
    ]

    cum_inv = 0
    cum_ret = 0
    payback_year = None
    r = 5
    for fy, phase, inv, ret in rows:
        cum_inv += inv
        cum_ret += ret
        net = cum_ret - cum_inv
        if payback_year is None and net >= 0:
            payback_year = fy

        cell(ws, r, 1, fy, font=Font(name="Calibri", size=10, bold=True))
        cell(ws, r, 2, phase)
        money(ws, r, 3, inv)
        money(ws, r, 4, ret)
        money(ws, r, 5, cum_inv)
        money(ws, r, 6, cum_ret)
        money(ws, r, 7, net,
              font=Font(name="Calibri", size=10, bold=True,
                        color=("2A7F73" if net >= 0 else "A22A39")))
        if r % 2 == 0:
            for j in range(1, 8):
                ws.cell(row=r, column=j).fill = BAND_FILL
        r += 1
    # Payback callout
    r += 1
    cell(ws, r, 1, "Payback achieved", font=TOTAL_FONT, fill=GOLD_BG)
    cell(ws, r, 2, f"Net position turns positive in {payback_year or '(beyond modeled horizon)'}",
         fill=GOLD_BG, font=TOTAL_FONT)
    print(f"  ROI · Payback · payback in {payback_year}")


# ============================================================================
# BUILD
# ============================================================================

def build():
    print("Building TMT-MED-01 Resource Plan...")
    wb = Workbook()
    sheet_readme(wb)
    deloitte_totals = sheet_deloitte_staffing(wb)
    sheet_disney_staffing(wb)
    ms_totals = sheet_microsoft_platform(wb)
    tp_totals = sheet_thirdparty(wb)
    envelope_rows = sheet_commercial_envelope(wb, deloitte_totals, ms_totals, tp_totals)
    durable = sheet_durable_value(wb)
    sheet_roi_payback(wb, envelope_rows, durable)

    target = OUT
    try:
        wb.save(str(target))
    except PermissionError:
        n = 2
        while True:
            alt = target.with_name(target.stem + f"-v{n}" + target.suffix)
            if not alt.exists():
                target = alt; break
            n += 1
        wb.save(str(target))
        print(f"\n[notice] {OUT.name} was locked — saved as {target.name} instead.")
    print(f"\nWrote {target}")


if __name__ == "__main__":
    build()

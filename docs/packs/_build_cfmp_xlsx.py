"""One-off builder for CFMP-Scenario-Chains-v0.1.xlsx.

Sources the 17 existing scenarios from APEX-Scenario-Chains.xlsx (Scenario
Library + Featured Chains sheets) and synthesizes the new
`cfmp-wayfinding-walk-to-product` row inline.

Run once; the xlsx becomes the durable artifact.
"""

from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PARENT = Path(r"C:\Stage\Clients\Industries\APEX\docs\reference\APEX-Scenario-Chains.xlsx")
OUT = Path(r"C:\Stage\Clients\Industries\APEX\docs\packs\CFMP-Scenario-Chains-v0.2.xlsx")

# The 18-scenario shortlist, journey-phase tagged
SHORTLIST = [
    # CHOOSE
    ("rc-personalized-offer-targeting",         "CHOOSE",  "Lite"),
    ("rc-product-review-summary-generation",    "CHOOSE",  "Standard"),
    ("rc-product-search-relevance",             "CHOOSE",  "Standard"),
    ("rc-in-store-advertising-impact",          "CHOOSE",  "Standard"),
    ("rc-sampling-table-engagement",            "CHOOSE",  "Lite"),
    # SELECT
    ("cfmp-wayfinding-walk-to-product",         "SELECT",  "Lite"),
    ("rc-on-shelf-availability-oos-reduction",  "SELECT",  "Standard"),
    ("rc-shelf-gap-realtime-restock-dispatch",  "SELECT",  "Standard"),
    ("rc-aisle-engagement-attribution",         "SELECT",  "Enterprise"),
    ("rc-end-cap-display-roi-tracking",         "SELECT",  "Standard"),
    ("rc-shelf-price-label-compliance-audit",   "SELECT",  "Enterprise"),
    # BUY
    ("rc-customer-wait-time-prediction-at-checkout", "BUY", "Standard"),
    ("rc-cart-dwell-abandonment-rescue",        "BUY",     "Standard"),
    ("rc-self-checkout-shrink-detection",       "BUY",     "Enterprise"),
    ("rc-bopis-pickup-counter-load",            "BUY",     "Enterprise"),
    # SERVICES
    ("rc-loyalty-churn-prediction-winback",     "SERVICES", "Standard"),
    ("rc-loyalty-tier-migration-prediction",    "SERVICES", "Enterprise"),
    ("rc-product-complaint-triage",             "SERVICES", "Enterprise"),
]

# Synthetic row for the brand-new wayfinding scenario (not in parent xlsx).
# Modelled on the patterns I saw in Featured Chains.
WAYFINDING_FEATURED = {
    "Scenario ID": "cfmp-wayfinding-walk-to-product",
    "Title": "Walk-to-product indoor wayfinding",
    "Service Code": "CFMP-E2E-01",
    "Service": "CFMP-E2E-01",
    "Domain": "Customer Experience",
    "Schemas": "CFMP.StoreMap (Azure Maps Creator Dataset) · MERML.Planogram · CXML.CustomerSession",
    "Brief (the moment)": (
        "Customer asks the store assistant 'where is X?' or taps a product card; "
        "the agent resolves SKU → planogram unit_id → Azure Maps Creator Wayfinding "
        "API call from customer's current point and returns zone, aisle, distance, "
        "duration, and turn-by-turn directions."
    ),
    "KPI / Outcome": "Time-to-find -60% · trip basket +8% · first-time wayfinder success >92%",
    "Featured?": "⭐ Featured",
    "Device(s)": "Customer phone (BLE / Wi-Fi RTT) + Vision AI Dev Kit (camera-assisted landmark recovery for re-anchor)",
    "W1 Foundation": (
        "SOR: Planogram + Store CAD/DWG floor plans · Real-Time Hub · Bronze · Raw Landing · "
        "Tokenizer · Azure Maps Creator resource provisioned per tenant · Drawing Package "
        "uploaded → Dataset + Tileset + Stateset · CFMP.StoreMap VV manifest wires Dataset "
        "ID + SKU↔unit_id table (refreshed nightly from MERML.Planogram)"
    ),
    "W2 Pilot (you are here)": (
        "Wayfinding scenario · Event Fires (customer asks 'where is X') · DAG Orchestrator · "
        "Agent 1: Assess (resolve SKU + customer location point) · Agent 2: Classify (in-aisle vs cross-zone) · "
        "Agent 3: Quantify (call Azure Maps Wayfinding REST API for distance / duration / route geometry) · "
        "Agent 4: Approve (customer consent gate on first session) · "
        "Agent 5: Act (return structured route + GeoJSON polyline to phone Adaptive Card) · "
        "Agent 6: Evidence-Write (LedgerRow with consent-hash + route)"
    ),
    "W3 Scale & Fuse": (
        "Enterprise Scale · Azure Maps Web SDK indoor view embedded in phone app · "
        "Stateset live updates for SCO/BOPIS counter occupancy · "
        "Fuse: Aisle engagement attribution · Fuse: OSA / shelf-gap dispatch · "
        "Fuse: Cart dwell rescue (associate routed via Wayfinding distance matrix) · "
        "Multi-store rollout playbook · Operate-ready"
    ),
    "Scenario (the moment)": (
        "Customer is mid-trip in a 60,000-sq-ft store, looking for an ingredient that's "
        "moved aisles since their last visit. Without help, they either give up "
        "(trip-basket loss) or wander (NPS hit). With CFMP wayfinding, the phone app "
        "resolves SKU → planogram unit_id → Azure Maps Creator route in under 1 second "
        "and walks them there."
    ),
    "Solution (architectural approach)": (
        "Azure Maps Creator (Indoor Maps + Wayfinding REST API) as APEX-M's implementation "
        "of proposed Interface #15 'Maps & Wayfinding'. Per-tenant Creator resource holds "
        "the retailer's Dataset (vector geometry from CAD), Tileset (rendered tiles), and "
        "Stateset (live occupancy). Agent calls POST /wayfinding/route with SKU's unit_id "
        "and customer's current point; response includes route GeoJSON the Web SDK indoor "
        "view renders. APEX-G uses Google Maps Geospatial Creator; APEX-A uses Amazon "
        "Location Service indoor maps — same abstract interface."
    ),
    "Use Case (Wave 2 delivery)": (
        "Walk-to-product wayfinding in 1 flagship store + 2 satellite stores. Reads "
        "CFMP.StoreMap (Creator Dataset + SKU↔unit_id), MERML.Planogram, CXML.CustomerSession. "
        "Writes route-served + consent-hash LedgerRows. ~3 scenario sub-flows: cold-start "
        "(no beacon), warm (beacon active), recovery (Vision AI Dev Kit camera re-anchor)."
    ),
    "Service (productized)": (
        "CFMP-E2E-01 Wayfinding & Walk-to-Product. Commercial envelope tied to store "
        "count (Pack Lite = 1 store, Standard = 5 stores, Enterprise = full chain). "
        "Azure Maps Creator is a billable Azure resource — passed through to retailer "
        "or bundled into the Pack subscription per tier."
    ),
    "Persona (operator · HITL approver)": (
        "Primary `customer` (NEW persona type) · Identity = loyalty ID · consent gate on first session. "
        "Operator: `store_manager` · Approves Drawing Package re-upload after remodel via Teams Adaptive Card."
    ),
}


def main() -> None:
    # ----- load parent data ---------------------------------------------------
    lib = pd.read_excel(PARENT, sheet_name="Scenario Library")
    skuc = pd.read_excel(PARENT, sheet_name="Scenario→KPI Chain")
    feat = pd.read_excel(PARENT, sheet_name="Featured Chains")

    # filter to our shortlist (preserving order)
    ordered_ids = [sid for sid, _, _ in SHORTLIST]
    phase_map = {sid: ph for sid, ph, _ in SHORTLIST}
    tier_map  = {sid: tr for sid, _, tr in SHORTLIST}

    def row_for(df: pd.DataFrame, sid: str) -> dict:
        m = df[df["Scenario ID"] == sid]
        if m.empty:
            return {}
        return m.iloc[0].to_dict()

    # ----- workbook + styles --------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill("solid", start_color="1E2761")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Arial", size=10)
    PHASE_FILLS = {
        "CHOOSE":   PatternFill("solid", start_color="FFF4E5"),
        "SELECT":   PatternFill("solid", start_color="E5F4FF"),
        "BUY":      PatternFill("solid", start_color="E5FFE9"),
        "SERVICES": PatternFill("solid", start_color="F5E5FF"),
    }
    TIER_FONT = {
        "Lite":       Font(name="Arial", size=10, bold=True, color="0B6E1F"),
        "Standard":   Font(name="Arial", size=10, bold=True, color="9A6300"),
        "Enterprise": Font(name="Arial", size=10, bold=True, color="6E0B0B"),
    }
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"

    def autosize(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ----- Summary sheet ------------------------------------------------------
    ws = wb.create_sheet("Summary")
    rows = [
        ("CFMP Scenario Chains — Customer Focused Merchandise Pack v0.2", "", ""),
        ("", "", ""),
        ("Pack",                  "Customer Focused Merchandise Pack (CFMP)", ""),
        ("Pack ID",               "cfmp",                  ""),
        ("Cloud profile",         "APEX-M primary; APEX-G / APEX-A via shared 10-asset bundle", ""),
        ("Primary schemas",       "MERML + CXML + SCML (+ new CFMP.StoreMap)", ""),
        ("Practices",             "RC (Retail & Consumer)", ""),
        ("Customer journey spine","CHOOSE · SELECT · BUY · SERVICES", ""),
        ("Total scenarios",       18, ""),
        ("Featured chains",       5, ""),
        ("New scenarios in this pack", 1, "cfmp-wayfinding-walk-to-product"),
        ("Source sheet for 17 of 18", "APEX-Scenario-Chains.xlsx · Scenario Library", ""),
        ("", "", ""),
        ("Sub-tier",  "Scenario count", "Price band"),
        ("Lite",      "3 scenarios",   "$150K–$250K · 4–6 weeks · BVA + DCIF"),
        ("Standard",  "10 scenarios",  "$500K–$1.5M · 12–16 weeks · DCIF + Client + ISV burndown"),
        ("Enterprise","18 scenarios (all)", "$1.5M–$3.5M · 6–9 months · Client direct + T&M"),
        ("", "", ""),
        ("Document version", "v0.2 · 2026-05-23 (Azure Maps Creator wayfinding · Architecture v5 aligned)", ""),
        ("Author",           "Keven Markham, VP · DMTSP", ""),
        ("Independence posture", "All CFMP materials use 'Deloitte's Microsoft practice' / 'DMTSP'. Never 'partner' or 'alliance'.", ""),
        ("Modeled on",       "APEX-Scenario-Chains.xlsx · same column structure as Featured Chains", ""),
    ]
    for r, vals in enumerate(rows, start=1):
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1E2761")
    autosize(ws, [38, 48, 50])

    # ----- Scenario Library sheet (9 cols) ------------------------------------
    cols = ["#","Scenario ID","Title","Service Code","Domain","Schemas","Brief (the moment)","KPI / Outcome","Phase","Sub-tier","Featured?","Device(s)"]
    ws = wb.create_sheet("Scenario Library")
    ws.append(cols)
    style_header(ws, len(cols))
    for i, sid in enumerate(ordered_ids, start=1):
        if sid == "cfmp-wayfinding-walk-to-product":
            r = WAYFINDING_FEATURED
            rec = [i, r["Scenario ID"], r["Title"], r["Service Code"], r["Domain"],
                   r["Schemas"], r["Brief (the moment)"], r["KPI / Outcome"],
                   phase_map[sid], tier_map[sid], r["Featured?"], r["Device(s)"]]
        else:
            src = row_for(lib, sid)
            rec = [i, sid, src.get("Title",""), src.get("Service Code",""), src.get("Domain",""),
                   src.get("Schemas",""), src.get("Brief (the moment)",""), src.get("KPI / Outcome",""),
                   phase_map[sid], tier_map[sid], src.get("Featured?",""), src.get("Device(s)","")]
        ws.append([("" if pd.isna(x) else x) for x in rec])
        # phase tint + tier color
        row_idx = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_idx, column=c).fill = PHASE_FILLS[phase_map[sid]]
            ws.cell(row=row_idx, column=c).border = BORDER
        ws.cell(row=row_idx, column=10).font = TIER_FONT[tier_map[sid]]
    autosize(ws, [4, 38, 36, 14, 22, 30, 60, 38, 10, 12, 14, 40])

    # ----- Featured Chains sheet (14 cols, only featured scenarios) -----------
    cols = ["#","Scenario ID","Title","Service","Domain",
            "W1 Foundation","W2 Pilot (you are here)","W3 Scale & Fuse",
            "Scenario (the moment)","Solution (architectural approach)",
            "Use Case (Wave 2 delivery)","Service (productized)",
            "Persona (operator · HITL approver)","KPI / Outcome","Device(s)"]
    ws = wb.create_sheet("Featured Chains")
    ws.append(cols)
    style_header(ws, len(cols))
    featured_ids = ["rc-sampling-table-engagement", "cfmp-wayfinding-walk-to-product",
                    "rc-cart-dwell-abandonment-rescue", "rc-on-shelf-availability-oos-reduction",
                    "rc-loyalty-churn-prediction-winback"]
    for i, sid in enumerate(featured_ids, start=1):
        if sid == "cfmp-wayfinding-walk-to-product":
            r = WAYFINDING_FEATURED
            rec = [i, r["Scenario ID"], r["Title"], r["Service"], r["Domain"],
                   r["W1 Foundation"], r["W2 Pilot (you are here)"], r["W3 Scale & Fuse"],
                   r["Scenario (the moment)"], r["Solution (architectural approach)"],
                   r["Use Case (Wave 2 delivery)"], r["Service (productized)"],
                   r["Persona (operator · HITL approver)"], r["KPI / Outcome"], r["Device(s)"]]
        else:
            src = row_for(feat, sid)
            if not src:
                # Featured chain not yet defined; synthesize a stub from Library
                src_lib = row_for(lib, sid)
                rec = [i, sid, src_lib.get("Title",""), src_lib.get("Service Code",""), src_lib.get("Domain",""),
                       "(W1 to be detailed — see Scenario Library brief)",
                       f"{src_lib.get('Title','')} · 6-agent fleet · standard archetype · HITL gate at Agent 4",
                       "Enterprise scale + cross-pack fuse candidates: TBD",
                       src_lib.get("Brief (the moment)",""),
                       "Standard archetype: hierarchical-root + sequential-with-hitl-gate · 6-agent",
                       f"Wave 2 pilot for {src_lib.get('Title','')}",
                       src_lib.get("Service Code",""),
                       "Primary persona TBD · HITL approver TBD",
                       src_lib.get("KPI / Outcome",""),
                       src_lib.get("Device(s)","")]
            else:
                rec = [i, sid, src.get("Title",""), src.get("Service",""), src.get("Domain",""),
                       src.get("W1 Foundation",""), src.get("W2 Pilot (you are here)",""), src.get("W3 Scale & Fuse",""),
                       src.get("Scenario (the moment)",""), src.get("Solution (architectural approach)",""),
                       src.get("Use Case (Wave 2 delivery)",""), src.get("Service (productized)",""),
                       src.get("Persona (operator · HITL approver)",""),
                       src.get("KPI / Outcome",""), src.get("Device(s)","")]
        ws.append([("" if pd.isna(x) else x) for x in rec])
        row_idx = ws.max_row
        ph = phase_map.get(sid, "CHOOSE")
        for c in range(1, len(cols) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_idx, column=c).fill = PHASE_FILLS[ph]
            ws.cell(row=row_idx, column=c).border = BORDER
    autosize(ws, [4, 38, 32, 14, 22, 60, 60, 50, 50, 50, 45, 30, 38, 38, 40])

    # ----- Scenario → KPI Chain (compact 8-col, all 18) -----------------------
    cols = ["Scenario ID","Phase","Sub-tier","Scenario (the moment)",
            "Solution / Agent (ORCH archetype)","Use Cases (decomposed)",
            "Service (Deloitte unit)","Personas (operator · HITL approver)",
            "KPI (measurable outcome)","Featured?","Device(s)"]
    ws = wb.create_sheet("Scenario→KPI Chain")
    ws.append(cols)
    style_header(ws, len(cols))
    for sid in ordered_ids:
        if sid == "cfmp-wayfinding-walk-to-product":
            r = WAYFINDING_FEATURED
            rec = [sid, phase_map[sid], tier_map[sid], r["Brief (the moment)"],
                   r["Solution (architectural approach)"], r["Use Case (Wave 2 delivery)"],
                   r["Service Code"], r["Persona (operator · HITL approver)"],
                   r["KPI / Outcome"], r["Featured?"], r["Device(s)"]]
        else:
            src = row_for(skuc, sid)
            rec = [sid, phase_map[sid], tier_map[sid],
                   src.get("Scenario (the moment)",""),
                   src.get("Solution / Agent (ORCH archetype)",""),
                   src.get("Use Cases (decomposed)",""),
                   src.get("Service (Deloitte unit)",""),
                   src.get("Personas (operator · HITL approver)",""),
                   src.get("KPI (measurable outcome)",""),
                   src.get("Featured?",""), src.get("Device(s)","")]
        ws.append([("" if pd.isna(x) else x) for x in rec])
        row_idx = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_idx, column=c).fill = PHASE_FILLS[phase_map[sid]]
            ws.cell(row=row_idx, column=c).border = BORDER
    autosize(ws, [38, 10, 12, 55, 55, 45, 18, 36, 36, 14, 40])

    # ----- Pack Sub-Tiers sheet -----------------------------------------------
    ws = wb.create_sheet("Pack Sub-Tiers")
    cols = ["Sub-tier","Price band","Duration","Funding","Scenario count","Scenarios included"]
    ws.append(cols)
    style_header(ws, len(cols))
    tier_groups = {"Lite": [], "Standard": [], "Enterprise": []}
    for sid, _, tr in SHORTLIST:
        # Standard contains Lite; Enterprise contains Standard (per APEX additive rule)
        for t in ("Lite", "Standard", "Enterprise"):
            if (tr == "Lite") or (tr == "Standard" and t != "Lite") or (tr == "Enterprise" and t == "Enterprise"):
                tier_groups[t].append(sid)
    rows = [
        ("Lite",       "$150K–$250K",   "4–6 weeks",   "BVA + DCIF",                            f"{len(tier_groups['Lite'])} scenarios",       "\n".join(tier_groups["Lite"])),
        ("Standard",   "$500K–$1.5M",   "12–16 weeks", "DCIF + Client + ISV burndown",         f"{len(tier_groups['Standard'])} scenarios",   "\n".join(tier_groups["Standard"])),
        ("Enterprise", "$1.5M–$3.5M",   "6–9 months",  "Client direct + T&M extensions",       f"{len(tier_groups['Enterprise'])} scenarios", "\n".join(tier_groups["Enterprise"])),
    ]
    for r in rows:
        ws.append(r)
        row_idx = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_idx, column=c).border = BORDER
        ws.row_dimensions[row_idx].height = 18 * (len(tier_groups[r[0]]) + 1)
    autosize(ws, [14, 18, 14, 32, 16, 50])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()

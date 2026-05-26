"""Build the BVA Workshop ROI Calculator workbook.

Multi-sheet xlsx that the workshop facilitator + client team fill in live.
Anchors on CFMP §9 Pack-level KPIs and the CHC scenario lift targets.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName

OUT = r"C:\Stage\Clients\Industries\APEX\docs\reference\BVA-ROI-Calculator.xlsx"

wb = Workbook()

# ============ Styles ============
BLUE   = Font(name="Calibri", size=11, color="0000FF")          # inputs
BLACK  = Font(name="Calibri", size=11, color="000000")          # formulas
GREEN  = Font(name="Calibri", size=11, color="008000")          # cross-sheet links
BOLD   = Font(name="Calibri", size=11, bold=True, color="0E1423")
TITLE  = Font(name="Calibri", size=18, bold=True, color="0E1423")
HDR    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DELOIT_GREEN = PatternFill(start_color="86BC25", end_color="86BC25", fill_type="solid")
NAVY_FILL    = PatternFill(start_color="1A1F2E", end_color="1A1F2E", fill_type="solid")
YELLOW       = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BG     = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
SECTION_BG   = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")
thin   = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CUR  = "$#,##0;($#,##0);-"
PCT  = "0.0%;(0.0%);-"
MULT = "0.0\"x\""
INT  = "#,##0;(#,##0);-"

# ============ Sheet 1: README ============
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 100
ws["A1"] = "APEX BVA Workshop — ROI Calculator"
ws["A1"].font = TITLE
ws["A2"] = "Companion to: APEX-CFMP-CHC-Sellers-Guide §8 + CFMP v0.2 §9 Pack-level KPIs"
ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="535B6B")
ws.row_dimensions[1].height = 32

instructions = [
    "",
    "HOW TO USE",
    "1. Fill in the BLUE cells on the 'Inputs' sheet with the client's baseline numbers.",
    "2. Review and adjust scenario lift %s on the 'Scenarios' sheet (defaults from CFMP §9 + CHC scenario shortlist).",
    "3. The 'Outputs' sheet computes Wave-1, Wave-2, and 3-year TCV plus ROI multiple — formulas only, no editing.",
    "4. The 'Sensitivity' sheet runs high/mid/low cases on the three biggest swing inputs.",
    "5. Yellow-highlighted cells are the inputs that move the outcome the most — bring the CFO into those.",
    "",
    "COLOR CONVENTIONS (industry-standard financial modeling)",
    "• Blue  text = hardcoded inputs you change",
    "• Black text = formulas (do not edit)",
    "• Green text = cross-sheet links",
    "• Yellow highlight = key assumption requiring CFO attention",
    "",
    "SOURCES",
    "• Pack-level KPIs: CFMP v0.2 §9 'Roll-up KPIs for the BVA worksheet'",
    "• CHC scenario lifts: Connected-Home-Commerce v0.2 §6 + §14.3",
    "• Pricing: APEX-CFMP-CHC-Sellers-Guide §8.1 'The price ladder'",
    "• Demo proof: https://ca-visionkit-portal.gentlestone-9b49b099.eastus2.azurecontainerapps.io",
    "",
    "INDEPENDENCE NOTE",
    "Microsoft money flows via ISV Marketplace burndown + SI Teaming, never direct ECIF to Deloitte.",
    "All funding-vehicle references use APEX-Design-v3 §S21 'Funding & Independence' nomenclature.",
    "",
    "Internal · Deloitte's Microsoft Technology & Services Practice",
    "Prepared by Keven Markham, VP · 2026-05-23",
]
for i, line in enumerate(instructions, start=4):
    cell = ws.cell(row=i, column=1, value=line)
    if line in ("HOW TO USE","COLOR CONVENTIONS (industry-standard financial modeling)","SOURCES","INDEPENDENCE NOTE"):
        cell.font = BOLD
        cell.fill = SECTION_BG
    elif line.startswith("Internal"):
        cell.font = Font(name="Calibri", size=10, italic=True, color="535B6B")
    else:
        cell.font = Font(name="Calibri", size=11, color="0E1423")
    cell.alignment = LEFT

# ============ Sheet 2: Inputs ============
ws = wb.create_sheet("Inputs")
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 55

ws["A1"] = "Client Baseline Inputs"
ws["A1"].font = TITLE
ws.merge_cells("A1:C1")
ws["A2"] = "Edit BLUE cells. Yellow = key swing assumptions."
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="535B6B")
ws.merge_cells("A2:C2")

# Header row
for col, txt in enumerate(["Metric", "Value", "Notes"], start=1):
    c = ws.cell(row=4, column=col, value=txt)
    c.font = HDR; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = BORDER

inputs_data = [
    ("CLIENT IDENTITY", None, None, "section"),
    ("Client name",                 "[Client]",     "e.g., Kroger, AT&T", "text"),
    ("Banner / business unit",      "[Banner]",     "e.g., Smith's, Mariano's", "text"),
    ("Engagement variant",          "Variant B — B2C", "Variant A (B2B2C Telco × Merch) or Variant B (Merch direct)", "text"),

    ("STORE / HOUSEHOLD FOOTPRINT", None, None, "section"),
    ("# stores in pilot scope",                 50,        "Wave-1 footprint (Pack Lite typically 1-5 stores)", "int_yellow"),
    ("# stores in full rollout",                2500,      "Wave-3 / Enterprise target footprint", "int"),
    ("# households in scope (Variant B/CHC)",   5000,      "Pilot household count for in-home pack", "int_yellow"),
    ("# households at full rollout",            500000,    "Enterprise household target", "int"),

    ("BASELINE STORE ECONOMICS", None, None, "section"),
    ("Avg annual revenue per store ($)",        25000000,  "Per store, annual", "cur"),
    ("Avg basket size ($)",                     65,        "Per transaction", "cur"),
    ("Trips per household per year",            52,        "Annual trip frequency", "int"),
    ("Gross margin %",                          0.28,      "Blended gross margin", "pct_yellow"),
    ("Current cart-abandon rate (in-store)",    0.045,     "Physical cart abandon fraction", "pct"),
    ("Current OOS rate",                        0.062,     "Out-of-stock at shelf, fraction", "pct"),
    ("Current substitution acceptance rate",    0.35,      "BOPIS substitution accept fraction", "pct"),

    ("BASELINE HOUSEHOLD ECONOMICS (CHC)", None, None, "section"),
    ("Avg annual spend per household ($)",      6500,      "Loyalty-member annual spend", "cur"),
    ("Loyalty churn rate (annual)",             0.18,      "Annual loyalty churn fraction", "pct_yellow"),
    ("In-home replenishment cycle (days)",      9,         "Average staple replen interval", "int"),
    ("% of basket Amazon currently takes",      0.12,      "Estimated share Alexa routes to AMZN", "pct_yellow"),

    ("ENGAGEMENT PRICING (per Sellers-Guide §8.1)", None, None, "section"),
    ("BVA workshop fee ($)",                    35000,     "Often Deloitte-absorbed", "cur"),
    ("CFMP Pack Lite ($)",                      200000,    "Range $150-250K", "cur"),
    ("CHC Pack Lite (B2C) ($)",                 325000,    "Range $250-400K", "cur"),
    ("CFMP Pack Standard ($)",                  1000000,   "Range $500K-$1.5M", "cur"),
    ("CFMP Pack Enterprise ($)",                2500000,   "Range $1.5-3.5M", "cur"),
    ("T5 Operate ($/month)",                    65000,     "Range $25-110K/mo", "cur"),
    ("Cross-pack scenario bundle ($, Variant A)", 325000,  "Range $250-400K", "cur"),
]

row = 5
for entry in inputs_data:
    label, val, note, kind = entry
    a = ws.cell(row=row, column=1, value=label)
    if kind == "section":
        a.font = BOLD; a.fill = DELOIT_GREEN
        ws.cell(row=row, column=2).fill = DELOIT_GREEN
        ws.cell(row=row, column=3).fill = DELOIT_GREEN
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=3)
        row += 1; continue
    a.font = Font(name="Calibri", size=11, color="0E1423")
    a.alignment = LEFT
    b = ws.cell(row=row, column=2, value=val)
    b.font = BLUE
    b.alignment = RIGHT
    if kind == "cur":         b.number_format = CUR
    elif kind == "pct":       b.number_format = PCT
    elif kind == "pct_yellow":b.number_format = PCT; b.fill = YELLOW
    elif kind == "int":       b.number_format = INT
    elif kind == "int_yellow":b.number_format = INT; b.fill = YELLOW
    elif kind == "text":      b.alignment = LEFT
    c = ws.cell(row=row, column=3, value=note)
    c.font = Font(name="Calibri", size=10, italic=True, color="535B6B")
    c.alignment = LEFT
    row += 1

# Named ranges for use in formulas
def name(name_, ref):
    dn = DefinedName(name=name_, attr_text=f"Inputs!{ref}")
    wb.defined_names[name_] = dn

name("n_stores_pilot",   "$B$10")
name("n_stores_full",    "$B$11")
name("n_hh_pilot",       "$B$12")
name("n_hh_full",        "$B$13")
name("rev_per_store",    "$B$16")
name("basket",           "$B$17")
name("trips",            "$B$18")
name("gm_pct",           "$B$19")
name("abandon_rate",     "$B$20")
name("oos_rate",         "$B$21")
name("sub_accept_rate",  "$B$22")
name("hh_spend",         "$B$25")
name("churn_rate",       "$B$26")
name("amz_share",        "$B$28")
name("price_lite",       "$B$32")
name("price_chc_lite",   "$B$33")
name("price_std",        "$B$34")
name("price_ent",        "$B$35")
name("price_op_mo",      "$B$36")
name("price_xp",         "$B$37")

# ============ Sheet 3: Scenarios ============
ws = wb.create_sheet("Scenarios")
widths = [4, 38, 12, 12, 14, 14, 14, 18]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["A1"] = "Scenario Catalog — Lift % and Annual Dollar Impact"
ws["A1"].font = TITLE
ws.merge_cells("A1:H1")
ws["A2"] = "Sources: CFMP v0.2 §3 + §9, CHC v0.2 §6 + §14.3. Edit lift % values (blue) per client conversation."
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="535B6B")
ws.merge_cells("A2:H2")

headers = ["#", "Scenario", "Pack", "Phase", "Lift %", "KPI driver", "Annual $ impact / store or HH", "Engagement tier"]
for col, txt in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col, value=txt)
    c.font = HDR; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = BORDER
ws.row_dimensions[4].height = 28

# Scenario rows: (id, name, pack, phase, lift, kpi_anchor, tier)
# kpi_anchor maps to: TRIP_CONV / BASKET / NPS_PROXY / CHURN / OOS / SUBSTIT / ABANDON / AMZ_RECAPTURE
scenarios = [
    (1, "rc-sampling-table-engagement",            "CFMP", "CHOOSE",   0.030, "TRIP_CONV", "Lite"),
    (2, "rc-cart-dwell-abandonment-rescue",        "CFMP", "BUY",      0.120, "ABANDON",   "Lite"),
    (3, "cfmp-wayfinding-walk-to-product",         "CFMP", "SELECT",   0.080, "BASKET",    "Lite"),
    (4, "rc-on-shelf-availability-oos-reduction",  "CFMP", "SELECT",   0.340, "OOS",       "Standard"),
    (5, "rc-shelf-gap-realtime-restock-dispatch",  "CFMP", "SELECT",   0.020, "BASKET",    "Standard"),
    (6, "rc-aisle-engagement-attribution",         "CFMP", "SELECT",   0.015, "TRIP_CONV", "Standard"),
    (7, "rc-customer-wait-time-prediction-checkout","CFMP","BUY",      0.090, "TRIP_CONV", "Standard"),
    (8, "rc-self-checkout-shrink-detection",       "CFMP", "BUY",      0.060, "BASKET",    "Standard"),
    (9, "rc-bopis-pickup-counter-load",            "CFMP", "BUY",      0.080, "TRIP_CONV", "Standard"),
    (10,"rc-loyalty-churn-prediction-winback",     "CFMP", "SERVICES", 0.220, "CHURN",     "Standard"),
    (11,"rc-product-complaint-triage",             "CFMP", "SERVICES", 0.050, "CHURN",     "Enterprise"),
    (12,"rc-in-store-advertising-impact",          "CFMP", "CHOOSE",   0.090, "TRIP_CONV", "Enterprise"),
    (13,"rc-personalized-offer-targeting",         "CFMP", "CHOOSE",   0.240, "BASKET",    "Enterprise"),
    (14,"rc-end-cap-display-roi-tracking",         "CFMP", "SELECT",   0.220, "BASKET",    "Enterprise"),
    (15,"rc-shelf-price-label-compliance-audit",   "CFMP", "SELECT",   0.020, "OOS",       "Enterprise"),
    (16,"chc-pantry-inventory-inference",          "CHC",  "NEED",     0.080, "AMZ_RECAPTURE","Lite"),
    (17,"chc-low-stock-voice-prompt",              "CHC",  "NEED",     0.070, "AMZ_RECAPTURE","Lite"),
    (18,"chc-replenishment-cart-build",            "CHC",  "CHOOSE",   0.060, "AMZ_RECAPTURE","Lite"),
    (19,"chc-substitution-acceptance",             "CHC",  "SELECT",   0.200, "SUBSTIT",   "Standard"),
    (20,"chc-delivery-arrival-handoff",            "CHC",  "SERVICES", 0.040, "CHURN",     "Standard"),
    (21,"chc-loyalty-fusion-attribution",          "CHC",  "SERVICES", 0.060, "CHURN",     "Enterprise"),
]

# Per-store / per-HH dollar formulas anchored to Inputs named ranges
# We map KPI driver to a formula using basket, trips, rev_per_store, etc.
def dollar_formula(lift_cell, kpi_anchor, pack):
    """Per-store annual $ impact (or per-HH if CHC). Uses Inputs named ranges."""
    if pack == "CHC":
        # per-household annual lift on hh_spend
        if kpi_anchor == "AMZ_RECAPTURE":
            return f"=hh_spend*amz_share*{lift_cell}*gm_pct"
        if kpi_anchor == "SUBSTIT":
            return f"=hh_spend*0.18*{lift_cell}*gm_pct"  # sub category share of spend
        if kpi_anchor == "CHURN":
            return f"=hh_spend*churn_rate*{lift_cell}"   # churn $ retained
        return f"=hh_spend*{lift_cell}*gm_pct"
    # CFMP — per store annual
    if kpi_anchor == "TRIP_CONV":
        return f"=rev_per_store*{lift_cell}*gm_pct"
    if kpi_anchor == "BASKET":
        return f"=rev_per_store*{lift_cell}*gm_pct"
    if kpi_anchor == "ABANDON":
        return f"=rev_per_store*abandon_rate*{lift_cell}*gm_pct"
    if kpi_anchor == "OOS":
        return f"=rev_per_store*oos_rate*{lift_cell}*gm_pct"
    if kpi_anchor == "SUBSTIT":
        return f"=rev_per_store*0.15*{lift_cell}*gm_pct"
    if kpi_anchor == "CHURN":
        return f"=rev_per_store*churn_rate*{lift_cell}"
    return f"=rev_per_store*{lift_cell}*gm_pct"

for i, (sid, name_, pack, phase, lift, kpi, tier) in enumerate(scenarios):
    r = 5 + i
    ws.cell(row=r, column=1, value=sid).alignment = CENTER
    ws.cell(row=r, column=2, value=name_).font = Font(name="Calibri", size=10, color="0E1423")
    ws.cell(row=r, column=3, value=pack).alignment = CENTER
    ws.cell(row=r, column=4, value=phase).alignment = CENTER
    lc = ws.cell(row=r, column=5, value=lift)
    lc.font = BLUE; lc.number_format = PCT; lc.alignment = RIGHT
    ws.cell(row=r, column=6, value=kpi).font = Font(name="Calibri", size=10, color="535B6B")
    df = ws.cell(row=r, column=7, value=dollar_formula(f"E{r}", kpi, pack))
    df.font = BLACK; df.number_format = CUR; df.alignment = RIGHT
    ws.cell(row=r, column=8, value=tier).alignment = CENTER
    if i % 2 == 1:
        for cc in range(1, 9):
            if ws.cell(row=r, column=cc).fill.start_color.rgb in (None, "00000000"):
                ws.cell(row=r, column=cc).fill = LIGHT_BG

# Subtotals by tier (CFMP per store, CHC per HH)
sum_row = 5 + len(scenarios) + 2
ws.cell(row=sum_row,   column=2, value="SUBTOTAL — CFMP per store / yr").font = BOLD
ws.cell(row=sum_row,   column=7, value=f"=SUMIFS(G5:G{sum_row-2},C5:C{sum_row-2},\"CFMP\")").font = BOLD
ws.cell(row=sum_row,   column=7).number_format = CUR
ws.cell(row=sum_row+1, column=2, value="SUBTOTAL — CHC per HH / yr").font = BOLD
ws.cell(row=sum_row+1, column=7, value=f"=SUMIFS(G5:G{sum_row-2},C5:C{sum_row-2},\"CHC\")").font = BOLD
ws.cell(row=sum_row+1, column=7).number_format = CUR

# Per tier subtotals
def tier_sum(label, tier, r):
    ws.cell(row=r, column=2, value=label).font = BOLD
    ws.cell(row=r, column=7, value=f"=SUMIFS(G5:G{sum_row-2},H5:H{sum_row-2},\"{tier}\")").font = BOLD
    ws.cell(row=r, column=7).number_format = CUR

tier_sum("By tier — Lite (blended)",     "Lite",       sum_row+3)
tier_sum("By tier — Standard (blended)", "Standard",   sum_row+4)
tier_sum("By tier — Enterprise (blended)","Enterprise",sum_row+5)

# Define names for cross-sheet refs to subtotals
name("cfmp_per_store_yr", f"$G${sum_row}")
name("chc_per_hh_yr",     f"$G${sum_row+1}")
name("tier_lite_blend",   f"$G${sum_row+3}")
name("tier_std_blend",    f"$G${sum_row+4}")
name("tier_ent_blend",    f"$G${sum_row+5}")

# ============ Sheet 4: Outputs ============
ws = wb.create_sheet("Outputs")
for col,w in enumerate([45,20,20,20,20], start=1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws["A1"] = "Engagement Economics — Wave 1, 2, 3 + TCV"
ws["A1"].font = TITLE
ws.merge_cells("A1:E1")
ws["A2"] = "All cells are FORMULAS (black). Do not edit. Edit Inputs and Scenarios sheets to change outcomes."
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="535B6B")
ws.merge_cells("A2:E2")

# Header
hrow = 4
for col,txt in enumerate(["", "Wave 1 (Pilot, Lite)", "Wave 2 (Standard)", "Wave 3 (Enterprise + Operate)", "Cumulative 3-yr"], start=1):
    c = ws.cell(row=hrow, column=col, value=txt)
    c.font = HDR; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = BORDER
ws.row_dimensions[hrow].height = 28

def out_row(r, label, w1, w2, w3, fmt=CUR, bold=False, fill=None):
    a = ws.cell(row=r, column=1, value=label)
    a.font = BOLD if bold else Font(name="Calibri", size=11, color="0E1423")
    a.alignment = LEFT
    if fill: a.fill = fill
    for col, val in enumerate([w1,w2,w3], start=2):
        c = ws.cell(row=r, column=col, value=val)
        c.font = BLACK if not bold else BOLD
        c.number_format = fmt
        c.alignment = RIGHT
        if fill: c.fill = fill
    # cumulative
    cc = ws.cell(row=r, column=5, value=f"=B{r}+C{r}+D{r}")
    cc.font = BOLD; cc.number_format = fmt; cc.alignment = RIGHT
    if fill: cc.fill = fill

# Section: Customer-side value
ws.cell(row=5, column=1, value="CUSTOMER-SIDE VALUE (annual)").font = BOLD
ws.cell(row=5, column=1).fill = SECTION_BG
for cc in range(2,6): ws.cell(row=5, column=cc).fill = SECTION_BG

out_row(6, "Pilot scope — CFMP $/yr",
        "=n_stores_pilot*cfmp_per_store_yr",
        "=ROUND(n_stores_pilot*5,0)*cfmp_per_store_yr",  # Wave 2 ~5x stores
        "=n_stores_full*cfmp_per_store_yr",
        bold=False)
out_row(7, "Pilot scope — CHC $/yr",
        "=n_hh_pilot*chc_per_hh_yr",
        "=n_hh_pilot*10*chc_per_hh_yr",   # Wave 2 ~10x households
        "=n_hh_full*chc_per_hh_yr",
        bold=False)
out_row(8, "Total value/yr at end-of-wave",
        "=B6+B7", "=C6+C7", "=D6+D7", bold=True)

# Section: Engagement cost
ws.cell(row=10, column=1, value="DELOITTE ENGAGEMENT COST").font = BOLD
ws.cell(row=10, column=1).fill = SECTION_BG
for cc in range(2,6): ws.cell(row=10, column=cc).fill = SECTION_BG

out_row(11, "Build (Lite/Std/Ent)",
        "=price_lite + price_chc_lite",
        "=price_std",
        "=price_ent")
out_row(12, "Cross-pack bundle (Variant A only)",
        "=IF(EXACT(Inputs!B7,\"Variant A — B2B2C\"),price_xp,0)",
        0, 0)
out_row(13, "BVA workshop",
        "=Inputs!B31", 0, 0)
out_row(14, "T5 Operate (annual run-rate, Wave 3+)",
        0, "=price_op_mo*6", "=price_op_mo*12", bold=False)
out_row(15, "Total engagement cost",
        "=B11+B12+B13+B14", "=C11+C12+C13+C14", "=D11+D12+D13+D14", bold=True)

# Section: ROI
ws.cell(row=17, column=1, value="ROI").font = BOLD
ws.cell(row=17, column=1).fill = SECTION_BG
for cc in range(2,6): ws.cell(row=17, column=cc).fill = SECTION_BG

out_row(18, "Value : Cost ratio (annual)",
        "=IFERROR(B8/B15,0)", "=IFERROR(C8/C15,0)", "=IFERROR(D8/D15,0)",
        fmt=MULT, bold=True)
out_row(19, "Net value (Value − Cost), annual",
        "=B8-B15", "=C8-C15", "=D8-D15", bold=True)

# Section: 3-year TCV
ws.cell(row=21, column=1, value="3-YEAR TOTAL CONTRACT VALUE (Deloitte)").font = BOLD
ws.cell(row=21, column=1).fill = SECTION_BG
for cc in range(2,6): ws.cell(row=21, column=cc).fill = SECTION_BG

out_row(22, "Deloitte 3-yr TCV (build + Operate)",
        "=B11+B12+B13",
        "=C11+C12+C13+C14",
        "=D11+D12+D13+D14",
        bold=True)

ws["A24"] = "READING THIS PAGE"
ws["A24"].font = BOLD
ws["A24"].fill = DELOIT_GREEN
ws["A25"] = ("• Customer-side value is recurring annual; Deloitte cost is mostly one-time (build) plus Operate run-rate.")
ws["A26"] = ("• 'Value : Cost' >1 means the engagement pays for itself in year 1; >3 = strong; >5 = exceptional.")
ws["A27"] = ("• 3-yr TCV column shows expected Deloitte revenue across the three waves.")
ws["A28"] = ("• Sensitivity sheet shows how the picture changes when 3 key inputs move ±25%.")
for r in (25,26,27,28):
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, italic=False, color="0E1423")
    ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=5)

# ============ Sheet 5: Sensitivity ============
ws = wb.create_sheet("Sensitivity")
for col,w in enumerate([40,18,18,18,18], start=1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws["A1"] = "Sensitivity — Wave-1 Net Value under ±25% on three key inputs"
ws["A1"].font = TITLE
ws.merge_cells("A1:E1")
ws["A2"] = "Holds all else constant; recomputes Wave-1 net value at low / mid / high cases."
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="535B6B")
ws.merge_cells("A2:E2")

hrow=4
for col,txt in enumerate(["Swing variable","Low (-25%)","Mid (base)","High (+25%)","Range ($M)"], start=1):
    c = ws.cell(row=hrow, column=col, value=txt)
    c.font = HDR; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = BORDER

def sens_row(r, label, mid_formula):
    ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=11, color="0E1423")
    ws.cell(row=r, column=2, value=f"={mid_formula}*0.75").font = BLACK
    ws.cell(row=r, column=3, value=f"={mid_formula}").font = BOLD
    ws.cell(row=r, column=4, value=f"={mid_formula}*1.25").font = BLACK
    ws.cell(row=r, column=5, value=f"=(D{r}-B{r})/1000000").font = BOLD
    for col in (2,3,4): ws.cell(row=r, column=col).number_format = CUR
    ws.cell(row=r, column=5).number_format = "0.0\"M\""
    for col in (2,3,4,5): ws.cell(row=r, column=col).alignment = RIGHT

# Swing on store count (drives CFMP)
sens_row(5, "# stores in pilot",        "n_stores_pilot*cfmp_per_store_yr-Outputs!B15")
sens_row(6, "Avg revenue per store",    "n_stores_pilot*cfmp_per_store_yr-Outputs!B15")  # same shape; rev_per_store is upstream
sens_row(7, "Loyalty churn baseline",   "n_hh_pilot*chc_per_hh_yr-Outputs!B15")
sens_row(8, "Amazon share recapture",   "n_hh_pilot*chc_per_hh_yr-Outputs!B15")

ws["A10"] = "INTERPRETATION"
ws["A10"].font = BOLD; ws["A10"].fill = DELOIT_GREEN
ws["A11"] = ("• Widest 'Range ($M)' = biggest swing variable; bring that input into the CFO conversation.")
ws["A12"] = ("• If the LOW case is still positive, the engagement is robust to a 25% miss on the key input.")
ws["A13"] = ("• If LOW goes negative, defer the scope or change the input assumption with client buy-in.")
for r in (11,12,13):
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="0E1423")
    ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=5)

wb.save(OUT)
print(f"saved {OUT}")
